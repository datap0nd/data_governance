import base64
import csv
from pathlib import Path

import pytest

from app import flow_worker


FIXTURE_DIR = Path(__file__).parent / "fixtures"


def _excel_fixture(name: str) -> bytes:
    return base64.b64decode(
        FIXTURE_DIR.joinpath(f"minimal.{name}.b64").read_text(encoding="ascii")
    )


@pytest.mark.parametrize("signature", [b"\x09\x00", b"\x09\x02", b"\x09\x04", b"\x09\x08"])
def test_standalone_legacy_biff_streams_are_detected_as_xls(tmp_path, signature):
    source = tmp_path / "legacy.xls"
    source.write_bytes(signature + b"\x00" * 64)

    assert flow_worker._detect_download_format(source) == "xls"


@pytest.mark.parametrize(
    ("suffix", "fixture", "detected", "source_encoding", "columns"),
    [
        (".xls", "xls", "xls", "xls", ["Code", "Units"]),
        (".xlt", "xls", "xlt", "xls", ["Code", "Units"]),
        # Derived from pandas' BSD-licensed test3.xlsb interoperability fixture.
        (".xlsb", "xlsb", "xlsb", "xlsb", ["Test"]),
    ],
)
def test_legacy_and_binary_excel_workbooks_are_preserved_and_normalized(
    tmp_path, suffix, fixture, detected, source_encoding, columns,
):
    source = tmp_path / f"outlook-download{suffix}"
    source.write_bytes(_excel_fixture(fixture))

    metadata = flow_worker._store_completed_download(
        source, tmp_path / f"report{suffix}", file_format="auto",
        csv_preamble="none", strict_headers=True, xlsx_header_mode="first_row",
    )

    assert metadata["detected_format"] == detected
    assert metadata["source_encoding"] == source_encoding
    assert metadata["columns"] == columns
    assert Path(metadata["original_file_path"]).suffix == suffix
    assert Path(metadata["original_file_path"]).read_bytes() == source.read_bytes()
    assert Path(metadata["file_path"]).suffix == ".csv"
    assert metadata["row_count"] == 1


@pytest.mark.parametrize("suffix", [".xlsx", ".xlsm", ".xltx", ".xltm"])
def test_modern_excel_extensions_keep_the_original_name_and_normalize(tmp_path, suffix):
    from openpyxl import Workbook

    source = tmp_path / f"outlook-download{suffix}"
    workbook = Workbook()
    workbook.active.append(["Code", "Units"])
    workbook.active.append(["A", 7])
    workbook.save(source)

    metadata = flow_worker._store_completed_download(
        source, tmp_path / f"report{suffix}", file_format="auto",
        csv_preamble="none", strict_headers=True, xlsx_header_mode="first_row",
    )

    assert metadata["detected_format"] == suffix.lstrip(".")
    assert metadata["source_encoding"] == "xlsx"
    assert metadata["columns"] == ["Code", "Units"]
    assert Path(metadata["original_file_path"]).suffix == suffix
    assert metadata["row_count"] == 1


def test_salesforce_html_xls_is_preserved_and_normalizes_a_large_table(tmp_path):
    data_rows = "".join(
        f"<tr><td>SF-{index}</td><td>{index}</td></tr>"
        for index in range(32_000)
    )
    payload = (
        "<!-- Salesforce legacy Excel export --><!DOCTYPE html><html><body>"
        "<table><tr><td>Report generated</td></tr></table>"
        "<table><tr><th>Code</th><th>Units</th></tr>"
        f"{data_rows}</table></body></html>"
    ).encode("utf-8-sig")
    source = tmp_path / "salesforce-report.xls"
    source.write_bytes(payload)
    target = tmp_path / "target"
    target.mkdir()

    metadata = flow_worker._store_completed_download(
        source, target / "salesforce-report.xls", file_format="auto",
        csv_preamble="none", strict_headers=True, xlsx_header_mode="first_row",
    )

    assert flow_worker._detect_download_format(source) == "html"
    assert metadata["detected_format"] == "xls"
    assert metadata["source_container"] == "html_xls"
    assert metadata["source_encoding"] == "utf-8-sig"
    assert metadata["columns"] == ["Code", "Units"]
    assert metadata["row_count"] == 32_000
    assert Path(metadata["original_file_path"]).read_bytes() == payload
    with Path(metadata["file_path"]).open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))
    assert rows[0] == ["Code", "Units"]
    assert rows[-1] == ["SF-31999", "31999"]


def test_utf16_xml_spreadsheet_disguised_as_xls_is_normalized(tmp_path):
    payload = """<?xml version="1.0"?>
    <Workbook xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet">
      <Worksheet><ss:Table>
        <ss:Row><ss:Cell><ss:Data>Account</ss:Data></ss:Cell><ss:Cell><ss:Data>Amount</ss:Data></ss:Cell></ss:Row>
        <ss:Row><ss:Cell><ss:Data>Acme &amp; Co</ss:Data></ss:Cell><ss:Cell><ss:Data>42</ss:Data></ss:Cell></ss:Row>
      </ss:Table></Worksheet>
    </Workbook>""".encode("utf-16")
    source = tmp_path / "salesforce-report.xls"
    source.write_bytes(payload)

    metadata = flow_worker._store_completed_download(
        source, tmp_path / "report.xls", file_format="auto",
        csv_preamble="none", strict_headers=True, xlsx_header_mode="first_row",
    )

    assert flow_worker._detect_download_format(source) == "html"
    assert metadata["source_encoding"] == "utf-16"
    with Path(metadata["file_path"]).open("r", encoding="utf-8-sig", newline="") as handle:
        assert list(csv.reader(handle)) == [["Account", "Amount"], ["Acme & Co", "42"]]


def test_html_xls_rejects_rows_wider_than_its_first_row_headers(tmp_path):
    source = tmp_path / "salesforce-report.xls"
    source.write_text(
        "<table><tr><th>Code</th><th>Units</th></tr>"
        "<tr><td>A</td><td>7</td><td>must-not-be-dropped</td></tr></table>",
        encoding="utf-8",
    )

    with pytest.raises(RuntimeError, match=r"table 1 row 2 has 3 columns.*2 first-row"):
        flow_worker._store_completed_download(
            source, tmp_path / "report.xls", file_format="auto",
            csv_preamble="none", strict_headers=True, xlsx_header_mode="first_row",
        )


def test_html_xls_without_a_data_table_is_still_rejected(tmp_path):
    source = tmp_path / "session-expired.xls"
    source.write_text(
        "<!DOCTYPE html><html><body>Please sign in again.</body></html>",
        encoding="utf-8",
    )

    with pytest.raises(RuntimeError, match="no usable first-row-header data table"):
        flow_worker._store_completed_download(
            source, tmp_path / "report.xls", file_format="auto",
            csv_preamble="none", strict_headers=True, xlsx_header_mode="first_row",
        )


def test_portal_html_named_xls_does_not_opt_into_outlook_table_extraction(tmp_path):
    source = tmp_path / "portal-download.xls"
    source.write_text(
        "<table><tr><th>Code</th></tr><tr><td>A</td></tr></table>",
        encoding="utf-8",
    )

    with pytest.raises(RuntimeError, match="HTML page"):
        flow_worker._store_completed_download(source, tmp_path / "report.xls")


def test_explicit_asap_excel_type_accepts_validated_html_based_xls(tmp_path):
    payload = (
        "<html><body><table>"
        "<tr><th>Code</th><th>Units</th></tr>"
        "<tr><td>A</td><td>7</td></tr>"
        "</table></body></html>"
    ).encode("utf-8")
    source = tmp_path / "asap-export.xls"
    source.write_bytes(payload)

    metadata = flow_worker._store_completed_download(
        source, tmp_path / "formatted.xlsx", file_format="xlsx",
        asap_download_type="excel_with_formatting",
    )

    assert metadata["detected_format"] == "xls"
    assert Path(metadata["original_file_path"]).read_bytes() == payload
    assert Path(metadata["file_path"]).read_text(encoding="utf-8-sig").splitlines() == [
        "Code,Units", "A,7",
    ]


def test_gscm_frame_workbook_normalizes_only_with_the_configured_trim(tmp_path):
    # The GSCM frame: a non-data first row, and a first column whose
    # header-row cell duplicates a real column label while its data cells are
    # row numbers. Untrimmed, no row in the sheet is a valid unique header -
    # the exact "no usable table" failure seen on live GSCM workbooks.
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "MENA_Actual_sales"
    sheet.append(["MENA Actual Sales Weekly", None, None, None])
    sheet.append(["Region", "Region", "Subsidiary", "Qty"])
    sheet.append([1, "MENA", "MENA", 5])
    sheet.append([2, "MENA", "MENA", 7])
    source = tmp_path / "gscm.xlsx"
    workbook.save(source)

    with pytest.raises(RuntimeError, match="usable table"):
        flow_worker._normalize_xlsx(
            source, tmp_path / "untrimmed.csv", requested_weeks=[],
        )

    output = tmp_path / "trimmed.csv"
    metadata = flow_worker._normalize_xlsx(
        source, output, requested_weeks=[], excel_trim="first_row_and_column",
    )

    assert metadata["excel_trim"] == "first_row_and_column"
    lines = output.read_text(encoding="utf-8-sig").splitlines()
    assert lines[0] == "Region,Subsidiary,Qty"
    assert lines[1:] == ["MENA,MENA,5", "MENA,MENA,7"]


def test_unknown_excel_trim_option_is_rejected_not_ignored(tmp_path):
    from openpyxl import Workbook

    workbook = Workbook()
    workbook.active.append(["Region", "Qty"])
    workbook.active.append(["MENA", 5])
    source = tmp_path / "plain.xlsx"
    workbook.save(source)

    with pytest.raises(RuntimeError, match="pre-processing"):
        flow_worker._normalize_xlsx(
            source, tmp_path / "out.csv", requested_weeks=[],
            excel_trim="drop_everything",
        )


def test_trim_takes_the_declared_first_row_as_header_and_numbers_repeats(tmp_path):
    # The realistic GSCM shape the heuristic can never accept: the header
    # legitimately repeats a label (one Qty per week). With the trim
    # configured, the operator has declared where the table starts, so the
    # first remaining row is the header and repeats are numbered, not fatal.
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["MENA Actual Sales Weekly", None, None, None])
    sheet.append(["No.", "Region", "Qty", "Qty"])
    sheet.append([1, "MENA", 5, 6])
    sheet.append([2, "EU", 7, 8])
    source = tmp_path / "gscm.xlsx"
    workbook.save(source)

    output = tmp_path / "trimmed.csv"
    metadata = flow_worker._normalize_xlsx(
        source, output, requested_weeks=[], excel_trim="first_row_and_column",
    )

    assert metadata["excel_trim"] == "first_row_and_column"
    assert metadata["xlsx_header_mode"] == "first_row"
    lines = output.read_text(encoding="utf-8-sig").splitlines()
    assert lines[0] == "Region,Qty,Qty_2"
    assert lines[1:] == ["MENA,5,6", "EU,7,8"]


def test_trim_skips_an_extra_sheet_with_a_blank_first_row(tmp_path):
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Title", None, None])
    sheet.append(["No.", "Region", "Qty"])
    sheet.append([1, "MENA", 5])
    notes = workbook.create_sheet("Criteria")
    notes.append([None, None])
    notes.append(["Generated by GSCM", None])
    source = tmp_path / "gscm.xlsx"
    workbook.save(source)

    output = tmp_path / "trimmed.csv"
    flow_worker._normalize_xlsx(
        source, output, requested_weeks=[], excel_trim="first_row_and_column",
    )

    lines = output.read_text(encoding="utf-8-sig").splitlines()
    assert lines == ["Region,Qty", "MENA,5"]


def test_no_usable_table_error_reports_a_row_by_row_sheet_scan(tmp_path):
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["MENA Actual Sales Weekly", None, None, None])
    sheet.append(["Region", "Region", "Subsidiary", "Qty"])
    sheet.append([1, "MENA", "MENA", 5])
    source = tmp_path / "gscm.xlsx"
    workbook.save(source)

    with pytest.raises(RuntimeError) as excinfo:
        flow_worker._normalize_xlsx(
            source, tmp_path / "out.csv", requested_weeks=[],
        )
    message = str(excinfo.value)
    assert "Sheet scan" in message
    assert "repeated labels: region" in message
    assert "1 cell(s)" in message
