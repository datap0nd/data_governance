import hashlib
import zipfile
from pathlib import Path
from types import SimpleNamespace

import pandas as pd
import pytest

from app import flow_sql, flow_worker


@pytest.mark.parametrize(
    ("identifier", "quoted"),
    [
        ("reporting", '"reporting"'),
        ("GSCM_Test", '"GSCM_Test"'),
        ("Import First and Second Activation", '"Import First and Second Activation"'),
        ('Team "Current" Imports', '"Team ""Current"" Imports"'),
        ('name"; DROP TABLE users; --', '"name""; DROP TABLE users; --"'),
    ],
)
def test_sql_identifier_quoting_supports_discovered_postgres_names(identifier, quoted):
    assert flow_sql._quote_identifier(identifier) == quoted


@pytest.mark.parametrize("identifier", ["", "invalid\x00identifier", "x" * 64, None])
def test_sql_identifier_quoting_rejects_values_postgres_cannot_identify(identifier):
    with pytest.raises(ValueError, match="Invalid SQL identifier"):
        flow_sql._quote_identifier(identifier)


def test_asap_csv_normalization_removes_title_and_blank_row(tmp_path):
    path = tmp_path / "download.csv"
    path.write_text(
        'Report title\n\n"Sell-in Region","Active"\n"MIDDLE EAST","116"\n',
        encoding="utf-8",
    )
    result = flow_worker._normalize_csv(path)
    assert result["preamble_rows_removed"] == 2
    assert result["columns"] == ["Sell-in Region", "Active"]
    assert path.read_text(encoding="utf-8-sig").splitlines() == [
        "Sell-in Region,Active", "MIDDLE EAST,116",
    ]
    assert flow_worker._csv_metadata(path)["row_count"] == 1


def test_asap_csv_normalization_detects_semicolon_delimiter(tmp_path):
    path = tmp_path / "download.csv"
    path.write_text(
        'Report title\n\n"Sell-in Region";"Active"\n"MIDDLE EAST";"116"\n',
        encoding="cp1252",
    )
    result = flow_worker._normalize_csv(path)
    assert result["source_delimiter"] == ";"
    assert result["columns"] == ["Sell-in Region", "Active"]
    assert path.read_text(encoding="utf-8-sig").splitlines() == [
        "Sell-in Region,Active", "MIDDLE EAST,116",
    ]


def test_excel_download_is_preserved_and_normalized_to_csv(tmp_path):
    from openpyxl import Workbook

    source = tmp_path / "browser-download.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Raw data"
    sheet.append(["Tutorial title"])
    sheet.append([])
    sheet.append(["Market", "Units"])
    sheet.append(["Global", 12])
    workbook.save(source)

    output = tmp_path / "bundle_global.xlsx"
    metadata = flow_worker._store_completed_download(
        source, output, file_format="xlsx",
    )

    normalized = Path(metadata["file_path"])
    assert output.is_file()
    assert metadata["original_file_path"] == str(output)
    assert metadata["row_count"] == 1
    assert normalized.read_text(encoding="utf-8-sig").splitlines() == [
        "Market,Units",
        "Global,12",
    ]


def test_raw_xlsx_passthrough_skips_unused_csv_normalization(
    tmp_path, monkeypatch,
):
    from openpyxl import Workbook

    source = tmp_path / "browser-download"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Market", "Units"])
    sheet.append(["Global", 12])
    workbook.save(source)
    output = tmp_path / "MTracker.xlsx"
    monkeypatch.setattr(
        flow_worker,
        "_normalize_xlsx",
        lambda *_args, **_kwargs: pytest.fail(
            "a raw-only flow must not scan the saved workbook into CSV"
        ),
    )
    monkeypatch.setattr(
        flow_worker,
        "_csv_metadata",
        lambda *_args, **_kwargs: pytest.fail(
            "raw XLSX metadata should be computed during its target copy"
        ),
    )

    metadata = flow_worker._store_completed_download(
        source, output, file_format="xlsx", require_normalized_csv=False,
    )

    assert output.read_bytes() == source.read_bytes()
    assert metadata["file_path"] == str(output)
    assert metadata["filename"] == output.name
    assert metadata["detected_format"] == "xlsx"
    assert metadata["row_count"] is None
    assert metadata["file_size"] == output.stat().st_size
    assert metadata["checksum"] == hashlib.sha256(output.read_bytes()).hexdigest()
    assert "normalization_error" not in metadata


class _StaleStatPath:
    """A saved file whose first ``stat`` sizes mimic a stale SMB cache."""

    def __init__(self, real: Path, stat_sizes):
        self._real = real
        self._stat_sizes = list(stat_sizes)

    def stat(self):
        if self._stat_sizes:
            return SimpleNamespace(st_size=self._stat_sizes.pop(0))
        return self._real.stat()

    def open(self, mode="rb"):
        return self._real.open(mode)

    def __str__(self):
        return str(self._real)


def test_copy_verification_survives_a_stale_smb_size_report(tmp_path):
    # Windows' SMB client can report a stale size right after the copied
    # handle closes. The bytes on the share are complete, so re-reading them
    # must count as proof instead of failing the whole flow run.
    output = tmp_path / "MTracker_subs.xlsx"
    payload = b"complete workbook bytes"
    output.write_bytes(payload)
    stale = _StaleStatPath(output, [0])

    flow_worker._verify_copied_file(
        stale, len(payload), hashlib.sha256(payload).hexdigest(),
        label="Downloaded Excel workbook",
    )

    assert not stale._stat_sizes


def test_copy_verification_still_rejects_a_truncated_target(tmp_path, monkeypatch):
    output = tmp_path / "MTracker_subs.xlsx"
    output.write_bytes(b"short")
    expected = b"the complete payload"
    monkeypatch.setattr(flow_worker.time, "sleep", lambda _seconds: None)

    with pytest.raises(
        RuntimeError, match="Downloaded Excel workbook was not copied completely",
    ):
        flow_worker._verify_copied_file(
            output, len(expected), hashlib.sha256(expected).hexdigest(),
            label="Downloaded Excel workbook",
        )


def test_raw_xlsx_passthrough_rejects_an_incomplete_zip_container(tmp_path):
    source = tmp_path / "browser-download"
    source.write_bytes(b"PK\x03\x04not-a-complete-workbook")
    output = tmp_path / "MTracker.xlsx"

    with pytest.raises(RuntimeError, match="complete XLSX ZIP container"):
        flow_worker._store_completed_download(
            source,
            output,
            file_format="xlsx",
            require_normalized_csv=False,
        )

    assert not output.exists()


def test_raw_xlsx_passthrough_rejects_a_corrupt_zip_member(tmp_path):
    source = tmp_path / "browser-download"
    with zipfile.ZipFile(source, "w", compression=zipfile.ZIP_STORED) as workbook:
        workbook.writestr("[Content_Types].xml", b"content-types")
        workbook.writestr("xl/workbook.xml", b"workbook")
        workbook.writestr("xl/worksheets/sheet1.xml", b"uncorrupted-sheet-data")
    damaged = bytearray(source.read_bytes())
    marker = damaged.index(b"uncorrupted-sheet-data")
    damaged[marker] ^= 0x01
    source.write_bytes(damaged)
    output = tmp_path / "MTracker.xlsx"

    with pytest.raises(RuntimeError, match="corrupt XLSX member"):
        flow_worker._store_completed_download(
            source, output, file_format="xlsx", require_normalized_csv=False,
        )

    assert not output.exists()


def test_raw_xlsx_exclusive_copy_preserves_an_existing_output(tmp_path):
    from openpyxl import Workbook

    source = tmp_path / "browser-download"
    workbook = Workbook()
    workbook.active.append(["Market", "Units"])
    workbook.active.append(["Global", 12])
    workbook.save(source)
    output = tmp_path / "MTracker.xlsx"
    output.write_bytes(b"user-owned-existing-workbook")

    with pytest.raises(FileExistsError):
        flow_worker._store_completed_download(
            source, output, file_format="xlsx", require_normalized_csv=False,
        )

    assert output.read_bytes() == b"user-owned-existing-workbook"


def test_completion_format_label_uses_the_detected_download_format():
    assert flow_worker._completed_export_format_label(
        [{"detected_format": "xlsx"}], "csv",
    ) == "XLSX"


def test_gscm_raw_workbook_is_kept_when_optional_normalization_fails(
    tmp_path, monkeypatch,
):
    from openpyxl import Workbook

    source = tmp_path / "browser-download.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Market", "Units"])
    sheet.append(["Global", 12])
    workbook.save(source)
    output = tmp_path / "GSCM_SIBP.xlsx"
    monkeypatch.setattr(
        flow_worker,
        "_normalize_xlsx",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            RuntimeError("Workbook contains no default style")
        ),
    )

    metadata = flow_worker._store_completed_download(
        source, output, file_format="xlsx", allow_raw_xlsx_fallback=True,
    )

    assert output.read_bytes() == source.read_bytes()
    assert metadata["file_path"] == str(output)
    assert metadata["detected_format"] == "xlsx"
    assert metadata["row_count"] is None
    assert "default style" in metadata["normalization_error"]


def test_excel_normalization_skips_multi_cell_filter_preamble(tmp_path):
    from openpyxl import Workbook

    source = tmp_path / "browser-download.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Raw data"
    sheet.append(["Regional FOTA"])
    sheet.append(["Week filter", "2025-W20"])
    sheet.append([])
    sheet.append(["Sell-out Region", "Sell-out Subsidiary", "Weekly", "202520"])
    sheet.append(["Middle East", "SEEG", "FOTA", 123])
    workbook.save(source)

    output = tmp_path / "ASAP_Fota__2025-W20.xlsx"
    metadata = flow_worker._store_completed_download(
        source, output, file_format="xlsx",
        requested_period=["2025-W20"],
    )

    normalized = Path(metadata["file_path"])
    assert metadata["preamble_rows_removed"] == 2
    assert metadata["columns"] == [
        "Sell-out Region", "Sell-out Subsidiary", "Weekly", "202520",
    ]
    assert normalized.read_text(encoding="utf-8-sig").splitlines() == [
        "Sell-out Region,Sell-out Subsidiary,Weekly,202520",
        "Middle East,SEEG,FOTA,123",
    ]


def test_excel_normalization_refuses_to_truncate_wider_data_row(tmp_path):
    from openpyxl import Workbook

    source = tmp_path / "browser-download.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Raw data"
    sheet.append(["Regional FOTA"])
    sheet.append(["Week filter", "2025-W20"])
    sheet.append([])
    sheet.append(["Sell-out Region", "Sell-out Subsidiary", "Weekly", "202520"])
    sheet.append(["Middle East", "SEEG", "FOTA", 123, "Middle East", "SEEG"])
    workbook.save(source)

    output = tmp_path / "ASAP_Fota__2025-W20.xlsx"
    with pytest.raises(RuntimeError, match="Refusing to discard data"):
        flow_worker._store_completed_download(
            source, output, file_format="xlsx",
        )


def test_excel_normalization_recovers_live_multi_week_metric_columns(tmp_path):
    from openpyxl import Workbook

    source = tmp_path / "browser-download.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Raw data"
    sheet.append(["Regional FOTA"])
    sheet.append(["Week filter", "2026-W30 to 2026-W31"])
    sheet.append([])
    sheet.append(["Sell-out Region", "Sell-out Subsidiary", "Metrics"])
    sheet.append(["Middle East", "SEEG", "Sell-out", 300, 310])
    workbook.save(source)

    output = tmp_path / "ASAP_Fota_2026-W30_2026-W31.xlsx"
    metadata = flow_worker._store_completed_download(
        source, output, file_format="xlsx",
        requested_period=["2026-W30", "2026-W31"],
    )

    normalized = Path(metadata["file_path"])
    assert metadata["recovered_week_columns"] == ["202630", "202631"]
    assert metadata["removed_metric_label"] == "sell_out"
    assert metadata["columns"] == [
        "Sell-out Region", "Sell-out Subsidiary", "202630", "202631",
    ]
    assert normalized.read_text(encoding="utf-8-sig").splitlines() == [
        "Sell-out Region,Sell-out Subsidiary,202630,202631",
        "Middle East,SEEG,300,310",
    ]


def test_excel_normalization_removes_live_weekly_sell_out_descriptor(tmp_path):
    from openpyxl import Workbook

    source = tmp_path / "browser-download.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Raw data"
    sheet.append(["Regional FOTA"])
    sheet.append(["Week filter", "2026-W30 to 2026-W31"])
    sheet.append([])
    sheet.append([
        "Sell-out Region", "Sell-out Subsidiary", "Weekly", "202630", "202631",
    ])
    sheet.append(["Middle East", "SEEG", "Sell-out", 300, 310])
    workbook.save(source)

    output = tmp_path / "ASAP_Fota_2026-W30_2026-W31.xlsx"
    metadata = flow_worker._store_completed_download(
        source, output, file_format="xlsx",
        requested_period=["2026-W30", "2026-W31"],
    )

    normalized = Path(metadata["file_path"])
    assert metadata["recovered_week_columns"] == ["202630", "202631"]
    assert metadata["removed_metric_label"] == "sell_out"
    assert metadata["columns"] == [
        "Sell-out Region", "Sell-out Subsidiary", "202630", "202631",
    ]
    assert normalized.read_text(encoding="utf-8-sig").splitlines() == [
        "Sell-out Region,Sell-out Subsidiary,202630,202631",
        "Middle East,SEEG,300,310",
    ]


def test_excel_normalization_rejects_metric_label_plus_one_value_for_two_weeks(tmp_path):
    from openpyxl import Workbook

    source = tmp_path / "browser-download.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Raw data"
    sheet.append(["Regional FOTA"])
    sheet.append(["Week filter", "2026-W30 to 2026-W31"])
    sheet.append([])
    sheet.append(["Sell-out Region", "Sell-out Subsidiary", "Metrics"])
    sheet.append(["Middle East", "SEEG", "Sell-out", 310])
    workbook.save(source)

    output = tmp_path / "ASAP_Fota_2026-W30_2026-W31.xlsx"
    with pytest.raises(RuntimeError, match="expected numeric week columns: 2.*observed.*1"):
        flow_worker._store_completed_download(
            source, output, file_format="xlsx",
            requested_period=["2026-W30", "2026-W31"],
        )


def test_excel_normalization_uses_complete_requested_period_not_filename_endpoints(tmp_path):
    from openpyxl import Workbook

    source = tmp_path / "browser-download.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Raw data"
    sheet.append(["Regional FOTA"])
    sheet.append(["Week filter", "2026-W22 to 2026-W26"])
    sheet.append([])
    sheet.append([
        "Sell-out Region", "Sell-out Subsidiary", "Weekly",
        "202622", "202623", "202624", "202625", "202626",
    ])
    sheet.append(["Middle East", "SEEG", "Sell-out", 22, 23, 24, 25, 26])
    workbook.save(source)

    output = tmp_path / "ASAP_Fota_2026-W22_2026-W26.xlsx"
    requested = [f"2026-W{week:02d}" for week in range(22, 27)]
    metadata = flow_worker._store_completed_download(
        source, output, file_format="xlsx",
        requested_period=requested,
    )

    assert metadata["recovered_week_columns"] == [
        "202622", "202623", "202624", "202625", "202626",
    ]
    assert metadata["columns"] == [
        "Sell-out Region", "Sell-out Subsidiary",
        "202622", "202623", "202624", "202625", "202626",
    ]


def test_excel_normalization_prefers_year_week_header(tmp_path):
    from openpyxl import Workbook

    source = tmp_path / "browser-download.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Raw data"
    sheet.append(["Regional FOTA"])
    sheet.append(["Week filter", "2025-W20"])
    sheet.append([])
    sheet.append(["Sell-out Region", "Sell-out Subsidiary", "Weekly", 202520])
    sheet.append(["Middle East", "SEEG", "FOTA", 123])
    workbook.save(source)

    output = tmp_path / "ASAP_Fota__2025-W20.xlsx"
    metadata = flow_worker._store_completed_download(
        source, output, file_format="xlsx",
    )

    normalized = Path(metadata["file_path"])
    assert metadata["preamble_rows_removed"] == 2
    assert metadata["columns"] == [
        "Sell-out Region", "Sell-out Subsidiary", "Weekly", "202520",
    ]
    assert normalized.read_text(encoding="utf-8-sig").splitlines() == [
        "Sell-out Region,Sell-out Subsidiary,Weekly,202520",
        "Middle East,SEEG,FOTA,123",
    ]


def test_excel_normalization_prefers_dimension_labels_when_week_heading_is_not_compact(tmp_path):
    from openpyxl import Workbook

    source = tmp_path / "browser-download.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Raw data"
    sheet.append(["Regional FOTA"])
    sheet.append(["Week filter", "2025-W20"])
    sheet.append([])
    sheet.append([
        "Sell-out Region", "Sell-out Subsidiary", "Country Code", "Week 20, 2025",
    ])
    sheet.append(["Middle East", "SEEG", "AE", 123])
    workbook.save(source)

    output = tmp_path / "ASAP_Fota__2025-W20.xlsx"
    metadata = flow_worker._store_completed_download(
        source, output, file_format="xlsx",
    )

    normalized = Path(metadata["file_path"])
    assert metadata["preamble_rows_removed"] == 2
    assert metadata["columns"] == [
        "Sell-out Region", "Sell-out Subsidiary", "Country Code", "Week 20, 2025",
    ]
    assert normalized.read_text(encoding="utf-8-sig").splitlines() == [
        'Sell-out Region,Sell-out Subsidiary,Country Code,"Week 20, 2025"',
        "Middle East,SEEG,AE,123",
    ]


def test_managed_snapshot_unions_different_bundle_columns_by_name(tmp_path, monkeypatch):
    global_path = tmp_path / "global.csv"
    country_path = tmp_path / "countries.csv"
    global_path.write_text("region,units,bundle_scope\nGlobal,10,Global\n", encoding="utf-8")
    country_path.write_text("country,vendor,bundle_scope\nAE,Vendor A,Countries\n", encoding="utf-8")
    executed = []
    copied = []
    transaction = SimpleNamespace(
        commit=lambda: executed.append("commit"), rollback=lambda: executed.append("rollback"),
    )

    class Result:
        def fetchall(self):
            return []

    class Cursor:
        def copy_expert(self, statement, stream):
            copied.append((statement, stream.read()))

        def close(self):
            return None

    class Connection:
        connection = SimpleNamespace(cursor=lambda: Cursor())

        def begin(self):
            return transaction

        def execute(self, statement, *_args, **_kwargs):
            rendered = str(statement)
            executed.append(rendered)
            return None if rendered.startswith("SET LOCAL") else Result()

        def close(self):
            return None

    monkeypatch.setattr(
        flow_sql, "_engine",
        lambda _database: SimpleNamespace(connect=lambda: Connection(), dispose=lambda: None),
    )
    result = flow_sql.load_artifacts(
        [{"file_path": str(global_path)}, {"file_path": str(country_path)}],
        {"database": "db", "schema": "reporting", "table": "asap_ti", "mode": "replace"},
    )

    create = next(item for item in executed if item.startswith('CREATE TABLE "reporting"'))
    assert all(f'"{column}" TEXT' in create for column in (
        "region", "units", "bundle_scope", "country", "vendor",
    ))
    assert 'COPY "reporting"."_metronome_stage_' in copied[0][0]
    assert '("region", "units", "bundle_scope")' in copied[0][0]
    assert '("country", "vendor", "bundle_scope")' in copied[1][0]
    assert result["files_loaded"] == 2
    assert result["rows_written"] == 2
    assert result["target_created"] is True
    assert executed[-1] == "commit"


def test_transformations_run_once_per_file_and_use_script_results(tmp_path):
    source = tmp_path / "input.csv"
    source.write_text("name,value\na,1\n", encoding="utf-8")
    script = tmp_path / "transform.py"
    script.write_text(
        "from argparse import ArgumentParser\nimport json\nimport os\nfrom pathlib import Path\n"
        "parser = ArgumentParser()\nparser.add_argument('--input', required=True)\n"
        "parser.add_argument('--output', required=True)\nargs = parser.parse_args()\n"
        "assert json.loads(os.environ['METRONOME_FLOW_PERIODS']) == ['2026-W01']\n"
        "Path(args.output).write_text(Path(args.input).read_text(), encoding='utf-8')\n",
        encoding="utf-8",
    )
    output = flow_worker._run_transformations(
        [{"file_path": str(source), "filename": source.name, "period_key": ["2026-W01"], "status": "saved"}],
        {"enabled": True, "script_path": str(script)},
    )
    assert len(output) == 1
    assert Path(output[0]["file_path"]).parent.name == "script_results"
    assert output[0]["status"] == "transformed"
    assert output[0]["row_count"] == 1


def test_transformation_requires_reserved_output_file(tmp_path):
    source = tmp_path / "input.csv"
    source.write_text("name,value\na,1\n", encoding="utf-8")
    script = tmp_path / "transform.py"
    script.write_text("print('done')\n", encoding="utf-8")
    with pytest.raises(RuntimeError, match="did not create"):
        flow_worker._run_transformations(
            [{"file_path": str(source), "filename": source.name, "status": "saved"}],
            {"enabled": True, "script_path": str(script)},
        )


def test_powershell_transformation_uses_named_path_parameters(tmp_path):
    command = flow_worker._script_command(
        tmp_path / "transform.ps1", tmp_path / "input.csv", tmp_path / "output.csv",
    )
    assert command[-4:] == [
        "-InputPath", str(tmp_path / "input.csv"),
        "-OutputPath", str(tmp_path / "output.csv"),
    ]


def test_csv_reader_rejects_duplicate_normalized_columns(tmp_path):
    path = tmp_path / "duplicate.csv"
    path.write_text("Sell-out Week,Sell out Week\n202630,202631\n", encoding="utf-8")
    with pytest.raises(RuntimeError, match="duplicate column"):
        flow_sql._read_artifact(path)


def test_csv_reader_uses_normalized_comma_format(tmp_path, monkeypatch):
    path = tmp_path / "normalized.csv"
    path.write_text(
        "Sell-out Week,Sell-out Qty,Active\n202630,123,116\n",
        encoding="utf-8-sig",
    )
    artifact = flow_sql._read_artifact(path)

    assert artifact["columns"] == ["sell_out_week", "sell_out_qty", "active"]
    assert artifact["row_count"] == 1
    assert artifact["path"] == path


def test_sql_preflight_reports_missing_and_unexpected_columns(tmp_path, monkeypatch):
    path = tmp_path / "wrong.csv"
    pd.DataFrame({"a": [1], "extra": [2]}).to_csv(path, index=False)

    class Result:
        def fetchall(self):
            return [("a", "NO", None, "NO", "NEVER"), ("required_b", "NO", None, "NO", "NEVER")]

    class Connection:
        def __init__(self):
            self.transaction = SimpleNamespace(rollback=lambda: None)

        def begin(self):
            return self.transaction

        def execute(self, statement, *_args, **_kwargs):
            if str(statement).startswith("SET LOCAL"):
                return None
            return Result()

        def close(self):
            return None

    engine = SimpleNamespace(connect=lambda: Connection(), dispose=lambda: None)
    monkeypatch.setattr(flow_sql, "_engine", lambda _database: engine)
    with pytest.raises(RuntimeError, match="missing: required_b; unexpected: extra"):
        flow_sql.load_artifacts(
            [{"file_path": str(path)}],
            {"database": "db", "schema": "reporting", "table": "target", "mode": "append"},
        )


def test_sql_copy_streams_csv_through_postgres_copy(tmp_path):
    copied = {}
    path = tmp_path / "normalized.csv"
    path.write_text("Sell-out Week,Active\n202627,116\n", encoding="utf-8-sig")

    class Cursor:
        def copy_expert(self, statement, stream):
            copied["statement"] = statement
            copied["content"] = stream.read()

        def close(self):
            copied["closed"] = True

    raw = SimpleNamespace(cursor=lambda: Cursor())
    connection = SimpleNamespace(connection=raw)
    artifact = flow_sql._read_artifact(path)

    flow_sql._copy_artifact(connection, artifact, '"bi_reporting"."this_is_test"')

    assert copied["statement"] == (
        'COPY "bi_reporting"."this_is_test" ("sell_out_week", "active") '
        "FROM STDIN WITH (FORMAT CSV, HEADER TRUE, ENCODING 'UTF8')"
    )
    assert copied["content"] == "Sell-out Week,Active\n202627,116\n"
    assert copied["closed"] is True


def test_sql_append_maps_normalized_csv_headers_to_exact_target_columns(tmp_path, monkeypatch):
    path = tmp_path / "normalized.csv"
    path.write_text("Active,Biz Sub\n116,Mobile\n", encoding="utf-8")
    executed = []
    copied = []
    transaction = SimpleNamespace(
        commit=lambda: executed.append("commit"),
        rollback=lambda: executed.append("rollback"),
    )

    class Result:
        def fetchall(self):
            return [
                ("Active", "NO", None, "NO", "NEVER"),
                ("Biz Sub", "NO", None, "NO", "NEVER"),
                ("Optional Legacy", "YES", None, "NO", "NEVER"),
            ]

    class Cursor:
        def copy_expert(self, statement, stream):
            copied.append((statement, stream.read()))

        def close(self):
            return None

    class Connection:
        connection = SimpleNamespace(cursor=lambda: Cursor())

        def begin(self):
            return transaction

        def execute(self, statement, *_args, **_kwargs):
            executed.append(str(statement))
            return None if str(statement).startswith("SET LOCAL") else Result()

        def close(self):
            return None

    monkeypatch.setattr(
        flow_sql, "_engine",
        lambda _database: SimpleNamespace(connect=lambda: Connection(), dispose=lambda: None),
    )

    result = flow_sql.load_artifacts(
        [{"file_path": str(path)}],
        {
            "database": "db", "schema": "Reporting Area",
            "table": "Import First and Second Activation", "mode": "append",
        },
    )

    assert result["rows_written"] == 1
    assert copied == [(
        'COPY "Reporting Area"."Import First and Second Activation" '
        '("Active", "Biz Sub") FROM STDIN '
        "WITH (FORMAT CSV, HEADER TRUE, ENCODING 'UTF8')",
        "Active,Biz Sub\n116,Mobile\n",
    )]
    assert executed[-1] == "commit"
    assert "rollback" not in executed


def test_sql_append_rejects_ambiguous_target_columns_after_normalization():
    with pytest.raises(RuntimeError, match="ambiguous column name.*sell_out_week"):
        flow_sql._target_columns_by_normalized_name(["Sell-out Week", "sell_out_week"])


def test_sql_append_maps_the_live_thirteen_column_shape_to_display_headers():
    csv_columns = [
        "active", "biz_sub", "item", "mkt_name", "sell_in_account",
        "sell_in_customer", "sell_in_region", "sell_in_subsidiary",
        "sell_out_country", "sell_out_month", "sell_out_region",
        "sell_out_subsidiary", "series",
    ]
    target_columns = [
        "Active", "Biz Sub", "Item", "MKT Name", "Sell-in Account",
        "Sell-in Customer", "Sell-in Region", "Sell-in Subsidiary",
        "Sell-out Country", "Sell-out Month", "Sell-out Region",
        "Sell-out Subsidiary", "Series",
    ]

    mapping = flow_sql._target_columns_by_normalized_name(target_columns)

    assert [mapping[column] for column in csv_columns] == target_columns


def test_sql_managed_snapshot_preserves_existing_table_and_commits(tmp_path, monkeypatch):
    path = tmp_path / "normalized.csv"
    path.write_text("Sell-out Week,Active\n202627,116\n", encoding="utf-8-sig")
    executed = []
    copied = []
    transaction = SimpleNamespace(
        commit=lambda: executed.append("commit"),
        rollback=lambda: executed.append("rollback"),
    )

    class Result:
        def fetchall(self):
            return [
                ("sell_out_week", "NO", None, "NO", "NEVER"),
                ("active", "NO", None, "NO", "NEVER"),
            ]

    class Cursor:
        def copy_expert(self, statement, stream):
            copied.append((statement, stream.read()))

        def close(self):
            return None

    class Connection:
        connection = SimpleNamespace(cursor=lambda: Cursor())

        def begin(self):
            return transaction

        def execute(self, statement, *_args, **_kwargs):
            executed.append(str(statement))
            return Result()

        def close(self):
            executed.append("close")

    engine = SimpleNamespace(connect=lambda: Connection(), dispose=lambda: executed.append("dispose"))
    monkeypatch.setattr(flow_sql, "_engine", lambda _database: engine)
    events = []

    result = flow_sql.load_artifacts(
        [{"file_path": str(path), "row_count": 1}],
        {
            "database": "db", "schema": "Reporting Area",
            "table": "Import First and Second Activation", "mode": "replace",
        },
        progress=events.append,
    )

    assert result["rows_written"] == 1
    assert result["target"] == "db.Reporting Area.Import First and Second Activation"
    assert executed.count("commit") == 1
    assert "rollback" not in executed
    # The load is bounded only so a wedged session cannot hold a transaction
    # open forever. A slow database is allowed to be slow.
    assert f"SET LOCAL lock_timeout = '{flow_sql.SQL_LOCK_TIMEOUT_SECONDS}s'" in executed
    assert f"SET LOCAL statement_timeout = '{flow_sql.SQL_STATEMENT_TIMEOUT_SECONDS}s'" in executed
    assert flow_sql.SQL_STATEMENT_TIMEOUT_SECONDS >= 30 * 60
    staging_create = next(
        item for item in executed
        if item.startswith('CREATE TABLE "Reporting Area"."_metronome_stage_')
    )
    staging_qualified = staging_create.removeprefix("CREATE TABLE ").split(" (", 1)[0]
    assert not any(
        item.startswith('DROP TABLE "Reporting Area"."Import First and Second Activation"')
        for item in executed
    )
    assert 'TRUNCATE TABLE "Reporting Area"."Import First and Second Activation"' in executed
    assert any(
        item.startswith(
            'INSERT INTO "Reporting Area"."Import First and Second Activation" '
            '("sell_out_week", "active") SELECT "sell_out_week", "active" FROM '
        )
        for item in executed
    )
    assert f"DROP TABLE {staging_qualified}" in executed
    assert copied[0][0].startswith(f"COPY {staging_qualified} ")
    assert copied[0][1] == "Sell-out Week,Active\n202627,116\n"
    assert [event["stage"] for event in events] == [
        "sql_artifact_validation", "sql_artifact_validation", "sql_connecting", "sql_connected",
        "sql_target_validation", "sql_target_validation", "sql_staging", "sql_staging",
        "sql_copy", "sql_copy", "sql_replace", "sql_replace", "sql_commit", "sql_commit",
    ]
    assert result["schema_replaced"] is False
    assert result["snapshot_refreshed"] is True
    assert result["target_created"] is False
    assert result["columns_added"] == []
    assert result["column_type"] == "TEXT"


def test_sql_managed_snapshot_adds_new_columns_without_dropping_existing_table(tmp_path, monkeypatch):
    path = tmp_path / "replacement.csv"
    path.write_text("New Column,Another\nvalue,2\n", encoding="utf-8")
    executed = []
    transaction = SimpleNamespace(commit=lambda: executed.append("commit"), rollback=lambda: None)

    class Result:
        def fetchall(self):
            return [("old_column", "YES", None, "NO", "NEVER")]

    class Cursor:
        def copy_expert(self, statement, stream):
            executed.append(statement)
            stream.read()

        def close(self):
            return None

    class Connection:
        connection = SimpleNamespace(cursor=lambda: Cursor())

        def begin(self):
            return transaction

        def execute(self, statement, *_args, **_kwargs):
            executed.append(str(statement))
            return None if str(statement).startswith("SET LOCAL") else Result()

        def close(self):
            return None

    monkeypatch.setattr(
        flow_sql, "_engine",
        lambda _database: SimpleNamespace(connect=lambda: Connection(), dispose=lambda: None),
    )

    result = flow_sql.load_artifacts(
        [{"file_path": str(path)}],
        {"database": "db", "schema": "reporting", "table": "target", "mode": "replace"},
    )

    assert result["rows_written"] == 1
    assert 'ALTER TABLE "reporting"."target" ADD COLUMN "new_column" TEXT' in executed
    assert 'ALTER TABLE "reporting"."target" ADD COLUMN "another" TEXT' in executed
    assert 'TRUNCATE TABLE "reporting"."target"' in executed
    assert not any(item.startswith('DROP TABLE "reporting"."target"') for item in executed)
    assert any(
        item.startswith(
            'INSERT INTO "reporting"."target" ("new_column", "another") '
            'SELECT "new_column", "another" FROM "reporting"."_metronome_stage_'
        )
        for item in executed
    )
    assert result["columns_added"] == ["new_column", "another"]
    assert result["target_created"] is False
    assert executed[-1] == "commit"


def test_sql_managed_snapshot_copy_failure_leaves_existing_target_untouched(tmp_path, monkeypatch):
    path = tmp_path / "replacement.csv"
    path.write_text("new_column\nvalue\n", encoding="utf-8")
    executed = []
    transaction = SimpleNamespace(
        commit=lambda: executed.append("commit"),
        rollback=lambda: executed.append("rollback"),
    )

    class Result:
        def fetchall(self):
            return [("old_column", "YES", None, "NO", "NEVER")]

    class Cursor:
        def copy_expert(self, _statement, _stream):
            raise RuntimeError("simulated COPY failure")

        def close(self):
            return None

    class Connection:
        connection = SimpleNamespace(cursor=lambda: Cursor())

        def begin(self):
            return transaction

        def execute(self, statement, *_args, **_kwargs):
            executed.append(str(statement))
            return None if str(statement).startswith("SET LOCAL") else Result()

        def close(self):
            return None

    monkeypatch.setattr(
        flow_sql, "_engine",
        lambda _database: SimpleNamespace(connect=lambda: Connection(), dispose=lambda: None),
    )
    events = []

    with pytest.raises(flow_sql.SqlHandoffError, match="PostgreSQL confirmed rollback"):
        flow_sql.load_artifacts(
            [{"file_path": str(path)}],
            {"database": "db", "schema": "reporting", "table": "target", "mode": "replace"},
            progress=events.append,
        )

    assert any(item.startswith('CREATE TABLE "reporting"."_metronome_stage_') for item in executed)
    assert not any(item.startswith('TRUNCATE TABLE "reporting"."target"') for item in executed)
    assert not any(item.startswith('ALTER TABLE "reporting"."target"') for item in executed)
    assert not any(item.startswith('DROP TABLE "reporting"."target"') for item in executed)
    assert "rollback" in executed
    assert "commit" not in executed
    assert events[-1]["stage"] == "sql_failed"
    assert events[-1]["outcome"] == "PostgreSQL confirmed rollback. No SQL changes were committed."


def test_sql_managed_snapshot_creates_missing_target_from_staging(tmp_path, monkeypatch):
    path = tmp_path / "first_snapshot.csv"
    path.write_text("New Column,Another\nvalue,2\n", encoding="utf-8")
    executed = []
    copied = []
    transaction = SimpleNamespace(
        commit=lambda: executed.append("commit"),
        rollback=lambda: executed.append("rollback"),
    )

    class Result:
        def fetchall(self):
            return []

    class Cursor:
        def copy_expert(self, statement, stream):
            copied.append((statement, stream.read()))

        def close(self):
            return None

    class Connection:
        connection = SimpleNamespace(cursor=lambda: Cursor())

        def begin(self):
            return transaction

        def execute(self, statement, *_args, **_kwargs):
            executed.append(str(statement))
            return None if str(statement).startswith("SET LOCAL") else Result()

        def close(self):
            return None

    monkeypatch.setattr(
        flow_sql, "_engine",
        lambda _database: SimpleNamespace(connect=lambda: Connection(), dispose=lambda: None),
    )

    result = flow_sql.load_artifacts(
        [{"file_path": str(path)}],
        {"database": "db", "schema": "reporting", "table": "new target", "mode": "replace"},
    )

    staging_create = next(
        item for item in executed
        if item.startswith('CREATE TABLE "reporting"."_metronome_stage_')
    )
    staging_qualified = staging_create.removeprefix("CREATE TABLE ").split(" (", 1)[0]
    assert copied[0][0].startswith(f"COPY {staging_qualified} ")
    assert f'ALTER TABLE {staging_qualified} RENAME TO "new target"' in executed
    assert not any(item.startswith('TRUNCATE TABLE "reporting"."new target"') for item in executed)
    assert result["target_created"] is True
    assert result["snapshot_refreshed"] is True
    assert result["columns_added"] == ["new_column", "another"]
    assert executed[-1] == "commit"


def test_sql_managed_snapshot_final_insert_failure_rolls_back_target_refresh(tmp_path, monkeypatch):
    path = tmp_path / "snapshot.csv"
    path.write_text("value\nnew\n", encoding="utf-8")
    executed = []
    transaction = SimpleNamespace(
        commit=lambda: executed.append("commit"),
        rollback=lambda: executed.append("rollback"),
    )

    class Result:
        def fetchall(self):
            return [("value", "YES", None, "NO", "NEVER")]

    class Cursor:
        def copy_expert(self, _statement, stream):
            stream.read()

        def close(self):
            return None

    class Connection:
        connection = SimpleNamespace(cursor=lambda: Cursor())

        def begin(self):
            return transaction

        def execute(self, statement, *_args, **_kwargs):
            rendered = str(statement)
            executed.append(rendered)
            if rendered.startswith('INSERT INTO "reporting"."target"'):
                raise RuntimeError("simulated final promotion failure")
            return None if rendered.startswith("SET LOCAL") else Result()

        def close(self):
            return None

    monkeypatch.setattr(
        flow_sql, "_engine",
        lambda _database: SimpleNamespace(connect=lambda: Connection(), dispose=lambda: None),
    )

    with pytest.raises(flow_sql.SqlHandoffError, match="PostgreSQL confirmed rollback"):
        flow_sql.load_artifacts(
            [{"file_path": str(path)}],
            {"database": "db", "schema": "reporting", "table": "target", "mode": "replace"},
        )

    assert 'TRUNCATE TABLE "reporting"."target"' in executed
    assert "rollback" in executed
    assert "commit" not in executed


def test_sql_copy_error_is_clean_and_rollback_is_logged(tmp_path, monkeypatch):
    path = tmp_path / "normalized.csv"
    path.write_text("value\nnot-an-integer\n", encoding="utf-8")
    rolled_back = []

    class DatabaseFailure(Exception):
        pgcode = "22P02"
        diag = SimpleNamespace(message_primary="invalid input syntax for type integer")

    class Result:
        def fetchall(self):
            return [("value", "NO", None, "NO", "NEVER")]

    class Cursor:
        def copy_expert(self, _statement, _stream):
            raise DatabaseFailure("noisy driver detail")

        def close(self):
            return None

    transaction = SimpleNamespace(commit=lambda: None, rollback=lambda: rolled_back.append(True))

    class Connection:
        connection = SimpleNamespace(cursor=lambda: Cursor())

        def begin(self):
            return transaction

        def execute(self, statement, *_args, **_kwargs):
            return None if str(statement).startswith("SET LOCAL") else Result()

        def close(self):
            return None

    monkeypatch.setattr(
        flow_sql, "_engine",
        lambda _database: SimpleNamespace(connect=lambda: Connection(), dispose=lambda: None),
    )
    events = []

    with pytest.raises(flow_sql.SqlHandoffError, match="SQLSTATE 22P02") as error:
        flow_sql.load_artifacts(
            [{"file_path": str(path)}],
            {"database": "db", "schema": "reporting", "table": "target", "mode": "append"},
            progress=events.append,
        )

    assert "invalid input syntax for type integer" in str(error.value)
    assert "PostgreSQL confirmed rollback" in str(error.value)
    assert rolled_back == [True]
    assert events[-1]["stage"] == "sql_failed"
    assert events[-1]["sql_stage"] == "copy"


def test_excel_normalization_merges_populated_sheets_in_one_streamed_pass(tmp_path):
    # Rows go straight from the workbook to the CSV writer, so a workbook whose
    # rows are split across sheets - which is how an export larger than Excel's
    # 1,048,576-row sheet limit has to arrive - still lands as one table.
    from openpyxl import Workbook

    source = tmp_path / "browser-download.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "Part 1"
    first.append(["Market", "Units"])
    first.append(["Global", 1])
    second = workbook.create_sheet("Part 2")
    second.append(["Market", "Units"])
    second.append(["Korea", 2])
    workbook.save(source)

    output = tmp_path / "bundle.xlsx"
    metadata = flow_worker._store_completed_download(source, output, file_format="xlsx")

    assert metadata["source_sheets"] == ["Part 1", "Part 2"]
    assert metadata["row_count"] == 2
    assert Path(metadata["file_path"]).read_text(encoding="utf-8-sig").splitlines() == [
        "Market,Units", "Global,1", "Korea,2",
    ]


def test_failed_excel_normalization_never_leaves_a_truncated_csv(tmp_path):
    # Streaming means rows reach disk before the workbook is fully validated.
    # The complete file must only ever appear at the target name.
    from openpyxl import Workbook

    source = tmp_path / "browser-download.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Raw data"
    sheet.append(["Regional FOTA"])
    sheet.append(["Week filter", "2025-W20"])
    sheet.append([])
    sheet.append(["Sell-out Region", "Sell-out Subsidiary", "Weekly", "202520"])
    sheet.append(["Middle East", "SEEG", "FOTA", 123])
    sheet.append(["Middle East", "SEEG", "FOTA", 123, "Middle East", "SEEG"])
    workbook.save(source)

    normalized = tmp_path / "bundle_normalized.csv"
    with pytest.raises(RuntimeError, match="Refusing to discard data"):
        flow_worker._normalize_xlsx(source, normalized, requested_weeks=[])

    assert not normalized.exists()
    # The worker never deletes, so the partial conversion stays as evidence.
    assert (tmp_path / ".bundle_normalized.csv.partial").is_file()


# --- The portal decides the format, not the flow's setting ---


def test_a_workbook_from_a_csv_configured_flow_is_read_as_a_workbook(tmp_path):
    # A dashboard link named "(xlsx)" on a flow set to CSV produced a run that
    # failed deep inside PostgreSQL with a mojibake column name: latin-1
    # accepts every byte, so the zip decoded into garbage and was written back
    # out as a "CSV".
    from openpyxl import Workbook

    source = tmp_path / "browser-download.bin"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Subsidiary", "Model", "Units"])
    sheet.append(["SEEG", "S24", 120])
    workbook.save(source)

    metadata = flow_worker._store_completed_download(
        source, tmp_path / "MTracker_subs.csv", file_format="csv",
    )

    assert metadata["detected_format"] == "xlsx"
    assert metadata["columns"] == ["Subsidiary", "Model", "Units"]
    # The workbook keeps a workbook's name; SQL reads the normalized CSV.
    assert Path(metadata["original_file_path"]).name == "MTracker_subs.xlsx"
    assert Path(metadata["file_path"]).name.endswith("_normalized.csv")


def test_an_html_page_is_refused_instead_of_being_loaded_as_data(tmp_path):
    source = tmp_path / "download.csv"
    source.write_bytes(b"<!DOCTYPE html><html><body>Session expired</body></html>")
    with pytest.raises(RuntimeError) as excinfo:
        flow_worker._store_completed_download(source, tmp_path / "out.csv")
    assert "HTML page" in str(excinfo.value)


def test_a_legacy_xls_is_refused_with_its_real_type(tmp_path):
    source = tmp_path / "download.csv"
    source.write_bytes(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 64)
    with pytest.raises(RuntimeError) as excinfo:
        flow_worker._store_completed_download(source, tmp_path / "out.csv")
    assert "looks like xls" in str(excinfo.value)


def test_a_real_csv_is_still_normalized_untouched(tmp_path):
    source = tmp_path / "download.csv"
    source.write_text("Subsidiary,Units\nSEEG,120\n", encoding="utf-8")
    metadata = flow_worker._store_completed_download(source, tmp_path / "out.csv")
    assert metadata["detected_format"] == "csv"
    assert metadata["columns"] == ["Subsidiary", "Units"]


def test_a_utf16_csv_is_text_not_binary(tmp_path):
    # UTF-16 exports are full of NUL bytes; the binary guard must not reject
    # them, because ASAP genuinely emits them.
    source = tmp_path / "download.csv"
    source.write_bytes("Region,Units\nMENA,7\n".encode("utf-16"))
    metadata = flow_worker._store_completed_download(source, tmp_path / "out.csv")
    assert metadata["detected_format"] == "csv"
    assert metadata["source_encoding"] == "utf-16"
    assert metadata["columns"] == ["Region", "Units"]


def test_the_csv_normalizer_refuses_binary_outright(tmp_path):
    source = tmp_path / "download.csv"
    source.write_bytes(b"PK\x03\x04" + b"\x00" * 64)
    with pytest.raises(RuntimeError) as excinfo:
        flow_worker._normalize_csv(source)
    assert "Refusing to read" in str(excinfo.value)
