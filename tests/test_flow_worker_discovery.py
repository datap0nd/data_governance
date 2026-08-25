import json
import re
from pathlib import Path

import pytest

from app import flow_worker
from app.flow_worker import (
    _asap_frame,
    _asap_goto,
    _asap_last_visible,
    _asap_wait_for_visible,
    _asap_wait_for_report_navigation,
    _menu_report_paths,
    _navigation_roots,
    _wait_for_navigation_roots,
)


def _record(text, x, y, *, href="", onclick="", tag="a", role="", disabled=False):
    return {
        "link": object(),
        "text": text,
        "href": href,
        "onclick": onclick,
        "tag": tag,
        "role": role,
        "disabled": disabled,
        "box": {"x": x, "y": y, "width": 70, "height": 24},
    }


def test_navigation_roots_are_inferred_from_top_link_row():
    records = [
        _record("ASAP", 10, 20),
        _record("Market", 160, 90),
        _record("Mobile", 235, 90),
        _record("Eco", 300, 90),
        _record("A report", 50, 250, href="/report/1"),
    ]

    assert [item["text"] for item in _navigation_roots(records)] == ["Market", "Mobile", "Eco"]


def test_fota_discovery_includes_visual_category_member_list():
    source = Path(flow_worker.__file__).read_text()

    assert 'add_definition("Category", "multi_select", category_options)' in source
    assert '{"weekly", "daily"}' in source


def test_flagship_discovery_includes_visual_measure_member_list():
    source = Path(flow_worker.__file__).read_text()

    assert 'add_definition("Measure", "multi_select", nearest_list_values(measure_label))' in source
    assert '{"dimension", "category", "measure"}' in source


def test_revealed_menu_columns_become_report_paths_without_target_hardcoding():
    root = _record("Mobile", 235, 90)
    before = [root, _record("Market", 160, 90)]
    after = [
        *before,
        _record("Installed Base", 45, 145),
        _record("Installed Base", 45, 180, href="javascript:open('base')"),
        _record("Installed Base (MENA)", 45, 215, href="javascript:open('mena')"),
        _record("Regional FOTA", 150, 145),
        _record("Regional FOTA", 150, 180, href="javascript:open('fota')"),
    ]

    assert _menu_report_paths(root, before, after) == [
        ["Mobile", "Installed Base", "Installed Base"],
        ["Mobile", "Installed Base", "Installed Base (MENA)"],
        ["Mobile", "Regional FOTA", "Regional FOTA"],
    ]


def test_navigation_waits_for_client_rendered_controls(monkeypatch):
    snapshots = [[], [_record("Market", 160, 90), _record("Mobile", 235, 90)]]
    monkeypatch.setattr("app.flow_worker._visible_anchor_records", lambda _page: snapshots.pop(0))

    class Page:
        def wait_for_timeout(self, _milliseconds):
            pass

    assert [item["text"] for item in _wait_for_navigation_roots(Page(), 1_000)] == ["Market", "Mobile"]


def test_menu_discovery_waits_for_loading_overlay_before_click(monkeypatch):
    events = []

    class Link:
        def click(self, **_kwargs):
            events.append("click")

    root = _record("Mobile", 235, 90)
    root["link"] = Link()
    monkeypatch.setattr(flow_worker, "_wait_for_navigation_roots", lambda _page: [root])
    monkeypatch.setattr(flow_worker, "_visible_anchor_records", lambda _page: [root])
    monkeypatch.setattr(flow_worker, "_navigation_roots", lambda _records: [root])
    monkeypatch.setattr(
        flow_worker,
        "_asap_wait_for_loading_clear",
        lambda _page: events.append("wait"),
    )
    monkeypatch.setattr(
        flow_worker,
        "_menu_report_paths",
        lambda *_args: [["Mobile", "Regional FOTA", "Regional FOTA"]],
    )

    class Page:
        def wait_for_timeout(self, _milliseconds):
            pass

    assert flow_worker._asap_discover_menu_reports(Page(), []) == [
        ["Mobile", "Regional FOTA", "Regional FOTA"],
    ]
    assert events[0:2] == ["wait", "click"]


def test_menu_discovery_fails_closed_when_one_root_reveals_no_reports(monkeypatch):
    events = []

    class Link:
        def click(self, **_kwargs):
            events.append("click")

        def hover(self, **_kwargs):
            events.append("hover")

    root = _record("Retail", 600, 90)
    root["link"] = Link()
    monkeypatch.setattr(flow_worker, "_wait_for_navigation_roots", lambda _page: [root])
    monkeypatch.setattr(flow_worker, "_visible_anchor_records", lambda _page: [root])
    monkeypatch.setattr(flow_worker, "_navigation_roots", lambda _records: [root])
    monkeypatch.setattr(flow_worker, "_asap_wait_for_loading_clear", lambda _page: None)
    moments = iter([0, 20, 30, 50])
    monkeypatch.setattr(flow_worker.time, "monotonic", lambda: next(moments))

    class Page:
        def wait_for_timeout(self, _milliseconds):
            pass

    with pytest.raises(RuntimeError, match="Retail.*no report controls revealed"):
        flow_worker._asap_discover_menu_reports(Page(), [])
    assert "hover" in events


def test_blank_menu_branch_is_recovered_from_the_live_portal_menu_tree(monkeypatch):
    class Link:
        def click(self, **_kwargs):
            pass

        def hover(self, **_kwargs):
            pass

    root = _record("Retail", 600, 90)
    root["link"] = Link()
    monkeypatch.setattr(flow_worker, "_wait_for_navigation_roots", lambda _page: [root])
    monkeypatch.setattr(flow_worker, "_visible_anchor_records", lambda _page: [root])
    monkeypatch.setattr(flow_worker, "_navigation_roots", lambda _records: [root])
    monkeypatch.setattr(flow_worker, "_asap_wait_for_loading_clear", lambda _page: None)
    monkeypatch.setattr(
        flow_worker, "_asap_portal_menu_paths",
        lambda _page, _diagnostics=None: [["Retail", "F8 Experience (Launching Daily)", "Flagship Experience"]],
    )
    moments = iter([0, 20, 30, 50])
    monkeypatch.setattr(flow_worker.time, "monotonic", lambda: next(moments))

    class Page:
        def wait_for_timeout(self, _milliseconds):
            pass

    assert flow_worker._asap_discover_menu_reports(Page(), []) == [
        ["Retail", "F8 Experience (Launching Daily)", "Flagship Experience"],
    ]


def test_portal_menu_tree_is_captured_before_any_header_click(monkeypatch):
    events = []

    class Link:
        def click(self, **_kwargs):
            events.append("click")

        def hover(self, **_kwargs):
            events.append("hover")

    root = _record("Retail", 600, 90)
    root["link"] = Link()
    monkeypatch.setattr(flow_worker, "_wait_for_navigation_roots", lambda _page: [root])
    monkeypatch.setattr(flow_worker, "_visible_anchor_records", lambda _page: [root])
    monkeypatch.setattr(flow_worker, "_navigation_roots", lambda _records: [root])
    monkeypatch.setattr(flow_worker, "_asap_wait_for_loading_clear", lambda _page: None)

    def portal_paths(_page, _diagnostics=None):
        assert events == []
        events.append("portal-tree")
        return [["Retail", "F8 Experience (Launching Daily)", "Flagship Experience"]]

    monkeypatch.setattr(flow_worker, "_asap_portal_menu_paths", portal_paths)
    moments = iter([0, 20, 30, 50])
    monkeypatch.setattr(flow_worker.time, "monotonic", lambda: next(moments))

    class Page:
        def wait_for_timeout(self, _milliseconds):
            pass

    assert flow_worker._asap_discover_menu_reports(Page(), ["*"]) == [
        ["Retail", "F8 Experience (Launching Daily)", "Flagship Experience"],
    ]
    assert events[0] == "portal-tree"


def test_portal_menu_tree_uses_role_specific_root_and_strips_ordering_prefixes():
    common = {
        "MSTR_MAIN_MENU_ID": "role-root",
        "MSTR_CUSTOM_WEB_CONTEXT_PATH": "/mstr",
    }
    session = {"MSTRWEB_AUTH_TOKEN_ENC": "live-token"}
    responses = {
        "role-root": {
            "children": [{"id": "retail", "name": "06.Retail"}],
        },
        "retail": {
            "id": "retail", "name": "06.Retail", "children": [{
                "id": "f8", "name": "01.F8 Experience (Launching Daily)",
                "child": [{"id": "flagship", "name": "01.Flagship Experience"}],
            }],
        },
    }
    calls = []

    class Response:
        ok = True

        def __init__(self, value):
            self.value = value

        def json(self):
            return self.value

    class Request:
        def post(self, url, **kwargs):
            folder_id = url.rsplit("=", 1)[-1]
            calls.append((folder_id, kwargs["form"]["authToken"], kwargs["form"]["depth"]))
            return Response(responses[folder_id])

    class Frame:
        def evaluate(self, _script):
            return {"common": common, "session": session}

    class Page:
        url = "https://asap.example/portal/main"
        frames = [Frame()]
        request = Request()

    assert flow_worker._asap_portal_menu_paths(Page()) == [
        ["Retail", "F8 Experience (Launching Daily)", "Flagship Experience"],
    ]
    assert calls == [
        ("role-root", "live-token", "1"),
        ("retail", "live-token", "2"),
    ]


def test_portal_session_is_read_from_frame_local_script_constants():
    class Frame:
        def evaluate(self, _script):
            return None

        def content(self):
            return """
                <script>
                const _COMMON_INFO = {"MSTR_MAIN_MENU_ID":"role-root","MSTR_CUSTOM_WEB_CONTEXT_PATH":"/mstr"};
                const _SESSION = {"MSTRWEB_AUTH_TOKEN_ENC":"live-token"};
                </script>
            """

    class Page:
        url = "https://asap.example/portal/main"
        frames = [Frame()]

    assert flow_worker._asap_portal_session(Page()) == {
        "web_base": "https://asap.example/mstr",
        "legacy_token": "live-token",
        "main_menu_id": "role-root",
    }


def test_portal_menu_tree_preserves_top_name_when_subtree_is_a_wrapper(monkeypatch):
    monkeypatch.setattr(
        flow_worker, "_asap_portal_session",
        lambda _page, _diagnostics=None: {
            "web_base": "https://asap.example/mstr",
            "legacy_token": "live-token",
            "main_menu_id": "role-root",
        },
    )
    responses = {
        "role-root": {"children": [{"id": "retail", "name": "06.Retail"}]},
        "retail": {"children": [{
            "id": "f8", "name": "01.F8 Experience (Launching Daily)",
            "children": [{"id": "flagship", "name": "01.Flagship Experience"}],
        }]},
    }

    class Response:
        ok = True

        def __init__(self, value):
            self.value = value

        def json(self):
            return self.value

    class Request:
        def post(self, url, **_kwargs):
            return Response(responses[url.rsplit("=", 1)[-1]])

    class Page:
        request = Request()

    assert flow_worker._asap_portal_menu_paths(Page()) == [
        ["Retail", "F8 Experience (Launching Daily)", "Flagship Experience"],
    ]


def test_portal_menu_root_retries_at_depth_two_when_depth_one_has_no_children(monkeypatch):
    monkeypatch.setattr(
        flow_worker, "_asap_portal_session",
        lambda _page, _diagnostics=None: {
            "web_base": "https://asap.example/mstr",
            "legacy_token": "live-token",
            "main_menu_id": "role-root",
        },
    )
    calls = []

    class Response:
        ok = True

        def __init__(self, value):
            self.value = value

        def json(self):
            return self.value

    class Request:
        def post(self, url, **kwargs):
            folder_id = url.rsplit("=", 1)[-1]
            depth = kwargs["form"]["depth"]
            calls.append((folder_id, depth))
            if folder_id == "role-root" and depth == "1":
                return Response({"id": "role-root", "name": "root"})
            if folder_id == "role-root":
                return Response({
                    "id": "role-root", "name": "root",
                    "children": [{"id": "retail", "name": "06.Retail"}],
                })
            return Response({
                "id": "retail", "name": "06.Retail",
                "children": [{"id": "flagship", "name": "01.Flagship Experience"}],
            })

    class Page:
        request = Request()

    diagnostics = {}
    assert flow_worker._asap_portal_menu_paths(Page(), diagnostics) == [
        ["Retail", "Flagship Experience"],
    ]
    assert calls[:2] == [("role-root", "1"), ("role-root", "2")]
    assert diagnostics["root_depth_retry"] is True


def test_portal_menu_tree_unwraps_the_live_data_envelope(monkeypatch):
    monkeypatch.setattr(
        flow_worker, "_asap_portal_session",
        lambda _page, _diagnostics=None: {
            "web_base": "https://asap.example/mstr",
            "legacy_token": "live-token",
            "main_menu_id": "role-root",
        },
    )
    responses = {
        "role-root": {
            "errorCode": 0, "message": "SUCCESS",
            "data": [{"id": "retail", "name": "06.Retail"}],
        },
        "retail": {
            "errorCode": "0", "message": "SUCCESS",
            "data": json.dumps({
                "id": "retail", "name": "06.Retail", "children": [{
                    "id": "f8", "name": "01.F8 Experience (Launching Daily)",
                    "children": [{"id": "flagship", "name": "01.Flagship Experience"}],
                }],
            }),
        },
    }

    class Response:
        ok = True

        def __init__(self, value):
            self.value = value

        def json(self):
            return self.value

    class Request:
        def post(self, url, **_kwargs):
            return Response(responses[url.rsplit("=", 1)[-1]])

    class Page:
        request = Request()

    assert flow_worker._asap_portal_menu_paths(Page()) == [
        ["Retail", "F8 Experience (Launching Daily)", "Flagship Experience"],
    ]


def test_portal_menu_tree_accepts_success_error_code(monkeypatch):
    monkeypatch.setattr(
        flow_worker, "_asap_portal_session",
        lambda _page, _diagnostics=None: {
            "web_base": "https://asap.example/mstr",
            "legacy_token": "live-token",
            "main_menu_id": "role-root",
        },
    )

    class Response:
        ok = True

        def __init__(self, value):
            self.value = value

        def json(self):
            return self.value

    class Request:
        def post(self, url, **_kwargs):
            folder_id = url.rsplit("=", 1)[-1]
            if folder_id == "role-root":
                return Response({
                    "errorCode": "success", "message": "SUCCESS",
                    "data": [{"id": "retail", "name": "06.Retail"}],
                })
            return Response({
                "errorCode": "success", "message": "SUCCESS",
                "data": {
                    "id": "retail", "name": "06.Retail",
                    "children": [{"id": "flagship", "name": "01.Flagship Experience"}],
                },
            })

    class Page:
        request = Request()

    assert flow_worker._asap_portal_menu_paths(Page()) == [
        ["Retail", "Flagship Experience"],
    ]


def test_asap_goto_waits_for_delayed_expired_session_redirect(monkeypatch, tmp_path):
    states = [False, False, True]
    calls = []
    monkeypatch.setattr("app.flow_worker._asap_login_visible", lambda _page: states.pop(0) if states else True)
    monkeypatch.setattr("app.flow_worker._visible_anchor_records", lambda _page: [])
    monkeypatch.setattr(
        "app.flow_worker._asap_authenticate_if_needed",
        lambda _page, profile: calls.append(profile) or True,
    )

    class Page:
        def goto(self, url, **_kwargs):
            calls.append(url)

        def wait_for_timeout(self, _milliseconds):
            pass

    assert _asap_goto(Page(), "https://portal.example/login", tmp_path) is True
    assert calls == ["https://portal.example/login", tmp_path]


def test_asap_frame_uses_newest_live_replacement():
    class Frame:
        def __init__(self, detached):
            self.detached = detached

        def is_detached(self):
            return self.detached

    class Element:
        def __init__(self, frame):
            self.frame = frame

        def content_frame(self):
            return self.frame

    old = Frame(True)
    current = Frame(False)

    class Locator:
        def wait_for(self, **_kwargs):
            pass

        def element_handles(self):
            return [old_element, current_element]

    old_element = Element(old)
    current_element = Element(current)

    class Page:
        def locator(self, _selector):
            return Locator()

    assert _asap_frame(Page()) is current


def test_asap_loading_wait_requires_four_clear_samples(monkeypatch):
    states = iter([True, False, False, False, False])
    waits = []
    monkeypatch.setattr(
        flow_worker, "_asap_loading_overlay_visible", lambda _page: next(states),
    )

    class Page:
        def wait_for_timeout(self, milliseconds):
            waits.append(milliseconds)

    flow_worker._asap_wait_for_loading_clear(Page(), timeout_ms=5_000)

    assert waits == [250, 250, 250, 250]


def test_asap_loading_overlay_detects_live_id_selector():
    selectors = []

    class Overlay:
        def count(self):
            return 1

        def nth(self, _index):
            return self

        def is_visible(self):
            return True

    class Frame:
        def locator(self, selector):
            selectors.append(selector)
            return Overlay()

    frame = Frame()

    class Page:
        main_frame = frame
        frames = [frame]

    assert flow_worker._asap_loading_overlay_visible(Page()) is True
    assert "#loading-spinner-container" in selectors[0]


def test_asap_results_accept_populated_raw_table_without_data_rows_marker(monkeypatch):
    class EmptyRows:
        @property
        def first(self):
            return self

        def count(self):
            return 0

    class Frame:
        def get_by_text(self, *_args, **_kwargs):
            return EmptyRows()

    frame = Frame()
    monkeypatch.setattr(flow_worker, "_asap_loading_overlay_visible", lambda _page: False)
    monkeypatch.setattr(flow_worker, "_asap_raw_table_ready", lambda candidate: candidate is frame)

    class Page:
        frames = [frame]

        def wait_for_timeout(self, _milliseconds):
            raise AssertionError("the populated raw table should be accepted immediately")

    assert flow_worker._asap_wait_for_results(Page(), timeout_ms=1_000) is frame


class _WaitPage:
    def __init__(self):
        self.waits = []

    def wait_for_timeout(self, milliseconds):
        self.waits.append(milliseconds)


def _empty_render_error():
    return RuntimeError(
        "ASAP report rows did not render within 600 seconds. "
        + flow_worker.ASAP_EMPTY_RESULT_DETAIL
    )


def test_asap_run_report_retries_once_after_silently_empty_rendering(monkeypatch):
    frame = object()
    attempts = []

    def fake_run(_page):
        attempts.append(1)
        if len(attempts) == 1:
            raise _empty_render_error()
        return frame

    monkeypatch.setattr(flow_worker, "_asap_run_report", fake_run)
    page = _WaitPage()
    retries = []

    assert flow_worker._asap_run_report_with_retry(page, on_retry=retries.append) is frame
    assert len(attempts) == 2
    assert len(retries) == 1
    assert page.waits == [1_500]


def test_asap_run_report_does_not_retry_while_overlay_still_visible(monkeypatch):
    attempts = []

    def fake_run(_page):
        attempts.append(1)
        raise RuntimeError(
            "ASAP report rows did not render within 600 seconds."
            " The ASAP loading overlay was still visible."
        )

    monkeypatch.setattr(flow_worker, "_asap_run_report", fake_run)

    with pytest.raises(RuntimeError, match="still visible"):
        flow_worker._asap_run_report_with_retry(_WaitPage())
    assert len(attempts) == 1


def test_asap_run_report_raises_after_second_empty_rendering(monkeypatch):
    attempts = []

    def fake_run(_page):
        attempts.append(1)
        raise _empty_render_error()

    monkeypatch.setattr(flow_worker, "_asap_run_report", fake_run)

    with pytest.raises(RuntimeError, match="did not render"):
        flow_worker._asap_run_report_with_retry(_WaitPage())
    assert len(attempts) == 2


def test_export_task_retry_restarts_a_failed_file_and_keeps_its_result():
    attempts = []

    def run_task(attempt):
        attempts.append(attempt)
        if len(attempts) == 1:
            raise RuntimeError("ASAP report menu item was not visible: installed base (MENA)")
        if len(attempts) == 2:
            raise RuntimeError(
                "The Edge download did not produce a stable finished file "
                "within 600 seconds in the local staging folder."
            )
        return {"status": "saved"}

    retries = []
    page = _WaitPage()
    result = flow_worker._export_task_with_retry(
        page, run_task, lambda attempt, exc: retries.append((attempt, str(exc))),
    )

    assert result == {"status": "saved"}
    assert attempts == [1, 2, 3]
    assert [attempt for attempt, _ in retries] == [1, 2]
    assert "menu item was not visible" in retries[0][1]
    assert "stable finished file" in retries[1][1]
    assert page.waits == [5_000, 5_000]


def test_export_task_retry_raises_the_original_error_after_three_attempts():
    attempts = []

    def run_task(attempt):
        attempts.append(attempt)
        raise RuntimeError("Target folder permissions changed mid-run.")

    with pytest.raises(RuntimeError, match="permissions changed"):
        flow_worker._export_task_with_retry(_WaitPage(), run_task, lambda *_args: None)
    assert len(attempts) == flow_worker.EXPORT_TASK_ATTEMPTS == 3


def test_export_task_retry_reports_every_distinct_attempt_error():
    errors = iter([
        RuntimeError("Workbook contains no default style"),
        RuntimeError("bookmark list was temporarily empty"),
    ])

    with pytest.raises(RuntimeError) as excinfo:
        flow_worker._export_task_with_retry(
            _WaitPage(), lambda _attempt: (_ for _ in ()).throw(next(errors)),
            lambda *_args: None, max_attempts=2,
        )

    message = str(excinfo.value)
    assert "Attempt 1: Workbook contains no default style" in message
    assert "Attempt 2: bookmark list was temporarily empty" in message


def test_export_task_does_not_reopen_report_after_edge_completed_download():
    attempts = []

    def run_task(attempt):
        attempts.append(attempt)
        raise flow_worker._CompletedDownloadProcessingError(
            "Edge completed the workbook; local processing failed"
        )

    with pytest.raises(
        flow_worker._CompletedDownloadProcessingError,
        match="Edge completed the workbook",
    ):
        flow_worker._export_task_with_retry(
            _WaitPage(), run_task, lambda *_args: None, max_attempts=2,
        )

    assert attempts == [1]


def test_edge_native_download_event_is_the_completion_authority(tmp_path):
    completed = tmp_path / "edge-download.xlsx"
    completed.write_bytes(b"PK\x03\x04workbook")

    class Download:
        def failure(self):
            return None

        def path(self):
            return str(completed)

    class Pending:
        value = Download()

    class ExpectDownload:
        def __enter__(self):
            return Pending()

        def __exit__(self, *_args):
            return False

    class Page:
        def expect_download(self, *, timeout):
            assert timeout == flow_worker.DOWNLOAD_MAX_TIMEOUT_SECONDS * 1_000
            return ExpectDownload()

    triggered = []
    result = flow_worker._edge_completed_download(Page(), lambda: triggered.append(True))

    assert result == completed
    assert triggered == [True]


def test_completed_edge_download_uses_the_native_path_for_gscm(tmp_path):
    completed = tmp_path / "native-gscm-download.xlsx"
    completed.write_bytes(b"PK\x03\x04workbook")
    calls = []

    class Download:
        def failure(self):
            calls.append("failure")
            return None

        def path(self):
            calls.append("path")
            return str(completed)

    result = flow_worker._completed_edge_download(
        Download(), "the GSCM Excel download",
    )

    assert result == completed
    assert calls == ["failure", "path"]


def test_completed_edge_download_reports_gscm_failure_without_a_staging_guess():
    class Download:
        def failure(self):
            return "net::ERR_FAILED"

        def path(self):
            raise AssertionError("a failed Edge download has no valid local path")

    with pytest.raises(RuntimeError, match=r"GSCM.*net::ERR_FAILED"):
        flow_worker._completed_edge_download(
            Download(), "the GSCM Excel download",
        )


def test_dashboard_edge_event_uses_a_bounded_staging_handoff(tmp_path, monkeypatch):
    staged = tmp_path / "dashboard.xlsx"
    staged.write_bytes(b"PK\x03\x04staged")
    before = flow_worker._download_staging_snapshot(tmp_path)
    calls = []

    def wait_for_staged(staging_dir, files_before, **kwargs):
        calls.append((staging_dir, files_before, kwargs))
        return staged

    monkeypatch.setattr(flow_worker, "_wait_for_staged_download", wait_for_staged)

    result = flow_worker._asap_dashboard_event_staged_download(
        tmp_path, before, "Download Main Data (xlsx)",
    )

    assert result == staged
    assert calls == [(
        tmp_path,
        before,
        {"start_timeout_seconds": flow_worker.DOWNLOAD_EVENT_STAGING_TIMEOUT_SECONDS},
    )]
    assert flow_worker.DOWNLOAD_EVENT_STAGING_TIMEOUT_SECONDS == 60


def test_dashboard_edge_event_reports_a_missing_staging_handoff(
    tmp_path, monkeypatch,
):
    monkeypatch.setattr(
        flow_worker, "_wait_for_staged_download",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(RuntimeError(
            "ASAP started an Edge download, but no new or updated file appeared"
        )),
    )

    with pytest.raises(RuntimeError) as excinfo:
        flow_worker._asap_dashboard_event_staged_download(
            tmp_path, {}, "Download Main Data (xlsx)",
        )

    message = str(excinfo.value)
    assert "native download-start event" in message
    assert "Download Main Data (xlsx)" in message
    assert "within 60 seconds" in message
    assert str(tmp_path) in message
    assert "permissions" in message


def test_gscm_load_buffers_are_one_minute_then_two_minutes():
    assert flow_worker.GSCM_EXPORT_TASK_ATTEMPTS == 2
    assert flow_worker.GSCM_INITIAL_LOAD_BUFFER_MS == 60_000
    assert flow_worker.GSCM_RETRY_LOAD_BUFFER_MS == 120_000


def test_asap_table_control_score_accepts_only_compact_top_right_controls():
    table = {"x": 100, "y": 200, "width": 800, "height": 500}

    info = {"x": 870, "y": 205, "width": 20, "height": 20}
    left_filter = {"x": 120, "y": 205, "width": 20, "height": 20}
    full_overlay = {"x": 100, "y": 200, "width": 800, "height": 500}

    assert flow_worker._asap_table_control_score(table, info) is not None
    assert flow_worker._asap_table_control_score(table, left_filter) is None
    assert flow_worker._asap_table_control_score(table, full_overlay) is None


def test_asap_run_control_accepts_input_value_rendering():
    class Locator:
        def __init__(self, visible):
            self.visible = visible

        @property
        def first(self):
            return self

        def count(self):
            return int(self.visible)

        def is_visible(self):
            return self.visible

    input_run = Locator(True)

    class Root:
        def get_by_role(self, _role, **_kwargs):
            return Locator(False)

        def locator(self, selector):
            assert "input[type='button'][value='RUN' i]" in selector
            return input_run

        def get_by_text(self, _text, **_kwargs):
            raise AssertionError("the visible RUN input should be accepted before text fallback")

    assert flow_worker._asap_run_control(Root()) is input_run


def test_asap_export_action_accepts_input_value_rendering():
    class Locator:
        def __init__(self, visible):
            self.visible = visible

        @property
        def first(self):
            return self

        def count(self):
            return int(self.visible)

        def is_visible(self):
            return self.visible

    input_export = Locator(True)

    class Root:
        def get_by_role(self, _role, **_kwargs):
            return Locator(False)

        def locator(self, selector):
            assert "input[type='button'][value='Export' i]" in selector
            return input_export

        def get_by_text(self, _text, **_kwargs):
            raise AssertionError("the visible Export input should be accepted before text fallback")

    assert flow_worker._asap_export_action(Root()) is input_export


def test_asap_xlsx_format_prefers_flat_excel_export_and_accepts_legacy_label():
    preferred, pattern = flow_worker._asap_export_format_names("xlsx")

    assert preferred == "Excel with plain text"
    assert pattern.fullmatch("Excel with plain text")
    assert pattern.fullmatch("Excel file format")
    assert not pattern.fullmatch("Excel with formatting")


def test_asap_download_retries_one_wizard_recognition_failure(monkeypatch, tmp_path):
    attempts = []

    def download(*_args):
        attempts.append(len(attempts) + 1)
        if len(attempts) == 1:
            raise RuntimeError(
                "ASAP Export Wizard opened, but its XLSX option or Export action "
                "was not recognized. Format option found: False. Export action found: False."
            )
        return tmp_path / "download.xlsx", []

    monkeypatch.setattr(flow_worker, "_asap_download", download)

    class Page:
        def __init__(self):
            self.waits = []

        def wait_for_timeout(self, milliseconds):
            self.waits.append(milliseconds)

    page = Page()
    result = flow_worker._asap_download_with_retry(
        page, object(), {}, tmp_path,
    )

    assert result == (tmp_path / "download.xlsx", [])
    assert attempts == [1, 2]
    assert page.waits == [1_500]


def test_asap_download_stops_after_second_wizard_recognition_failure(monkeypatch, tmp_path):
    attempts = []
    error = (
        "ASAP Export Wizard opened, but its XLSX option or Export action was not "
        "recognized. Format option found: False. Export action found: False."
    )

    def download(*_args):
        attempts.append(len(attempts) + 1)
        raise RuntimeError(error)

    monkeypatch.setattr(flow_worker, "_asap_download", download)

    class Page:
        def wait_for_timeout(self, _milliseconds):
            pass

    with pytest.raises(RuntimeError, match="ASAP Export Wizard opened"):
        flow_worker._asap_download_with_retry(Page(), object(), {}, tmp_path)

    assert attempts == [1, 2]


def test_asap_download_does_not_retry_other_failures(monkeypatch, tmp_path):
    attempts = []

    def download(*_args):
        attempts.append(len(attempts) + 1)
        raise RuntimeError("ASAP export started, but no download was emitted.")

    monkeypatch.setattr(flow_worker, "_asap_download", download)

    class Page:
        def wait_for_timeout(self, _milliseconds):
            raise AssertionError("non-retryable failures must not wait")

    with pytest.raises(RuntimeError, match="no download was emitted"):
        flow_worker._asap_download_with_retry(Page(), object(), {}, tmp_path)

    assert attempts == [1]


def test_asap_download_retries_once_with_replaced_frame(monkeypatch, tmp_path):
    attempts = []
    original_frame = object()
    replacement_frame = object()

    def download(_page, frame, _job, _staging_dir):
        attempts.append(frame)
        if len(attempts) == 1:
            raise flow_worker.PlaywrightError("Frame was detached")
        return tmp_path / "download.csv", []

    monkeypatch.setattr(flow_worker, "_asap_download", download)
    monkeypatch.setattr(flow_worker, "_asap_frame", lambda _page: replacement_frame)

    class Page:
        def __init__(self):
            self.waits = []

        def wait_for_timeout(self, milliseconds):
            self.waits.append(milliseconds)

    page = Page()
    result = flow_worker._asap_download_with_retry(
        page, original_frame, {}, tmp_path,
    )

    assert result == (tmp_path / "download.csv", [])
    assert attempts == [original_frame, replacement_frame]
    assert page.waits == [1_500]


def test_asap_download_does_not_retry_other_playwright_errors(monkeypatch, tmp_path):
    attempts = []

    def download(*_args):
        attempts.append(len(attempts) + 1)
        raise flow_worker.PlaywrightError("Target page, context or browser has been closed")

    monkeypatch.setattr(flow_worker, "_asap_download", download)

    class Page:
        def wait_for_timeout(self, _milliseconds):
            raise AssertionError("non-frame Playwright errors must not wait")

    with pytest.raises(flow_worker.PlaywrightError, match="browser has been closed"):
        flow_worker._asap_download_with_retry(Page(), object(), {}, tmp_path)

    assert attempts == [1]


def test_raw_table_excel_menu_item_is_a_direct_download_action():
    class Locator:
        def __init__(self, visible):
            self.visible = visible

        @property
        def first(self):
            return self

        def count(self):
            return int(self.visible)

        def is_visible(self):
            return self.visible

        def filter(self, **_kwargs):
            return self

    class Page:
        frames = []

        def __init__(self):
            self.context = type("Context", (), {"pages": [self]})()

        def get_by_role(self, _role, **_kwargs):
            return Locator(False)

        def get_by_text(self, pattern):
            assert pattern.fullmatch("Excel")
            return Locator(True)

        def wait_for_timeout(self, _milliseconds):
            raise AssertionError("the visible Excel menu item should be accepted immediately")

    action = flow_worker._asap_wait_for_raw_menu_download_action(Page(), "xlsx")

    assert action.is_visible()
    assert flow_worker._asap_wait_for_raw_menu_download_action(Page(), "csv") is None


def test_raw_table_export_control_accepts_direct_export_options_popup():
    popup = object()

    class Page:
        frames = []

        def __init__(self):
            self.context = type("Context", (), {"pages": [self, popup]})()

        def wait_for_timeout(self, _milliseconds):
            raise AssertionError("the new Export Options popup should be accepted immediately")

    page = Page()
    menu_export, wizard_opened = flow_worker._asap_wait_for_raw_menu_export_or_wizard(
        page, {page},
    )

    assert menu_export is None
    assert wizard_opened is True


def test_raw_table_excel_dialog_exposes_final_export_action():
    class Locator:
        def __init__(self, visible):
            self.visible = visible

        @property
        def first(self):
            return self

        def count(self):
            return int(self.visible)

        def is_visible(self):
            return self.visible

        def filter(self, **_kwargs):
            return self

    class Page:
        frames = []

        def __init__(self):
            self.context = type("Context", (), {"pages": [self]})()

        def get_by_text(self, pattern):
            assert pattern.fullmatch("Export to Excel")
            return Locator(True)

        def get_by_role(self, role, **kwargs):
            assert role == "button"
            assert kwargs == {"name": "Export", "exact": True}
            return Locator(True)

        def locator(self, _selector):
            return Locator(False)

        def wait_for_timeout(self, _milliseconds):
            raise AssertionError("the visible Export button should be accepted immediately")

    action = flow_worker._asap_wait_for_raw_export_confirmation(Page(), "xlsx")

    assert action.is_visible()
    assert flow_worker._asap_wait_for_raw_export_confirmation(Page(), "csv") is None


def test_export_view_waits_for_loading_overlay_before_and_after_click(monkeypatch):
    waits = []

    class Control:
        def click(self, **kwargs):
            assert kwargs == {"timeout": 30_000}

    class VisibleMarker:
        def is_visible(self):
            return True

    class Markers:
        def all(self):
            return [VisibleMarker()]

    class ActiveFrame:
        def get_by_text(self, *_args, **_kwargs):
            return Markers()

    active_frame = ActiveFrame()
    frame_values = iter(["fresh-before-click", active_frame])
    monkeypatch.setattr(
        flow_worker, "_asap_wait_for_loading_clear",
        lambda _page: waits.append("clear"),
    )
    monkeypatch.setattr(flow_worker, "_asap_frame", lambda _page: next(frame_values))

    candidate_frames = []
    monkeypatch.setattr(
        flow_worker, "_asap_export_view_candidates",
        lambda _page, frame: candidate_frames.append(frame) or [("Target view", Control())],
    )

    class Page:
        def wait_for_timeout(self, milliseconds):
            waits.append(milliseconds)

    returned_frame, selected = flow_worker._asap_activate_export_view(
        Page(), object(), "Target view",
    )

    assert candidate_frames == ["fresh-before-click"]
    assert waits[:3] == ["clear", 500, "clear"]
    assert returned_frame is active_frame
    assert selected == "Target view"


def test_asap_navigation_waits_for_replacement_and_requested_breadcrumb(monkeypatch):
    class Frame:
        def is_detached(self):
            return False

    previous = Frame()
    current = Frame()
    monkeypatch.setattr(flow_worker, "_asap_frame", lambda _page: current)

    class Item:
        def is_visible(self):
            return True

    class Locator:
        def all(self):
            return [Item()]

    class Target:
        def __init__(self):
            self.waited = None

        def wait_for(self, **kwargs):
            self.waited = kwargs

    class Page:
        def get_by_text(self, _label, exact=True):
            assert exact is True
            return Locator()

        def wait_for_timeout(self, _milliseconds):
            raise AssertionError("replacement and breadcrumb should be accepted immediately")

    target = Target()
    result = _asap_wait_for_report_navigation(
        Page(), previous, target,
        ["Market", "TechInsights Smartphone M/S", "All Products Demand"],
        timeout_ms=1_000,
    )

    assert result is current
    assert target.waited == {"state": "hidden", "timeout": 1_000}


def test_asap_navigation_accepts_stable_requested_breadcrumb_in_reused_frame(monkeypatch):
    class Frame:
        def is_detached(self):
            return False

    current = Frame()
    monkeypatch.setattr(flow_worker, "_asap_frame", lambda _page: current)
    monkeypatch.setattr(flow_worker, "_asap_frame_signature", lambda _frame: "same-body")
    monkeypatch.setattr(
        flow_worker, "_asap_text_visible_across_frames", lambda _page, _label: True,
    )

    class Target:
        def wait_for(self, **_kwargs):
            pass

    class Page:
        def __init__(self):
            self.waits = []

        def wait_for_timeout(self, milliseconds):
            self.waits.append(milliseconds)

    page = Page()
    result = _asap_wait_for_report_navigation(
        page, current, Target(), ["Market", "TechInsights Smartphone M/S"],
        previous_signature="same-body", timeout_ms=1_000,
    )

    assert result is current
    assert page.waits == [250, 250]


def test_asap_menu_wait_allows_delayed_hover_items(monkeypatch):
    expected = object()
    states = iter([None, None, expected])
    monkeypatch.setattr(flow_worker, "_asap_first_visible", lambda _locator: next(states))

    class Page:
        def __init__(self):
            self.waits = []

        def wait_for_timeout(self, milliseconds):
            self.waits.append(milliseconds)

    page = Page()
    assert _asap_wait_for_visible(page, object(), timeout_ms=1_000) is expected
    assert page.waits == [100, 100]


def test_asap_last_visible_chooses_leaf_when_group_reuses_label():
    class Item:
        def __init__(self, name, visible=True):
            self.name = name
            self.visible = visible

        def is_visible(self):
            return self.visible

    group = Item("group")
    hidden = Item("hidden", visible=False)
    leaf = Item("leaf")

    class Locator:
        def all(self):
            return [group, hidden, leaf]

    assert _asap_last_visible(Locator()) is leaf


def test_asap_menu_wait_can_prefer_last_duplicate_label(monkeypatch):
    expected = object()
    monkeypatch.setattr(flow_worker, "_asap_last_visible", lambda _locator: expected)

    class Page:
        def wait_for_timeout(self, _milliseconds):
            raise AssertionError("visible duplicate leaf should be found immediately")

    assert _asap_wait_for_visible(
        Page(), object(), timeout_ms=1_000, prefer_last=True,
    ) is expected


def test_asap_navigation_accepts_stable_changed_body_when_iframe_is_reused(monkeypatch):
    class Frame:
        def is_detached(self):
            return False

    current = Frame()
    monkeypatch.setattr(flow_worker, "_asap_frame", lambda _page: current)
    monkeypatch.setattr(flow_worker, "_asap_frame_signature", lambda _frame: "new-body")

    class Item:
        def is_visible(self):
            return False

    class Locator:
        def all(self):
            return [Item()]

    class Target:
        def wait_for(self, **_kwargs):
            pass

    class Page:
        def __init__(self):
            self.waits = []

        def get_by_text(self, _label, exact=True):
            assert exact is True
            return Locator()

        def wait_for_timeout(self, milliseconds):
            self.waits.append(milliseconds)

    page = Page()
    result = _asap_wait_for_report_navigation(
        page, current, Target(),
        ["Market", "TechInsights Smartphone M/S", "All Products Demand"],
        previous_signature="old-body", timeout_ms=1_000,
    )

    assert result is current
    assert page.waits == [250, 250]


def test_targeted_scan_surfaces_path_error_when_no_report_is_discovered(monkeypatch, tmp_path):
    monkeypatch.setattr(flow_worker, "_asap_goto", lambda *_args: False)
    monkeypatch.setattr(
        flow_worker, "_asap_open_report",
        lambda *_args: (_ for _ in ()).throw(RuntimeError("breadcrumb proof missing")),
    )
    job = {
        "site": {"base_url": "https://portal.example", "auth_url": None},
        "discovery": {
            "max_duration_minutes": 1,
            "scope": ["*"],
            "report_paths": [["Market", "Group", "Target Report"]],
        },
    }

    with pytest.raises(RuntimeError, match="Target Report: breadcrumb proof missing"):
        flow_worker.discover_asap_catalog(
            object(), job, lambda *_args: None, tmp_path,
        )


RUN_ID = 77


def _run_dir(target) -> Path:
    """The run folder execute_job creates for RUN_ID inside the target folder."""
    return Path(target) / flow_worker.flow_retention.run_folder_name(RUN_ID)


def test_execute_job_downloads_every_export_view_before_returning(monkeypatch, tmp_path):
    activated = []
    staged = []

    monkeypatch.setattr(flow_worker, "_asap_open_report", lambda *_args: object())
    monkeypatch.setattr(
        flow_worker, "_asap_activate_export_view",
        lambda _page, frame, label: (activated.append(label) or frame, label),
    )
    monkeypatch.setattr(flow_worker, "_asap_apply_configuration", lambda *_args: None)
    monkeypatch.setattr(flow_worker, "_has_named_control", lambda *_args: False)

    def fake_download(_page, _frame, _job, staging_dir):
        path = Path(staging_dir) / f"source-{len(staged) + 1}.csv"
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text("value,units\nitem,1\n", encoding="utf-8")
        staged.append(path)
        return path, []

    monkeypatch.setattr(flow_worker, "_asap_download", fake_download)
    progress = []
    job = {
        "flow": {"id": 1, "name": "ASAP TI"},
        "site": {"adapter": "asap_portal"},
        "report": {
            "id": 1, "name": "TechInsights Smartphone M/S", "filters": [],
            "automation": {"export_views": [
                {"label": "Export Wizard (Global/Region)", "filter_keys": []},
                {"label": "Export Wizard (Selected Countries)", "filter_keys": []},
            ]},
            "export_views": [
                "Export Wizard (Global/Region)", "Export Wizard (Selected Countries)",
            ],
        },
        "selections": {},
        "downloads": {
            "periods": [None], "target_folder": str(tmp_path), "file_format": "csv",
            "filename_template": "{flow}_{export}.csv",
        },
    }

    artifacts, _timings = flow_worker.execute_job(
        object(), job, lambda *args: progress.append(args), tmp_path,
        tmp_path / "staging", run_id=RUN_ID,
    )

    assert activated == job["report"]["export_views"]
    assert len(artifacts) == 2
    assert [item["export_view"] for item in artifacts] == activated
    assert [item["bundle_index"] for item in artifacts] == [1, 2]
    # Each export view lands in its own file, named after the view. The file
    # itself carries only the portal's own columns - Metronome adds none.
    paths = [Path(artifact["file_path"]) for artifact in artifacts]
    assert len({path.name for path in paths}) == 2
    for path, label in zip(paths, activated):
        assert re.sub(r"\W+", "_", label).strip("_").casefold() in path.stem.casefold()
        assert "Metronome" not in path.read_text(encoding="utf-8-sig")


def test_execute_job_skips_files_already_saved_by_the_resumed_run(tmp_path):
    events = []
    job = {
        "site": {"adapter": "web_export"},
        "report": {
            "name": "Weekly", "url": "https://reports.example.test/weekly",
            "ready_text": None, "open_export_text": None,
            "download_text": "Download CSV", "export_views": [],
        },
        "downloads": {
            "target_folder": str(tmp_path),
            "periods": [["2026-W30"], ["2026-W31"]],
            "filename_template": "weekly_{week}.csv",
            "file_format": "csv",
        },
        "resume": {"from_run_id": 41, "completed": [
            {"export_view": None, "period_key": ["2026-W30"]},
            {"export_view": None, "period_key": ["2026-W31"]},
        ]},
    }

    artifacts, timings = flow_worker.execute_job(
        object(), job,
        lambda _status, progress, _artifacts=None: events.append(progress),
        tmp_path, run_id=RUN_ID,
    )

    assert artifacts == []
    skips = [item for item in events if item.get("stage") == "resume_skip"]
    assert len(skips) == 2
    assert "run #41" in skips[0]["message"]
    assert skips[0]["item_index"] == 1 and skips[1]["item_index"] == 2
    assert timings[-1]["phase"] == "total"


def test_export_task_key_matches_completed_entries_across_json_roundtrip():
    task_key = flow_worker._export_task_key(None, ["2026-W30", "2026-W31"])
    completed = flow_worker._resume_completed_keys({
        "resume": {"completed": [{"export_view": None, "period_key": ["2026-W30", "2026-W31"]}]},
    })
    assert task_key in completed
    assert flow_worker._export_task_key("Global", ["2026-W30", "2026-W31"]) not in completed


def test_execute_job_appends_into_the_callers_artifact_list(tmp_path):
    shared = []
    job = {
        "site": {"adapter": "web_export"},
        "report": {
            "name": "Weekly", "url": "https://reports.example.test/weekly",
            "ready_text": None, "open_export_text": None,
            "download_text": "Download CSV", "export_views": [],
        },
        "downloads": {
            "target_folder": str(tmp_path),
            "periods": [["2026-W30"]],
            "filename_template": "weekly_{week}.csv",
            "file_format": "csv",
        },
        "resume": {"from_run_id": 9, "completed": [
            {"export_view": None, "period_key": ["2026-W30"]},
        ]},
    }

    artifacts, _timings = flow_worker.execute_job(
        object(), job, lambda *_args: None, tmp_path, artifacts=shared, run_id=RUN_ID,
    )

    # In-place semantics: a failure that unwinds execute_job leaves every
    # already-saved file visible in the caller's list for the failed report.
    assert artifacts is shared


# --- Filter discovery: native select snapshot ---

class _FilterLocator:
    def all(self):
        return []

    def all_inner_texts(self):
        return []

    def count(self):
        return 0


class _FilterFramePage:
    def __init__(self, frames):
        self.frames = frames


class _FilterFrame:
    def __init__(self, records):
        self._records = records
        self.page = _FilterFramePage([])

    def evaluate(self, _script):
        return self._records

    def locator(self, _selector):
        return _FilterLocator()

    def get_by_text(self, *_args, **_kwargs):
        return _FilterLocator()

    def is_detached(self):
        return False


def _discover(records, monkeypatch):
    monkeypatch.setattr(flow_worker, "_asap_discover_week_slider", lambda _frame: None)
    diagnostics = {}
    definitions = flow_worker._asap_discover_filters(_FilterFrame(records), diagnostics)
    return definitions, diagnostics


def test_select_snapshot_resolves_container_labels_and_multiplicity(monkeypatch):
    definitions, diagnostics = _discover([
        {"index": 0, "id": "dimSel", "name": "", "label": "Dimension",
         "label_source": "container", "multiple": True, "select2": True, "visible": True,
         "options": ["Sell-out Region", "Sell-out Subsidiary", "Country"]},
        {"index": 1, "id": "", "name": "", "label": "Channel Type",
         "label_source": "sibling", "multiple": False, "select2": False, "visible": True,
         "options": ["Open", "Carrier"]},
    ], monkeypatch)

    dimension = next(item for item in definitions if item["label"] == "Dimension")
    assert dimension["control_type"] == "multi_select"
    assert dimension["options"] == ["Sell-out Region", "Sell-out Subsidiary", "Country"]
    assert dimension["automation"]["label_source"] == "container"
    assert dimension["automation"]["select2"] is True
    channel = next(item for item in definitions if item["label"] == "Channel Type")
    assert channel["control_type"] == "select"
    assert diagnostics["selects_seen"] == 2
    assert diagnostics["selects_unlabeled"] == 0
    assert diagnostics["definitions"] == 2


def test_unlabeled_selects_are_never_dropped(monkeypatch):
    definitions, diagnostics = _discover([
        # No label anywhere but an id: the id names the filter.
        {"index": 0, "id": "yearSelect", "name": "", "label": "", "label_source": "",
         "multiple": False, "select2": False, "visible": True, "options": ["2025", "2026"]},
        # Nothing at all: a stable synthesized name keeps the options.
        {"index": 3, "id": "", "name": "", "label": "", "label_source": "",
         "multiple": False, "select2": True, "visible": True, "options": ["A", "B"]},
    ], monkeypatch)

    year = next(item for item in definitions if item["label"] == "yearSelect")
    assert year["options"] == ["2025", "2026"]
    assert year["automation"]["label_source"] == "select_attribute"
    synthesized = next(item for item in definitions if item["label"] == "Filter 4")
    assert synthesized["options"] == ["A", "B"]
    assert synthesized["automation"]["label_source"] == "synthesized"
    assert diagnostics["selects_unlabeled"] == 1


def test_three_part_options_still_identify_data_configuration(monkeypatch):
    definitions, _diagnostics = _discover([
        {"index": 0, "id": "", "name": "", "label": "", "label_source": "",
         "multiple": False, "select2": True, "visible": True,
         "options": ["MENA - Global - Global", "Global - Global - CIS"]},
    ], monkeypatch)

    assert definitions[0]["label"] == "Data Configuration"
    assert definitions[0]["options"] == ["MENA - Global - Global", "Global - Global - CIS"]
    # Automation must survive the merge's label normalization: the lookup key
    # is derived from the normalized label, not the raw one.
    assert definitions[0]["automation"]["select2"] is True
    assert definitions[0]["automation"]["label_source"] == "three_part_options"


def test_period_slider_replaces_generated_handle_titles(monkeypatch):
    monkeypatch.setattr(flow_worker, "_asap_discover_week_slider", lambda _frame: None)
    monkeypatch.setattr(
        flow_worker,
        "_asap_discover_period_slider",
        lambda _frame: (["202623", "202624", "202625"], {"kind": "range_slider"}),
    )
    records = [
        {"index": 0, "id": "", "name": "", "label": (
            "7D4EBFE00EB4F6835F8D27DAD15D436C4_1B9EE11C7B49DB0A67E429A149EE9322"
        ),
         "label_source": "sibling", "multiple": False, "select2": False,
         "visible": True, "options": ["202623", "202625"]},
        {"index": 1, "id": "", "name": "", "label": "202623",
         "label_source": "sibling", "multiple": False, "select2": False,
         "visible": True, "options": ["202623", "202625"]},
    ]

    definitions = flow_worker._asap_discover_filters(_FilterFrame(records), {})

    assert [definition["label"] for definition in definitions] == ["Period"]
    assert definitions[0]["options"] == ["202623", "202624", "202625"]
    assert definitions[0]["automation"] == {"kind": "range_slider"}


def test_synthesized_labels_are_frame_qualified_to_avoid_merging(monkeypatch):
    monkeypatch.setattr(flow_worker, "_asap_discover_week_slider", lambda _frame: None)
    unlabeled = {"index": 0, "id": "", "name": "", "label": "", "label_source": "",
                 "multiple": False, "select2": False, "visible": True}
    outer = _FilterFrame([{**unlabeled, "options": ["A", "B"]}])
    inner = _FilterFrame([{**unlabeled, "options": ["C", "D"]}])
    outer.page = _FilterFramePage([outer, inner])

    diagnostics = {}
    definitions = flow_worker._asap_discover_filters(outer, diagnostics)

    labels = sorted(item["label"] for item in definitions)
    assert labels == ["Filter 1", "Filter 2.1"]
    first = next(item for item in definitions if item["label"] == "Filter 1")
    second = next(item for item in definitions if item["label"] == "Filter 2.1")
    assert first["options"] == ["A", "B"]
    assert second["options"] == ["C", "D"]


def test_container_labels_require_a_single_select_container():
    source = Path(flow_worker.__file__).read_text()
    # A shared section title would name several filters identically and merge
    # their option lists into one definition.
    assert 'querySelectorAll("select").length === 1' in source


def test_select_snapshot_sweeps_every_frame_and_dedupes(monkeypatch):
    record = {"index": 0, "id": "s1", "name": "", "label": "Region",
              "label_source": "container", "multiple": False, "select2": False,
              "visible": True, "options": ["Global"]}
    extra = {"index": 1, "id": "s2", "name": "", "label": "Channel",
             "label_source": "container", "multiple": False, "select2": False,
             "visible": True, "options": ["Open"]}
    inner = _FilterFrame([record, extra])
    outer = _FilterFrame([record])
    outer.page = _FilterFramePage([outer, inner])

    diagnostics = {}
    records = flow_worker._asap_native_select_records(outer, diagnostics)

    assert [item["id"] for item in records] == ["s1", "s2"]
    assert diagnostics["frames_scanned"] == 2
    assert diagnostics["selects_seen"] == 2


def test_empty_discovery_reports_diagnostics_instead_of_passing_silently(monkeypatch):
    definitions, diagnostics = _discover([], monkeypatch)
    assert definitions == []
    assert diagnostics["selects_seen"] == 0
    assert diagnostics["definitions"] == 0

    source = Path(flow_worker.__file__).read_text()
    assert '"stage": "filter_inspection_empty"' in source
    assert '"discovery_diagnostics": view_diagnostics' in source
    # Inactive report tabs keep hidden selects in the DOM; the snapshot must
    # exclude fully hidden containers while keeping Select2's hidden owners.
    assert "if (!selfVisible && !select2Visible && !parentVisible) return;" in source


def test_mega_menu_groups_come_from_styled_headers_not_first_links():
    """The Advanced menu titles its columns with styled text (SCM Insights,
    Outreach, AI Insights), not anchors. The first report link of a column
    must never be promoted to group ('Advanced > Z8 Command Center > M
    Tracker' was wrong: Z8 Command Center is itself a report)."""
    root = _record("Advanced", 700, 90)
    before = [root]
    after = [
        root,
        _record("Gross ASP Monitoring", 125, 475, href="javascript:open('gasp')"),
        _record("Stretch Achievement", 125, 519, href="javascript:open('stretch')"),
        _record("Sales Tracker", 125, 563, href="javascript:open('sales')"),
        _record("Sales Leadtime", 125, 607, href="javascript:open('leadtime')"),
        _record("Price Sensing", 125, 651, href="javascript:open('price')"),
        _record("Outreach Market", 660, 475, href="javascript:open('outreach')"),
        _record("Z8 Command Center", 1190, 475, href="javascript:open('z8')"),
        _record("M Tracker", 1190, 519, href="javascript:open('mt')"),
        _record("RV Tracker", 1190, 563, href="javascript:open('rv')"),
        _record("TP Monitoring", 1190, 607, href="javascript:open('tp')"),
        _record("WOS Monitoring", 1190, 651, href="javascript:open('wos')"),
    ]
    headers = [
        {"text": "SCM Insights", "x": 125, "y": 417},
        {"text": "Outreach", "x": 660, "y": 417},
        {"text": "AI Insights", "x": 1190, "y": 417},
    ]

    paths = _menu_report_paths(root, before, after, headers)

    assert ["Advanced", "SCM Insights", "Gross ASP Monitoring"] in paths
    assert ["Advanced", "SCM Insights", "Price Sensing"] in paths
    assert ["Advanced", "Outreach", "Outreach Market"] in paths
    assert ["Advanced", "AI Insights", "Z8 Command Center"] in paths
    assert ["Advanced", "AI Insights", "M Tracker"] in paths
    assert ["Advanced", "AI Insights", "WOS Monitoring"] in paths
    # The old failure mode: first link swallowed its siblings as children.
    assert not any(path[1] == "Z8 Command Center" for path in paths)
    assert not any(path[1] == "Gross ASP Monitoring" for path in paths)
    assert len(paths) == 11


def test_clickable_first_links_without_headers_stay_flat_reports():
    """If the header snapshot sees nothing, a column whose first item is a
    real link must yield flat report paths - wrong grouping is worse than a
    two-level path, and runtime navigation only needs root + leaf."""
    root = _record("Advanced", 700, 90)
    after = [
        root,
        _record("Z8 Command Center", 1190, 475, href="javascript:open('z8')"),
        _record("M Tracker", 1190, 519, href="javascript:open('mt')"),
    ]

    paths = _menu_report_paths(root, [root], after)

    assert paths == [
        ["Advanced", "Z8 Command Center"],
        ["Advanced", "M Tracker"],
    ]


def test_inert_hrefs_do_not_make_a_link_clickable():
    assert flow_worker._menu_link_target({"href": "#", "onclick": ""}) == ""
    assert flow_worker._menu_link_target({"href": "javascript:void(0)", "onclick": ""}) == ""
    assert flow_worker._menu_link_target({"href": "javascript:open('x')", "onclick": ""}) != ""


def test_semantic_menu_buttons_are_actionable_without_inline_javascript():
    root = _record("Retail", 600, 90)
    after = [
        root,
        _record("Flagship Experience", 650, 180, tag="button"),
        _record("Retail Analytics", 650, 215, role="menuitem"),
    ]

    assert _menu_report_paths(root, [root], after) == [
        ["Retail", "Flagship Experience"],
        ["Retail", "Retail Analytics"],
    ]


def test_disabled_semantic_menu_controls_are_not_reports():
    assert flow_worker._menu_link_target({
        "href": "", "onclick": "", "tag": "button", "role": "", "disabled": True,
    }) == ""


def test_revealed_menu_headers_diff_by_text_and_position():
    before = [{"text": "Filters", "x": 10, "y": 10}]
    after = [
        {"text": "Filters", "x": 10, "y": 10},
        {"text": "AI Insights", "x": 1190, "y": 417},
    ]
    revealed = flow_worker._revealed_menu_headers(before, after)
    assert revealed == [{"text": "AI Insights", "x": 1190, "y": 417}]


# --- HTML dashboard download links + scan payload caps ---


def _stub_dashboard_execution(monkeypatch):
    monkeypatch.setattr(flow_worker, "_asap_open_report", lambda *_args: object())
    monkeypatch.setattr(flow_worker, "_asap_wait_for_loading_clear", lambda _page: None)
    monkeypatch.setattr(flow_worker, "_asap_frame", lambda _page: object())
    monkeypatch.setattr(flow_worker, "_asap_apply_configuration", lambda *_args: None)
    monkeypatch.setattr(flow_worker, "_has_named_control", lambda *_args: False)
    monkeypatch.setattr(
        flow_worker,
        "_asap_activate_export_view",
        lambda *_args: (_ for _ in ()).throw(
            AssertionError("dashboards have no export views")
        ),
    )


def _mtracker_dashboard_job(
    target: Path, *, transformation=False, sql=False, file_format="xlsx",
):
    return {
        "site": {"adapter": "asap_portal"},
        "flow": {"name": "M Tracker pull"},
        "report": {
            "id": 7,
            "name": "M Tracker",
            "url": "https://portal.example.test/mtracker",
            "ready_text": None,
            "open_export_text": None,
            "download_text": "Download Main Data",
            "filters": [],
            "export_views": [],
            "download_links": ["Download Main Data (xlsx)"],
            "automation": {"kind": "html_dashboard", "download_links": []},
        },
        "downloads": {
            "target_folder": str(target),
            "periods": [None],
            "filename_template": f"mtracker.{file_format}",
            "file_format": file_format,
        },
        "transformation": {"enabled": transformation},
        "sql_handoff": {"enabled": sql},
    }

def test_download_links_are_swept_across_frames_and_deduped():
    outer = _FilterFrame([
        {"label": "Download CSV", "href": "/files/report.csv", "download_attr": False},
        {"label": "Open Help", "href": "/help", "download_attr": False},  # matched by JS only
    ])
    inner = _FilterFrame([
        {"label": "Download CSV", "href": "/files/report.csv", "download_attr": False},
        {"label": "Download raw data", "href": "", "download_attr": True},
    ])
    outer.page = _FilterFramePage([outer, inner])

    links = flow_worker._asap_discover_download_links(outer)

    labels = [item["label"] for item in links]
    assert labels.count("Download CSV") == 1
    assert "Download raw data" in labels
    raw = next(item for item in links if item["label"] == "Download raw data")
    assert raw["download_attr"] is True


def test_download_link_snapshot_targets_download_semantics():
    source = Path(flow_worker.__file__).read_text()
    assert "csv|xlsx|xls|zip|txt|pdf" in source
    assert '"kind"] = "html_dashboard"' in source
    assert '"stage": "html_dashboard_links"' in source
    # A dashboard with download links but no prompts is not a discovery
    # failure, so the empty-filters warning must account for the links.
    assert "if not discovered and not download_links:" in source


def test_filter_definitions_are_capped_below_server_scan_limits():
    from app.flow_worker import _merge_asap_filter_definition
    from app.routers import flows

    definitions = []
    _merge_asap_filter_definition(
        definitions, "X" * 300, "multi_select",
        [f"Option {i}" for i in range(flow_worker.ASAP_MAX_FILTER_OPTIONS + 500)],
    )
    definition = definitions[0]
    assert len(definition["label"]) == flow_worker.ASAP_MAX_FILTER_LABEL
    assert len(definition["options"]) == flow_worker.ASAP_MAX_FILTER_OPTIONS

    # The capped definition must validate against the server's scan model -
    # one oversized field rejects the whole progress post with a 422.
    flows.DiscoveredFilter(
        filter_key=definition["filter_key"], label=definition["label"],
        control_label=definition["control_label"], control_type=definition["control_type"],
        options=definition["options"], required=False, position=0,
    )
    flows.ScanProgress(status="failed", error="x" * flow_worker.ASAP_MAX_ERROR_CHARS)
    assert flow_worker.ASAP_MAX_REPORT_FILTERS == 200  # DiscoveredReport.filters max_length


def test_api_errors_carry_the_server_validation_detail():
    import httpx

    class Client:
        def request(self, method, path, json=None, timeout=None):
            return httpx.Response(
                422,
                request=httpx.Request("POST", "http://localhost" + path),
                json={"detail": [{"loc": ["body", "reports", 0, "filters", 3, "options"],
                                  "msg": "List should have at most 2000 items"}]},
            )

    with pytest.raises(RuntimeError, match="at most 2000 items"):
        flow_worker._api(Client(), "POST", "/api/flows/worker/w/scans/47/progress", {})


def test_progress_posters_truncate_error_fields():
    source = Path(flow_worker.__file__).read_text()
    assert source.count("error[:ASAP_MAX_ERROR_CHARS] if error else error") == 2
    assert "traceback_text[:100_000]" in source


def test_execute_job_downloads_each_selected_dashboard_link(tmp_path, monkeypatch):
    clicked = []
    monkeypatch.setattr(flow_worker, "_asap_open_report", lambda *_a: object())
    monkeypatch.setattr(flow_worker, "_asap_wait_for_loading_clear", lambda _page: None)
    monkeypatch.setattr(flow_worker, "_asap_frame", lambda _page: object())
    monkeypatch.setattr(flow_worker, "_asap_apply_configuration", lambda *_a: None)
    monkeypatch.setattr(flow_worker, "_has_named_control", lambda _root, _text: False)
    monkeypatch.setattr(
        flow_worker, "_asap_activate_export_view",
        lambda *_a: (_ for _ in ()).throw(AssertionError("dashboards have no export views")),
    )

    def fake_link_download(_page, label, staging, job=None):
        # The job is threaded through so the run can match on the catalogued
        # href when the label alone cannot find the control.
        assert job is not None
        clicked.append(label)
        staged = Path(staging) / f"{label}.csv"
        staged.parent.mkdir(parents=True, exist_ok=True)
        staged.write_text("a,b\n1,2\n")
        return staged

    monkeypatch.setattr(flow_worker, "_asap_download_dashboard_link", fake_link_download)
    monkeypatch.setattr(
        flow_worker, "_store_completed_download",
        lambda staged, output, **_kw: {"file_path": str(output), "filename": output.name},
    )

    job = {
        "site": {"adapter": "asap_portal"},
        "flow": {"name": "Digital shelf pull"},
        "report": {
            "name": "Digital Shelf", "url": "https://portal.example.test",
            "ready_text": None, "open_export_text": None, "download_text": "Download CSV",
            "export_views": [], "download_links": ["Download CSV", "Download raw data"],
            "automation": {"category_path": ["Online", "Digital Shelf", "Digital Shelf"]},
        },
        "downloads": {
            "target_folder": str(tmp_path), "periods": [None],
            "filename_template": "shelf_{export}.csv", "file_format": "csv",
        },
    }
    events = []

    artifacts, _timings = flow_worker.execute_job(
        _WaitPage(), job, lambda _s, progress, _a=None: events.append(progress),
        tmp_path, tmp_path / "staging", run_id=RUN_ID,
    )

    assert clicked == ["Download CSV", "Download raw data"]
    assert [item["export_view"] for item in artifacts] == ["Download CSV", "Download raw data"]
    assert all(item["status"] == "saved" for item in artifacts)
    assert any("from the dashboard" in (item.get("message") or "") for item in events)


def test_dashboard_xlsx_without_downstream_finishes_from_the_raw_workbook(
    tmp_path, monkeypatch,
):
    from openpyxl import Workbook

    target = tmp_path / "target"
    staging = tmp_path / "staging"
    target.mkdir()
    staging.mkdir()
    source = staging / "8f4a2b8e-6c9d-4d69-9ea4-17ad9d2f33e1"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Market", "Units"])
    sheet.append(["Global", 12])
    workbook.save(source)
    _stub_dashboard_execution(monkeypatch)
    monkeypatch.setattr(
        flow_worker,
        "_asap_download_dashboard_link",
        lambda *_args, **_kwargs: source,
    )
    monkeypatch.setattr(
        flow_worker,
        "_normalize_xlsx",
        lambda *_args, **_kwargs: pytest.fail(
            "a dashboard with no transformation or SQL must not normalize XLSX"
        ),
    )
    events = []

    artifacts, _timings = flow_worker.execute_job(
        _WaitPage(),
        _mtracker_dashboard_job(target, file_format="csv"),
        lambda _status, detail, _artifacts=None: events.append(detail),
        tmp_path / "profile",
        staging, run_id=RUN_ID,
    )

    saved = _run_dir(target) / "mtracker.xlsx"
    assert saved.is_file()
    assert artifacts[0]["file_path"] == str(saved)
    assert artifacts[0]["filename"] == saved.name
    assert artifacts[0]["detected_format"] == "xlsx"
    assert artifacts[0]["row_count"] is None
    assert not (_run_dir(target) / "mtracker.csv").exists()
    stages = [event["stage"] for event in events]
    assert "file_transfer" in stages
    assert "file_normalization" not in stages
    assert "file_metadata" not in stages


@pytest.mark.parametrize("downstream", ["transformation", "sql_handoff"])
def test_dashboard_xlsx_normalizes_only_for_configured_downstream_processing(
    downstream, tmp_path, monkeypatch,
):
    from openpyxl import Workbook

    target = tmp_path / "target"
    staging = tmp_path / "staging"
    target.mkdir()
    staging.mkdir()
    source = staging / "1f73a67f-a23b-4e44-9511-5c112edb2370"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Raw data"
    sheet.append(["Market", "Units"])
    sheet.append(["Global", 12])
    workbook.save(source)
    _stub_dashboard_execution(monkeypatch)
    monkeypatch.setattr(
        flow_worker,
        "_asap_download_dashboard_link",
        lambda *_args, **_kwargs: source,
    )
    events = []
    raw_seen_before_normalization = []
    normalized_seen_before_metadata = []

    def progress(_status, detail, _artifacts=None):
        events.append(detail)
        if detail["stage"] == "file_normalization":
            raw_seen_before_normalization.append((_run_dir(target) / "mtracker.xlsx").is_file())
        if detail["stage"] == "file_metadata":
            normalized_seen_before_metadata.append(
                (_run_dir(target) / "mtracker_normalized.csv").is_file()
            )

    job = _mtracker_dashboard_job(
        target,
        transformation=downstream == "transformation",
        sql=downstream == "sql_handoff",
    )
    artifacts, _timings = flow_worker.execute_job(
        _WaitPage(), job, progress, tmp_path / "profile", staging, run_id=RUN_ID,
    )

    normalized = _run_dir(target) / "mtracker_normalized.csv"
    assert (_run_dir(target) / "mtracker.xlsx").is_file()
    assert normalized.is_file()
    assert artifacts[0]["file_path"] == str(normalized)
    assert artifacts[0]["row_count"] == 1
    assert raw_seen_before_normalization == [True]
    assert normalized_seen_before_metadata == [True]
    stages = [event["stage"] for event in events]
    assert stages.index("file_export") < stages.index("file_transfer")
    assert stages.index("file_transfer") < stages.index("file_normalization")
    assert stages.index("file_normalization") < stages.index("file_metadata")


def test_asap_export_view_workbook_still_normalizes_without_downstream(
    tmp_path, monkeypatch,
):
    from openpyxl import Workbook

    target = tmp_path / "target"
    staging = tmp_path / "staging"
    target.mkdir()
    staging.mkdir()
    source = staging / "regular-export"
    workbook = Workbook()
    workbook.active.append(["Market", "Units"])
    workbook.active.append(["Global", 12])
    workbook.save(source)
    _stub_dashboard_execution(monkeypatch)
    monkeypatch.setattr(
        flow_worker,
        "_asap_activate_export_view",
        lambda _page, frame, label: (frame, label),
    )
    monkeypatch.setattr(
        flow_worker,
        "_asap_download_with_retry",
        lambda *_args, **_kwargs: (source, []),
    )
    job = _mtracker_dashboard_job(target)
    job["report"]["download_links"] = []
    job["report"]["export_views"] = ["Export Wizard"]
    job["report"]["automation"] = {
        "export_views": [{"label": "Export Wizard", "filter_keys": []}],
    }

    artifacts, _timings = flow_worker.execute_job(
        _WaitPage(), job, lambda *_args: None, tmp_path / "profile", staging, run_id=RUN_ID,
    )

    assert (_run_dir(target) / "mtracker.xlsx").is_file()
    assert (_run_dir(target) / "mtracker_normalized.csv").is_file()
    assert artifacts[0]["file_path"] == str(_run_dir(target) / "mtracker_normalized.csv")


def test_dashboard_transfer_progress_failure_is_not_redownloaded(
    tmp_path, monkeypatch,
):
    from openpyxl import Workbook

    target = tmp_path / "target"
    staging = tmp_path / "staging"
    target.mkdir()
    staging.mkdir()
    source = staging / "completed-browser-file"
    workbook = Workbook()
    workbook.active.append(["Market", "Units"])
    workbook.active.append(["Global", 12])
    workbook.save(source)
    _stub_dashboard_execution(monkeypatch)
    downloads = []

    def download(*_args, **_kwargs):
        downloads.append(1)
        return source

    monkeypatch.setattr(flow_worker, "_asap_download_dashboard_link", download)
    events = []

    def progress(_status, detail, _artifacts=None):
        events.append(detail)
        if detail["stage"] == "file_transfer":
            raise RuntimeError("progress endpoint unavailable during transfer")

    with pytest.raises(
        flow_worker._CompletedDownloadProcessingError,
        match="ASAP will not be reopened.*progress endpoint unavailable",
    ):
        flow_worker.execute_job(
            _WaitPage(), _mtracker_dashboard_job(target), progress,
            tmp_path / "profile", staging, run_id=RUN_ID,
        )

    assert downloads == [1]
    assert not (_run_dir(target) / "mtracker.xlsx").exists()
    assert not any(event["stage"] == "export_retry" for event in events)


def test_dashboard_completed_download_processing_failure_is_not_redownloaded(
    tmp_path, monkeypatch,
):
    from openpyxl import Workbook

    target = tmp_path / "target"
    staging = tmp_path / "staging"
    target.mkdir()
    staging.mkdir()
    source = staging / "completed-browser-file"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Market", "Units"])
    sheet.append(["Global", 12])
    workbook.save(source)
    _stub_dashboard_execution(monkeypatch)
    downloads = []

    def download(*_args, **_kwargs):
        downloads.append(1)
        return source

    monkeypatch.setattr(flow_worker, "_asap_download_dashboard_link", download)
    events = []

    def progress(_status, detail, _artifacts=None):
        events.append(detail)
        if detail["stage"] == "file_normalization":
            raise RuntimeError("progress endpoint unavailable during normalization")

    with pytest.raises(
        flow_worker._CompletedDownloadProcessingError,
        match="ASAP will not be reopened.*progress endpoint unavailable",
    ):
        flow_worker.execute_job(
            _WaitPage(),
            _mtracker_dashboard_job(target, sql=True),
            progress,
            tmp_path / "profile",
            staging, run_id=RUN_ID,
        )

    assert downloads == [1]
    assert (_run_dir(target) / "mtracker.xlsx").is_file()
    assert not (_run_dir(target) / "mtracker_normalized.csv").exists()
    assert not any(event["stage"] == "export_retry" for event in events)


def test_execute_job_stores_dashboard_csv_without_closing_download_popup(
    tmp_path, monkeypatch,
):
    """A half-detached popup must not block the real configured-folder copy."""
    target = tmp_path / "target"
    staging = tmp_path / "staging"
    target.mkdir()
    staging.mkdir()

    class Context:
        def __init__(self):
            self.pages = []
            self.listeners = []

        def on(self, event, callback):
            self.listeners.append((event, callback))

        def remove_listener(self, event, callback):
            self.listeners.remove((event, callback))

    class Page:
        def __init__(self):
            self.context = Context()
            self.context.pages.append(self)
            self.listeners = []

        def on(self, event, callback):
            self.listeners.append((event, callback))

        def remove_listener(self, event, callback):
            self.listeners.remove((event, callback))

    class Popup:
        def __init__(self):
            self.listeners = []
            self.is_closed_calls = 0
            self.close_calls = 0

        def on(self, event, callback):
            self.listeners.append((event, callback))

        def remove_listener(self, event, callback):
            self.listeners.remove((event, callback))

        def is_closed(self):
            self.is_closed_calls += 1
            return False

        def close(self):
            self.close_calls += 1
            raise AssertionError("a half-detached download popup cannot be closed safely")

    page = Page()
    popup = Popup()

    class Control:
        clicks = 0

        def click(self, **_kwargs):
            self.clicks += 1
            page_callback = next(
                callback for event, callback in page.context.listeners if event == "page"
            )
            page_callback(popup)
            # Playwright's ``allowAndName`` staging artifact is GUID-named and
            # has no suggested-filename suffix after Edge finalizes it.
            (staging / "8f4a2b8e-6c9d-4d69-9ea4-17ad9d2f33e1").write_text(
                "a,b\n1,2\n", encoding="utf-8",
            )

    control = Control()
    monkeypatch.setattr(flow_worker, "_asap_open_report", lambda *_args: object())
    monkeypatch.setattr(flow_worker, "_asap_wait_for_loading_clear", lambda _page: None)
    monkeypatch.setattr(flow_worker, "_asap_frame", lambda _page: object())
    monkeypatch.setattr(flow_worker, "_asap_apply_configuration", lambda *_args: None)
    monkeypatch.setattr(flow_worker, "_has_named_control", lambda *_args: False)
    monkeypatch.setattr(
        flow_worker, "_asap_dashboard_link_locator",
        lambda *_args, **_kwargs: control,
    )

    job = {
        "site": {"adapter": "asap_portal"},
        "flow": {"name": "M Tracker pull"},
        "report": {
            "id": 7,
            "name": "M Tracker",
            "url": "https://portal.example.test/mtracker",
            "ready_text": None,
            "open_export_text": None,
            "download_text": "Download Main Data",
            "filters": [],
            "export_views": [],
            "download_links": ["Download Main Data (csv)"],
            "automation": {"download_links": []},
        },
        "downloads": {
            "target_folder": str(target),
            "periods": [None],
            "filename_template": "mtracker.csv",
            "file_format": "csv",
        },
    }

    artifacts, _timings = flow_worker.execute_job(
        page, job, lambda *_args: None, tmp_path / "profile", staging, run_id=RUN_ID,
    )

    saved = _run_dir(target) / "mtracker.csv"
    assert saved.is_file()
    assert artifacts[0]["file_path"] == str(saved)
    assert artifacts[0]["status"] == "saved"
    assert control.clicks == 1
    assert popup.is_closed_calls == 0
    assert popup.close_calls == 0
    assert popup.listeners == []
    assert page.listeners == []
    assert page.context.listeners == []


def test_gscm_retry_reloads_portal_after_an_empty_favorite_dialog(
    tmp_path, monkeypatch,
):
    """A retry must discard the stale Nexacro dialog, not reuse its empty grid."""
    staged = tmp_path / "gscm.xlsx"
    staged.write_bytes(b"PK\x03\x04workbook")
    order = []

    class Page(_WaitPage):
        stale_dialog = False

    page = Page()

    def open_bookmark(_page, _job, report_progress=None):
        order.append("open")
        if order.count("open") == 1:
            _page.stale_dialog = True
            raise RuntimeError(
                "GSCM's Public bookmark tab stayed empty after three refreshes."
            )
        if _page.stale_dialog:
            raise RuntimeError("GSCM's Setting > Favorite dialog did not open.")
        return "bookmark-row"

    def reload_portal(_page, _job):
        order.append("reload")
        _page.stale_dialog = False

    def edge_download(_page, trigger):
        trigger()
        order.append("download")
        return staged

    monkeypatch.setattr(flow_worker.flow_gscm, "open_bookmark", open_bookmark)
    monkeypatch.setattr(flow_worker.flow_gscm, "reload_portal", reload_portal)
    monkeypatch.setattr(
        flow_worker.flow_gscm, "trigger_excel_export",
        lambda *_args, **_kwargs: order.append("export"),
    )
    monkeypatch.setattr(flow_worker, "_edge_completed_download", edge_download)
    monkeypatch.setattr(
        flow_worker, "_store_completed_download",
        lambda _staged, output, **_kwargs: {
            "file_path": str(output), "filename": output.name,
        },
    )
    job = {
        "site": {
            "adapter": "gscm_portal",
            "base_url": "https://mdscm.sec.samsung.net/",
        },
        "flow": {"name": "GSCM bookmark"},
        "report": {
            "id": 7,
            "name": "Public > SCM > Saved report",
            "url": "https://mdscm.sec.samsung.net/nexa/index.html",
            "export_views": [],
            "automation": {
                "favorite_tab": "Public",
                "favorite_name": "Saved report",
                "favorite_folder_path": ["SCM"],
            },
        },
        "downloads": {
            "target_folder": str(tmp_path),
            "periods": [None],
            "filename_template": "gscm.xlsx",
            "file_format": "xlsx",
        },
    }

    artifacts, _timings = flow_worker.execute_job(
        page, job, lambda *_args, **_kwargs: None, tmp_path, tmp_path / "staging", run_id=RUN_ID,
    )

    assert order == ["open", "reload", "open", "export", "download"]
    assert artifacts[0]["status"] == "saved"


def test_dashboard_link_download_falls_back_to_an_intermediate_page():
    source = Path(flow_worker.__file__).read_text()
    assert "_asap_download_dashboard_link" in source
    # The staging watch catches popup-emitted downloads. Cleanup detaches its
    # listeners without closing a half-detached page before target storage.
    assert "_download_staging_snapshot(staging_dir)" in source
    assert 'context.on("page", _track_page)' in source
    assert 'candidate.on("download", _capture_download)' in source
    assert 'candidate.remove_listener("download", _capture_download)' in source
    assert "_asap_wait_for_dashboard_download_signal" in source
    assert "popup_control.click(timeout=30_000)" in source


def test_dashboard_download_signal_pumps_playwright_until_edge_reports_start(tmp_path):
    downloads = []

    class Page:
        waits = 0

        def wait_for_timeout(self, milliseconds):
            assert milliseconds == 250
            self.waits += 1
            downloads.append(object())

    page = Page()
    signal = flow_worker._asap_wait_for_dashboard_download_signal(
        page, tmp_path, flow_worker._download_staging_snapshot(tmp_path),
        downloads, [], timeout_seconds=1,
    )

    assert signal == "download"
    assert page.waits == 1


def test_dashboard_download_signal_detects_a_file_without_an_edge_event(tmp_path):
    class Page:
        waits = 0

        def wait_for_timeout(self, _milliseconds):
            self.waits += 1
            (tmp_path / "finished.xlsx").write_bytes(b"PK\x03\x04data")

    page = Page()
    signal = flow_worker._asap_wait_for_dashboard_download_signal(
        page, tmp_path, flow_worker._download_staging_snapshot(tmp_path),
        [], [], timeout_seconds=1,
    )

    assert signal == "staging"
    assert page.waits == 1


def test_dashboard_download_signal_exposes_an_intermediate_popup(tmp_path):
    class Page:
        def wait_for_timeout(self, _milliseconds):
            raise AssertionError("a visible popup needs no extra wait when its grace is zero")

    signal = flow_worker._asap_wait_for_dashboard_download_signal(
        Page(), tmp_path, flow_worker._download_staging_snapshot(tmp_path),
        [], [object()], timeout_seconds=1, popup_grace_seconds=0,
    )

    assert signal == "popup"


def test_dashboard_uses_staging_after_edge_event_even_if_folder_signal_won(
    tmp_path, monkeypatch,
):
    """A native event starts the bounded staging handoff; it does not finish it."""
    staged_path = tmp_path / "edge-staged.xlsx"
    staged_path.write_bytes(b"PK\x03\x04staged")
    captured_download = object()
    staged_calls = []

    class Control:
        def click(self, **_kwargs):
            pass

    class Context:
        def __init__(self):
            self.pages = []
            self.listeners = []

        def on(self, event, callback):
            self.listeners.append((event, callback))

        def remove_listener(self, event, callback):
            self.listeners.remove((event, callback))

    class Page:
        def __init__(self):
            self.context = Context()
            self.context.pages.append(self)
            self.listeners = []

        def on(self, event, callback):
            self.listeners.append((event, callback))

        def remove_listener(self, event, callback):
            self.listeners.remove((event, callback))

    page = Page()

    monkeypatch.setattr(
        flow_worker, "_asap_dashboard_link_locator",
        lambda *_args, **_kwargs: Control(),
    )

    def signal(_page, _staging, _before, downloads, _opened, **_kwargs):
        downloads.append(captured_download)
        # The real helper checks staging before callbacks, so both can become
        # observable in the same poll and it can report "staging" here.
        return "staging"

    monkeypatch.setattr(
        flow_worker, "_asap_wait_for_dashboard_download_signal", signal,
    )
    monkeypatch.setattr(
        flow_worker, "_asap_dashboard_event_staged_download",
        lambda staging, before, label: (
            staged_calls.append((staging, before, label)) or staged_path
        ),
    )
    monkeypatch.setattr(
        flow_worker, "_completed_edge_download",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            AssertionError("ASAP dashboards must not block on the Download object")
        ),
    )

    result = flow_worker._asap_download_dashboard_link(
        page, "Download Main Data (xlsx)", tmp_path / "staging",
        {"report": {"automation": {"download_links": []}}},
    )

    assert result == staged_path
    assert len(staged_calls) == 1
    assert staged_calls[0][0] == tmp_path / "staging"
    assert staged_calls[0][2] == "Download Main Data (xlsx)"
    assert page.listeners == []
    assert page.context.listeners == []


def test_dashboard_accepts_download_emitted_while_inspecting_popup(
    tmp_path, monkeypatch,
):
    """A late popup event starts staging before a second control is clicked."""
    staged_path = tmp_path / "popup-staged.xlsx"
    staged_path.write_bytes(b"PK\x03\x04staged")
    captured_download = object()
    staged_calls = []

    class EmptyLocator:
        @property
        def first(self):
            return self

        def count(self):
            return 0

        def is_visible(self):
            return False

    class Popup:
        def __init__(self):
            self.listeners = []
            self.is_closed_calls = 0
            self.close_calls = 0

        def on(self, event, callback):
            self.listeners.append((event, callback))

        def remove_listener(self, event, callback):
            self.listeners.remove((event, callback))

        def is_closed(self):
            self.is_closed_calls += 1
            return False

        def close(self):
            self.close_calls += 1
            raise AssertionError("download popups must not be closed before target storage")

        def wait_for_load_state(self, *_args, **_kwargs):
            callback = next(
                callback for event, callback in self.listeners if event == "download"
            )
            callback(captured_download)

        def get_by_role(self, *_args, **_kwargs):
            return EmptyLocator()

    popup = Popup()

    class Context:
        def __init__(self):
            self.pages = []
            self.listeners = []

        def on(self, event, callback):
            self.listeners.append((event, callback))

        def remove_listener(self, event, callback):
            self.listeners.remove((event, callback))

    class Page:
        def __init__(self):
            self.context = Context()
            self.context.pages.append(self)
            self.listeners = []

        def on(self, event, callback):
            self.listeners.append((event, callback))

        def remove_listener(self, event, callback):
            self.listeners.remove((event, callback))

    page = Page()

    class InitialControl:
        clicks = 0

        def click(self, **_kwargs):
            self.clicks += 1
            callback = next(
                callback for event, callback in page.context.listeners if event == "page"
            )
            callback(popup)

    control = InitialControl()
    monkeypatch.setattr(
        flow_worker, "_asap_dashboard_link_locator",
        lambda target, *_args, **_kwargs: control if target is page else None,
    )
    monkeypatch.setattr(
        flow_worker, "_asap_wait_for_dashboard_download_signal",
        lambda _page, _staging, _before, _downloads, opened, **_kwargs: (
            "popup" if opened else "timeout"
        ),
    )
    monkeypatch.setattr(flow_worker, "_asap_export_action", lambda _popup: None)
    monkeypatch.setattr(
        flow_worker, "_asap_dashboard_event_staged_download",
        lambda staging, before, label: (
            staged_calls.append((staging, before, label)) or staged_path
        ),
    )
    monkeypatch.setattr(
        flow_worker, "_completed_edge_download",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            AssertionError("the late event must finish through staging")
        ),
    )

    result = flow_worker._asap_download_dashboard_link(
        page, "Download Main Data (xlsx)", tmp_path / "staging",
        {"report": {"automation": {"download_links": []}}},
    )

    assert result == staged_path
    assert len(staged_calls) == 1
    assert staged_calls[0][2] == "Download Main Data (xlsx)"
    assert control.clicks == 1
    # The popup is queried once while resolving its intermediate control, but
    # must not be queried again or closed during post-download cleanup.
    assert popup.is_closed_calls == 1
    assert popup.close_calls == 0
    assert popup.listeners == []


def test_dashboard_accepts_download_emitted_at_second_click_timeout_boundary(
    tmp_path, monkeypatch,
):
    """A final-poll event must win even when the signal reports timeout."""
    staged_path = tmp_path / "second-click-staged.xlsx"
    staged_path.write_bytes(b"PK\x03\x04staged")
    captured_download = object()
    staged_calls = []

    class Context:
        def __init__(self):
            self.pages = []
            self.listeners = []

        def on(self, event, callback):
            self.listeners.append((event, callback))

        def remove_listener(self, event, callback):
            self.listeners.remove((event, callback))

    class Page:
        def __init__(self):
            self.context = Context()
            self.context.pages.append(self)
            self.listeners = []

        def on(self, event, callback):
            self.listeners.append((event, callback))

        def remove_listener(self, event, callback):
            self.listeners.remove((event, callback))

    class Popup:
        def __init__(self):
            self.listeners = []
            self.is_closed_calls = 0
            self.close_calls = 0

        def on(self, event, callback):
            self.listeners.append((event, callback))

        def remove_listener(self, event, callback):
            self.listeners.remove((event, callback))

        def is_closed(self):
            self.is_closed_calls += 1
            return False

        def close(self):
            self.close_calls += 1
            raise AssertionError("download popups must not be closed before target storage")

        def wait_for_load_state(self, *_args, **_kwargs):
            pass

    page = Page()
    popup = Popup()

    class InitialControl:
        clicks = 0

        def click(self, **_kwargs):
            self.clicks += 1
            callback = next(
                callback for event, callback in page.context.listeners if event == "page"
            )
            callback(popup)

    class PopupControl:
        clicks = 0

        def click(self, **_kwargs):
            self.clicks += 1

    initial_control = InitialControl()
    popup_control = PopupControl()
    monkeypatch.setattr(
        flow_worker, "_asap_dashboard_link_locator",
        lambda target, *_args, **_kwargs: (
            initial_control if target is page else popup_control
        ),
    )

    signals = []

    def signal(_page, _staging, _before, downloads, _opened, **_kwargs):
        signals.append(1)
        if len(signals) == 1:
            return "popup"
        downloads.append(captured_download)
        return "timeout"

    monkeypatch.setattr(
        flow_worker, "_asap_wait_for_dashboard_download_signal", signal,
    )
    monkeypatch.setattr(
        flow_worker, "_asap_dashboard_event_staged_download",
        lambda staging, before, label: (
            staged_calls.append((staging, before, label)) or staged_path
        ),
    )
    monkeypatch.setattr(
        flow_worker, "_completed_edge_download",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            AssertionError("the second-click event must finish through staging")
        ),
    )

    result = flow_worker._asap_download_dashboard_link(
        page, "Download Main Data (xlsx)", tmp_path / "staging",
        {"report": {"automation": {"download_links": []}}},
    )

    assert result == staged_path
    assert len(signals) == 2
    assert len(staged_calls) == 1
    assert staged_calls[0][2] == "Download Main Data (xlsx)"
    assert initial_control.clicks == 1
    assert popup_control.clicks == 1
    # The popup is queried once while resolving its intermediate control, but
    # must not be queried again or closed during post-download cleanup.
    assert popup.is_closed_calls == 1
    assert popup.close_calls == 0
    assert popup.listeners == []


# --- Resolving a dashboard download control at run time ---


class _DashboardLocator:
    def __init__(self, matches, visible=True):
        self.matches = matches
        self.visible = visible

    @property
    def first(self):
        return self

    def count(self):
        return len(self.matches)

    def nth(self, index):
        return _DashboardLocator([self.matches[index]], self.visible)

    def is_visible(self):
        return self.visible and bool(self.matches)


class _DashboardFrame:
    """A frame whose only usable handle on a control is the marker script.

    This is the shape that broke in production: the portal's role and text
    queries find nothing, because the control's label lives in an attribute
    rather than in rendered text.
    """

    def __init__(self, marked="Download Main Data (xlsx)", inventory=(), appears_after=0):
        self.marked = marked
        self.inventory = list(inventory)
        self.appears_after = appears_after
        self.evaluations = 0

    def get_by_role(self, *_args, **_kwargs):
        return _DashboardLocator([])

    def get_by_text(self, *_args, **_kwargs):
        return _DashboardLocator([])

    def evaluate(self, script, argument=None):
        if "data-metronome-dl" in script:
            self.evaluations += 1
            if self.evaluations <= self.appears_after:
                return None
            return self.marked
        if "out.push(text.slice" in script:
            return list(self.inventory)
        raise AssertionError("unexpected script")

    def locator(self, selector):
        assert selector == "[data-metronome-dl='1']"
        return _DashboardLocator(["control"])


class _DashboardPage:
    def __init__(self, frame):
        self.main_frame = frame
        self.frames = [frame]
        self.waits = 0

    def wait_for_timeout(self, _ms):
        self.waits += 1


def test_dashboard_link_is_found_by_marker_when_text_queries_cannot_match():
    page = _DashboardPage(_DashboardFrame())
    control = flow_worker._asap_dashboard_link_locator(
        page, "Download Main Data (xlsx)", "", timeout_seconds=1,
    )
    assert control is not None


def test_dashboard_link_lookup_waits_for_a_late_rendering_dashboard():
    # The control appears a beat after the frame reports loaded. A single
    # immediate pass is what produced "download link was not visible".
    frame = _DashboardFrame(appears_after=2)
    page = _DashboardPage(frame)
    control = flow_worker._asap_dashboard_link_locator(
        page, "Download Main Data (xlsx)", "", timeout_seconds=30,
    )
    assert control is not None
    assert page.waits >= 1


def test_dashboard_link_failure_lists_the_controls_that_were_visible():
    frame = _DashboardFrame(marked=None, inventory=["Download Detail", "Print"])
    page = _DashboardPage(frame)
    assert flow_worker._asap_dashboard_link_locator(
        page, "Download Main Data (xlsx)", "", timeout_seconds=0,
    ) is None
    assert "'Download Detail'" in flow_worker._asap_dashboard_link_inventory(page)


def test_catalogued_href_is_available_to_the_run():
    job = {"report": {"automation": {"download_links": [
        {"label": "Download Main Data (xlsx)", "href": "/rs/export?id=7"},
        {"label": "Download Detail", "href": "/rs/detail.csv"},
    ]}}}
    assert flow_worker._asap_dashboard_link_href(job, "Download Main Data (xlsx)") == "/rs/export?id=7"
    assert flow_worker._asap_dashboard_link_href(job, "Missing") == ""
    assert flow_worker._asap_dashboard_link_href(None, "Download Detail") == ""


def test_dashboard_download_passes_the_job_so_href_matching_can_work():
    source = Path(flow_worker.__file__).read_text()
    assert "_asap_download_dashboard_link(\n                        page, download_link, staging, job,\n                    )" in source


# --- Run folders and retention operations ---


def test_run_folder_name_formats_run_id_and_date():
    from datetime import date

    assert flow_worker.flow_retention.run_folder_name(57, date(2026, 8, 25)) == "#57_25-08-2026"
    assert flow_worker.flow_retention.RUN_FOLDER_RE.fullmatch("#57_25-08-2026")
    assert flow_worker.flow_retention.RUN_FOLDER_RE.fullmatch("#57_25-08-2026 (2)")
    assert not flow_worker.flow_retention.RUN_FOLDER_RE.fullmatch("#57_2026-08-25")


def test_create_run_folder_writes_marker_and_reuses_its_own_folder(tmp_path):
    retention = flow_worker.flow_retention
    created = retention.create_run_folder(tmp_path, 57, 4)
    assert created.name == retention.run_folder_name(57)
    marker = retention.read_marker(created)
    assert marker["run_id"] == 57 and marker["flow_id"] == 4
    # A re-claimed run re-enters its own folder instead of creating another.
    assert retention.create_run_folder(tmp_path, 57, 4) == created


def test_create_run_folder_never_reuses_a_user_folder_with_the_same_name(tmp_path):
    retention = flow_worker.flow_retention
    user_folder = tmp_path / retention.run_folder_name(57)
    user_folder.mkdir()
    (user_folder / "keep.txt").write_text("user data", encoding="utf-8")
    created = retention.create_run_folder(tmp_path, 57, 4)
    assert created != user_folder
    assert created.name == f"{retention.run_folder_name(57)} (2)"
    assert (user_folder / "keep.txt").read_text(encoding="utf-8") == "user data"
    assert retention.read_marker(user_folder) is None


def _op_for(target, folder_name, op_id=1, source_run_id=54):
    return {
        "op_id": op_id,
        "source_run_id": source_run_id,
        "original_path": str(target / folder_name),
        "tombstone_path": str(target / f".{folder_name}.op{op_id}.deleting"),
    }


def _marked_run_folder(target, run_id, name=None):
    folder = target / (name or flow_worker.flow_retention.run_folder_name(run_id))
    folder.mkdir()
    (folder / flow_worker.flow_retention.MARKER_NAME).write_text(
        json.dumps({"run_id": run_id, "flow_id": 4}), encoding="utf-8",
    )
    (folder / "report.csv").write_text("a,b\n1,2\n", encoding="utf-8")
    return folder


def test_execute_ops_deletes_only_the_marked_recorded_folder(tmp_path):
    retention = flow_worker.flow_retention
    victim = _marked_run_folder(tmp_path, 54)
    decoys = [tmp_path / "Archive", tmp_path / "#old", tmp_path / "#54_2026-08-25"]
    for decoy in decoys:
        decoy.mkdir()
    loose = tmp_path / "report.xlsx"
    loose.write_text("keep", encoding="utf-8")

    results = retention.execute_ops(tmp_path, [_op_for(tmp_path, victim.name)])

    assert [item["outcome"] for item in results] == ["deleted"]
    assert not victim.exists()
    assert all(decoy.is_dir() for decoy in decoys)
    assert loose.is_file()


def test_execute_ops_safety_gate_refuses_everything_not_provably_ours(tmp_path):
    retention = flow_worker.flow_retention
    inner = tmp_path / "target"
    inner.mkdir()

    unmarked = tmp_path / "target" / retention.run_folder_name(50)
    unmarked.mkdir()
    foreign = _marked_run_folder(inner, 51)  # marker says 51, op will say 99
    elsewhere = _marked_run_folder(tmp_path, 52)  # outside the target folder
    badname = inner / "not_a_run_folder"
    badname.mkdir()
    plainfile = inner / retention.run_folder_name(53)
    plainfile.write_text("a file, not a folder", encoding="utf-8")
    linked = inner / retention.run_folder_name(55)
    linked.symlink_to(elsewhere)

    ops = [
        _op_for(inner, unmarked.name, op_id=1, source_run_id=50),
        _op_for(inner, foreign.name, op_id=2, source_run_id=99),
        {"op_id": 3, "source_run_id": 52, "original_path": str(elsewhere),
         "tombstone_path": str(tmp_path / f".{elsewhere.name}.op3.deleting")},
        _op_for(inner, badname.name, op_id=4, source_run_id=54),
        _op_for(inner, plainfile.name, op_id=5, source_run_id=53),
        _op_for(inner, linked.name, op_id=6, source_run_id=55),
    ]
    results = retention.execute_ops(inner, ops)

    assert [item["outcome"] for item in results] == ["skipped"] * 6
    assert unmarked.is_dir() and foreign.is_dir() and elsewhere.is_dir()
    assert badname.is_dir() and plainfile.is_file() and linked.is_symlink()
    # The out-of-target op is refused on its tombstone location too.
    assert "tombstone" in results[2]["detail"] or "target folder" in results[2]["detail"]


def test_execute_ops_skips_junctions(tmp_path, monkeypatch):
    retention = flow_worker.flow_retention
    victim = _marked_run_folder(tmp_path, 54)
    monkeypatch.setattr(retention, "_is_junction", lambda path: path == victim)
    results = retention.execute_ops(tmp_path, [_op_for(tmp_path, victim.name)])
    assert [item["outcome"] for item in results] == ["skipped"]
    assert victim.is_dir()


def test_execute_ops_reconciles_after_a_crash_between_rename_and_delete(tmp_path):
    retention = flow_worker.flow_retention
    op = _op_for(tmp_path, retention.run_folder_name(54))
    tombstone = Path(op["tombstone_path"])
    tombstone.mkdir()
    (tombstone / "leftover.csv").write_text("half-deleted", encoding="utf-8")

    results = retention.execute_ops(tmp_path, [op])

    assert [item["outcome"] for item in results] == ["deleted"]
    assert not tombstone.exists()


def test_execute_ops_reports_deleted_when_both_paths_are_already_gone(tmp_path):
    retention = flow_worker.flow_retention
    results = retention.execute_ops(tmp_path, [_op_for(tmp_path, retention.run_folder_name(54))])
    assert [item["outcome"] for item in results] == ["deleted"]


def test_execute_ops_leaves_the_folder_intact_when_rename_fails(tmp_path, monkeypatch):
    retention = flow_worker.flow_retention
    victim = _marked_run_folder(tmp_path, 54)

    def refuse(_self, _target):
        raise PermissionError("file is open in Excel")

    monkeypatch.setattr(Path, "rename", refuse)
    results = retention.execute_ops(tmp_path, [_op_for(tmp_path, victim.name)])
    assert [item["outcome"] for item in results] == ["failed"]
    assert victim.is_dir() and (victim / "report.csv").is_file()


def test_execute_ops_quarantines_when_delete_fails_after_rename(tmp_path, monkeypatch):
    retention = flow_worker.flow_retention
    victim = _marked_run_folder(tmp_path, 54)
    op = _op_for(tmp_path, victim.name)

    def refuse(_path):
        raise PermissionError("file is open in Excel")

    monkeypatch.setattr(flow_worker.flow_retention.shutil, "rmtree", refuse)
    results = retention.execute_ops(tmp_path, [op])
    assert [item["outcome"] for item in results] == ["quarantined"]
    assert not victim.exists()
    assert Path(op["tombstone_path"]).is_dir()

    # The next assigned run reconciles the DB-known tombstone and finishes.
    monkeypatch.undo()
    retry = flow_worker.flow_retention.execute_ops(tmp_path, [op])
    assert [item["outcome"] for item in retry] == ["deleted"]
    assert not Path(op["tombstone_path"]).exists()


def test_execute_job_places_files_in_a_run_folder_and_prunes_assigned_ops(
    tmp_path, monkeypatch,
):
    target = tmp_path / "target"
    staging = tmp_path / "staging"
    target.mkdir()
    staging.mkdir()
    old = _marked_run_folder(target, 53)
    survivor = _marked_run_folder(target, 55)
    decoy = target / "Keep Me"
    decoy.mkdir()
    _stub_dashboard_execution(monkeypatch)
    monkeypatch.setattr(
        flow_worker, "_asap_download_dashboard_link",
        lambda _page, _link, staging_dir, _job: (
            staging_dir.mkdir(parents=True, exist_ok=True) or None,
            (staging_dir / "mtracker.csv").write_text("a,b\n1,2\n", encoding="utf-8"),
        ) and (staging_dir / "mtracker.csv"),
    )
    registered = []

    def register_folder(folder):
        registered.append(folder)
        return {"ops": [_op_for(target, old.name, op_id=9, source_run_id=53)]}

    events = []
    artifacts, _timings = flow_worker.execute_job(
        _WaitPage(), _mtracker_dashboard_job(target), lambda _s, detail, *rest: events.append(detail),
        tmp_path / "profile", staging, run_id=RUN_ID, register_folder=register_folder,
    )

    run_dir = _run_dir(target)
    assert registered == [str(run_dir)]
    assert [Path(item["file_path"]).parent for item in artifacts] == [run_dir]
    assert not old.exists()
    assert survivor.is_dir() and decoy.is_dir()
    stages = [event.get("stage") for event in events]
    assert "run_folder" in stages and "run_folder_retention" in stages
    retention_event = next(e for e in events if e.get("stage") == "run_folder_retention")
    assert retention_event["retention_results"] == [
        {"op_id": 9, "outcome": "deleted", "detail": ""},
    ]
