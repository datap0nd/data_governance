"""Authenticated browser worker for Metronome Flows.

Run this under the Windows user that is authorized for the configured website.
Every run first saves immutable artifacts into an owned run folder. Flows may
either keep those folders visible in the configured target or publish a
validated deliverable bundle directly into that target with exact-name
replacement. The only recursive deletion is a server-assigned retention
operation against an owned run folder (keeping the newest three per store).
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import os
import re
import socket
import subprocess
import sys
import threading
import time
import traceback
import zipfile
from contextlib import contextmanager
from datetime import date, datetime, timedelta
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Callable
from urllib.parse import urlsplit

import httpx
from playwright.sync_api import Error as PlaywrightError
from playwright.sync_api import Frame, Page, TimeoutError as PlaywrightTimeoutError, sync_playwright

# Keep direct-file execution compatible with the isolated Windows embedded
# runtime as well as the preferred ``python -m app.flow_worker`` launcher.
_CODE_DIR = Path(__file__).resolve().parent.parent
if str(_CODE_DIR) not in sys.path:
    sys.path.insert(0, str(_CODE_DIR))

from app.flow_paths import assert_job_paths
from app import flow_layout

try:
    from app import flow_gscm, flow_outlook, flow_publish, flow_replay, flow_retention
    from app.flow_asap_exports import (
        ASAP_DOWNLOAD_TYPES,
        ASAP_EXPORT_CHECKBOXES,
        resolve_asap_download_type,
    )
    from app.flow_credentials import load_asap_credentials
    from app.flow_download_security import looks_like_sign_in
except ModuleNotFoundError:  # setup.ps1 also invokes this file directly
    import flow_gscm
    import flow_outlook
    import flow_publish
    import flow_replay
    import flow_retention
    from flow_asap_exports import (
        ASAP_DOWNLOAD_TYPES,
        ASAP_EXPORT_CHECKBOXES,
        resolve_asap_download_type,
    )
    from flow_credentials import load_asap_credentials
    from flow_download_security import looks_like_sign_in


ASAP_FRAME_SELECTOR = "iframe#content-frame"
ASAP_PORTAL_ADAPTER = "asap_portal"
GSCM_PORTAL_ADAPTER = flow_gscm.GSCM_PORTAL_ADAPTER
OUTLOOK_ATTACHMENT_ADAPTER = flow_outlook.OUTLOOK_ATTACHMENT_ADAPTER
LOCAL_FILE_ADAPTER = "local_file"
AUTH_MARKER = ".asap_authenticated"
GSCM_AUTH_MARKER = ".gscm_authenticated"
ASAP_LOADING_OVERLAY_SELECTOR = (
    "#loading-spinner-container, .loading-spinner-container, .loading-overlay"
)
# A wide report over a long period can render for a long time. Wait it out.
ASAP_REPORT_RESULT_TIMEOUT_MS = 30 * 60 * 1_000
EXPORT_TASK_ATTEMPTS = 3
GSCM_EXPORT_TASK_ATTEMPTS = 2
GSCM_INITIAL_LOAD_BUFFER_MS = 60_000
GSCM_RETRY_LOAD_BUFFER_MS = 120_000
# Server-side scan payload limits (DiscoveredFilter / DiscoveredReport /
# ScanProgress in app.routers.flows). A single oversized field rejects the
# whole progress post with an opaque 422, so the worker caps everything it
# sends below those limits.
ASAP_MAX_FILTER_OPTIONS = 3_000
ASAP_MAX_FILTER_LABEL = 200
ASAP_MAX_REPORT_FILTERS = 200
ASAP_MAX_ERROR_CHARS = 10_000
ASAP_MAX_DOWNLOAD_LINKS = 50
# Progress messages land in flow_run_events and are rendered verbatim by the
# run-log page; an unbounded screen dump once produced a 100,000-character
# entry nobody could read.
PROGRESS_MESSAGE_MAX_CHARS = 4_000
# How many failure screenshots to keep in <profile>/diagnostics.
FAILURE_SCREENSHOT_KEEP = 20
# An embedded dashboard renders after its frame reports loaded, so its
# download controls appear later than the report navigation completes.
ASAP_DASHBOARD_LINK_TIMEOUT_SECONDS = 120
ASAP_EMPTY_RESULT_DETAIL = (
    "The loading overlay cleared, but neither a Data rows marker nor a populated raw table appeared."
)
XLSX_HEADER_LABEL_HINTS = frozenset({
    "sell_out_region",
    "sell_out_subsidiary",
    "sell_out_country",
    "country_code",
    "operator",
    "province",
    "latitude",
    "longitude",
    "category",
    "biz_sub",
    "series",
    "mkt_name",
    "item",
})
REQUESTED_WEEK = re.compile(r"^(20\d{2})-W(\d{2})$", re.IGNORECASE)
OOXML_EXCEL_EXTENSIONS = frozenset({".xlsx", ".xlsm", ".xltx", ".xltm"})
LEGACY_EXCEL_EXTENSIONS = frozenset({".xls", ".xlt"})
XLSB_EXCEL_EXTENSIONS = frozenset({".xlsb"})
EXCEL_EXTENSIONS_BY_FORMAT = {
    "xlsx": OOXML_EXCEL_EXTENSIONS,
    "xls": LEGACY_EXCEL_EXTENSIONS,
    "xlsb": XLSB_EXCEL_EXTENSIONS,
}
EXCEL_DOWNLOAD_FORMATS = frozenset(EXCEL_EXTENSIONS_BY_FORMAT)
LOCAL_FILE_FORMAT_BY_EXTENSION = {
    ".csv": "csv",
    **{suffix: "xls" for suffix in LEGACY_EXCEL_EXTENSIONS},
    **{suffix: "xlsb" for suffix in XLSB_EXCEL_EXTENSIONS},
    **{suffix: "xlsx" for suffix in OOXML_EXCEL_EXTENSIONS},
}


class _CompletedDownloadProcessingError(RuntimeError):
    """A native browser download finished, so report navigation must not retry."""


WORKER_LOCK_FILENAME = ".worker.lock"
# How long a starting worker waits for a stopping predecessor to release the
# profile lock before treating the holder as an orphan and exiting nonzero.
LOCK_RETRY_SECONDS = 60


@contextmanager
def _exclusive_worker_lock(profile_dir: Path):
    """Prevent the service and login task from running duplicate workers."""
    profile_dir.mkdir(parents=True, exist_ok=True)
    lock_path = profile_dir / WORKER_LOCK_FILENAME
    handle = lock_path.open("a+b")
    handle.seek(0, os.SEEK_END)
    if handle.tell() == 0:
        handle.write(b"0")
        handle.flush()
    handle.seek(0)
    acquired = False
    try:
        if os.name == "nt":
            import msvcrt
            msvcrt.locking(handle.fileno(), msvcrt.LK_NBLCK, 1)
        else:
            import fcntl
            fcntl.flock(handle.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
        acquired = True
    except OSError:
        pass
    try:
        yield acquired
    finally:
        if acquired:
            handle.seek(0)
            if os.name == "nt":
                import msvcrt
                msvcrt.locking(handle.fileno(), msvcrt.LK_UNLCK, 1)
            else:
                import fcntl
                fcntl.flock(handle.fileno(), fcntl.LOCK_UN)
        handle.close()


class _Timings:
    def __init__(self):
        self.started = time.perf_counter()
        self.items: list[dict[str, Any]] = []

    def measure(self, phase: str, *, report_id: int | None = None, item_count: int | None = None):
        timings = self
        class Measurement:
            def __enter__(self):
                self.started = time.perf_counter()
                return self

            def __exit__(self, exc_type, exc, traceback):
                item = {
                    "phase": phase,
                    "duration_ms": round((time.perf_counter() - self.started) * 1000),
                    "status": "failed" if exc_type else "succeeded",
                }
                if report_id is not None:
                    item["report_id"] = report_id
                if item_count is not None:
                    item["item_count"] = item_count
                timings.items.append(item)
        return Measurement()

    def finish(self, *, item_count: int | None = None, status: str = "succeeded") -> list[dict[str, Any]]:
        total = {
            "phase": "total", "duration_ms": round((time.perf_counter() - self.started) * 1000),
            "status": status,
        }
        if item_count is not None:
            total["item_count"] = item_count
        return [*self.items, total]


def _api(client: httpx.Client, method: str, path: str, body: dict | None = None) -> dict:
    # Progress calls are part of the execution record, so a short-lived SQLite
    # write collision or local service restart must not kill the browser run.
    # Terminal updates are idempotent on the server and are therefore safe to
    # retry when the local API returns a transient 5xx response.
    last_error: Exception | None = None
    for attempt in range(1, 6):
        try:
            response = client.request(method, path, json=body, timeout=60)
            response.raise_for_status()
            return response.json()
        except (httpx.TransportError, httpx.HTTPStatusError) as exc:
            last_error = exc
            retryable = isinstance(exc, httpx.TransportError) or (
                exc.response is not None and exc.response.status_code >= 500
            )
            if not retryable or attempt == 5:
                # A 4xx body names the exact field the server rejected.
                # Without it a validation failure surfaces as an opaque
                # "422 Unprocessable Content" with no way to diagnose.
                detail = ""
                if isinstance(exc, httpx.HTTPStatusError) and exc.response is not None:
                    detail = _clean_text(exc.response.text)[:500]
                if detail:
                    raise RuntimeError(f"{exc} Server detail: {detail}") from exc
                raise
            time.sleep(attempt)
    raise RuntimeError("Local API request failed after retries.") from last_error


def _safe_output_path(folder: Path, filename: str) -> Path:
    """Return an unused output path without deleting or overwriting anything."""
    candidate = folder / filename
    if not candidate.exists():
        return candidate
    stem, suffix = candidate.stem, candidate.suffix
    for index in range(2, 10000):
        candidate = folder / f"{stem} ({index}){suffix}"
        if not candidate.exists():
            return candidate
    raise RuntimeError("Could not find an unused filename after 9,999 attempts.")


def _render_filename(
    template: str, job: dict, period: str | list[str] | None, index: int,
    export_view: str | None = None,
) -> str:
    raw_week = period or job["downloads"].get("periods", [None])[0] or ""
    if isinstance(raw_week, list):
        start = raw_week[0] if raw_week else ""
        end = raw_week[-1] if raw_week else ""
        week = start if start == end else f"{start}_{end}"
        year = start.split("-W", 1)[0] if start else ""
        start_number = start.split("-W", 1)[1] if "-W" in start else ""
        end_number = end.split("-W", 1)[1] if "-W" in end else ""
        week_number = start_number if start_number == end_number else f"{start_number}-{end_number}"
    else:
        week = raw_week
        start = end = raw_week
        year, week_number = (week.split("-W", 1) + [""])[:2] if week else ("", "")
    def short_period(value: str) -> str:
        return f"W{value.split('-W', 1)[1]}" if "-W" in value else value
    values = {
        "flow": job["flow"]["name"],
        "report": job["report"]["name"],
        "export": re.sub(r"[^A-Za-z0-9]+", "_", export_view or "export").strip("_").casefold(),
        "week": week,
        "start_period": short_period(start),
        "end_period": short_period(end),
        "year": year,
        "week_number": week_number,
        "index": str(index),
        "date": job.get('_runtime_task_date') or date.today().isoformat(),
    }
    name = template
    for key, value in values.items():
        name = name.replace("{" + key + "}", str(value))
    name = re.sub(r"[<>:\"/\\|?*\x00-\x1f\s]+", "_", name).strip(" ._")
    downloads = job.get("downloads") or {}
    semantic_type = downloads.get("asap_download_type")
    if semantic_type:
        spec = resolve_asap_download_type(
            semantic_type, legacy_file_format=downloads.get("file_format"),
        )
        if not name.casefold().endswith(spec.compatible_suffixes):
            name += spec.preferred_suffix
    else:
        expected = f".{downloads.get('file_format') or 'csv'}"
        if not name.casefold().endswith(expected):
            name += expected
    return name


def _click_named(page: Page | Frame, text: str):
    candidates = [
        page.get_by_role("button", name=text, exact=True),
        page.get_by_role("link", name=text, exact=True),
        page.get_by_text(text, exact=True),
    ]
    for locator in candidates:
        try:
            if locator.count() and locator.first.is_visible():
                locator.first.click()
                return
        except Exception:
            continue
    raise RuntimeError(f"Could not find visible control: {text}")


def _select_native_options_by_text(
    page: Page | Frame, values: list[str], expected_options: list[str] | None = None,
) -> bool:
    """Select an enhanced native control by its option labels.

    ASAP uses Select2 for Data Configuration. Its visible widget has no
    accessible name, while the owning ``select`` and ``option`` elements are
    hidden. Playwright can still use ``select_option`` on that native control,
    which also emits the change event Select2 and the report listen for.
    """
    expected = [str(item) for item in (expected_options or values)]
    selects = page.locator("select")
    for index in range(selects.count()):
        select = selects.nth(index)
        labels = [re.sub(r"\s+", " ", text).strip() for text in select.locator("option").all_text_contents()]
        if not all(item in labels for item in expected):
            continue
        selected = values if len(values) > 1 else values[0]
        select.select_option(label=selected, force=True)
        actual = [
            re.sub(r"\s+", " ", text).strip()
            for text in select.locator("option:checked").all_text_contents()
        ]
        if set(actual) != set(values):
            raise RuntimeError(
                f"ASAP native selection mismatch. Requested: {values}. Selected: {actual}."
            )
        return True
    return False


def _set_filter(page: Page | Frame, definition: dict, value: Any):
    if value in (None, "", []):
        return
    label = definition["control_label"]
    automation = definition.get("automation") or {}
    locator = page.locator(automation["locator"]) if automation.get("locator") else page.get_by_label(label, exact=True)
    control_type = definition["control_type"]
    values = value if isinstance(value, list) else [value]
    values = [str(item) for item in values]
    try:
        if control_type == "multi_select":
            locator.select_option(values)
        elif control_type == "select":
            locator.select_option(values[0])
        else:
            locator.fill(values[0])
        return
    except Exception:
        pass

    if control_type in {"select", "multi_select"}:
        try:
            if _select_native_options_by_text(page, values, definition.get("options")):
                return
        except Exception:
            pass

    # Enterprise report portals commonly render comboboxes and tree selectors
    # instead of native <select> elements. Keep this semantic and text based.
    trigger = page.get_by_role("combobox", name=label, exact=True)
    if not trigger.count():
        trigger = page.get_by_text(label, exact=True).first
    if not trigger.count():
        raise RuntimeError(f"Could not find report filter: {label}")
    trigger.click()
    for index, selected in enumerate(values):
        option = page.get_by_role("option", name=selected, exact=True)
        if not option.count():
            option = page.get_by_text(selected, exact=True)
        if not option.count():
            raise RuntimeError(f"Could not find {label} option: {selected}")
        option.first.click()
        if control_type == "multi_select" and index < len(values) - 1:
            trigger.click()


def _apply_configuration(page: Page, job: dict, period: str | list[str] | None):
    selections = dict(job.get("selections") or {})
    for definition in job["report"].get("filters", []):
        key = definition["filter_key"]
        value = period if definition["control_type"] == "week" and period else selections.get(key)
        _set_filter(page, definition, value)


def _week_to_asap(value: str) -> str:
    match = re.fullmatch(r"(\d{4})-W(\d{2})", value)
    if not match:
        raise RuntimeError(f"ASAP week must use YYYY-Www: {value}")
    return "".join(match.groups())


def _asap_week_dates(value: str) -> tuple[str, str]:
    """Return ASAP's Sunday-to-Saturday dates for one ISO-numbered week."""
    match = re.fullmatch(r"(\d{4})-W(\d{2})", value)
    if not match:
        raise RuntimeError(f"ASAP week must use YYYY-Www: {value}")
    try:
        monday = date.fromisocalendar(int(match.group(1)), int(match.group(2)), 1)
    except ValueError as exc:
        raise RuntimeError(f"ASAP week does not exist: {value}") from exc
    sunday = monday - timedelta(days=1)
    saturday = sunday + timedelta(days=6)
    return sunday.strftime("%Y%m%d"), saturday.strftime("%Y%m%d")


def _asap_range_scope(frame: Frame, label: str):
    """Return the smallest visible prompt containing a two-handle range control."""
    labels = frame.get_by_text(re.compile(rf"^{re.escape(label)}:?$", re.I))
    for label_index in range(labels.count()):
        prompt = labels.nth(label_index)
        try:
            if not prompt.is_visible():
                continue
        except Exception:
            continue
        ancestor = prompt
        for _depth in range(8):
            ancestor = ancestor.locator("xpath=parent::*")
            if not ancestor.count():
                break
            handles = ancestor.locator("[role=slider],input[type=range]")
            visible = []
            for handle_index in range(handles.count()):
                handle = handles.nth(handle_index)
                try:
                    if handle.is_visible():
                        visible.append(handle)
                except Exception:
                    continue
            if len(visible) == 2:
                return ancestor, visible
    raise RuntimeError(f"ASAP {label} range slider was not found.")


def _asap_slider_value(handle, value_pattern: str) -> str | None:
    """Read one semantic slider value without trusting its screen position."""
    for attribute in (
        "aria-valuetext", "aria-valuenow", "value", "data-value", "data-val",
        "data-current-value",
    ):
        try:
            raw = _clean_text(handle.get_attribute(attribute))
        except Exception:
            raw = ""
        match = re.search(value_pattern, raw)
        if match:
            return match.group(0)
    try:
        match = re.search(value_pattern, _clean_text(handle.inner_text()))
    except Exception:
        match = None
    return match.group(0) if match else None


def _asap_slider_ordinal(value: str, kind: str) -> int:
    if kind == "week":
        match = re.fullmatch(r"(\d{4})(\d{2})", value)
        if not match:
            raise RuntimeError(f"ASAP Week slider exposed an invalid value: {value}")
        try:
            return date.fromisocalendar(int(match.group(1)), int(match.group(2)), 1).toordinal() // 7
        except ValueError as exc:
            raise RuntimeError(f"ASAP Week slider exposed a nonexistent week: {value}") from exc
    if kind == "date":
        try:
            return datetime.strptime(value, "%Y%m%d").date().toordinal()
        except ValueError as exc:
            raise RuntimeError(f"ASAP Date slider exposed an invalid value: {value}") from exc
    raise RuntimeError(f"Unsupported ASAP slider kind: {kind}")


def _asap_move_slider(handle, target: str, kind: str, *, current: str | None = None,
                      read_value=None) -> None:
    """Move one focused slider handle by keyboard and verify its exact value."""
    pattern = r"20\d{4}" if kind == "week" else r"20\d{6}"
    read_value = read_value or (lambda: _asap_slider_value(handle, pattern))
    current = current or read_value()
    if current is None:
        handle.press("Home")
        current = read_value()
    if current is None:
        raise RuntimeError(f"ASAP {kind.title()} slider did not expose its current value.")
    delta = _asap_slider_ordinal(target, kind) - _asap_slider_ordinal(current, kind)
    if abs(delta) > 1_000:
        raise RuntimeError(
            f"ASAP {kind.title()} slider target is more than 1,000 steps from its current value."
        )
    key = "ArrowRight" if delta > 0 else "ArrowLeft"
    for _step in range(abs(delta)):
        handle.press(key)
    actual = None
    for _attempt in range(10):
        actual = read_value()
        if actual == target:
            break
        time.sleep(0.05)
    if actual != target:
        raise RuntimeError(
            f"ASAP {kind.title()} slider mismatch. Requested: {target}. Selected: {actual or 'unknown'}."
        )


def _asap_range_values(scope, handles: list, kind: str) -> list[str | None]:
    """Read both range values from handle semantics or the control's visible labels."""
    pattern = r"20\d{4}" if kind == "week" else r"20\d{6}"
    values = [_asap_slider_value(handle, pattern) for handle in handles]
    if None not in values:
        return values
    try:
        visible = re.findall(rf"(?<!\d){pattern}(?!\d)", _clean_text(scope.inner_text()))
    except Exception:
        visible = []
    if len(visible) == len(handles):
        return [value or visible[index] for index, value in enumerate(values)]
    return values


def _asap_set_range(frame: Frame, label: str, start: str, end: str, kind: str) -> None:
    """Set and read back an ASAP two-handle range without coordinate guessing."""
    if _asap_slider_ordinal(end, kind) < _asap_slider_ordinal(start, kind):
        raise RuntimeError(f"ASAP {label} range ends before it starts: {start} to {end}")
    scope, handles = _asap_range_scope(frame, label)
    current = _asap_range_values(scope, handles, kind)
    # When advancing a collapsed one-period range, the upper handle must move
    # first or the lower handle is constrained by the old upper value.
    if current[1] and _asap_slider_ordinal(start, kind) > _asap_slider_ordinal(current[1], kind):
        order = ((1, end), (0, start))
    elif current[0] and _asap_slider_ordinal(end, kind) < _asap_slider_ordinal(current[0], kind):
        order = ((0, start), (1, end))
    else:
        order = ((0, start), (1, end))
    for index, target in order:
        _asap_move_slider(
            handles[index], target, kind, current=current[index],
            read_value=lambda index=index: _asap_range_values(scope, handles, kind)[index],
        )
        current = _asap_range_values(scope, handles, kind)
    actual = _asap_range_values(scope, handles, kind)
    if actual != [start, end]:
        raise RuntimeError(
            f"ASAP {label} range did not match the flow. Requested: {[start, end]}. Selected: {actual}."
        )


def _asap_week_options(start: str, end: str) -> list[str]:
    """Expand compact ASAP week bounds into every valid YYYYWW value."""
    start_match = re.fullmatch(r"(\d{4})(\d{2})", start)
    end_match = re.fullmatch(r"(\d{4})(\d{2})", end)
    if not start_match or not end_match:
        return []
    try:
        current = date.fromisocalendar(int(start_match.group(1)), int(start_match.group(2)), 1)
        final = date.fromisocalendar(int(end_match.group(1)), int(end_match.group(2)), 1)
    except ValueError:
        return []
    if final < current or (final - current).days > 7 * 104:
        return []
    values = []
    while current <= final:
        year, week, _weekday = current.isocalendar()
        values.append(f"{year:04d}{week:02d}")
        current += timedelta(days=7)
    return values


def _asap_discover_labeled_week_slider(
    frame: Frame, label: str, *, date_range_label: str | None = None,
) -> tuple[list[str], dict] | None:
    """Read a complete YYYYWW range and restore its original handles."""
    try:
        scope, handles = _asap_range_scope(frame, label)
    except RuntimeError:
        return None
    original = _asap_range_values(scope, handles, "week")
    if None in original:
        return None
    handles[0].press("Home")
    minimum = _asap_range_values(scope, handles, "week")[0]
    _asap_move_slider(
        handles[0], original[0], "week", current=minimum,
        read_value=lambda: _asap_range_values(scope, handles, "week")[0],
    )
    handles[1].press("End")
    maximum = _asap_range_values(scope, handles, "week")[1]
    _asap_move_slider(
        handles[1], original[1], "week", current=maximum,
        read_value=lambda: _asap_range_values(scope, handles, "week")[1],
    )
    if _asap_range_values(scope, handles, "week") != original:
        raise RuntimeError(f"ASAP {label} slider could not be restored after discovery.")
    options = _asap_week_options(minimum or "", maximum or "")
    if not options:
        return None
    automation = {"kind": "range_slider"}
    if date_range_label:
        automation["date_range_label"] = date_range_label
    return options, automation


def _asap_discover_week_slider(frame: Frame) -> tuple[list[str], dict] | None:
    """Read the complete Week range and its coupled Date range behavior."""
    return _asap_discover_labeled_week_slider(
        frame, "Week", date_range_label="Date",
    )


def _asap_discover_period_slider(frame: Frame) -> tuple[list[str], dict] | None:
    """Read reports whose visible range is titled Period rather than Week."""
    return _asap_discover_labeled_week_slider(frame, "Period")


def _asap_frame(page: Page) -> Frame:
    handle = page.locator(ASAP_FRAME_SELECTOR)
    handle.wait_for(state="attached", timeout=120_000)
    # ASAP replaces the report iframe during navigation. For a short interval
    # both the detached predecessor and its replacement can remain in the DOM.
    # Always choose the newest live frame so discovery cannot inspect the
    # report that was open before the requested menu item was clicked.
    for element in reversed(handle.element_handles()):
        frame = element.content_frame()
        if frame is not None and not frame.is_detached():
            return frame
    raise RuntimeError("ASAP report frame did not become available.")


def _asap_first_visible(locator):
    for item in locator.all():
        try:
            if item.is_visible():
                return item
        except Exception:
            continue
    return None


def _asap_last_visible(locator):
    """Return the last visible match when a menu group and leaf share a label."""
    for item in reversed(locator.all()):
        try:
            if item.is_visible():
                return item
        except Exception:
            continue
    return None


def _asap_loading_overlay_visible(page: Page) -> bool:
    """Return whether ASAP is currently blocking report interaction.

    ASAP leaves report controls visible beneath its loading overlay. Playwright
    therefore considers the control actionable until the overlay intercepts
    the actual pointer event. Inspect every live frame because the overlay can
    move with MicroStrategy's replacement iframe.
    """
    roots = list(dict.fromkeys([page.main_frame, *page.frames]))
    for root in roots:
        try:
            overlays = root.locator(ASAP_LOADING_OVERLAY_SELECTOR)
            for index in range(overlays.count()):
                if overlays.nth(index).is_visible():
                    return True
        except Exception:
            continue
    return False


def _asap_wait_for_loading_clear(page: Page, timeout_ms: int = ASAP_REPORT_RESULT_TIMEOUT_MS):
    """Wait for a sustained clear state before clicking an ASAP control."""
    deadline = time.monotonic() + (timeout_ms / 1_000)
    clear_polls = 0
    while time.monotonic() < deadline:
        if _asap_loading_overlay_visible(page):
            clear_polls = 0
        else:
            clear_polls += 1
            # A single clear DOM sample can occur between replacement frames.
            # Require one full second without the blocking overlay.
            if clear_polls >= 4:
                return
        page.wait_for_timeout(250)
    raise RuntimeError(
        f"ASAP loading overlay did not clear within {timeout_ms // 1_000} seconds."
    )


def _asap_wait_for_visible(
    page: Page, *locators, timeout_ms: int = 15_000, prefer_last: bool = False,
):
    """Wait for a portal menu item that may appear after hover animation."""
    visible_match = _asap_last_visible if prefer_last else _asap_first_visible
    deadline = time.monotonic() + (timeout_ms / 1_000)
    while time.monotonic() < deadline:
        for locator in locators:
            item = visible_match(locator)
            if item is not None:
                return item
        page.wait_for_timeout(100)
    return None


def _asap_frame_signature(frame: Frame | None) -> str | None:
    if frame is None or frame.is_detached():
        return None
    try:
        text = _clean_text(frame.locator("body").inner_text(timeout=2_000))
    except Exception:
        return None
    if not text:
        return None
    return hashlib.sha256(text.encode("utf-8", errors="replace")).hexdigest()


def _asap_text_visible_across_frames(page: Page, label: str) -> bool:
    """Find an exact ASAP breadcrumb in either the portal shell or report frame."""
    roots = [getattr(page, "main_frame", page), *getattr(page, "frames", [])]
    for root in list(dict.fromkeys(roots)):
        try:
            if _asap_first_visible(root.get_by_text(label, exact=True)) is not None:
                return True
        except Exception:
            continue
    return False


def _asap_wait_for_report_navigation(
    page: Page, previous_frame: Frame | None, target_control, path: list[str],
    previous_signature: str | None = None, timeout_ms: int = 120_000,
) -> Frame:
    """Wait until the clicked menu item closes and the requested report becomes stable."""
    try:
        target_control.wait_for(state="hidden", timeout=min(timeout_ms, 30_000))
    except PlaywrightTimeoutError as exc:
        raise RuntimeError(
            f"ASAP did not leave the menu item for the requested report: {path[-1]}"
        ) from exc

    deadline = time.monotonic() + (timeout_ms / 1_000)
    stable_signature = None
    stable_polls = 0
    last_path_visibility: dict[str, bool] = {}
    last_frame_replaced = False
    last_content_changed = False
    while time.monotonic() < deadline:
        try:
            current = _asap_frame(page)
            frame_replaced = previous_frame is None or current is not previous_frame
            if previous_frame is not None:
                frame_replaced = frame_replaced or previous_frame.is_detached()
            visible_path = True
            last_path_visibility = {}
            for label in path[-2:]:
                label_visible = _asap_text_visible_across_frames(page, label)
                last_path_visibility[label] = label_visible
                if not label_visible:
                    visible_path = False
            current_signature = _asap_frame_signature(current)
            content_changed = bool(
                previous_signature and current_signature
                and current_signature != previous_signature
            )
            last_frame_replaced = frame_replaced
            last_content_changed = content_changed
            if visible_path:
                # Reopening the report already preserved in ASAP can reuse
                # both the iframe and its body. The exact live breadcrumb is
                # still strong proof once that body is stable for three polls.
                if current_signature == stable_signature:
                    stable_polls += 1
                else:
                    stable_signature = current_signature
                    stable_polls = 1
                if frame_replaced or stable_polls >= 3:
                    return current
            # Some ASAP renderings do not expose breadcrumb labels to the
            # accessibility tree even though the exact menu link was clicked.
            # For a reused iframe, the closed target control plus a changed,
            # stable report body is the reliable navigation proof.
            if content_changed and not visible_path:
                if current_signature == stable_signature:
                    stable_polls += 1
                else:
                    stable_signature = current_signature
                    stable_polls = 1
                if stable_polls >= 3:
                    return current
        except Exception:
            pass
        page.wait_for_timeout(250)
    raise RuntimeError(
        "ASAP did not finish navigating to the requested report breadcrumb: "
        + " > ".join(path)
        + f". Visible path labels: {last_path_visibility}; "
        + f"iframe replaced: {last_frame_replaced}; body changed: {last_content_changed}."
    )


def _asap_raw_table_ready(frame: Frame) -> bool:
    """Recognize TechInsights raw-data views that omit the Data rows marker."""
    try:
        hint = frame.get_by_text(re.compile(r"to download raw data", re.I)).first
        if not hint.count() or not hint.is_visible():
            return False
        tables = frame.locator("table:visible")
        for index in range(tables.count()):
            rows = tables.nth(index).locator("tr")
            visible_rows = 0
            for row_index in range(rows.count()):
                if rows.nth(row_index).is_visible():
                    visible_rows += 1
                    if visible_rows >= 2:
                        return True
    except Exception:
        return False
    return False


def _asap_wait_for_results(
    page: Page, timeout_ms: int = ASAP_REPORT_RESULT_TIMEOUT_MS,
) -> Frame:
    """Return the live report frame once ASAP has rendered result rows.

    Running a report replaces ``iframe#content-frame`` in the current ASAP UI.
    A Frame captured before clicking RUN can therefore remain detached while the
    replacement frame already shows the completed report. Re-resolve the iframe
    during the wait and return the replacement so export uses the same live UI.
    """
    deadline = time.monotonic() + (timeout_ms / 1_000)
    last_error: Exception | None = None
    last_loading_state = False
    while time.monotonic() < deadline:
        try:
            last_loading_state = _asap_loading_overlay_visible(page)
            # MicroStrategy may briefly retain the old iframe element and add
            # the replacement as another frame. Inspect every current frame,
            # newest first, instead of trusting the first matching element.
            for frame in reversed(page.frames):
                rows = frame.get_by_text("Data rows:", exact=False).first
                if rows.count() and rows.is_visible():
                    return frame
                # TechInsights export views render the populated raw table
                # directly and never add MicroStrategy's Data rows label.
                # Require both the report's raw-data instruction and at least
                # two visible table rows, with no blocking overlay remaining.
                if not last_loading_state and _asap_raw_table_ready(frame):
                    return frame
        except Exception as exc:
            # Frame replacement can race any locator operation. The next poll
            # resolves the new element instead of retaining the detached frame.
            last_error = exc
        page.wait_for_timeout(500)
    detail = f" Last frame error: {last_error}" if last_error else ""
    detail += (
        " The ASAP loading overlay was still visible."
        if last_loading_state else
        f" {ASAP_EMPTY_RESULT_DETAIL}"
    )
    raise RuntimeError(f"ASAP report rows did not render within {timeout_ms // 1000} seconds.{detail}")


def _asap_run_report(page: Page) -> Frame:
    """Click RUN in the live report frame and wait for the rendered rows."""
    _asap_wait_for_loading_clear(page)
    run_frame = _asap_frame(page)
    _click_named(run_frame, "RUN")
    page.wait_for_timeout(1_000)
    frame = _asap_wait_for_results(page)
    return frame


def _asap_run_report_with_retry(page: Page, on_retry=None) -> Frame:
    """Retry one silently-empty report run.

    After RUN, MicroStrategy occasionally clears its loading overlay without
    ever rendering the Data rows marker or a populated raw table: the run
    finished with an empty rendering rather than a slow one. Clicking RUN a
    second time is safe at the point where that specific timeout is raised —
    no export has started and RUN re-executes the same configuration. A wait
    that ends with the loading overlay still visible is never retried; the
    report is genuinely still executing and a second RUN could interrupt it.
    """
    for attempt in range(2):
        try:
            return _asap_run_report(page)
        except RuntimeError as exc:
            message = str(exc)
            if (
                attempt == 1
                or not message.startswith("ASAP report rows did not render within")
                or ASAP_EMPTY_RESULT_DETAIL not in message
            ):
                raise
            if on_retry is not None:
                on_retry(exc)
            page.wait_for_timeout(1_500)
    raise AssertionError("unreachable")


def _asap_login_visible(page: Page) -> bool:
    try:
        password = page.locator('input[type="password"]:visible')
        return password.count() > 0 and password.first.is_visible()
    except Exception:
        return False


def _sso_fill_and_submit(page: Page, credentials: dict, portal_label: str) -> None:
    """Fill and submit the Samsung SSO form with the stored credential.

    ASAP and GSCM redirect to the same SSO host, so one fill serves both; the
    caller waits for its own portal's readiness proof afterwards.
    """
    visible_inputs = page.locator('input:visible')
    username = page.locator('input[type="text"]:visible, input:not([type]):visible').first
    password = page.locator('input[type="password"]:visible').first
    if not username.count() and visible_inputs.count() >= 2:
        username = visible_inputs.nth(0)
    username.fill(credentials["username"])
    password.fill(credentials["password"])
    submit = page.get_by_role("button", name=re.compile(r"^login$", re.I)).first
    if not submit.count():
        submit = page.locator('button[type="submit"]:visible, input[type="submit"]:visible').first
    if not submit.count():
        raise RuntimeError(f"{portal_label} sign-in form was found, but its Login action was not recognized.")
    submit.click()


def _asap_authenticate_if_needed(page: Page, profile_dir: Path) -> bool:
    """Recover an expired ASAP session using the local DPAPI credential."""
    if not _asap_login_visible(page):
        return False
    # Browser profiles are isolated by mode, but both use the one account-level
    # DPAPI credential selected by METRONOME_FLOW_PROFILE during setup.
    credentials = load_asap_credentials()
    if not credentials:
        raise RuntimeError(
            "ASAP sign-in is required. Configure the encrypted BI desktop credential in Flows > Catalog."
        )
    _sso_fill_and_submit(page, credentials, "ASAP")
    roots = _wait_for_navigation_roots(page, 120_000)
    if not roots:
        error = page.locator("text=/incorrect user id|incorrect password|try again/i").first
        detail = _clean_text(error.text_content()) if error.count() else "ASAP did not open after automatic sign-in."
        raise RuntimeError(f"ASAP automatic sign-in failed: {detail}")
    return True


#: How long GSCM's Nexacro shell gets to render after an automatic SSO submit.
GSCM_AUTO_LOGIN_TIMEOUT_MS = 180_000


def _gscm_authenticate_if_needed(page: Page, profile_dir: Path) -> bool:
    """Recover an expired GSCM session using the same stored SSO credential.

    GSCM sits behind the same Samsung SSO as ASAP, and ASAP has recovered its
    session unattended with this credential since it was stored - so GSCM uses
    the identical fill instead of demanding a manual re-bootstrap. Returns
    False when there is nothing to do (no visible form, or no stored
    credential); raises when the credential was submitted but the portal never
    rendered - a wrong password or a Knox MFA challenge.
    """
    if not _asap_login_visible(page):
        return False
    credentials = load_asap_credentials()
    if not credentials:
        return False
    _sso_fill_and_submit(page, credentials, "GSCM")
    deadline = time.monotonic() + GSCM_AUTO_LOGIN_TIMEOUT_MS / 1000
    while time.monotonic() < deadline:
        if flow_gscm.portal_shell_rendered(page):
            flow_gscm.clear_screen(page)
            return True
        page.wait_for_timeout(1_000)
    raise RuntimeError(
        "GSCM automatic sign-in submitted the stored credential, but the portal "
        "did not render within 180 seconds. The password may be wrong, or a "
        "Knox MFA prompt is blocking it - run the flow in headed mode once and "
        "complete the prompt in the visible window."
    )


def _asap_goto(page: Page, url: str, profile_dir: Path) -> bool:
    page.goto(url, wait_until="domcontentloaded", timeout=120_000)
    # ASAP may render /expiredSession before redirecting to its SSO host. Wait
    # for either terminal state so a delayed login form is never missed.
    deadline = time.monotonic() + 60
    while time.monotonic() < deadline:
        if _asap_login_visible(page):
            return _asap_authenticate_if_needed(page, profile_dir)
        if _navigation_roots(_visible_anchor_records(page)):
            return False
        page.wait_for_timeout(500)
    if _asap_login_visible(page):
        return _asap_authenticate_if_needed(page, profile_dir)
    return False


def _asap_open_report(page: Page, job: dict, profile_dir: Path) -> Frame:
    report = job["report"]
    automation = report.get("automation") or {}
    path = automation.get("category_path") or []
    if len(path) < 2:
        raise RuntimeError("ASAP reports need a category path with a menu and report name.")
    _asap_goto(page, job["site"].get("auth_url") or report["url"], profile_dir)
    def navigate(previous_frame: Frame | None, previous_signature: str | None):
        root = _asap_wait_for_visible(
            page,
            page.get_by_role("link", name=path[0], exact=True),
            page.get_by_text(path[0], exact=True),
        )
        if root is None:
            raise RuntimeError(f"ASAP menu was not visible: {path[0]}")
        root.click()
        if len(path) > 2:
            visible_group = _asap_wait_for_visible(
                page,
                page.get_by_role("link", name=path[-2], exact=True),
                page.get_by_text(path[-2], exact=True),
            )
            if visible_group is not None:
                visible_group.hover()
        repeated_label = len(path) > 2 and path[-1].casefold() == path[-2].casefold()
        target = _asap_wait_for_visible(
            page,
            page.get_by_role("link", name=path[-1], exact=True),
            page.get_by_text(path[-1], exact=True),
            prefer_last=repeated_label,
        )
        if target is None:
            raise RuntimeError(f"ASAP report menu item was not visible: {path[-1]}")
        target.click()
        return _asap_wait_for_report_navigation(
            page, previous_frame, target, path, previous_signature,
        )

    try:
        previous_frame = _asap_frame(page)
    except Exception:
        previous_frame = None
    previous_signature = _asap_frame_signature(previous_frame)
    frame = navigate(previous_frame, previous_signature)

    expected = automation.get("report_tab") or report.get("ready_text")
    last_error = None
    for attempt in range(2):
        try:
            if expected:
                # Export-view links can live in the outer ASAP shell or in a
                # nested MicroStrategy frame. Waiting in only content-frame
                # made the same report appear missing depending on portal
                # rendering timing.
                deadline = time.monotonic() + 120
                found_expected = False
                while time.monotonic() < deadline and not found_expected:
                    for root in list(dict.fromkeys([page.main_frame, frame, *page.frames])):
                        try:
                            matches = root.get_by_text(expected, exact=True)
                            if any(item.is_visible() for item in matches.all()):
                                found_expected = True
                                break
                        except Exception:
                            continue
                    if not found_expected:
                        page.wait_for_timeout(250)
                if not found_expected:
                    raise RuntimeError(f"ASAP report did not show its expected view: {expected}")
            if frame.get_by_text("500", exact=True).count():
                raise RuntimeError("ASAP returned an internal server error while loading the report.")
            return frame
        except (PlaywrightTimeoutError, RuntimeError) as exc:
            last_error = exc
            if attempt:
                break
            page.go_back(wait_until="domcontentloaded", timeout=120_000)
            try:
                previous_frame = _asap_frame(page)
            except Exception:
                previous_frame = None
            previous_signature = _asap_frame_signature(previous_frame)
            frame = navigate(previous_frame, previous_signature)
    raise RuntimeError(f"ASAP report did not load after one retry: {last_error}")


def _asap_member_selected(option) -> bool | None:
    """Read selection state from ASAP's ARIA, class, or non-hovered row styling."""
    try:
        return option.evaluate(
            r"""node => {
                for (let current = node, depth = 0; current && depth < 5; current = current.parentElement, depth++) {
                    const aria = current.getAttribute('aria-selected') ?? current.getAttribute('aria-checked');
                    if (aria === 'true') return true;
                    if (aria === 'false') return false;
                    const classes = String(current.className || '');
                    if (/(^|[-_ ])(?:selected|checked)(?:$|[-_ ])/i.test(classes) || /itemSelected/i.test(classes)) return true;
                    // ASAP uses a blue background for both selection and hover.
                    // Styling is only evidence after the pointer has left the row.
                    if (!current.matches(':hover')) {
                        const style = getComputedStyle(current);
                        const match = style.backgroundColor.match(/rgba?\((\d+),\s*(\d+),\s*(\d+)/);
                        if (match) {
                            const [r, g, b] = match.slice(1).map(Number);
                            if (b > r + 35 && b > g + 15 && b > 120) return true;
                        }
                    }
                }
                return null;
            }"""
        )
    except Exception:
        return None


def _asap_list_scope(frame: Frame, label: str, requested: list[str]):
    """Resolve the semantic or smallest structural owner of a prompt list."""
    # Prefer the WAI-ARIA listbox contract when the portal exposes it. This is
    # both stricter and more resilient than page-wide text matching.
    try:
        listboxes = frame.get_by_role("listbox", name=label, exact=True)
        for index in range(listboxes.count()):
            listbox = listboxes.nth(index)
            if not listbox.is_visible():
                continue
            if all(any(
                listbox.get_by_text(value, exact=True).nth(option_index).is_visible()
                for option_index in range(listbox.get_by_text(value, exact=True).count())
            ) for value in requested):
                return listbox
    except Exception:
        pass

    # Legacy MicroStrategy controls do not always expose listbox semantics.
    # In that case, narrow from every matching label to the nearest ancestor
    # containing every requested member, then use only descendants of it.
    labels = frame.get_by_text(label, exact=True)
    deadline = time.monotonic() + 60
    while time.monotonic() < deadline:
        matches = []
        for index in range(labels.count()):
            heading = labels.nth(index)
            try:
                if not heading.is_visible():
                    continue
                # Lightweight test doubles and older adapters do not expose
                # locator traversal. Their frame is already the list scope.
                if not callable(getattr(heading, "locator", None)):
                    return frame
                ancestor = heading
                for depth in range(1, 9):
                    parents = ancestor.locator("xpath=parent::*")
                    if not parents.count():
                        break
                    ancestor = parents.first
                    contains_requested = True
                    for value in requested:
                        candidates = ancestor.get_by_text(value, exact=True)
                        if not any(
                            candidates.nth(option_index).is_visible()
                            for option_index in range(candidates.count())
                        ):
                            contains_requested = False
                            break
                    if not contains_requested:
                        continue
                    box = ancestor.bounding_box()
                    area = box["width"] * box["height"] if box else float("inf")
                    matches.append((depth, area, ancestor))
                    break
            except Exception:
                continue
        if matches:
            return min(matches, key=lambda match: (match[0], match[1]))[2]
        frame.page.wait_for_timeout(250)
    raise RuntimeError(
        f"Could not isolate the ASAP {label} prompt containing: {requested}. "
        "The report was not run with an ambiguous filter."
    )


def _asap_live_list_values(scope, label: str) -> list[str]:
    """Read list members rendered inside an isolated ASAP prompt."""
    values = []
    try:
        options = scope.locator("[role='option'], option, [aria-selected], [aria-checked]")
        for index in range(options.count()):
            option = options.nth(index)
            if not option.is_visible():
                continue
            value = _clean_text(option.inner_text() or option.get_attribute("value"))
            if value and value not in values:
                values.append(value)
    except Exception:
        pass
    if values:
        return values

    # Visual-only legacy lists have no option role. The already-isolated
    # container is the boundary, so its distinct rendered lines are the safest
    # generic fallback. Selection-state inspection later discards non-options.
    try:
        lines = list(dict.fromkeys(
            _clean_text(line) for line in scope.inner_text().splitlines() if _clean_text(line)
        ))
    except Exception:
        return []
    for value in lines:
        if value.casefold().rstrip(":") == label.casefold().rstrip(":"):
            continue
        if re.fullmatch(r"\(all\)(?:\s*\(\d+\s+values?\))?", value, re.I):
            continue
        if "type to search" in value.casefold():
            continue
        values.append(value)
    return values


def _asap_select_list_values(
    frame: Frame, label: str, values: list[str], available_values: list[str] | None = None,
):
    if not values:
        return

    requested = list(dict.fromkeys(values))
    scope = _asap_list_scope(frame, label, requested)

    def park_pointer():
        """Remove hover styling before reading a row's rendered state."""
        frame.page.mouse.move(2, 2)
        frame.page.wait_for_timeout(100)

    def visible_option(value: str):
        candidates = scope.get_by_text(value, exact=True)
        option = next(
            (candidates.nth(index) for index in range(candidates.count()) if candidates.nth(index).is_visible()),
            None,
        )
        if option is None:
            raise RuntimeError(f"Could not find {label} option: {value}")
        return option

    available = list(dict.fromkeys([
        *(available_values or []), *_asap_live_list_values(scope, label), *requested,
    ]))

    # A previous interaction may have left the pointer over a blue hover row.
    # Never use that transient styling as the initial selected-state snapshot.
    park_pointer()

    def selected_states() -> dict[str, bool]:
        result = {}
        for value in available:
            try:
                state = _asap_member_selected(visible_option(value))
            except RuntimeError:
                continue
            if state is not None:
                result[value] = state
        return result

    def stable_exact_selection(samples: int = 3) -> tuple[bool, dict[str, bool]]:
        states = {}
        for sample in range(samples):
            states = selected_states()
            actual = {value for value, selected in states.items() if selected}
            if actual != set(requested) or any(states.get(value) is not True for value in requested):
                return False, states
            if sample + 1 < samples:
                frame.page.wait_for_timeout(150)
        return True, states

    def member_state(value: str) -> bool | None:
        try:
            return _asap_member_selected(visible_option(value))
        except RuntimeError:
            return None

    def set_member_state(value: str, selected: bool, attempts: int = 3) -> bool:
        """Click one member until the rendered portal state confirms the result."""
        for _attempt in range(attempts):
            state = member_state(value)
            if (selected and state is True) or (not selected and state is not True):
                return True
            # Playwright's delay is the time between mouse-down and mouse-up.
            # Keep this a single ordinary left click, with no keyboard modifier.
            visible_option(value).click(button="left", click_count=1, delay=100)
            park_pointer()
            for _sample in range(10):
                frame.page.wait_for_timeout(150)
                state = member_state(value)
                if (selected and state is True) or (not selected and state is not True):
                    return True
        return False

    def reconcile_requested_selection(rounds: int = 4) -> tuple[bool, dict[str, bool]]:
        """Make the rendered selection exactly match requested using plain clicks."""
        states = {}
        for _round in range(rounds):
            exact, states = stable_exact_selection()
            if exact:
                return True, states
            extras = [
                value for value, selected in states.items()
                if selected and value not in requested
            ]
            # MicroStrategy commonly represents an unselected row as neither
            # true nor false. Missing therefore means anything not confirmed
            # true, including an absent or unknown state.
            missing = [value for value in requested if states.get(value) is not True]
            for value in extras:
                set_member_state(value, False)
            for value in missing:
                set_member_state(value, True)
            frame.page.wait_for_timeout(300)
        return stable_exact_selection()

    if label.casefold().rstrip(":") == "dimension":
        # Dimension is the one ASAP prompt that opens with a retained member.
        # Its list rows toggle with a normal left click. Clear every visibly
        # selected member first, then select the Metronome values the same way.
        # Do not use Ctrl or manipulate the hidden native owner for this prompt.
        initial_states = selected_states()
        if not initial_states:
            raise RuntimeError(
                "ASAP Dimension did not expose a verifiable selected state. "
                "The report was not run with an unverified filter."
            )
        cleared_states = initial_states
        for _round in range(4):
            retained = [value for value, selected in cleared_states.items() if selected]
            if not retained:
                break
            for value in retained:
                set_member_state(value, False)
            frame.page.wait_for_timeout(300)
            cleared_states = selected_states()
        retained = [value for value, selected in cleared_states.items() if selected]
        if retained:
            raise RuntimeError(
                f"ASAP Dimension could not be cleared with plain clicks. Still selected: {retained}."
            )

        exact, final_states = reconcile_requested_selection()
        actual = [value for value, selected in final_states.items() if selected]
        missing = [value for value in requested if final_states.get(value) is not True]
        extras = [value for value in actual if value not in requested]
        if not exact or missing or extras:
            raise RuntimeError(
                f"ASAP Dimension selection did not match the flow. "
                f"Requested: {requested}. Selected: {actual}."
            )
        return

    # The rendered list toggles each value with a normal left click. Do not
    # assume a click landed: MicroStrategy can drop a rapid final click and it
    # reports an unselected row as an unknown state rather than explicit false.
    # Confirm each change and retry it before the final exact-set check.
    exact, final_states = reconcile_requested_selection()
    if final_states:
        actual = [value for value, selected in final_states.items() if selected]
        if not exact:
            raise RuntimeError(
                f"ASAP {label} selection did not match the flow. "
                f"Requested: {requested}. Selected: {actual}."
            )
        return
    raise RuntimeError(
        f"ASAP {label} did not expose a verifiable selected state. "
        "The report was not run with an unverified filter."
    )


def _asap_apply_configuration(frame: Frame, job: dict, period: str | list[str] | None):
    selections = dict(job.get("selections") or {})
    for definition in job["report"].get("filters", []):
        key = definition["filter_key"]
        value = period if definition["control_type"] == "week" and period else selections.get(key)
        if value in (None, "", []):
            continue
        values = value if isinstance(value, list) else [value]
        values = [_week_to_asap(str(item)) if definition["control_type"] == "week" else str(item) for item in values]
        automation = definition.get("automation") or {}
        declared_range_slider = (
            definition["control_type"] == "week"
            and automation.get("kind") == "range_slider"
        )
        range_slider = declared_range_slider
        if definition["control_type"] == "week" and not range_slider:
            try:
                _asap_range_scope(frame, definition["control_label"])
                range_slider = True
            except Exception:
                # This is only capability probing. The established visible-list
                # path below remains authoritative when no live slider is found.
                pass
        if range_slider:
            _asap_set_range(frame, definition["control_label"], values[0], values[-1], "week")
            date_range_label = _clean_text(
                automation.get("date_range_label")
                if declared_range_slider else "Date"
            )
            if date_range_label:
                start_date = _asap_week_dates(str(value[0] if isinstance(value, list) else value))[0]
                end_date = _asap_week_dates(str(value[-1] if isinstance(value, list) else value))[1]
                _asap_set_range(frame, date_range_label, start_date, end_date, "date")
            continue
        visible_list_only = (
            definition["control_label"].casefold().rstrip(":")
            in {"dimension", "category", "measure"}
            or definition["control_type"] == "week"
        )
        if visible_list_only:
            _asap_select_list_values(
                frame, definition["control_label"], values, definition.get("options") or values,
            )
            continue
        if definition["control_type"] == "select":
            _set_filter(frame, definition, values[0])
        else:
            if _select_native_options_by_text(frame, values, definition.get("options") or values):
                continue
            _asap_select_list_values(
                frame, definition["control_label"], values, definition.get("options") or values,
            )


# Download budgets are backstops for a browser that will never produce a file,
# not an opinion about how fast a portal should be. A multi-million-row export
# can sit for many minutes while the server builds the workbook before the
# first byte ever reaches the staging folder.
DOWNLOAD_START_TIMEOUT_SECONDS = 15 * 60
DOWNLOAD_STALL_TIMEOUT_SECONDS = 10 * 60
DOWNLOAD_MAX_TIMEOUT_SECONDS = 60 * 60
# Once Edge emits its native download event, the browser has accepted the
# transfer.  The worker-configured staging file must then appear promptly.  A
# missing browser-to-staging handoff is an error, not a reason to block on the
# Playwright Download object's terminal state for the full export budget.
DOWNLOAD_EVENT_STAGING_TIMEOUT_SECONDS = 60


def _download_file_state(path: Path) -> tuple[int, int]:
    stat = path.stat()
    return stat.st_mtime_ns, stat.st_size


def _download_staging_snapshot(staging_dir: Path) -> dict[Path, tuple[int, int]]:
    staging_dir.mkdir(parents=True, exist_ok=True)
    snapshot = {}
    for candidate in staging_dir.iterdir():
        try:
            if candidate.is_file():
                snapshot[candidate.resolve()] = _download_file_state(candidate)
        except OSError:
            continue
    return snapshot


def _download_staging_changed(
    staging_dir: Path, files_before: dict[Path, tuple[int, int]],
) -> bool:
    current = _download_staging_snapshot(staging_dir)
    return any(files_before.get(path) != state for path, state in current.items())


def _asap_wait_for_dashboard_download_signal(
    page: Page,
    staging_dir: Path,
    files_before: dict[Path, tuple[int, int]],
    downloads: list,
    opened: list,
    *,
    timeout_seconds: int = DOWNLOAD_START_TIMEOUT_SECONDS,
    popup_grace_seconds: int | None = 20,
) -> str:
    """Pump Playwright while waiting for a dashboard download to start.

    A plain ``time.sleep`` loop sees local files but does not dispatch a popup
    or download event to Playwright's sync callbacks. HTML dashboards often
    open their real download page a beat after the first click, so the old
    folder-only wait could sit for 15 minutes even though Edge had already
    completed the transfer or an intermediate popup was ready.
    """
    started = time.monotonic()
    popup_seen_at = None
    while time.monotonic() - started < timeout_seconds:
        if _download_staging_changed(staging_dir, files_before):
            return "staging"
        if downloads:
            return "download"
        if opened and popup_grace_seconds is not None:
            popup_seen_at = popup_seen_at or time.monotonic()
            if time.monotonic() - popup_seen_at >= popup_grace_seconds:
                return "popup"
        # Unlike time.sleep, this keeps context ``page`` and ``download``
        # callbacks flowing while Edge and the dashboard work independently.
        page.wait_for_timeout(250)
    return "timeout"


def _asap_frame_was_detached(exc: Exception) -> bool:
    message = str(exc).casefold()
    return "frame was detached" in message or "frame has been detached" in message


def _asap_wait_for_download_start(
    staging_dir: Path,
    files_before: dict[Path, tuple[int, int]],
    timeout_seconds: int = 15,
) -> bool:
    """Detect a download that started while its initiating frame was replaced."""
    deadline = time.monotonic() + timeout_seconds
    while time.monotonic() < deadline:
        current = _download_staging_snapshot(staging_dir)
        if any(files_before.get(path) != state for path, state in current.items()):
            return True
        time.sleep(0.25)
    return False


def _wait_for_staged_download(
    staging_dir: Path,
    files_before: dict[Path, tuple[int, int]],
    timeout_seconds: int = DOWNLOAD_MAX_TIMEOUT_SECONDS,
    start_timeout_seconds: int = DOWNLOAD_START_TIMEOUT_SECONDS,
    stall_timeout_seconds: int = DOWNLOAD_STALL_TIMEOUT_SECONDS,
) -> Path:
    """Return a new or overwritten stable local file without Playwright path calls."""
    started = time.monotonic()
    deadline = started + timeout_seconds
    last_activity = started
    observed: dict[Path, tuple[int, int]] = {}
    last_candidate = None
    last_candidate_state = None
    stable_checks = 0
    while time.monotonic() < deadline:
        candidates = []
        for candidate in staging_dir.iterdir():
            try:
                if not candidate.is_file():
                    continue
                resolved = candidate.resolve()
                state = _download_file_state(candidate)
            except OSError:
                continue
            if files_before.get(resolved) == state:
                continue
            if observed.get(resolved) != state:
                observed[resolved] = state
                last_activity = time.monotonic()
            if not candidate.name.casefold().endswith((".crdownload", ".tmp")):
                candidates.append((candidate, state))
        if candidates:
            candidate, state = max(candidates, key=lambda item: item[1][0])
            if candidate == last_candidate and state[1] > 0 and state == last_candidate_state:
                stable_checks += 1
                if stable_checks >= 3:
                    return candidate
            else:
                last_candidate = candidate
                last_candidate_state = state
                stable_checks = 0
        else:
            last_candidate = None
            last_candidate_state = None
            stable_checks = 0
        now = time.monotonic()
        if not observed and now - started >= start_timeout_seconds:
            raise RuntimeError(
                "ASAP started an Browser download, but no new or updated file appeared "
                f"in the local staging folder within {start_timeout_seconds} seconds: {staging_dir}"
            )
        if observed and now - last_activity >= stall_timeout_seconds:
            observed_summary = ", ".join(
                f"{path.name} ({state[1]} bytes)" for path, state in observed.items()
            )
            raise RuntimeError(
                f"The Browser download stopped changing for {stall_timeout_seconds} seconds in "
                f"the local staging folder: {staging_dir}. Observed: {observed_summary}"
            )
        time.sleep(0.5)
    observed_summary = ", ".join(
        f"{path.name} ({state[1]} bytes)" for path, state in observed.items()
    ) or "none"
    raise RuntimeError(
        f"The Browser download did not produce a stable finished file within {timeout_seconds} "
        f"seconds in the local staging folder: {staging_dir}. Observed: {observed_summary}"
    )


def _completed_edge_download(download, description: str) -> Path:
    """Validate one native browser Download and return its completed local path."""
    # ``failure`` waits for Edge's terminal download state. ``path`` then
    # returns the browser-managed completed file, not a guessed directory
    # candidate based on timestamps.
    failure = download.failure()
    if failure:
        raise RuntimeError(f"Browser reported that {description} failed: {failure}")
    completed_path = download.path()
    if not completed_path:
        raise RuntimeError(
            f"Browser reported that {description} completed but returned no local file path."
        )
    completed = Path(completed_path)
    if not completed.is_file() or completed.stat().st_size <= 0:
        raise RuntimeError(
            f"Browser reported that {description} completed, but its local file is "
            f"missing or empty: {completed}"
        )
    return completed


def _asap_dashboard_event_staged_download(
    staging_dir: Path,
    files_before: dict[Path, tuple[int, int]],
    label: str,
) -> Path:
    """Finish a dashboard download through its worker staging directory.

    The native Edge event proves only that the browser accepted the transfer.
    Dashboard downloads are configured to land in ``staging_dir`` and all
    downstream stability, normalization, and retry logic depends on that file.
    Do not call ``Download.failure()`` or ``Download.path()`` here: either can
    wait indefinitely for the remote dashboard response to reach a terminal
    browser state.
    """
    try:
        return _wait_for_staged_download(
            staging_dir,
            files_before,
            start_timeout_seconds=DOWNLOAD_EVENT_STAGING_TIMEOUT_SECONDS,
        )
    except RuntimeError as exc:
        if "no new or updated file appeared" not in str(exc):
            raise
        raise RuntimeError(
            "Browser emitted a native download-start event for ASAP dashboard "
            f"link {label!r}, but no file reached the worker staging folder "
            f"within {DOWNLOAD_EVENT_STAGING_TIMEOUT_SECONDS} seconds: "
            f"{staging_dir}. Check browser's download-directory policy and the "
            "worker staging-folder permissions."
        ) from exc


def _edge_completed_download(page: Page, trigger_download) -> Path:
    """Use Edge's native download lifecycle as the completion authority."""
    try:
        with page.expect_download(
            timeout=DOWNLOAD_MAX_TIMEOUT_SECONDS * 1_000,
        ) as pending:
            trigger_download()
        download = pending.value
    except PlaywrightTimeoutError as exc:
        raise RuntimeError(
            "Browser did not emit its native download event after the GSCM Excel action."
        ) from exc
    return _completed_edge_download(download, "the GSCM Excel download")


def _asap_table_control_score(
    table_box: dict[str, float] | None, control_box: dict[str, float] | None,
) -> float | None:
    """Rank a compact control rendered at the raw table's top-right corner."""
    if not table_box or not control_box:
        return None
    width = control_box.get("width", 0)
    height = control_box.get("height", 0)
    if not (4 <= width <= 64 and 4 <= height <= 64):
        return None
    table_left = table_box["x"]
    table_top = table_box["y"]
    table_right = table_left + table_box["width"]
    center_x = control_box["x"] + width / 2
    center_y = control_box["y"] + height / 2
    if center_x < table_left + table_box["width"] * 0.65 or center_x > table_right + 24:
        return None
    if center_y < table_top - 40 or center_y > table_top + 72:
        return None
    return abs(center_x - (table_right - 14)) + abs(center_y - (table_top + 14))


def _asap_run_control(root: Page | Frame):
    """Return ASAP's RUN control across button, input, and text renderings."""
    for build_locator in (
        lambda: root.get_by_role("button", name=re.compile(r"^RUN$", re.I)),
        lambda: root.locator(
            "input[type='button'][value='RUN' i]:visible,"
            "input[type='submit'][value='RUN' i]:visible"
        ),
        lambda: root.get_by_text("RUN", exact=True),
    ):
        try:
            locator = build_locator()
            if locator.count() and locator.first.is_visible():
                return locator.first
        except Exception:
            continue
    return None


def _asap_export_action(root: Page | Frame):
    """Return a visible Export action across button, input, and text renderings."""
    for build_locator in (
        lambda: root.get_by_role("button", name="Export", exact=True),
        lambda: root.locator(
            "input[type='button'][value='Export' i]:visible,"
            "input[type='submit'][value='Export' i]:visible"
        ),
        lambda: root.get_by_text("Export", exact=True),
    ):
        try:
            locator = build_locator()
            if locator.count() and locator.first.is_visible():
                return locator.first
        except Exception:
            continue
    return None


def _asap_export_format_names(download_type: str) -> tuple[str, re.Pattern]:
    """Return one canonical Export Wizard label and an exact matcher."""
    spec = resolve_asap_download_type(
        download_type if download_type in {item.key for item in ASAP_DOWNLOAD_TYPES} else None,
        legacy_file_format=download_type,
    )
    return spec.label, re.compile(rf"^{re.escape(spec.label)}$", re.I)


def _asap_visible_labeled_control(root: Page | Frame, label: str):
    pattern = re.compile(rf"^{re.escape(label)}$", re.I)
    for locator in (
        root.get_by_label(label, exact=True),
        root.get_by_role("radio", name=pattern),
        root.get_by_role("checkbox", name=pattern),
        root.get_by_text(pattern),
    ):
        try:
            if locator.count() and locator.first.is_visible():
                return locator.first
        except Exception:
            continue
    return None


def _asap_control_checked(control) -> bool | None:
    if control is None:
        return None
    candidates = [control]
    try:
        candidates.extend([
            control.locator("input[type=radio],input[type=checkbox]").first,
            control.locator("xpath=preceding::input[@type='radio' or @type='checkbox'][1]").first,
        ])
    except Exception:
        pass
    for candidate in candidates:
        try:
            if candidate.count() and candidate.is_checked():
                return True
            if candidate.count():
                candidate.is_checked()
                return False
        except Exception:
            continue
    return None


def _asap_set_checked(control, desired: bool, label: str) -> None:
    if control is None:
        raise RuntimeError(f"ASAP Export Wizard did not expose {label!r}.")
    current = _asap_control_checked(control)
    if current is desired:
        return
    try:
        (control.check if desired else control.uncheck)()
    except Exception:
        control.click()
    verified = _asap_control_checked(control)
    if verified is not desired:
        raise RuntimeError(
            f"ASAP Export Wizard did not retain {label!r} as "
            f"{'checked' if desired else 'unchecked'}."
        )


def _asap_resolve_export_wizard(page: Page, timeout_ms: int = 10_000) -> dict | None:
    """Resolve one wizard root and all recognized controls without exporting."""
    deadline = time.monotonic() + timeout_ms / 1000
    while time.monotonic() < deadline:
        for candidate in reversed(page.context.pages):
            for root in [candidate, *reversed(candidate.frames)]:
                action = _asap_export_action(root)
                if action is None:
                    continue
                formats = {
                    item.key: control
                    for item in ASAP_DOWNLOAD_TYPES
                    if (control := _asap_visible_labeled_control(root, item.label)) is not None
                }
                if not formats:
                    continue
                options = {
                    key: _asap_visible_labeled_control(root, label)
                    for key, label in ASAP_EXPORT_CHECKBOXES.items()
                }
                return {
                    "root": root, "action": action,
                    "formats": formats, "options": options,
                    "page": root if isinstance(root, Page) else root.page,
                }
        page.wait_for_timeout(100)
    return None


def _asap_toolbar_export_control(page: Page):
    """Find the ordinary Export Options control beside RUN."""
    for root in reversed(page.frames):
        try:
            run = _asap_run_control(root)
            if run is None:
                continue
            control = run.locator(
                "xpath=following::*[self::button or self::a or self::input or @role='button'][1]"
            )
            if not control.count() or not control.first.is_visible():
                control = run.locator(
                    "xpath=following::*[contains(@class,'btn') or contains(@class,'button')][1]"
                )
            if control.count() and control.first.is_visible():
                return control.first
            images = run.locator("xpath=following::img")
            for index in range(min(images.count(), 6)):
                if images.nth(index).is_visible():
                    return images.nth(index)
        except Exception:
            continue
    return None


def _asap_close_probe_wizard(page: Page, wizard: dict | None, pages_before: tuple) -> None:
    if wizard is not None:
        root = wizard["root"]
        for label in ("Cancel", "Close"):
            for locator in (
                root.get_by_role("button", name=label, exact=True),
                root.get_by_text(label, exact=True),
            ):
                try:
                    if locator.count() and locator.first.is_visible():
                        locator.first.click()
                        page.wait_for_timeout(100)
                        break
                except Exception:
                    continue
    for candidate in list(page.context.pages):
        if candidate in pages_before or candidate is page:
            continue
        try:
            candidate.close()
        except Exception:
            pass
    try:
        page.keyboard.press("Escape")
    except Exception:
        pass


def _asap_probe_export_capabilities(page: Page, timeout_ms: int = 8_000) -> dict:
    """Inventory one Export Wizard without triggering a download."""
    control = _asap_toolbar_export_control(page)
    if control is None:
        raise RuntimeError("ASAP Export Options control was not found for capability discovery.")
    pages_before = tuple(page.context.pages)
    wizard = None
    try:
        control.click()
        wizard = _asap_resolve_export_wizard(page, timeout_ms=timeout_ms)
        if wizard is None:
            raise RuntimeError("ASAP Export Wizard did not expose recognized format controls.")
        original_type = next(
            (
                key for key, item in wizard["formats"].items()
                if _asap_control_checked(item) is True
            ),
            None,
        )
        if original_type is None:
            raise RuntimeError(
                "ASAP Export Wizard formats were visible, but its original selected type "
                "could not be determined safely."
            )
        original_options = {
            key: _asap_control_checked(item)
            for key, item in wizard["options"].items()
        }
        by_type = {}
        for spec in ASAP_DOWNLOAD_TYPES:
            if spec.key not in wizard["formats"]:
                continue
            _asap_set_checked(wizard["formats"][spec.key], True, spec.label)
            page.wait_for_timeout(50)
            refreshed = _asap_resolve_export_wizard(page, timeout_ms=1_000) or wizard
            wizard = refreshed
            by_type[spec.key] = {
                key: {
                    "available": item is not None,
                    "checked": _asap_control_checked(item),
                }
                for key, item in wizard["options"].items()
            }
        if original_type and original_type in wizard["formats"]:
            _asap_set_checked(
                wizard["formats"][original_type], True,
                resolve_asap_download_type(original_type).label,
            )
            wizard = _asap_resolve_export_wizard(page, timeout_ms=1_000) or wizard
            for key, checked in original_options.items():
                if checked is not None and wizard["options"].get(key) is not None:
                    _asap_set_checked(
                        wizard["options"][key], checked, ASAP_EXPORT_CHECKBOXES[key],
                    )
        return {
            "status": "detected",
            "download_types": [
                item.key for item in ASAP_DOWNLOAD_TYPES if item.key in wizard["formats"]
            ],
            "options_by_type": by_type,
            "original_type": original_type,
        }
    finally:
        _asap_close_probe_wizard(page, wizard, pages_before)


def _asap_raw_table_information_control(root: Page | Frame, canvas):
    """Find an unlabeled hover control by its documented table-corner position."""
    try:
        table_box = canvas.bounding_box()
    except Exception:
        return None
    candidates = root.locator(
        "button:visible,a:visible,[role='button']:visible,input[type='button']:visible,"
        "img:visible,[class*='info' i]:visible,[class*='icon' i]:visible"
    )
    ranked = []
    for index in range(candidates.count()):
        candidate = candidates.nth(index)
        try:
            score = _asap_table_control_score(table_box, candidate.bounding_box())
            if score is None:
                continue
            signal = " ".join(filter(None, (
                candidate.get_attribute("title"),
                candidate.get_attribute("aria-label"),
                candidate.get_attribute("class"),
                candidate.get_attribute("alt"),
            ))).casefold()
            if re.search(r"info|export|more", signal):
                score -= 100
            ranked.append((score, candidate))
        except Exception:
            continue
    return min(ranked, key=lambda item: item[0])[1] if ranked else None


def _asap_wait_for_raw_menu_download_action(
    page: Page, download_type: str, timeout_ms: int = 10_000,
):
    """Return a raw-table menu item that starts the requested download directly."""
    if resolve_asap_download_type(
        download_type if download_type in {item.key for item in ASAP_DOWNLOAD_TYPES} else None,
        legacy_file_format=download_type,
    ).key != "excel_plain_text":
        return None
    excel_pattern = re.compile(r"^(?:Microsoft )?Excel$", re.I)
    deadline = time.monotonic() + timeout_ms / 1000
    while time.monotonic() < deadline:
        for candidate in reversed(page.context.pages):
            for root in [candidate, *reversed(candidate.frames)]:
                for locator in (
                    root.get_by_role("menuitem", name=excel_pattern),
                    root.get_by_text(excel_pattern),
                ):
                    try:
                        if locator.count() and locator.first.is_visible():
                            return locator.first
                    except Exception:
                        continue
        page.wait_for_timeout(100)
    return None


def _asap_wait_for_raw_menu_export_or_wizard(
    page: Page, pages_before: tuple, timeout_ms: int = 10_000,
):
    """Return the intermediate menu action, or report that Export Options opened directly."""
    deadline = time.monotonic() + timeout_ms / 1000
    wizard_pattern = re.compile(
        "^(?:" + "|".join(re.escape(item.label) for item in ASAP_DOWNLOAD_TYPES) + ")$",
        re.I,
    )
    while time.monotonic() < deadline:
        current_pages = page.context.pages
        if any(candidate not in pages_before for candidate in current_pages):
            return None, True
        for candidate in reversed(current_pages):
            for root in [candidate, *reversed(candidate.frames)]:
                try:
                    wizard_marker = root.get_by_text(wizard_pattern)
                    if wizard_marker.count() and wizard_marker.first.is_visible():
                        return None, True
                except Exception:
                    pass
                for locator in (
                    root.get_by_role("menuitem", name="Export", exact=True),
                    root.get_by_role("button", name="Export", exact=True),
                    root.get_by_text("Export", exact=True),
                ):
                    try:
                        if locator.count() and locator.first.is_visible():
                            return locator.first, False
                    except Exception:
                        continue
        page.wait_for_timeout(200)
    return None, False


def _asap_wait_for_raw_export_confirmation(
    page: Page, download_type: str, timeout_ms: int = 10_000,
):
    """Find the final Export button in ASAP's compact raw-table dialog."""
    if resolve_asap_download_type(
        download_type if download_type in {item.key for item in ASAP_DOWNLOAD_TYPES} else None,
        legacy_file_format=download_type,
    ).key != "excel_plain_text":
        return None
    title_pattern = re.compile(r"^Export to (?:Microsoft )?Excel$", re.I)
    deadline = time.monotonic() + timeout_ms / 1000
    while time.monotonic() < deadline:
        for candidate in reversed(page.context.pages):
            for root in [candidate, *reversed(candidate.frames)]:
                try:
                    title = root.get_by_text(title_pattern)
                    if not title.count() or not title.first.is_visible():
                        continue
                except Exception:
                    continue
                action = _asap_export_action(root)
                if action is not None:
                    return action
        page.wait_for_timeout(100)
    return None


def _asap_direct_excel_allowed(download_type: str, checkbox_values: dict[str, Any]) -> bool:
    """The compact shortcut cannot enforce Export Wizard checkbox choices."""
    return (
        resolve_asap_download_type(download_type).key == "excel_plain_text"
        and all(checkbox_values.get(key) is None for key in ASAP_EXPORT_CHECKBOXES)
    )


def _asap_download(page: Page, frame: Frame, job: dict, staging_dir: Path):
    export_control = None
    opens_export_menu = False
    downloads_config = job.get("downloads", {})
    file_format = str(downloads_config.get("file_format") or "csv").casefold()
    download_type = resolve_asap_download_type(
        downloads_config.get("asap_download_type"), legacy_file_format=file_format,
    )
    checkbox_values = {
        key: downloads_config.get(key) for key in ASAP_EXPORT_CHECKBOXES
    }
    direct_export_allowed = _asap_direct_excel_allowed(
        download_type.key, checkbox_values,
    )
    # The current MicroStrategy report has two compact controls beside RUN.
    # The first is Export Options and the second is subtotal. Their icon-only
    # markup has no stable accessible label, so resolve the first visible
    # button-like control after RUN within the same toolbar.
    for root in reversed(page.frames):
        try:
            run = _asap_run_control(root)
            if run is None:
                continue
            export_control = run.locator("xpath=following::*[self::button or self::a or self::input or @role='button'][1]")
            if not export_control.count() or not export_control.first.is_visible():
                export_control = run.locator("xpath=following::*[contains(@class,'btn') or contains(@class,'button')][1]")
            if export_control.count() and export_control.first.is_visible():
                export_control = export_control.first
                break
            # Older MicroStrategy templates render both toolbar icons as
            # unlabelled images. Choose the first visible image after RUN and
            # before the report canvas, which is Export Options.
            images = run.locator("xpath=following::img")
            export_control = next(
                (images.nth(index) for index in range(min(images.count(), 6)) if images.nth(index).is_visible()),
                None,
            )
            if export_control is not None:
                break
            export_control = None
        except Exception:
            continue
        if export_control is not None:
            break
    if export_control is None:
        # Some raw-data views have no prompt RUN toolbar. Their export action
        # appears only after the rendered table is hovered and its information
        # icon is opened, matching ASAP's own download tutorial.
        for root in reversed(page.frames):
            try:
                canvas = next((
                    locator.first
                    for selector in (
                        "table:visible", "[role=grid]:visible",
                        "[class*='grid' i]:visible", "[class*='table' i]:visible",
                        "[class*='report' i]:visible",
                    )
                    if (locator := root.locator(selector)).count()
                    and locator.first.is_visible()
                ), root.locator("table:visible").first)
                if canvas.count() and canvas.is_visible():
                    canvas.hover()
                    page.wait_for_timeout(750)
                candidates = [
                    root.get_by_role(
                        "button", name=re.compile(r"^(?:information|info|more info)$", re.I),
                    ),
                    root.locator(
                        "[title*='information' i]:visible,[title='Info' i]:visible,"
                        "[aria-label*='information' i]:visible,[aria-label='Info' i]:visible,"
                        "img[alt*='information' i]:visible,img[alt='Info' i]:visible"
                    ),
                ]
                export_control = next(
                    (
                        locator.first for locator in candidates
                        if locator.count() and locator.first.is_visible()
                    ),
                    None,
                )
                if export_control is None and canvas.count() and canvas.is_visible():
                    export_control = _asap_raw_table_information_control(root, canvas)
                if export_control is not None:
                    opens_export_menu = True
                    break
            except Exception:
                continue
    if export_control is None:
        raise RuntimeError(
            "Could not find either the ASAP Export Options control beside RUN "
            "or the raw-table information control."
        )
    pages_before = tuple(page.context.pages)
    export_control.click()
    direct_download_action = None
    if opens_export_menu:
        menu_export, wizard_opened = _asap_wait_for_raw_menu_export_or_wizard(
            page, pages_before,
        )
        if menu_export is None and not wizard_opened:
            raise RuntimeError("ASAP raw-table information menu opened, but its Export action was not visible.")
        if menu_export is not None:
            menu_export.click()
            if direct_export_allowed:
                direct_download_action = _asap_wait_for_raw_menu_download_action(
                    page, download_type.key,
                )
    # ASAP sometimes opens the wizard as a page and sometimes as a modal/frame
    # in the existing page. Search both shapes instead of requiring a popup.
    wizard_pages = set()
    wizard = None if direct_download_action is not None else _asap_resolve_export_wizard(
        page, timeout_ms=60_000,
    )
    if direct_download_action is None and wizard is None:
        raise RuntimeError(
            f"ASAP Export Wizard opened, but {download_type.label!r} and its Export action "
            "were not recognized in the same dialog."
        )
    format_option = None if wizard is None else wizard["formats"].get(download_type.key)
    export_action = None if wizard is None else wizard["action"]
    if wizard is not None:
        wizard_pages.add(wizard["page"])
    if direct_download_action is None and format_option is None:
        available = ", ".join(
            item.label for item in ASAP_DOWNLOAD_TYPES
            if wizard is not None and item.key in wizard["formats"]
        ) or "none"
        raise RuntimeError(
            f"ASAP Export Wizard does not offer {download_type.label!r}. Detected: {available}."
        )
    if direct_download_action is None:
        _asap_set_checked(format_option, True, download_type.label)
        wizard = _asap_resolve_export_wizard(page, timeout_ms=2_000) or wizard
        export_action = wizard["action"]
        wizard_pages.add(wizard["page"])
        missing_options = [
            ASAP_EXPORT_CHECKBOXES[key]
            for key, desired in checkbox_values.items()
            if desired is not None and wizard["options"].get(key) is None
        ]
        if missing_options:
            detected_options = [
                label for key, label in ASAP_EXPORT_CHECKBOXES.items()
                if wizard["options"].get(key) is not None
            ]
            raise RuntimeError(
                "ASAP Export Wizard is missing requested control(s): "
                + ", ".join(missing_options)
                + ". Detected checkbox controls: "
                + (", ".join(detected_options) or "none")
                + "."
            )
        for key, desired in checkbox_values.items():
            if desired is None:
                continue
            _asap_set_checked(
                wizard["options"].get(key), bool(desired), ASAP_EXPORT_CHECKBOXES[key],
            )
    # The export wizard may render in one popup while Edge attributes the
    # resulting download to another ASAP page. Listening only on the guessed
    # popup can therefore leave a visibly completed browser download waiting
    # until timeout. Subscribe to every current portal page before clicking.
    files_before = _download_staging_snapshot(staging_dir)
    downloads = []
    observed_pages = list(page.context.pages)

    def capture_download(download):
        downloads.append(download)

    for candidate in observed_pages:
        candidate.on("download", capture_download)
    try:
        try:
            (direct_download_action or export_action).click()
            if direct_download_action is not None:
                confirmation = _asap_wait_for_raw_export_confirmation(
                    page, download_type.key,
                )
                if confirmation is None:
                    raise RuntimeError(
                        "ASAP raw-table Excel dialog opened, but its final Export action was not visible."
                    )
                confirmation.click()
        except PlaywrightError as exc:
            if not _asap_frame_was_detached(exc) or not _asap_wait_for_download_start(
                staging_dir, files_before,
            ):
                raise
            # The frame can disappear after the browser has accepted the
            # export click. Recover the emitted file rather than clicking a
            # second time and creating a duplicate CSV.
            staged_file = _wait_for_staged_download(staging_dir, files_before)
            export_pages = [
                candidate for candidate in wizard_pages
                if candidate not in pages_before and candidate is not page
            ]
            return staged_file, export_pages
        deadline = time.monotonic() + 180
        while not downloads and time.monotonic() < deadline:
            page.wait_for_timeout(100)
        if not downloads:
            raise RuntimeError("ASAP export started, but Browser did not expose the completed download within 3 minutes.")
        staged_file = _wait_for_staged_download(staging_dir, files_before)
        export_pages = [
            candidate for candidate in wizard_pages
            if candidate not in pages_before and candidate is not page
        ]
        return staged_file, export_pages
    finally:
        for candidate in observed_pages:
            candidate.remove_listener("download", capture_download)


ASAP_DASHBOARD_LINK_MARK_JS = """
([label, href, mode]) => {
    const norm = s => (s || "").replace(/\\s+/g, " ").trim();
    const target = norm(label).toLowerCase();
    document.querySelectorAll("[data-metronome-dl]").forEach(
        el => el.removeAttribute("data-metronome-dl"));
    let best = null;
    for (const el of document.querySelectorAll("a,button,[role=button],[onclick]")) {
        if (!el || el.offsetWidth <= 0 || el.offsetHeight <= 0) continue;
        const text = norm(el.innerText || el.getAttribute("title")
            || el.getAttribute("aria-label") || "").toLowerCase();
        const elHref = el.getAttribute("href") || "";
        if (href && elHref && elHref === href) { best = el; break; }
        if (mode === "exact" && text === target) { best = el; break; }
        if (mode === "contains" && target && (text.includes(target) || target.includes(text)) && text) {
            best = best || el;
        }
    }
    if (!best) return null;
    best.setAttribute("data-metronome-dl", "1");
    return norm(best.innerText || best.getAttribute("title")
        || best.getAttribute("aria-label") || "") || "(unnamed control)";
}
"""

ASAP_DASHBOARD_LINK_INVENTORY_JS = """
() => {
    const norm = s => (s || "").replace(/\\s+/g, " ").trim();
    const out = [];
    for (const el of document.querySelectorAll("a,button,[role=button],[onclick]")) {
        if (!el || el.offsetWidth <= 0 || el.offsetHeight <= 0) continue;
        const text = norm(el.innerText || el.getAttribute("title")
            || el.getAttribute("aria-label") || "");
        if (!text) continue;
        out.push(text.slice(0, 80));
        if (out.length >= 60) break;
    }
    return out;
}
"""


def _asap_dashboard_link_href(job: dict | None, label: str) -> str:
    """The href the scan catalogued for this label, if it recorded one.

    Matching on href is exact where matching on text is not: the label stored
    in the catalog may have come from a title or aria-label attribute rather
    than from rendered text, in which case no text query can ever find it.
    """
    records = ((job or {}).get("report", {}).get("automation") or {}).get("download_links") or []
    for record in records:
        if isinstance(record, dict) and _clean_text(record.get("label")) == _clean_text(label):
            return str(record.get("href") or "")
    return ""


def _asap_dashboard_link_inventory(page: Page) -> str:
    """Every visible clickable control, for a failure the user can act on."""
    seen: list[str] = []
    for root in list(dict.fromkeys([page.main_frame, *page.frames])):
        try:
            for text in root.evaluate(ASAP_DASHBOARD_LINK_INVENTORY_JS) or []:
                if text not in seen:
                    seen.append(text)
        except Exception:
            continue
    return ", ".join(repr(item) for item in seen[:40]) or "none"


def _asap_dashboard_link_locator(
    page: Page, label: str, href: str = "",
    timeout_seconds: int = ASAP_DASHBOARD_LINK_TIMEOUT_SECONDS,
):
    """Resolve one catalogued dashboard download control across every frame.

    An embedded dashboard renders after its frame reports loaded, so a single
    immediate pass fails on a control that appears a second later. Poll until
    the deadline, and widen the match on each pass:

    1. Playwright's role and text queries, which are precise when the label is
       the control's rendered text.
    2. The catalogued href, which identifies the control even when its label
       came from a ``title`` or ``aria-label`` and no text query can match it.
    3. Case-insensitive whole-label, then either side containing the other -
       enough to survive a portal that re-cases or decorates its own labels.
    """
    deadline = time.monotonic() + timeout_seconds
    while True:
        roots = list(dict.fromkeys([page.main_frame, *page.frames]))
        for root in roots:
            for build_locator in (
                lambda: root.get_by_role("link", name=label, exact=True),
                lambda: root.get_by_role("button", name=label, exact=True),
                lambda: root.get_by_text(label, exact=True),
            ):
                try:
                    locator = build_locator()
                    for index in range(locator.count()):
                        item = locator.nth(index)
                        if item.is_visible():
                            return item
                except Exception:
                    continue
        for mode in ("exact", "contains"):
            for root in roots:
                try:
                    matched = root.evaluate(
                        ASAP_DASHBOARD_LINK_MARK_JS, [label, href, mode],
                    )
                except Exception:
                    continue
                if not matched:
                    continue
                item = root.locator("[data-metronome-dl='1']").first
                try:
                    if item.count() and item.is_visible():
                        return item
                except Exception:
                    continue
        if time.monotonic() >= deadline:
            return None
        page.wait_for_timeout(1_000)


def _asap_download_dashboard_link(
    page: Page, label: str, staging_dir: Path, job: dict | None = None,
) -> Path:
    """Click one HTML-dashboard download link and capture the resulting file.

    Some dashboards download in place; others open a popup or a new tab that
    emits the file, and a few use an intermediate page with its own download
    control. A native Edge event is a download-start signal; completion still
    comes from a stable file in the worker staging directory. Portal builds
    that emit no browser event are handled by the same directory monitor.
    Download-page listeners are detached after the staged file is stable, but
    popup pages are deliberately left alone. Edge can leave a download popup
    half-detached while its native Download object is nonterminal, and a later
    ``Page.close()`` has no timeout and can block before the caller copies the
    staged file to the configured target folder.
    """
    staging_dir.mkdir(parents=True, exist_ok=True)
    href = _asap_dashboard_link_href(job, label)
    control = _asap_dashboard_link_locator(page, label, href)
    if control is None:
        raise RuntimeError(
            f"ASAP dashboard download link was not visible: {label}. "
            f"Visible controls on the dashboard: {_asap_dashboard_link_inventory(page)}. "
            "Rescan this report if the dashboard renamed its download links."
        )
    files_before = _download_staging_snapshot(staging_dir)
    opened: list = []
    downloads: list = []
    observed_pages: list = []

    def _capture_download(download):
        downloads.append(download)

    def _observe_page(candidate):
        if candidate in observed_pages:
            return
        observed_pages.append(candidate)
        candidate.on("download", _capture_download)

    def _track_page(popup):
        opened.append(popup)
        _observe_page(popup)

    context = page.context
    context.on("page", _track_page)
    for candidate in list(context.pages):
        _observe_page(candidate)
    try:
        control.click(timeout=30_000)
        signal = _asap_wait_for_dashboard_download_signal(
            page, staging_dir, files_before, downloads, opened,
        )
        # A native event is authoritative proof that the transfer started, but
        # dashboard completion comes from the stable worker staging file. This
        # also wins a simultaneous event/folder race so the post-event handoff
        # budget is consistently enforced.
        if downloads:
            return _asap_dashboard_event_staged_download(
                staging_dir, files_before, label,
            )
        if signal == "staging":
            return _wait_for_staged_download(staging_dir, files_before)
        if signal == "timeout":
            raise RuntimeError(
                "ASAP dashboard download did not emit an Browser download event, "
                "open an intermediate page, or create a local staging file within "
                f"{DOWNLOAD_START_TIMEOUT_SECONDS} seconds: {label}."
            )

        # No file or Browser download event followed the first click, but a popup
        # remained open for the grace period. Its own control is the trigger.
        popup_control = None
        for popup in opened:
            try:
                if popup.is_closed():
                    continue
                popup.wait_for_load_state("domcontentloaded", timeout=15_000)
                popup_control = _asap_dashboard_link_locator(
                    popup, label, href, timeout_seconds=15,
                ) or next(
                    (
                        candidate for candidate in [
                            _asap_export_action(popup),
                            popup.get_by_role("link", name=re.compile(r"download", re.I)).first,
                        ] if candidate is not None and candidate.count() and candidate.is_visible()
                    ),
                    None,
                )
            except Exception:
                popup_control = None
            if popup_control is not None:
                break
        # The popup may emit its file while we are waiting for its DOM to load.
        # Re-check completion before declaring its control missing or clicking
        # again, otherwise a successful transfer becomes a false failure (or a
        # duplicate download).
        if downloads:
            return _asap_dashboard_event_staged_download(
                staging_dir, files_before, label,
            )
        if _download_staging_changed(staging_dir, files_before):
            return _wait_for_staged_download(staging_dir, files_before)
        if popup_control is None:
            raise RuntimeError(
                "ASAP dashboard opened an intermediate page, but no download "
                f"control was visible and no file started: {label}."
            )
        popup_control.click(timeout=30_000)
        signal = _asap_wait_for_dashboard_download_signal(
            page, staging_dir, files_before, downloads, opened,
            popup_grace_seconds=None,
        )
        if downloads:
            return _asap_dashboard_event_staged_download(
                staging_dir, files_before, label,
            )
        if _download_staging_changed(staging_dir, files_before):
            return _wait_for_staged_download(staging_dir, files_before)
        if signal == "timeout":
            raise RuntimeError(
                "ASAP dashboard intermediate page did not emit an Browser download "
                "event or create a local staging file within "
                f"{DOWNLOAD_START_TIMEOUT_SECONDS} seconds: {label}."
            )
        return _wait_for_staged_download(staging_dir, files_before)
    finally:
        try:
            context.remove_listener("page", _track_page)
        except Exception:
            pass
        for candidate in observed_pages:
            try:
                candidate.remove_listener("download", _capture_download)
            except Exception:
                pass
        # Do not query or close ``opened`` here. This function must return the
        # staged path before any download-associated popup can strand the sync
        # Playwright dispatcher. The persistent context owns eventual page
        # cleanup, matching the regular ASAP export path's no-close behavior.


def _asap_download_with_retry(
    page: Page, frame: Frame, job: dict, staging_dir: Path,
):
    """Retry one transient export-wizard or detached-frame failure.

    A long ASAP run repeatedly opens the same MicroStrategy export wizard. In
    practice, the popup can occasionally open without exposing either of its
    controls to Playwright. Reopening that same export once is safe because no
    download has started at the point where this specific error is raised.
    """
    retryable_prefix = "ASAP Export Wizard opened, but "
    for attempt in range(2):
        try:
            return _asap_download(page, frame, job, staging_dir)
        except PlaywrightError as exc:
            if attempt == 1 or not _asap_frame_was_detached(exc):
                raise
            # A pre-download report-frame replacement is safe to retry. The
            # final export click is handled inside _asap_download so a CSV is
            # never requested twice after its browser download has begun.
            page.wait_for_timeout(1_500)
            frame = _asap_frame(page)
        except RuntimeError as exc:
            message = str(exc)
            if (
                attempt == 1
                or not message.startswith(retryable_prefix)
                or not (
                    "option or Export action was not recognized." in message
                    or "were not recognized in the same dialog." in message
                )
            ):
                raise
            page.wait_for_timeout(1_500)
    raise AssertionError("unreachable")


def _slug_key(value: str, fallback: str) -> str:
    key = re.sub(r"[^a-z0-9]+", "_", value.casefold()).strip("_")
    if not key or not key[0].isalpha():
        key = f"field_{key}" if key else fallback
    return key[:100]


def _unique_visible_text(locator, limit: int = 2000) -> list[str]:
    values = []
    for item in locator.all_inner_texts():
        value = re.sub(r"\s+", " ", item).strip()
        if value and value not in values:
            values.append(value)
        if len(values) >= limit:
            break
    return values


def _clean_text(value: str | None) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def _normalize_asap_filter_label(label: str, control_type: str, options: list[str]) -> str:
    """Repair labels omitted by custom MicroStrategy prompt controls."""
    if (
        control_type == "select"
        and len(options) >= 1
        and len([part for part in label.split(" - ") if part.strip()]) == 3
        and all(len([part for part in value.split(" - ") if part.strip()]) == 3 for value in options)
    ):
        return "Data Configuration"
    return label


def _merge_asap_filter_definition(
    definitions: list[dict], label: str, control_type: str, options: list[str],
) -> None:
    """Add a discovered prompt or merge a second, more complete rendering."""
    label = _clean_text(label).rstrip(":")
    raw_options = list(dict.fromkeys(_clean_text(value) for value in options if _clean_text(value)))
    # Identify unnamed custom controls before discarding text duplicated by the
    # displayed selected value. Otherwise a two-option popup becomes a
    # one-option list and can no longer be recognized as Data Configuration.
    label = _normalize_asap_filter_label(label, control_type, raw_options)
    # The catalog API rejects the whole scan payload when one filter exceeds
    # its field limits, so both the label and the option list are capped at
    # the single point every discovery strategy funnels through.
    label = label[:ASAP_MAX_FILTER_LABEL]
    options = [
        value[:ASAP_MAX_FILTER_LABEL * 2] for value in raw_options
        if value and value != label and not re.fullmatch(r"\(all\)(?:\s*\(\d+\s+values?\))?", value, re.I)
        and "type to search" not in value.casefold()
    ][:ASAP_MAX_FILTER_OPTIONS]
    if not label or not options:
        return
    key = _slug_key(label, f"filter_{len(definitions) + 1}")
    existing = next((item for item in definitions if item["filter_key"] == key), None)
    if existing is not None:
        # Unions must honor the cap too: capping only the first discovery let
        # merged lists grow past the server limit and 422 the whole scan.
        existing["options"] = list(dict.fromkeys(
            [*existing["options"], *options]
        ))[:ASAP_MAX_FILTER_OPTIONS]
        return
    definitions.append({
        "filter_key": key, "label": label, "control_label": label,
        "control_type": control_type, "options": options, "automation": {},
        "required": False, "position": len(definitions),
    })


def _visible_anchor_records(page: Page) -> list[dict]:
    records = []
    selector = "a:visible,button:visible,[role=button]:visible,[role=menuitem]:visible,[onclick]:visible"
    for link in page.locator(selector).all():
        text = _clean_text(link.inner_text())
        box = link.bounding_box()
        if not text or not box:
            continue
        records.append({
            "link": link,
            "text": text,
            "href": link.get_attribute("href") or "",
            "onclick": link.get_attribute("onclick") or "",
            "tag": (link.evaluate("element => element.tagName") or "").casefold(),
            "role": (link.get_attribute("role") or "").casefold(),
            "disabled": not link.is_enabled(),
            "box": box,
        })
    return records


def _wait_for_navigation_roots(page: Page, timeout_ms: int = 120_000) -> list[dict]:
    """Wait for the client-rendered ASAP shell without assuming a menu label."""
    deadline = time.monotonic() + timeout_ms / 1000
    while time.monotonic() < deadline:
        roots = _navigation_roots(_visible_anchor_records(page))
        if roots:
            return roots
        page.wait_for_timeout(500)
    return []


def _navigation_roots(records: list[dict]) -> list[dict]:
    """Find the dense horizontal row of top-level ASAP navigation links."""
    candidates = [
        item for item in records
        if item["box"]["y"] < 180 and item["box"]["height"] < 70 and len(item["text"]) <= 50
    ]
    if not candidates:
        return []
    buckets: dict[int, list[dict]] = {}
    for item in candidates:
        center = item["box"]["y"] + item["box"]["height"] / 2
        buckets.setdefault(round(center / 12), []).append(item)
    row = max(buckets.values(), key=lambda items: len({round(item["box"]["x"] / 20) for item in items}))
    if len(row) < 2:
        return []
    ignored = {"asap", "home", "logout", "log out", "help", "profile"}
    return [
        item for item in sorted(row, key=lambda item: item["box"]["x"])
        if item["text"].casefold() not in ignored
    ]


ASAP_MENU_HEADER_SNAPSHOT_JS = """
() => {
    const results = [];
    document.querySelectorAll("h1,h2,h3,h4,h5,h6,strong,b,label,span,p,div").forEach(el => {
        if (!el || el.offsetWidth <= 0 || el.offsetHeight <= 0) return;
        if (el.closest("a,button,[role=button],[role=menuitem],[onclick]")) return;
        if (el.querySelector("a,button,select,input")) return;
        const text = (el.innerText || "").trim();
        if (!text || text.length > 50 || text.includes("\\n")) return;
        const box = el.getBoundingClientRect();
        results.push({ text: text, x: box.x, y: box.y });
    });
    return results;
}
"""


def _visible_menu_header_records(page: Page) -> list[dict]:
    """Snapshot short non-link text elements that can title a menu column.

    Mega-menu group headers ('SCM Insights', 'AI Insights') are styled text,
    not anchors, so the anchor snapshot never sees them. Without them the
    first report link of a column gets promoted to group and swallows its
    siblings as children.
    """
    try:
        records = page.evaluate(ASAP_MENU_HEADER_SNAPSHOT_JS) or []
    except Exception:
        return []
    return [
        record for record in records
        if isinstance(record, dict) and _clean_text(record.get("text"))
    ]


def _revealed_menu_headers(before: list[dict], after: list[dict]) -> list[dict]:
    seen = {
        (item["text"].casefold(), round(item["x"]), round(item["y"])) for item in before
    }
    return [
        item for item in after
        if (item["text"].casefold(), round(item["x"]), round(item["y"])) not in seen
    ]


def _menu_link_target(item: dict) -> str:
    """The link's navigation payload or semantic UI action marker."""
    href = (item.get("href") or "").strip()
    if href.casefold() in {"#", "javascript:void(0)", "javascript:void(0);", "javascript:;"}:
        href = ""
    target = (href + (item.get("onclick") or "").strip()).casefold()
    if target:
        return target
    if not item.get("disabled") and (
        str(item.get("tag") or "").casefold() == "button"
        or str(item.get("role") or "").casefold() in {"button", "menuitem"}
    ):
        # Modern portal menus attach handlers with addEventListener or their
        # component framework, leaving no inline onclick/href to inspect.
        return "semantic-ui-action"
    return ""


def _menu_report_paths(
    root: dict, before: list[dict], after: list[dict],
    headers: list[dict] | None = None,
) -> list[list[str]]:
    """Convert links revealed by one mega-menu into category/report paths."""
    before_keys = {
        (item["text"].casefold(), item["href"], item["onclick"], round(item["box"]["x"]), round(item["box"]["y"]))
        for item in before
    }
    revealed = [
        item for item in after
        if (item["text"].casefold(), item["href"], item["onclick"],
            round(item["box"]["x"]), round(item["box"]["y"])) not in before_keys
        and item["text"].casefold() != root["text"].casefold()
    ]
    if not revealed:
        return []

    columns: list[list[dict]] = []
    for item in sorted(revealed, key=lambda value: value["box"]["x"]):
        target = next((column for column in columns if abs(
            sum(entry["box"]["x"] for entry in column) / len(column) - item["box"]["x"]
        ) <= 70), None)
        (target if target is not None else columns.append([]) or columns[-1]).append(item)

    header_records = [
        item for item in headers or []
        if _clean_text(item.get("text")).casefold() != root["text"].casefold()
    ]
    paths = []
    for column in columns:
        column.sort(key=lambda item: (item["box"]["y"], item["box"]["x"]))
        first = column[0]
        column_x = sum(entry["box"]["x"] for entry in column) / len(column)
        # A styled text header directly above the column names the group.
        above = [
            item for item in header_records
            if abs(item["x"] - column_x) <= 70 and 0 < first["box"]["y"] - item["y"] <= 160
        ]
        header = max(above, key=lambda item: item["y"]) if above else None
        first_target = _menu_link_target(first)
        if header is not None:
            group = _clean_text(header["text"])
            leaves = [item for item in column if _menu_link_target(item)] or column
        elif len(column) > 1 and not first_target:
            # Legacy menus repeat the group as a non-clickable anchor above
            # its links. A clickable first item is a report, never a group -
            # promoting it would swallow its siblings as children.
            group = first["text"]
            leaves = column[1:]
        else:
            group = None
            leaves = [item for item in column if _menu_link_target(item)]
        for item in leaves:
            path = [root["text"], group, item["text"]] if group else [root["text"], item["text"]]
            if path not in paths:
                paths.append(path)
    return paths


def _asap_extract_js_object(html_text: str, variable: str) -> dict | None:
    """Brace-match one JSON object embedded by the signed-in ASAP portal."""
    match = re.search(rf"(?:const|var|let)\s+{re.escape(variable)}\s*=\s*(\{{)", html_text)
    if not match:
        return None
    start = match.start(1)
    depth = 0
    for index in range(start, len(html_text)):
        char = html_text[index]
        if char == "{":
            depth += 1
        elif char == "}":
            depth -= 1
            if depth == 0:
                try:
                    value = json.loads(html_text[start:index + 1])
                except ValueError:
                    return None
                return value if isinstance(value, dict) else None
    return None


def _asap_portal_session(page: Page, diagnostics: dict | None = None) -> dict | None:
    """Read only the session fields needed for the portal's own menu tree."""
    diagnostics = diagnostics if diagnostics is not None else {}
    common = session = None
    try:
        frames = list(page.frames)
    except Exception:
        frames = []
    diagnostics["frame_count"] = len(frames)
    diagnostics["frames"] = []
    for frame_index, frame in enumerate(frames):
        frame_detail: dict[str, Any] = {"index": frame_index}
        try:
            frame_detail["url"] = str(frame.url)[:300]
        except Exception:
            pass
        try:
            data = frame.evaluate(
                "() => {"
                " const common = typeof _COMMON_INFO !== 'undefined' ? _COMMON_INFO : window._COMMON_INFO;"
                " const session = typeof _SESSION !== 'undefined' ? _SESSION : window._SESSION;"
                " return common && session ? {common, session} : null;"
                "}"
            )
        except Exception as exc:
            data = None
            frame_detail["evaluate_error"] = type(exc).__name__
        frame_detail["evaluate_session"] = bool(
            isinstance(data, dict) and data.get("common") and data.get("session")
        )
        if isinstance(data, dict) and data.get("common") and data.get("session"):
            common, session = data["common"], data["session"]
            frame_detail["source"] = "evaluate"
            diagnostics["frames"].append(frame_detail)
            diagnostics["session_source"] = f"frame_{frame_index}_evaluate"
            break
        try:
            frame_html = frame.content()
        except Exception as exc:
            frame_detail["content_error"] = type(exc).__name__
            diagnostics["frames"].append(frame_detail)
            continue
        frame_detail["content_common_marker"] = "_COMMON_INFO" in frame_html
        frame_detail["content_session_marker"] = "_SESSION" in frame_html
        frame_common = _asap_extract_js_object(frame_html, "_COMMON_INFO")
        frame_session = _asap_extract_js_object(frame_html, "_SESSION")
        if frame_common and frame_session:
            common, session = frame_common, frame_session
            frame_detail["source"] = "content"
            diagnostics["frames"].append(frame_detail)
            diagnostics["session_source"] = f"frame_{frame_index}_content"
            break
        diagnostics["frames"].append(frame_detail)
    if common is None or session is None:
        try:
            html_text = page.content()
        except Exception as exc:
            diagnostics["page_content_error"] = type(exc).__name__
            return None
        diagnostics["page_common_marker"] = "_COMMON_INFO" in html_text
        diagnostics["page_session_marker"] = "_SESSION" in html_text
        common = _asap_extract_js_object(html_text, "_COMMON_INFO")
        session = _asap_extract_js_object(html_text, "_SESSION")
        if common and session:
            diagnostics["session_source"] = "page_content"
    if not common or not session:
        diagnostics["session_result"] = "objects_missing"
        return None
    legacy_token = session.get("MSTRWEB_AUTH_TOKEN_ENC")
    menu_id = common.get("MSTR_MAIN_MENU_ID")
    diagnostics["has_legacy_token"] = bool(legacy_token)
    diagnostics["has_main_menu_id"] = bool(menu_id)
    if not legacy_token or not menu_id:
        diagnostics["session_result"] = "required_fields_missing"
        return None
    parts = urlsplit(page.url)
    if not parts.scheme or not parts.netloc:
        diagnostics["session_result"] = "page_url_invalid"
        return None
    web_path = str(common.get("MSTR_CUSTOM_WEB_CONTEXT_PATH") or "/mstr").rstrip("/")
    diagnostics["session_result"] = "available"
    return {
        "web_base": f"{parts.scheme}://{parts.netloc}{web_path}",
        "legacy_token": legacy_token,
        "main_menu_id": menu_id,
    }


def _asap_clean_portal_menu_name(value: Any) -> str:
    return _clean_text(re.sub(r"^\d+\.", "", _clean_text(str(value or ""))))


def _asap_portal_menu_paths(page: Page, diagnostics: dict | None = None) -> list[list[str]]:
    """Walk the role-specific menu tree used by the visible ASAP header.

    This is a live browser-session fallback for branches whose mega-menu is
    blank. It never reads the retained Metronome catalog and never persists
    the session token embedded by the portal.
    """
    diagnostics = diagnostics if diagnostics is not None else {}
    context = _asap_portal_session(page, diagnostics)
    if not context:
        return []

    def fetch(folder_id: str, depth: int) -> dict | None:
        try:
            response = page.request.post(
                f"{context['web_base']}/menuInfo.do?folderId={folder_id}",
                headers={"Accept": "application/json", "X-Requested-With": "XMLHttpRequest"},
                form={"authToken": context["legacy_token"], "depth": str(depth)},
                timeout=20_000,
            )
            value = response.json() if response.ok else None
        except Exception as exc:
            diagnostics.setdefault("fetch_errors", []).append({
                "folder": "root" if folder_id == str(context["main_menu_id"]) else "branch",
                "error": type(exc).__name__,
            })
            return None
        if not response.ok:
            diagnostics.setdefault("fetch_statuses", []).append({
                "folder": "root" if folder_id == str(context["main_menu_id"]) else "branch",
                "status": getattr(response, "status", None),
            })
        if isinstance(value, dict) and "data" in value:
            error_code = value.get("errorCode")
            if error_code not in (None, 0, "0") and str(error_code).casefold() != "success":
                diagnostics.setdefault("menu_error_codes", []).append(str(error_code)[:100])
                return None
            value = value.get("data")
            if isinstance(value, str):
                try:
                    value = json.loads(value)
                except ValueError:
                    diagnostics.setdefault("fetch_errors", []).append({
                        "folder": "root" if folder_id == str(context["main_menu_id"]) else "branch",
                        "error": "InvalidWrappedJson",
                    })
                    return None
            if isinstance(value, list):
                value = {"children": value}
        return value if isinstance(value, dict) else None

    def children(node: dict) -> list[dict]:
        values = node.get("children") or node.get("child") or []
        return [value for value in values if isinstance(value, dict)]

    paths: list[list[str]] = []

    def walk(node: dict, parents: list[str]) -> None:
        name = _asap_clean_portal_menu_name(node.get("name"))
        path = [*parents, name] if name else list(parents)
        descendants = children(node)
        if descendants:
            for descendant in descendants:
                walk(descendant, path)
        elif node.get("id") and len(path) >= 2 and path not in paths:
            paths.append(path)

    root = fetch(str(context["main_menu_id"]), 1)
    if not root:
        diagnostics["menu_result"] = "root_fetch_failed"
        return []
    diagnostics["root_keys"] = sorted(str(key) for key in root.keys())[:30]
    top_nodes = children(root)
    if not top_nodes:
        # Some portal builds interpret depth as the number of descendant
        # levels below the requested node, returning only the root at depth 1.
        # Retry the same role-specific root at depth 2 before giving up.
        diagnostics["root_depth_retry"] = True
        deeper_root = fetch(str(context["main_menu_id"]), 2)
        if deeper_root:
            diagnostics["depth_2_root_keys"] = sorted(
                str(key) for key in deeper_root.keys()
            )[:30]
            deeper_nodes = children(deeper_root)
            if deeper_nodes:
                root = deeper_root
                top_nodes = deeper_nodes
    diagnostics["top_node_count"] = len(top_nodes)
    diagnostics["top_names"] = [_asap_clean_portal_menu_name(top.get("name")) for top in top_nodes]
    for top in top_nodes:
        subtree = fetch(str(top.get("id")), 2) if top.get("id") else None
        top_name = _asap_clean_portal_menu_name(top.get("name"))
        subtree_name = _asap_clean_portal_menu_name((subtree or {}).get("name"))
        if subtree and top_name and subtree_name.casefold() != top_name.casefold():
            walk(subtree, [top_name])
        else:
            walk(subtree or top, [])
    diagnostics["menu_result"] = "paths_found" if paths else "no_leaf_paths"
    diagnostics["path_count"] = len(paths)
    return paths


def _asap_discover_menu_reports(page: Page, scope: list[str]) -> list[list[str]]:
    # Discovery is deliberately page-driven. The old implementation required a
    # configured label such as "Mobile", which made every navigation rename a
    # deployment incident. Scope remains in the job schema for compatibility,
    # but the scanner now inventories every top-level menu it can see.
    roots = _wait_for_navigation_roots(page)
    if not roots:
        raise RuntimeError(
            f"ASAP top-level navigation did not render within 120 seconds "
            f"(URL: {page.url}, title: {_clean_text(page.title())})."
        )
    # Capture the role-specific portal tree while the signed-in landing page
    # still owns the MicroStrategy session handoff. Clicking a blank branch can
    # replace the document with a report route that no longer exposes it.
    portal_diagnostics: dict[str, Any] = {}
    portal_paths = _asap_portal_menu_paths(page, portal_diagnostics)
    root_names = list(dict.fromkeys(root["text"] for root in roots))
    paths: list[list[str]] = []
    missing_roots: list[str] = []
    for root_name in root_names:
        # The portal renders its navigation shell before removing the blocking
        # loading overlay. Playwright can therefore resolve a menu link while
        # the overlay still intercepts the click. Wait for a sustained clear
        # state before resolving the live link used for this interaction.
        _asap_wait_for_loading_clear(page)
        # Mega-menu contents change the number and order of matching elements.
        # Re-resolve the navigation trigger instead of retaining an nth-based
        # Playwright locator from the initial DOM snapshot.
        current_roots = _navigation_roots(_visible_anchor_records(page))
        root = next((item for item in current_roots if item["text"] == root_name), None)
        if root is None:
            missing_roots.append(f"{root_name} (top-level control disappeared)")
            continue
        before = _visible_anchor_records(page)
        before_headers = _visible_menu_header_records(page)
        root["link"].click(timeout=15_000)
        def revealed_paths_after_interaction(timeout_seconds: int) -> list[list[str]]:
            after = before
            revealed_headers: list[dict] = []
            stable_signature: tuple[tuple[str, ...], ...] | None = None
            stable_polls = 0
            reveal_deadline = time.monotonic() + timeout_seconds
            while time.monotonic() < reveal_deadline:
                page.wait_for_timeout(200)
                after = _visible_anchor_records(page)
                revealed_headers = _revealed_menu_headers(
                    before_headers, _visible_menu_header_records(page),
                )
                revealed_paths = _menu_report_paths(root, before, after, revealed_headers)
                signature = tuple(tuple(path) for path in revealed_paths)
                if signature and signature == stable_signature:
                    stable_polls += 1
                else:
                    stable_signature = signature
                    stable_polls = 0
                # ASAP builds large mega-menus incrementally. The first
                # non-empty snapshot can contain only a few columns, so wait
                # until it remains unchanged for roughly one second.
                if signature and stable_polls >= 5:
                    return revealed_paths
            return _menu_report_paths(root, before, after, revealed_headers)

        root_paths = revealed_paths_after_interaction(10)
        if not root_paths:
            # Some ASAP navigation builds reveal mega-menus on hover and use a
            # click only for direct navigation. Support both without assuming
            # one interaction model for every top-level branch.
            try:
                root["link"].hover(timeout=10_000)
                root_paths = revealed_paths_after_interaction(5)
            except Exception:
                root_paths = []
        if not root_paths:
            missing_roots.append(f"{root_name} (no report controls revealed)")
        for path in root_paths:
            if path not in paths:
                paths.append(path)
        # Close the menu before measuring the next root. Clicking the active
        # trigger is reversible and avoids confusing links from two menus.
        try:
            root["link"].click(timeout=5_000)
            page.wait_for_timeout(150)
        except Exception:
            pass
    if missing_roots:
        recovered_roots: set[str] = set()
        for missing in missing_roots:
            root_name = missing.split(" (", 1)[0]
            recovered = [
                path for path in portal_paths
                if path and path[0].casefold() == root_name.casefold()
            ]
            if recovered:
                recovered_roots.add(root_name.casefold())
                for path in recovered:
                    if path not in paths:
                        paths.append(path)
        missing_roots = [
            missing for missing in missing_roots
            if missing.split(" (", 1)[0].casefold() not in recovered_roots
        ]
    if missing_roots:
        portal_root_names = sorted({path[0] for path in portal_paths if path})
        fallback_detail = (
            "Portal-session menu fallback returned no paths: "
            + json.dumps(portal_diagnostics, ensure_ascii=True, separators=(",", ":")) + "."
            if not portal_root_names else
            "Portal-session menu fallback roots: " + ", ".join(portal_root_names) + "."
        )
        raise RuntimeError(
            "ASAP menu discovery was incomplete. Refusing to mark unseen reports stale. "
            "Missing branches: " + "; ".join(missing_roots) + ". " + fallback_detail
        )
    if not paths:
        raise RuntimeError("ASAP navigation was detected, but no report links were revealed.")
    return paths


def _asap_owned_popup_roots(frame: Frame, control) -> list:
    """Resolve only popups owned by the active combobox when possible."""
    roots = []
    owner_ids = _clean_text(
        control.get_attribute("aria-controls") or control.get_attribute("aria-owns")
    ).split()
    for owner_id in owner_ids:
        owned = frame.locator(f"[id={json.dumps(owner_id)}]")
        for index in range(owned.count()):
            candidate = owned.nth(index)
            if candidate.is_visible():
                roots.append(candidate)
    if roots:
        return roots

    # Legacy Select2 can omit ownership attributes. Its open dropdown is still
    # a bounded listbox/dropdown, so use that container rather than scanning
    # every list item in the report frame.
    candidates = frame.locator(
        "[role=listbox]:visible,.select2-dropdown:visible,.select2-results:visible"
    )
    return [
        candidates.nth(index) for index in range(candidates.count())
        if candidates.nth(index).is_visible()
    ]


ASAP_SELECT_SNAPSHOT_JS = """
() => {
    const isVisible = (el) => !!el && el.offsetWidth > 0 && el.offsetHeight > 0;
    const cleanLabel = (text) => {
        const line = String(text || "").trim().split("\\n")[0].trim();
        return line.length > 0 && line.length <= 80 ? line : "";
    };
    const results = [];
    document.querySelectorAll("select").forEach((select, index) => {
        const sibling = select.nextElementSibling;
        const select2 = sibling && String(sibling.className || "").toLowerCase().includes("select2")
            ? sibling : null;
        const selfVisible = isVisible(select);
        const select2Visible = isVisible(select2);
        const parentVisible = isVisible(select.parentElement);
        // Select2 deliberately hides its owning select, so activity is judged
        // by the rendered widget or the immediate container. A select whose
        // whole container is hidden belongs to another loaded report tab.
        if (!selfVisible && !select2Visible && !parentVisible) return;
        let label = "";
        let source = "";
        if (select.id) {
            const forLabel = document.querySelector('label[for="' + CSS.escape(select.id) + '"]');
            const text = forLabel ? cleanLabel(forLabel.innerText) : "";
            if (text) { label = text; source = "label_for"; }
        }
        if (!label) {
            const aria = cleanLabel(select.getAttribute("aria-label"));
            if (aria) { label = aria; source = "aria_label"; }
        }
        if (!label) {
            // Walk a few ancestors for a field title, but only accept a
            // container that holds exactly this one select: a shared section
            // title would name several filters identically, and identical
            // labels merge into one definition with mixed option lists.
            let holder = select.parentElement;
            for (let depth = 0; holder && depth < 4 && !label; depth += 1) {
                if (holder.querySelectorAll("select").length === 1) {
                    const title = holder.querySelector("label,legend,span.title,span.label,p,h1,h2,h3,h4,h5,h6");
                    const text = title ? cleanLabel(title.innerText) : "";
                    if (text) { label = text; source = "container"; }
                }
                holder = holder.parentElement;
            }
        }
        if (!label) {
            let previous = select.previousElementSibling;
            while (previous && !label) {
                if (["LABEL", "SPAN", "DIV", "P", "H1", "H2", "H3", "H4", "H5", "H6"].includes(previous.tagName)) {
                    const text = cleanLabel(previous.innerText);
                    if (text) { label = text; source = "sibling"; }
                }
                previous = previous.previousElementSibling;
            }
        }
        results.push({
            index: index,
            id: select.id || "",
            name: select.getAttribute("name") || "",
            label: label,
            label_source: source,
            multiple: !!select.multiple,
            select2: !!select2,
            visible: selfVisible || select2Visible,
            options: Array.from(select.options).map(option => String(option.text || "").trim()).filter(Boolean),
        });
    });
    return results;
}
"""


def _asap_filter_frames(frame: Frame) -> list[Frame]:
    """The active report frame first, then every other live frame.

    Some ASAP reports render their prompt sidebar in the outer portal shell
    or in a nested MicroStrategy frame rather than in the frame that shows
    the grid. A single-frame scan silently misses every filter of such a
    report, so the select sweep covers all of them.
    """
    frames = [frame]
    try:
        for candidate in frame.page.frames:
            if candidate not in frames and not candidate.is_detached():
                frames.append(candidate)
    except Exception:
        pass
    return frames


def _asap_native_select_records(frame: Frame, diagnostics: dict | None = None) -> list[dict]:
    """Snapshot every active <select> across the report's frames."""
    records: list[dict] = []
    seen = set()
    frames_scanned = 0
    for root in _asap_filter_frames(frame):
        try:
            found = root.evaluate(ASAP_SELECT_SNAPSHOT_JS) or []
        except Exception:
            continue
        frames_scanned += 1
        for record in found:
            if not isinstance(record, dict):
                continue
            identity = (
                record.get("id"), record.get("name"), record.get("label"),
                tuple(record.get("options") or []),
            )
            if identity in seen:
                continue
            seen.add(identity)
            record["frame_index"] = frames_scanned - 1
            records.append(record)
    if diagnostics is not None:
        diagnostics["frames_scanned"] = frames_scanned
        diagnostics["selects_seen"] = len(records)
    return records


ASAP_DOWNLOAD_LINK_SNAPSHOT_JS = """
() => {
    const results = [];
    document.querySelectorAll("a,button,[role=button]").forEach(el => {
        if (!el || el.offsetWidth <= 0 || el.offsetHeight <= 0) return;
        const text = (el.innerText || el.getAttribute("title") || el.getAttribute("aria-label") || "").trim();
        const href = el.getAttribute("href") || "";
        const hasDownloadAttr = el.getAttribute("download") !== null;
        const fileHref = /\\.(csv|xlsx|xls|zip|txt|pdf)([?#]|$)/i.test(href);
        if (!hasDownloadAttr && !fileHref && !/download/i.test(text)) return;
        if (text.length > 120) return;
        results.push({
            label: text || "Download",
            href: href.slice(0, 2000),
            download_attr: hasDownloadAttr,
        });
    });
    return results;
}
"""


def _asap_discover_download_links(frame: Frame) -> list[dict]:
    """Catalog download hyperlinks inside embedded HTML dashboards.

    Some ASAP panes embed plain HTML dashboards instead of MicroStrategy
    prompt reports; their data leaves through download hyperlinks rather
    than the Export Wizard. Sweep every live frame for visible controls that
    carry a download attribute, a file-typed href, or download wording.
    """
    links: list[dict] = []
    seen = set()
    for root in _asap_filter_frames(frame):
        try:
            found = root.evaluate(ASAP_DOWNLOAD_LINK_SNAPSHOT_JS) or []
        except Exception:
            continue
        for record in found:
            if not isinstance(record, dict):
                continue
            label = _clean_text(record.get("label"))[:120]
            href = str(record.get("href") or "")[:2000]
            identity = (label.casefold(), href)
            if not label or identity in seen:
                continue
            seen.add(identity)
            links.append({
                "label": label, "href": href,
                "download_attr": bool(record.get("download_attr")),
            })
            if len(links) >= ASAP_MAX_DOWNLOAD_LINKS:
                return links
    return links


def _asap_discover_filters(frame: Frame, diagnostics: dict | None = None) -> list[dict]:
    definitions = []
    if diagnostics is None:
        diagnostics = {}

    def add_definition(
        label: str, control_type: str, options: list[str], automation: dict | None = None,
    ):
        _merge_asap_filter_definition(definitions, label, control_type, options)
        if automation:
            # The merge normalizes labels (unnamed three-part selects become
            # Data Configuration), so the lookup key must be derived the same
            # way or automation for renamed filters silently vanishes.
            cleaned = _clean_text(label).rstrip(":")
            raw_options = list(dict.fromkeys(
                _clean_text(value) for value in options if _clean_text(value)
            ))
            normalized = _normalize_asap_filter_label(cleaned, control_type, raw_options)
            key = _slug_key(normalized, f"filter_{len(definitions)}")
            definition = next(
                (item for item in definitions if item["filter_key"] == key), None,
            )
            if definition is not None:
                definition["automation"] = {**definition.get("automation", {}), **automation}

    def nearest_list_values(label_locator, *, require_search_marker: bool = False) -> list[str]:
        """Read the smallest visible MicroStrategy control containing a label."""
        ancestor = label_locator
        short_fallback: list[str] = []
        for _ in range(7):
            ancestor = ancestor.locator("xpath=parent::*")
            if not ancestor.count():
                break
            lines = list(dict.fromkeys(
                _clean_text(line) for line in ancestor.first.inner_text().splitlines() if _clean_text(line)
            ))
            has_marker = any(
                "type to search" in line.casefold() or re.search(r"\(\d+\s+values?\)", line, re.I)
                for line in lines
            )
            if has_marker or not require_search_marker:
                if len(lines) >= 3:
                    return lines[1:500]
                # Two lines are a label plus a single member. Keep climbing in
                # case a wider ancestor holds the full member list, but retain
                # this as a fallback so a legitimate one-value list survives.
                if len(lines) == 2 and not short_fallback:
                    short_fallback = lines[1:]
        return short_fallback

    def wait_for_popup_options(control, timeout_ms: int = 5_000) -> list[str]:
        """Collect leaf options from the active asynchronous popup."""
        selector = (
            "[role=option]:visible,option:visible,"
            "li.select2-results__option:visible,li.select2-result:visible"
        )
        collected = []
        deadline = time.monotonic() + timeout_ms / 1000
        stable_since = None
        while time.monotonic() < deadline:
            current = []
            for root in _asap_owned_popup_roots(frame, control):
                current = list(dict.fromkeys([
                    *current, *_unique_visible_text(root.locator(selector), 500),
                ]))
            merged = list(dict.fromkeys([*collected, *current]))
            if merged != collected:
                collected = merged
                stable_since = time.monotonic()
            elif collected and stable_since is not None and time.monotonic() - stable_since >= 1.5:
                break
            frame.page.wait_for_timeout(250)
        return collected
    # Native controls are preferred because they expose complete option lists
    # without opening the control or changing report state. Select2
    # deliberately hides its owning select, and ASAP rarely wires labels
    # through label[for]/aria - they live in the surrounding field container
    # or a preceding sibling - so the snapshot resolves labels in the DOM and
    # a real prompt is NEVER dropped for want of one: a select with options
    # but no label gets a stable synthesized name instead, because runtime
    # selection locates native selects by their options, not their label.
    unlabeled = 0
    label_sources: dict[str, int] = {}
    for record in _asap_native_select_records(frame, diagnostics):
        options = list(dict.fromkeys(
            _clean_text(value) for value in record.get("options") or [] if _clean_text(value)
        ))
        label = _clean_text(record.get("label"))
        source = record.get("label_source") or ""
        # ASAP's Select2 Data Configuration owner has no accessible name. Its
        # three-part region choices identify the prompt without hardcoding any
        # actual region value into the repository.
        if not label and len(options) >= 2 and all(
            len([part for part in value.split(" - ") if part.strip()]) == 3
            for value in options
        ):
            label = options[0]
            source = "three_part_options"
        if not label:
            label = _clean_text(record.get("id")) or _clean_text(record.get("name"))
            source = "select_attribute" if label else source
        if not label and options:
            unlabeled += 1
            # The name is qualified by frame so two unlabeled selects at the
            # same index in different frames never share a label - identical
            # labels would merge into one definition with mixed options.
            frame_index = record.get("frame_index") or 0
            position = record.get("index", len(definitions)) + 1
            label = f"Filter {position}" if not frame_index else f"Filter {frame_index + 1}.{position}"
            source = "synthesized"
        if not label:
            unlabeled += 1
            continue
        label_sources[source or "unknown"] = label_sources.get(source or "unknown", 0) + 1
        control_type = "multi_select" if record.get("multiple") else "select"
        add_definition(label, control_type, options, automation={
            "select_id": record.get("id") or None,
            "select_name": record.get("name") or None,
            "label_source": source or None,
            "select2": bool(record.get("select2")),
        })
    diagnostics["selects_unlabeled"] = unlabeled
    diagnostics["label_sources"] = label_sources

    # New ASAP renders some selects as asynchronous Select2 comboboxes. Open
    # them long enough for the remote results to settle, then restore the page
    # with Escape. A fixed 150 ms snapshot missed late-arriving values.
    custom_selects = frame.locator(
        "[role=combobox]:visible,button[aria-haspopup=listbox]:visible,input[aria-haspopup=listbox]:visible"
    )
    for control in custom_selects.all():
        label = _clean_text(control.get_attribute("aria-label") or control.get_attribute("name"))
        if not label:
            lines = nearest_list_values(control)
            label = lines[0] if lines else ""
        try:
            control.click(timeout=3_000)
            options = wait_for_popup_options(control)
        except Exception:
            options = []
        finally:
            try:
                frame.page.keyboard.press("Escape")
            except Exception:
                pass
        add_definition(label, "select", options)

    # Searchable member selectors expose their label and values as plain divs,
    # without heading or option roles. Their search/count marker is the stable
    # structural signal across reports.
    search_markers = frame.get_by_text(re.compile(r"type to search|\(\d+\s+values?\)", re.I))
    for marker in search_markers.all():
        block = marker
        for _ in range(6):
            block = block.locator("xpath=parent::*")
            if not block.count():
                break
            lines = list(dict.fromkeys(
                _clean_text(line) for line in block.first.inner_text().splitlines() if _clean_text(line)
            ))
            marker_index = next((i for i, line in enumerate(lines) if
                                 "type to search" in line.casefold() or re.search(r"\(\d+\s+values?\)", line, re.I)), -1)
            if marker_index > 0 and len(lines) > marker_index + 1:
                label = lines[marker_index - 1]
                control_type = "week" if "week" in label.casefold() else "multi_select"
                add_definition(label, control_type, lines[marker_index + 1:500])
                break

    # Dimension pickers use a named member list but no ARIA roles. The adapter
    # recognizes the UI control label, while every option remains page-driven.
    for dimension_label in frame.get_by_text(re.compile(r"^dimension:?$", re.I)).all():
        add_definition("Dimension", "multi_select", nearest_list_values(dimension_label))

    # Flagship Experience exposes Measure with the same visual-only member
    # list as Dimension. It has no native select, ARIA options, or search/count
    # marker, so it must be anchored by its visible portal title as well.
    for measure_label in frame.get_by_text(re.compile(r"^measure:?$", re.I)).all():
        add_definition("Measure", "multi_select", nearest_list_values(measure_label))

    # Regional FOTA renders Category as the same visual-only member list used
    # by Dimension. It has no search/count marker or native select, so the
    # generic discovery paths above cannot see it. Capture the two live report
    # members explicitly; execution still resolves and verifies the visible
    # list before changing its exact selection.
    category_options = []
    for category_label in frame.get_by_text(re.compile(r"^category:?$", re.I)).all():
        for value in nearest_list_values(category_label):
            if value.casefold() in {"weekly", "daily"} and value not in category_options:
                category_options.append(value)
    if category_options:
        add_definition("Category", "multi_select", category_options)

    # Regional FOTA exposes Week and Date as coupled two-handle sliders rather
    # than searchable member lists. Catalog the complete Week domain while
    # recording that execution must also drive the derived Date range.
    week_slider = _asap_discover_week_slider(frame)
    if week_slider:
        options, automation = week_slider
        add_definition("Week", "week", options, automation)

    # Some reports title the same YYYYWW control Period. Its two slider
    # handles also expose generated hex ids and their current value as hidden
    # native-select labels. Remove those structural artifacts and retain the
    # one portal-authored title shown to the user.
    period_slider = _asap_discover_period_slider(frame)
    if period_slider:
        definitions[:] = [
            definition for definition in definitions
            if not re.fullmatch(
                r"(?:[0-9a-f]{24,}(?:[_-][0-9a-f]{24,})*|20\d{4})",
                definition["label"], re.I,
            )
        ]
        for position, definition in enumerate(definitions):
            definition["position"] = position
        options, automation = period_slider
        add_definition("Period", "week", options, automation)

    # The Installed Base report exposes its week prompt as a searchable
    # MicroStrategy member list. Depending on render timing the count/search
    # marker may not be returned as its own text node, so anchor discovery on
    # the stable semantic label as well.
    for week_label in frame.get_by_text(re.compile(r"^sell-out week:?$", re.I)).all():
        week_values = [
            value for value in nearest_list_values(week_label)
            if re.fullmatch(r"20\d{4}", value)
        ]
        add_definition("Sell-out Week", "week", week_values)

    # MicroStrategy list selectors are represented by a heading followed by a
    # member list. Detect the labels from the report's own prompt headings and
    # capture visible members without selecting them.
    prompt_labels = _unique_visible_text(frame.locator("h1:visible,h2:visible,h3:visible,h4:visible,[role=heading]:visible"), 300)
    ignored = {"run", "data rows", "data columns", "export wizard", "select dimensions"}
    for label in prompt_labels:
        normalized = label.casefold().rstrip(":")
        if not label or any(token in normalized for token in ignored):
            continue
        if any(item["label"].casefold() == normalized for item in definitions):
            continue
        heading = frame.get_by_text(label, exact=True).first
        following = heading.locator("xpath=following::*[@role='option' or self::option or self::li][position() <= 500]")
        options = _unique_visible_text(following, 500) if following.count() else []
        if not options:
            continue
        control_type = "week" if "week" in normalized else "multi_select"
        add_definition(label, control_type, options)
    diagnostics["definitions"] = len(definitions)
    return definitions


def _asap_export_view_candidates(page: Page, frame: Frame) -> list[tuple[str, Any]]:
    """Return every visible export view without changing the report."""
    visible = []
    # The current portal places this control in the outer ASAP shell, while
    # some older reports render it inside a nested MicroStrategy frame. The
    # control can also arrive after the report iframe itself is attached.
    deadline = time.monotonic() + 12
    while time.monotonic() < deadline and not visible:
        roots = list(dict.fromkeys([page.main_frame, frame, *page.frames]))
        for root in roots:
            candidates = root.get_by_text(re.compile(r"^Export Wizard(?:\s*\([^)]*\))?$", re.I))
            for candidate in candidates.all():
                try:
                    if candidate.is_visible():
                        label = _clean_text(candidate.inner_text())
                        if label and label.casefold() not in {item[0].casefold() for item in visible}:
                            visible.append((label, candidate))
                except Exception:
                    continue
        if not visible:
            page.wait_for_timeout(250)
    return visible


def _asap_activate_export_view(
    page: Page, frame: Frame, requested_label: str | None = None,
) -> tuple[Frame, str | None]:
    """Open one exact export-oriented view using visible ASAP semantics."""
    # The report shell exposes its tabs before the default report has finished
    # loading. Wait until the overlay is stably gone, then resolve the tab
    # again so the click cannot target a control from a replaced frame.
    _asap_wait_for_loading_clear(page)
    visible = _asap_export_view_candidates(page, _asap_frame(page))
    if not visible:
        if requested_label:
            raise RuntimeError(f"ASAP export view is no longer visible: {requested_label}")
        return frame, None
    if requested_label:
        selected = next(
            (item for item in visible if item[0].casefold() == requested_label.casefold()), None,
        )
        if selected is None:
            available = ", ".join(item[0] for item in visible)
            raise RuntimeError(
                f"ASAP export view is no longer visible: {requested_label}. Available: {available}"
            )
        label, control = selected
    else:
        label, control = next(
            (item for item in visible if "detail" in item[0].casefold()), visible[0],
        )
    control.click(timeout=30_000)
    page.wait_for_timeout(500)
    _asap_wait_for_loading_clear(page)
    active_frame = _asap_frame(page)
    deadline = time.monotonic() + 10
    while time.monotonic() < deadline:
        markers = active_frame.get_by_text(
            re.compile(r"^(?:RUN|Dimension:?|Data Configuration:?|Sell-out Week:?)$", re.I)
        )
        if any(item.is_visible() for item in markers.all()):
            break
        page.wait_for_timeout(250)
    return active_frame, label


def discover_asap_catalog(page: Page, job: dict, report_progress, profile_dir: Path) -> tuple[list[dict], list[dict], bool]:
    timings = _Timings()
    deadline = time.monotonic() + 60 * int(job["discovery"].get("max_duration_minutes") or 90)
    site = job["site"]
    report_progress("running", {"stage": "navigation", "message": "Opening ASAP for catalog discovery."})
    with timings.measure("navigation"):
        _asap_goto(page, site.get("auth_url") or site.get("base_url"), profile_dir)
    target_paths = [path for path in job["discovery"].get("report_paths", []) if len(path) >= 2]
    with timings.measure("report_discovery"):
        paths = target_paths or _asap_discover_menu_reports(
            page, job["discovery"].get("scope") or ["Mobile"]
        )
    if job["discovery"].get("mode") == "partial" and not target_paths:
        # A partial scan inventories the menu only: names and paths, without
        # opening a single report. It is the fast way to see what exists;
        # filters come from a targeted scan of the one report being
        # configured. Entries are marked so the catalog merges them over
        # what a full scan already discovered instead of replacing it.
        report_progress("running", {
            "stage": "partial_scan",
            "message": (
                f"Partial scan: cataloguing {len(paths)} report(s) from the portal menu "
                f"without opening them. Use Scan report in the flow builder for filters."
            ),
        })
        partial_reports = [
            {
                "discovery_key": " > ".join(path), "name": path[-1],
                "report_url": site.get("base_url") or site.get("auth_url"),
                "ready_text": None, "download_text": "Export CSV",
                "automation": {"category_path": path, "scan_mode": "partial"},
                "filters": [],
            }
            for path in paths
        ]
        return (
            partial_reports,
            timings.finish(item_count=len(partial_reports), status="succeeded"),
            True,
        )
    reports = []
    failures = []
    complete = not target_paths
    for index, path in enumerate(paths, start=1):
        if time.monotonic() >= deadline:
            complete = False
            report_progress("running", {
                "stage": "time_budget_reached",
                "message": "The 90-minute scan budget was reached. Keeping partial discoveries without marking unseen entries stale.",
                "report_index": index, "report_count": len(paths),
            })
            break
        report_progress("running", {
            "stage": "filter_inspection", "message": f"Inspecting report {index} of {len(paths)}.",
            "current_report": path[-1], "report_index": index, "report_count": len(paths),
        })
        lightweight_job = {
            "site": site,
            "report": {
                "name": path[-1], "url": site.get("base_url") or site.get("auth_url"),
                "ready_text": None, "automation": {"category_path": path},
            },
        }
        try:
            with timings.measure("report_navigation"):
                frame = _asap_open_report(page, lightweight_job, profile_dir)
                export_view_labels = [
                    label for label, _control in _asap_export_view_candidates(page, frame)
                ]
            filters = []
            export_views = []
            view_diagnostics = {}
            export_capability_views: dict[str, dict] = {}
            download_links: list[dict] = []
            ready_text = export_view_labels[0] if export_view_labels else None
            report_title = path[-1]
            with timings.measure("filter_inspection"):
                labels_to_scan = export_view_labels or [None]
                for view_index, export_view_label in enumerate(labels_to_scan):
                    if view_index:
                        frame = _asap_open_report(page, lightweight_job, profile_dir)
                    active_frame, selected_label = _asap_activate_export_view(
                        page, frame, export_view_label,
                    )
                    diagnostics: dict = {}
                    discovered = _asap_discover_filters(active_frame, diagnostics)
                    if view_index == 0:
                        # Embedded HTML dashboards expose their data through
                        # download hyperlinks instead of the Export Wizard.
                        download_links = _asap_discover_download_links(active_frame)
                        diagnostics["download_links"] = len(download_links)
                    view_diagnostics[selected_label or export_view_label or "default"] = diagnostics
                    view_filter_keys = []
                    for definition in discovered:
                        view_filter_keys.append(definition["filter_key"])
                        existing = next(
                            (item for item in filters if item["filter_key"] == definition["filter_key"]),
                            None,
                        )
                        if existing is None:
                            filters.append(definition)
                        else:
                            existing["options"] = list(dict.fromkeys([
                                *existing.get("options", []), *definition.get("options", []),
                            ]))[:ASAP_MAX_FILTER_OPTIONS]
                    if not discovered and not download_links:
                        # An empty result is a discovery defect until proven
                        # otherwise - surface it in the scan log with enough
                        # detail to diagnose instead of passing silently.
                        report_progress("running", {
                            "stage": "filter_inspection_empty",
                            "message": (
                                f"No filters were discovered for {path[-1]}"
                                + (f" ({selected_label})" if selected_label else "")
                                + f". Diagnostics: {diagnostics}"
                            ),
                            "report_index": index, "report_count": len(paths),
                        })
                    if selected_label:
                        export_views.append({
                            "label": selected_label, "filter_keys": view_filter_keys,
                        })
                    capability_key = selected_label or export_view_label or "__default__"
                    remaining_ms = max(0, int((deadline - time.monotonic()) * 1000))
                    try:
                        if remaining_ms < 1_000:
                            raise RuntimeError("Catalog scan budget was exhausted before Export Wizard inspection.")
                        export_capability_views[capability_key] = _asap_probe_export_capabilities(
                            page, timeout_ms=min(8_000, remaining_ms),
                        )
                    except Exception as capability_exc:
                        message = str(capability_exc)[:1_000]
                        export_capability_views[capability_key] = {
                            "status": "unknown", "error": message,
                        }
                        diagnostics["export_capabilities_error"] = message
                        report_progress("running", {
                            "stage": "export_capability_warning",
                            "message": (
                                f"{path[-1]}"
                                + (f" ({capability_key})" if capability_key != "__default__" else "")
                                + f": Export Wizard options could not be inventoried; keeping the report. {message}"
                            ),
                            "report_index": index, "report_count": len(paths),
                        })
                    if view_index == 0:
                        report_title = active_frame.locator("title").text_content() or path[-1]
            if len(filters) > ASAP_MAX_REPORT_FILTERS:
                view_diagnostics["report"] = {
                    "filters_truncated": len(filters) - ASAP_MAX_REPORT_FILTERS,
                }
                filters = filters[:ASAP_MAX_REPORT_FILTERS]
            for position, definition in enumerate(filters):
                definition["position"] = position
            automation = {
                "category_path": path, "report_tab": ready_text,
                "report_title": report_title, "export_text": ready_text,
                "export_views": export_views,
                "asap_export_capabilities": {
                    "status": (
                        "detected" if export_capability_views and all(
                            item.get("status") == "detected"
                            for item in export_capability_views.values()
                        ) else "partial"
                    ),
                    "views": export_capability_views,
                },
                "discovery_diagnostics": view_diagnostics,
            }
            if download_links:
                automation["download_links"] = download_links
                if not export_views:
                    automation["kind"] = "html_dashboard"
                report_progress("running", {
                    "stage": "html_dashboard_links",
                    "message": (
                        f"{path[-1]}: {len(download_links)} download link(s) found in the "
                        f"embedded dashboard: "
                        + ", ".join(link["label"] for link in download_links[:5])
                        + ("..." if len(download_links) > 5 else "")
                    ),
                    "report_index": index, "report_count": len(paths),
                })
            discovery_key = " > ".join(path)
            reports.append({
                "discovery_key": discovery_key, "name": path[-1],
                "report_url": site.get("base_url") or site.get("auth_url"),
                "ready_text": ready_text,
                "download_text": download_links[0]["label"] if download_links and not export_views else "Export CSV",
                "automation": automation,
                "filters": filters,
            })
        except Exception as exc:
            complete = False
            failures.append(f"{' > '.join(path)}: {exc}")
            timings.items.append({
                "phase": "report_inspection", "duration_ms": 0, "status": "failed",
                "metadata": {"path": path, "error": str(exc)},
            })
            _asap_goto(page, site.get("auth_url") or site.get("base_url"), profile_dir)
    if target_paths and not reports and failures:
        raise RuntimeError("ASAP targeted catalog scan failed. " + " | ".join(failures))
    return reports, timings.finish(item_count=len(reports), status="succeeded" if complete else "partial"), complete


def _validate_strict_headers(header: list[str], *, source_label: str) -> None:
    if not header or any(not value for value in header):
        raise RuntimeError(f"{source_label} contains a blank column header in its first row.")
    folded = [value.casefold() for value in header]
    duplicates = sorted({value for value in folded if folded.count(value) > 1})
    if duplicates:
        raise RuntimeError(
            f"{source_label} contains duplicate column headers: {', '.join(duplicates)}."
        )


def _text_encoding_candidates(raw: bytes) -> list[str]:
    """Return likely text encodings without letting latin-1 hide UTF-16."""
    encodings = ["utf-8-sig"]
    if raw.startswith((b"\xff\xfe", b"\xfe\xff")):
        encodings.append("utf-16")
    elif b"\x00" in raw[:512]:
        even_nuls = raw[:512:2].count(0)
        odd_nuls = raw[1:512:2].count(0)
        encodings.extend(
            ["utf-16-le", "utf-16-be"]
            if odd_nuls >= even_nuls
            else ["utf-16-be", "utf-16-le"]
        )
    encodings.extend(["cp1252", "latin-1"])
    return encodings


def _decode_downloaded_text(raw: bytes, *, source_label: str) -> tuple[str, str]:
    for encoding in _text_encoding_candidates(raw):
        try:
            return raw.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    raise RuntimeError(f"Could not decode {source_label}.")


def _parse_strict_delimited_rows(decoded: str) -> tuple[list[list[str]], str] | None:
    """Choose an Outlook delimiter by rectangularity, not raw punctuation.

    ``csv.Sniffer`` commonly mistakes commas inside Salesforce text values for
    separators even when the export is tab- or semicolon-delimited. Only accept
    candidates explicitly present in the header and structurally valid across
    all populated rows. A genuinely malformed CSV therefore still fails rather
    than being reinterpreted as a one-column file.
    """
    candidates: list[tuple[int, int, int, list[list[str]], str]] = []
    preference = {",": 0, ";": 1, "|": 2, "\t": 3}
    for delimiter in (",", ";", "\t", "|"):
        try:
            rows = list(csv.reader(io.StringIO(decoded), delimiter=delimiter))
        except csv.Error:
            continue
        if not rows or len(rows[0]) < 2:
            continue
        header = [str(value).strip() for value in rows[0]]
        try:
            _validate_strict_headers(header, source_label="Downloaded CSV")
        except RuntimeError:
            continue
        populated = [
            row for row in rows[1:] if any(str(value).strip() for value in row)
        ]
        if not populated or any(
            len(row) > len(header)
            and any(str(value).strip() for value in row[len(header):])
            for row in populated
        ):
            continue
        exact_width = sum(len(row) == len(header) for row in populated)
        # A delimiter seen only in the header is probably punctuation inside a
        # real CSV header, not the file separator.
        if exact_width * 2 < len(populated):
            continue
        candidates.append((
            preference[delimiter], len(header), exact_width, rows, delimiter,
        ))
    if not candidates:
        return None
    _preference, _width, _exact, rows, delimiter = max(
        candidates, key=lambda item: item[:3],
    )
    return rows, delimiter


def _asap_csv_header_index(rows: list[list[str]]) -> int:
    """Choose the last rectangular section after ASAP title/filter details.

    ASAP separates report title, filter details, and the real table with blank
    rows. A filter block can be longer than a small report, so globally choosing
    the largest rectangle would sometimes promote ``Filter,Value`` to the data
    header. Prefer the last usable section, then the strongest header candidate
    inside that section.
    """
    sections: list[tuple[int, int]] = []
    start = None
    for index, row in enumerate(rows):
        populated = any(str(value).strip() for value in row)
        if populated and start is None:
            start = index
        elif not populated and start is not None:
            sections.append((start, index))
            start = None
    if start is not None:
        sections.append((start, len(rows)))

    section_candidates = []
    for section_number, (section_start, section_end) in enumerate(sections):
        candidates = []
        for index in range(section_start, min(section_end, 200)):
            header = [str(value).strip() for value in rows[index]]
            populated = [value for value in header if value]
            if (
                len(populated) < 2
                or len({value.casefold() for value in populated}) != len(populated)
            ):
                continue
            width = len(header)
            data_rows = rows[index + 1:section_end]
            if not data_rows or any(
                any(str(value).strip() for value in row[width:]) for row in data_rows
            ):
                continue
            exact = sum(len(row) == width for row in data_rows)
            if exact * 2 < len(data_rows):
                continue
            nonempty = sum(
                sum(bool(str(value).strip()) for value in row[:width])
                for row in data_rows
            )
            candidates.append((len(data_rows), exact, nonempty, -index, index))
        if candidates:
            section_candidates.append((section_number, max(candidates)[-1]))
    return max(section_candidates)[-1] if section_candidates else 0


def _normalize_csv(
    path: Path, *, output: Path | None = None,
    preamble: str = "asap", strict_headers: bool = False,
) -> dict:
    """Normalize a delimited file with source-specific header resolution."""
    if preamble not in {"asap", "none"}:
        raise ValueError(f"Unsupported CSV preamble mode: {preamble}")
    detected = _detect_download_format(path)
    if detected != "csv":
        raise RuntimeError(
            f"Refusing to read {path.name} as CSV: it looks like {detected}."
        )
    raw = path.read_bytes()
    decoded, encoding_used = _decode_downloaded_text(
        raw, source_label=f"downloaded CSV {path.name}",
    )
    parsed = _parse_strict_delimited_rows(decoded) if strict_headers else None
    if parsed is not None:
        rows, delimiter = parsed
    else:
        lines = decoded.splitlines()
        sample = "\n".join(lines[:20])
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = max((",", ";", "\t", "|"), key=lambda item: sample.count(item))
        rows = list(csv.reader(io.StringIO(decoded), delimiter=delimiter))
    if not rows:
        raise RuntimeError("The downloaded CSV is empty.")
    header_index = 0
    if preamble == "asap":
        header_index = _asap_csv_header_index(rows)
    header = [str(value).strip() for value in rows[header_index]]
    if strict_headers:
        _validate_strict_headers(header, source_label="Downloaded CSV")
        data_rows = rows[header_index + 1:]
        if not any(any(str(value).strip() for value in row) for row in data_rows):
            raise RuntimeError("Downloaded CSV contains a header but no data rows.")
        for row_index, row in enumerate(data_rows, start=header_index + 2):
            overflow = row[len(header):]
            populated_overflow = sum(bool(str(value).strip()) for value in overflow)
            if populated_overflow:
                raise RuntimeError(
                    f"Downloaded CSV row {row_index} has {len(row)} columns, exceeding "
                    f"the {len(header)} first-row headers while parsing with "
                    f"{('tab' if delimiter == chr(9) else repr(delimiter))} as the delimiter. "
                    "Refusing to discard "
                    f"{populated_overflow} populated extra cell(s)."
                )
    elif len(header) < 2:
        raise RuntimeError("Downloaded CSV did not contain a usable delimited header.")
    destination = Path(output) if output is not None else path
    mode = "x" if output is not None else "w"
    with destination.open(mode, encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle, lineterminator="\n")
        writer.writerows(rows[header_index:])
    return {
        "preamble_rows_removed": header_index,
        "source_encoding": encoding_used,
        "source_delimiter": "tab" if delimiter == "\t" else delimiter,
        "columns": header,
    }


def _html_tag_name(tag: str) -> str:
    """Treat XML Spreadsheet namespace prefixes like ordinary HTML tags."""
    return str(tag).casefold().rsplit(":", 1)[-1]


def _html_span(attrs: list[tuple[str, str | None]], name: str) -> int:
    values = {
        _html_tag_name(key): value for key, value in attrs if value is not None
    }
    try:
        return max(1, min(1_000, int(values.get(name) or 1)))
    except (TypeError, ValueError):
        return 1


class _ExcelHtmlTableParser(HTMLParser):
    """Collect HTML and XML Spreadsheet tables without executing any markup."""

    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.tables: list[list[list[tuple[str, int, int]]]] = []
        self._stack: list[dict[str, Any]] = []

    @staticmethod
    def _cell_value(parts: list[str]) -> str:
        value = "".join(parts).replace("\xa0", " ").replace("\r", "")
        lines = [re.sub(r"[^\S\n]+", " ", line).strip() for line in value.split("\n")]
        return "\n".join(line for line in lines if line).strip()

    def _finish_cell(self, context: dict[str, Any]) -> None:
        cell = context.get("cell")
        row = context.get("row")
        if cell is None or row is None:
            return
        row.append((
            self._cell_value(cell["parts"]), cell["colspan"], cell["rowspan"],
        ))
        context["cell"] = None

    def _finish_row(self, context: dict[str, Any]) -> None:
        self._finish_cell(context)
        row = context.get("row")
        if row is not None:
            context["rows"].append(row)
        context["row"] = None

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        tag = _html_tag_name(tag)
        if tag == "table":
            self._stack.append({"rows": [], "row": None, "cell": None})
            return
        if not self._stack:
            return
        context = self._stack[-1]
        if tag in {"tr", "row"}:
            if context["row"] is not None:
                self._finish_row(context)
            context["row"] = []
        elif tag in {"td", "th", "cell"}:
            if context["row"] is None:
                context["row"] = []
            self._finish_cell(context)
            context["cell"] = {
                "parts": [],
                "colspan": _html_span(attrs, "colspan"),
                "rowspan": _html_span(attrs, "rowspan"),
            }
        elif tag == "br" and context["cell"] is not None:
            context["cell"]["parts"].append("\n")

    def handle_startendtag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        self.handle_starttag(tag, attrs)
        self.handle_endtag(tag)

    def handle_endtag(self, tag: str) -> None:
        tag = _html_tag_name(tag)
        if not self._stack:
            return
        context = self._stack[-1]
        if tag in {"td", "th", "cell"}:
            self._finish_cell(context)
        elif tag in {"tr", "row"}:
            self._finish_row(context)
        elif tag == "table":
            self._finish_row(context)
            self.tables.append(context["rows"])
            self._stack.pop()

    def handle_data(self, data: str) -> None:
        # A nested presentation table may sit inside an outer cell. Attribute
        # text to the nearest active cell while collecting the nested table
        # independently as another candidate.
        for context in reversed(self._stack):
            if context["cell"] is not None:
                context["cell"]["parts"].append(data)
                break


def _expand_html_table(
    raw_rows: list[list[tuple[str, int, int]]],
) -> list[list[str]]:
    """Expand colspan/rowspan into a rectangular sequence of explicit cells."""
    rows: list[list[str]] = []
    carried: dict[int, tuple[str, int]] = {}
    for raw_row in raw_rows:
        cells = {column: value for column, (value, _remaining) in carried.items()}
        next_carried = {
            column: (value, remaining - 1)
            for column, (value, remaining) in carried.items()
            if remaining > 1
        }
        column = 0
        for value, colspan, rowspan in raw_row:
            targets = []
            while len(targets) < colspan:
                while column in cells:
                    column += 1
                targets.append(column)
                column += 1
            for offset, target in enumerate(targets):
                expanded_value = value if offset == 0 else ""
                cells[target] = expanded_value
                if rowspan > 1:
                    next_carried[target] = (expanded_value, rowspan - 1)
        if cells:
            row = [cells.get(index, "") for index in range(max(cells) + 1)]
            while row and not str(row[-1]).strip():
                row.pop()
            rows.append(row)
        carried = next_carried
    return [row for row in rows if any(str(value).strip() for value in row)]


def _validated_html_table(
    raw_rows: list[list[tuple[str, int, int]]], *, table_number: int,
) -> tuple[list[str], list[list[str]]]:
    rows = _expand_html_table(raw_rows)
    source_label = f"Downloaded HTML-as-XLS table {table_number}"
    if not rows:
        raise RuntimeError(f"{source_label} is empty.")
    header = [str(value).strip() for value in rows[0]]
    _validate_strict_headers(header, source_label=source_label)
    data_rows: list[list[str]] = []
    for row_number, row in enumerate(rows[1:], start=2):
        overflow = row[len(header):]
        populated_overflow = sum(bool(str(value).strip()) for value in overflow)
        if populated_overflow:
            raise RuntimeError(
                f"{source_label} row {row_number} has {len(row)} columns, exceeding "
                f"the {len(header)} first-row headers. Refusing to discard "
                f"{populated_overflow} populated extra cell(s)."
            )
        values = row[:len(header)] + [""] * max(0, len(header) - len(row))
        if any(str(value).strip() for value in values):
            data_rows.append(values)
    if not data_rows:
        raise RuntimeError(f"{source_label} contains a header but no data rows.")
    return header, data_rows


def _normalize_html_excel(source: Path, output: Path) -> dict:
    """Normalize a Salesforce-style HTML/XML table delivered as legacy XLS."""
    decoded, encoding_used = _decode_downloaded_text(
        source.read_bytes(), source_label=f"HTML-as-XLS attachment {source.name}",
    )
    parser = _ExcelHtmlTableParser()
    parser.feed(decoded.lstrip("\ufeff"))
    parser.close()
    candidates: list[tuple[int, int, list[str], list[list[str]]]] = []
    failures: list[tuple[int, str]] = []
    for table_number, raw_rows in enumerate(parser.tables, start=1):
        try:
            header, rows = _validated_html_table(raw_rows, table_number=table_number)
        except RuntimeError as exc:
            failures.append((len(raw_rows), str(exc)))
            continue
        candidates.append((len(rows), table_number, header, rows))
    if not candidates:
        detail = max(failures, default=(0, "no table elements were found"))[1]
        raise RuntimeError(
            "The legacy .xls attachment contains HTML/XML but no usable first-row-header "
            f"data table. {detail}"
        )
    _row_count, table_number, header, data_rows = max(
        candidates, key=lambda item: (item[0], len(item[2]), -item[1]),
    )
    partial = output.with_name(f".{output.name}.partial")
    with partial.open("x", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle, lineterminator="\n")
        writer.writerow(header)
        writer.writerows(data_rows)
    partial.replace(output)
    return {
        "preamble_rows_removed": 0,
        "source_encoding": encoding_used,
        "source_delimiter": None,
        "source_sheets": [f"HTML table {table_number}"],
        "source_container": "html_xls",
        "columns": header,
    }


def _validate_asap_html_export(source: Path) -> dict:
    raw = source.read_bytes()
    decoded, encoding_used = _decode_downloaded_text(
        raw, source_label=f"ASAP HTML export {source.name}",
    )
    if looks_like_sign_in(decoded):
        raise RuntimeError("The HTML download contains a sign-in or expired-session page.")
    parser = _ExcelHtmlTableParser()
    parser.feed(decoded.lstrip("\ufeff"))
    parser.close()
    usable = []
    for raw_rows in parser.tables:
        rows = _expand_html_table(raw_rows)
        if len(rows) < 2:
            continue
        width = max((len(row) for row in rows), default=0)
        if width >= 2 and any(
            sum(bool(str(value).strip()) for value in row) >= 2 for row in rows[1:]
        ):
            usable.append(rows)
    if not usable:
        raise RuntimeError(
            "The HTML download did not contain a usable multi-row report table. "
            "Refusing to accept a portal message or error page as an export."
        )
    largest = max(usable, key=lambda rows: (len(rows), max(map(len, rows))))
    return {
        "source_encoding": encoding_used,
        "source_sheets": [],
        "columns": [str(value).strip() for value in largest[0]],
        "row_count": max(0, len(largest) - 1),
        "html_table_count": len(usable),
    }


def _validate_asap_plain_text_export(source: Path) -> dict:
    raw = source.read_bytes()
    decoded, encoding_used = _decode_downloaded_text(
        raw, source_label=f"ASAP Plain text export {source.name}",
    )
    if looks_like_sign_in(decoded):
        raise RuntimeError("The Plain text download contains a sign-in or expired-session message.")
    control_count = sum(
        1 for character in decoded
        if ord(character) < 32 and character not in "\t\r\n\f"
    )
    if control_count > max(2, len(decoded) // 200):
        raise RuntimeError("The Plain text download contains binary control data.")
    stripped = decoded.lstrip("\ufeff \t\r\n")
    if not stripped or re.search(r"<(?:html|table|body|form)\b|<!doctype", stripped, re.I):
        raise RuntimeError("The Plain text download is empty or contains portal markup.")
    parsed = _parse_strict_delimited_rows(decoded)
    if parsed is not None:
        rows, delimiter = parsed
        if len([row for row in rows if any(str(value).strip() for value in row)]) >= 2:
            return {
                "source_encoding": encoding_used,
                "source_delimiter": "tab" if delimiter == "\t" else delimiter,
                "row_count": max(0, len(rows) - 1),
                "columns": [str(value).strip() for value in rows[0]],
            }
    lines = [line.strip() for line in decoded.splitlines() if line.strip()]
    whitespace_rows = [re.split(r"\s{2,}", line) for line in lines]
    widths = [len(row) for row in whitespace_rows if len(row) >= 2]
    if len(lines) < 2 or len(widths) < 2 or len(set(widths[-min(20, len(widths)):])) != 1:
        raise RuntimeError(
            "The Plain text download did not contain a plausible multi-row table. "
            "Refusing to accept an arbitrary portal message as an export."
        )
    return {
        "source_encoding": encoding_used,
        "source_delimiter": "whitespace",
        "row_count": len(lines) - 1,
        "columns": whitespace_rows[0],
    }


def _excel_cell_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return "" if value is None else value


def _validated_requested_weeks(period: Any) -> list[str]:
    """Return the exact contiguous week list supplied by the Metronome job."""
    if period is None:
        return []
    values = period if isinstance(period, list) else [period]
    weeks: list[str] = []
    mondays: list[date] = []
    for value in values:
        match = REQUESTED_WEEK.fullmatch(str(value).strip())
        if not match:
            raise RuntimeError(
                f"Metronome supplied an invalid requested week {value!r}; expected YYYY-Www."
            )
        year, week = (int(item) for item in match.groups())
        try:
            monday = date.fromisocalendar(year, week, 1)
        except ValueError as exc:
            raise RuntimeError(
                f"Metronome supplied a nonexistent requested week: {year:04d}-W{week:02d}."
            ) from exc
        compact = f"{year:04d}{week:02d}"
        if compact in weeks:
            raise RuntimeError(f"Metronome supplied a duplicate requested week: {compact}.")
        weeks.append(compact)
        mondays.append(monday)
    if any(current - previous != timedelta(days=7) for previous, current in zip(mondays, mondays[1:])):
        raise RuntimeError(
            f"Metronome supplied requested weeks that are not ordered and contiguous: {weeks}."
        )
    return weeks


#: Header cells whose value names a descriptor column rather than a dimension.
XLSX_DESCRIPTOR_LABELS = frozenset({"weekly", "metric", "metrics"})


def _xlsx_descriptor_indexes(header: list[str]) -> list[int]:
    return [
        index for index, value in enumerate(header)
        if str(value).strip().casefold() in XLSX_DESCRIPTOR_LABELS
    ]


def _xlsx_expand_multi_week_metric_header(
    requested_weeks: list[str], header: list[str], max_data_width: int,
    descriptor_labels: dict[int, set[str]], worksheet_title: str,
) -> tuple[list[str], list[str], str | None, int | None]:
    """Recover ASAP's multi-level week headings without guessing column count.

    In the live flat Excel matrix, the lower header row ends in ``Metrics``.
    Each selected week occupies one value column, but only the first value
    column has that lower-level label. The prior normalizer used the lower
    header width and silently truncated every later week. Metronome passes the
    complete in-memory period list that it set and read back in ASAP, so the
    physical value-column count must prove a one-to-one mapping to that list.

    The sheet's rows are described by ``max_data_width`` and the distinct
    normalized values seen under each descriptor column, never by the rows
    themselves: a multi-million-row export must not be held in memory to
    decide a header. The returned fourth element is the column index the
    caller must drop from every data row, or ``None``.
    """
    if len(requested_weeks) < 2:
        return header, [], None, None

    # The current Regional FOTA export can also arrive with a fully expanded
    # header: ``Weekly, 202630, 202631``. ``Weekly`` is not a dimension in the
    # file. It is a constant descriptor whose row value is ``Sell-out``. Keep
    # the explicit week columns, but remove that descriptor only when both the
    # requested period list and every populated descriptor value prove the
    # shape. This avoids teaching the downstream flow-specific script to
    # tolerate an ambiguous extra column.
    explicit_week_columns = [
        (index, str(value).strip()) for index, value in enumerate(header)
        if re.fullmatch(r"20\d{2}(?:0[1-9]|[1-4]\d|5[0-3])", str(value).strip())
    ]
    if explicit_week_columns:
        explicit_weeks = [value for _, value in explicit_week_columns]
        if explicit_weeks != requested_weeks:
            raise RuntimeError(
                "Downloaded multi-week Excel columns do not match Metronome's requested weeks "
                f"on sheet {worksheet_title!r}. Requested weeks: {requested_weeks}; explicit "
                f"week columns: {explicit_weeks}."
            )
        descriptor_indexes = _xlsx_descriptor_indexes(header)
        if len(descriptor_indexes) == 1:
            descriptor_index = descriptor_indexes[0]
            labels = descriptor_labels.get(descriptor_index, set())
            recognized_descriptor = (
                next(iter(labels))
                if len(labels) == 1 and labels <= {"sell_out", "fota"}
                else None
            )
            if recognized_descriptor:
                return (
                    [value for index, value in enumerate(header) if index != descriptor_index],
                    explicit_weeks,
                    recognized_descriptor,
                    descriptor_index,
                )
        return header, [], None, None

    metric_indexes = [
        index for index, value in enumerate(header)
        if str(value).strip().casefold() in {"metric", "metrics"}
    ]
    if len(metric_indexes) != 1 or metric_indexes[0] != len(header) - 1:
        return header, [], None, None
    metric_index = metric_indexes[0]
    metric_labels = descriptor_labels.get(metric_index, set())
    recognized_metric_label = (
        next(iter(metric_labels))
        if len(metric_labels) == 1 and metric_labels <= {"sell_out", "fota"}
        else None
    )
    if recognized_metric_label:
        expected_width = metric_index + 1 + len(requested_weeks)
        if max_data_width != expected_width:
            actual_values = max(0, max_data_width - metric_index - 1)
            raise RuntimeError(
                "Downloaded multi-week Excel matrix does not contain one numeric value column "
                f"per requested week on sheet {worksheet_title!r}. Requested weeks: "
                f"{requested_weeks}; metric label: {recognized_metric_label!r}; expected numeric "
                f"week columns: {len(requested_weeks)}; observed numeric week columns: "
                f"{actual_values}."
            )
        return (
            [*header[:metric_index], *requested_weeks],
            requested_weeks,
            recognized_metric_label,
            metric_index,
        )
    expected_width = metric_index + len(requested_weeks)
    if max_data_width == expected_width:
        return [*header[:metric_index], *requested_weeks], requested_weeks, None, None
    if max_data_width <= len(header):
        raise RuntimeError(
            "Downloaded multi-week Excel matrix exposes one Metrics column but no distinct "
            f"value column per requested week on sheet {worksheet_title!r}. Requested weeks: "
            f"{requested_weeks}; header width: {len(header)}; maximum data width: {max_data_width}."
        )
    raise RuntimeError(
        "Downloaded multi-week Excel matrix width does not match its requested weeks on sheet "
        f"{worksheet_title!r}. Requested weeks: {requested_weeks}; Metrics starts at column "
        f"{metric_index + 1}; expected width: {expected_width}; maximum data width: "
        f"{max_data_width}."
    )


class _TrimmedWorksheet:
    """A worksheet view with its first row and first column removed.

    GSCM's toolbar export frames every workbook with a blank first column and
    a non-data first row. Applying the flow's configured trim here - before
    header detection - lets the standard normalization see a clean table.
    The trim is a per-flow setting (``downloads.excel_trim``), never a
    hardcoded assumption, so each run records that the step was applied.
    """

    def __init__(self, worksheet):
        self._worksheet = worksheet

    @property
    def title(self):
        return self._worksheet.title

    def iter_rows(
        self, *, min_row: int | None = None, max_row: int | None = None,
        values_only: bool = True,
    ):
        for row in self._worksheet.iter_rows(
            min_row=int(min_row or 1) + 1,
            max_row=None if max_row is None else int(max_row) + 1,
            values_only=values_only,
        ):
            yield tuple(row)[1:]


def _trim_worksheet(worksheet, excel_trim: str):
    """Apply the flow's configured Excel trim, validating the option name."""
    trim = str(excel_trim or "none").strip().casefold()
    if trim == "none":
        return worksheet
    if trim == "first_row_and_column":
        return _TrimmedWorksheet(worksheet)
    raise RuntimeError(f"Unsupported Excel pre-processing option: {excel_trim!r}.")


def _xlsx_rows(worksheet):
    """Yield one populated, right-trimmed row at a time.

    Streaming matters: a Regional export can carry millions of rows, and
    holding them as Python lists costs roughly a kilobyte each.
    """
    for raw_row in worksheet.iter_rows(values_only=True):
        values = [_excel_cell_value(value) for value in raw_row]
        while values and str(values[-1]).strip() == "":
            values.pop()
        if values:
            yield values


def _deduped_header(header: list) -> list:
    """Number repeated header labels instead of rejecting them.

    Portal reports legitimately repeat column labels - one ``Qty`` per week,
    or merged group headers expanded across their span. When the operator has
    declared where the header row is, the labels are kept and numbered
    (``Qty``, ``Qty_2``) so the CSV and its SQL table stay unambiguous.
    """
    seen: dict[str, int] = {}
    deduped = []
    for value in header:
        label = str(value).strip()
        key = re.sub(r"\W+", "_", label).strip("_").casefold()
        if not key:
            deduped.append(label)
            continue
        count = seen.get(key, 0) + 1
        seen[key] = count
        deduped.append(label if count == 1 else f"{label}_{count}")
    return deduped


def _sheet_header_scan(worksheet, limit: int = 12) -> str:
    """A compact, label-level account of why no row was a usable header.

    Reports populated-cell counts and repeated labels only - enough to fix a
    layout straight from the run log, without copying data values into it.
    """
    notes = []
    for index, row in enumerate(_xlsx_rows(worksheet)):
        if index >= limit:
            notes.append("…")
            break
        labels = [
            re.sub(r"\W+", "_", str(value).strip()).strip("_").casefold()
            for value in row
        ]
        counts: dict[str, int] = {}
        for label in labels:
            if label:
                counts[label] = counts.get(label, 0) + 1
        populated = sum(counts.values())
        repeated = sorted(label for label, count in counts.items() if count > 1)
        if populated < 2:
            notes.append(f"row {index + 1}: {populated} cell(s)")
        elif repeated:
            notes.append(
                f"row {index + 1}: {populated} cells, repeated labels: "
                + ", ".join(repeated[:4])
            )
        else:
            notes.append(f"row {index + 1}: {populated} cells, header-shaped")
    return "; ".join(notes) or "no populated rows"


def _xlsx_header_index(preview: list[list[Any]]) -> int | None:
    """Pick the row that defines this sheet's columns."""
    header_candidates = []
    for index, row in enumerate(preview):
        populated = sum(bool(str(value).strip()) for value in row)
        if populated < 2:
            continue
        candidate_names = [
            re.sub(r"\W+", "_", str(value).strip()).strip("_").casefold()
            or f"col_{column_index}"
            for column_index, value in enumerate(row)
        ]
        # Dense data rows can be wider than the actual Excel header.
        # Repeated dimension values make those rows invalid column
        # definitions, so never let them outrank a unique header row.
        if len(candidate_names) != len(set(candidate_names)):
            continue
        # Weekly exports expose their value column as a compact YYYYWW
        # label. Prefer a candidate containing that strong header
        # signal before falling back to density. Otherwise a wide data
        # row whose values happen to be unique can still win.
        has_year_week_header = any(
            re.fullmatch(r"20\d{2}(?:0[1-9]|[1-4]\d|5[0-3])", str(value).strip())
            for value in row
        )
        header_label_hits = len(set(candidate_names) & XLSX_HEADER_LABEL_HINTS)
        header_candidates.append(
            (header_label_hits, has_year_week_header, populated, len(row), -index, index)
        )
    return max(header_candidates)[-1] if header_candidates else None


def _xlsx_sheet_plan(
    worksheet, requested_weeks: list[str], *,
    header_mode: str = "auto", strict_headers: bool = False,
) -> dict | None:
    """Resolve one sheet's header and row transform in two streaming passes.

    Pass one buffers only the first rows, enough to find the header, then
    measures the rest of the sheet: its widest row and the distinct values
    under each descriptor column. Pass two (the caller's) rewrites rows
    straight to the CSV. Neither pass keeps the sheet in memory.
    """
    if header_mode not in {"auto", "first_row"}:
        raise ValueError(f"Unsupported Excel header mode: {header_mode}")
    preview = []
    for row in _xlsx_rows(worksheet):
        preview.append(row)
        if len(preview) >= 50:
            break
    if not preview:
        return None
    if header_mode == "first_row":
        raw_first = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        physical_header = [_excel_cell_value(value) for value in raw_first]
        while physical_header and not str(physical_header[-1]).strip():
            physical_header.pop()
        if not physical_header:
            if strict_headers:
                raise RuntimeError(
                    f"Downloaded Excel sheet {worksheet.title!r} has no headers in its first row."
                )
            # A trimmed portal workbook can carry an extra empty or notes
            # sheet whose first row is blank; skip it, don't fail the file.
            return None
        if not strict_headers:
            physical_header = _deduped_header(physical_header)
        header_index = 0
        preview[0] = physical_header
    else:
        header_index = _xlsx_header_index(preview)
    if header_index is None:
        return None
    header = [str(value).strip() for value in preview[header_index]]
    if strict_headers:
        _validate_strict_headers(
            header, source_label=f"Downloaded Excel sheet {worksheet.title!r}",
        )
    elif len(header) < 2:
        return None

    descriptor_indexes = _xlsx_descriptor_indexes(header)
    descriptor_labels: dict[int, set[str]] = {index: set() for index in descriptor_indexes}
    max_data_width = 0
    if len(requested_weeks) >= 2:
        # Only a multi-week export has a header to reconstruct, and only that
        # reconstruction needs to see the whole sheet. Every other workbook -
        # single-week ASAP, GSCM, anything without a period - converts in one
        # pass, which is what keeps a multi-million-row export affordable.
        for index, row in enumerate(_xlsx_rows(worksheet)):
            if index <= header_index:
                continue
            max_data_width = max(max_data_width, len(row))
            for descriptor_index in descriptor_indexes:
                if len(row) > descriptor_index and str(row[descriptor_index]).strip():
                    descriptor_labels[descriptor_index].add(
                        re.sub(r"\W+", "_", str(row[descriptor_index]).strip()).strip("_").casefold()
                    )

    header, week_columns, metric_label, drop_index = _xlsx_expand_multi_week_metric_header(
        requested_weeks, header, max_data_width, descriptor_labels, worksheet.title,
    )
    return {
        "header": header,
        "header_index": header_index,
        "drop_index": drop_index,
        "week_columns": week_columns,
        "metric_label": metric_label,
    }


class _XlrdWorksheet:
    """Expose an xlrd sheet through the small interface our normalizer uses."""

    def __init__(self, sheet, workbook, xlrd_module):
        self._sheet = sheet
        self._workbook = workbook
        self._xlrd = xlrd_module
        self.title = sheet.name

    def iter_rows(
        self, *, min_row: int | None = None, max_row: int | None = None,
        values_only: bool = True,
    ):
        if not values_only:
            raise ValueError("Metronome's Excel normalizer only exposes cell values.")
        start = max(0, int(min_row or 1) - 1)
        stop = min(self._sheet.nrows, int(max_row) if max_row is not None else self._sheet.nrows)
        for row_index in range(start, stop):
            values = []
            for cell in self._sheet.row(row_index):
                if cell.ctype in {self._xlrd.XL_CELL_EMPTY, self._xlrd.XL_CELL_BLANK}:
                    value = None
                elif cell.ctype == self._xlrd.XL_CELL_DATE:
                    try:
                        value = self._xlrd.xldate_as_datetime(
                            cell.value, self._workbook.datemode,
                        )
                    except (OverflowError, ValueError):
                        value = cell.value
                elif cell.ctype == self._xlrd.XL_CELL_BOOLEAN:
                    value = bool(cell.value)
                elif cell.ctype == self._xlrd.XL_CELL_ERROR:
                    value = self._xlrd.error_text_from_code.get(
                        cell.value, f"#ERROR_{cell.value}",
                    )
                else:
                    value = cell.value
                values.append(value)
            yield tuple(values)


class _XlrdWorkbook:
    def __init__(self, workbook, xlrd_module):
        self._workbook = workbook
        self.worksheets = [
            _XlrdWorksheet(workbook.sheet_by_index(index), workbook, xlrd_module)
            for index in range(workbook.nsheets)
        ]

    def close(self):
        self._workbook.release_resources()


class _XlsbWorksheet:
    """Re-open an XLSB sheet for every streaming pass used by the normalizer."""

    def __init__(self, workbook, title: str):
        self._workbook = workbook
        self.title = title

    def iter_rows(
        self, *, min_row: int | None = None, max_row: int | None = None,
        values_only: bool = True,
    ):
        if not values_only:
            raise ValueError("Metronome's Excel normalizer only exposes cell values.")
        start = max(1, int(min_row or 1))
        stop = int(max_row) if max_row is not None else None
        with self._workbook.get_sheet(self.title) as worksheet:
            for row_number, cells in enumerate(worksheet.rows(), start=1):
                if row_number < start:
                    continue
                if stop is not None and row_number > stop:
                    break
                yield tuple(cell.v for cell in cells)


class _XlsbWorkbook:
    def __init__(self, workbook):
        self._workbook = workbook
        self.worksheets = [
            _XlsbWorksheet(workbook, title) for title in workbook.sheets
        ]

    def close(self):
        self._workbook.close()


def _open_excel_workbook(source: Path, workbook_format: str):
    """Open an Excel family with a consistent read-only worksheet interface."""
    if workbook_format == "xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "Modern Excel normalization requires openpyxl. Re-run setup.ps1."
            ) from exc
        try:
            # data_only reads cached formula results and never executes VBA.
            return load_workbook(source, read_only=True, data_only=True)
        except Exception as exc:
            raise RuntimeError(
                f"Downloaded modern Excel workbook could not be opened: {source.name}. "
                "It may be damaged or password-protected/encrypted."
            ) from exc
    if workbook_format == "xls":
        try:
            import xlrd
        except ImportError as exc:
            raise RuntimeError(
                "Legacy XLS/XLT normalization requires xlrd. Re-run setup.ps1."
            ) from exc
        try:
            return _XlrdWorkbook(
                xlrd.open_workbook(str(source), on_demand=True), xlrd,
            )
        except Exception as exc:
            raise RuntimeError(
                f"Downloaded legacy Excel workbook could not be opened: {source.name}. "
                "It may be damaged, password-protected/encrypted, or a modern encrypted workbook."
            ) from exc
    if workbook_format == "xlsb":
        try:
            from pyxlsb import open_workbook
        except ImportError as exc:
            raise RuntimeError(
                "Binary XLSB normalization requires pyxlsb. Re-run setup.ps1."
            ) from exc
        try:
            return _XlsbWorkbook(open_workbook(str(source)))
        except Exception as exc:
            raise RuntimeError(
                f"Downloaded binary Excel workbook could not be opened: {source.name}. "
                "It may be damaged or password-protected/encrypted."
            ) from exc
    raise RuntimeError(f"Unsupported Excel workbook format: {workbook_format}")


def _normalize_xlsx(
    source: Path, output: Path, *, requested_weeks: list[str],
    header_mode: str = "auto", strict_headers: bool = False,
    workbook_format: str | None = None, excel_trim: str = "none",
    worksheet_name: str | None = None,
) -> dict:
    """Convert populated Excel-family sheets into one normalized UTF-8 CSV.

    Rows are streamed from the workbook straight to the CSV writer, so a
    multi-million-row export costs a constant amount of memory.
    ``excel_trim`` applies the flow's configured pre-processing (for GSCM:
    drop the frame's first row and first column) before header detection.
    """
    excel_trim = str(excel_trim or "none").strip().casefold()
    if excel_trim != "none" and header_mode == "auto":
        # Selecting the trim declares that the frame is gone and the table
        # starts immediately: the first remaining row IS the header. The
        # heuristic header hunt - and its unique-label rule, built for ASAP
        # exports - no longer applies, so a GSCM header that repeats labels
        # (one Qty per week) converts instead of being rejected.
        header_mode = "first_row"
    workbook_format = str(workbook_format or _detect_download_format(source)).casefold()
    workbook = _open_excel_workbook(source, workbook_format)
    worksheets = list(workbook.worksheets)
    if worksheet_name is not None:
        matches = [sheet for sheet in worksheets if sheet.title == worksheet_name]
        if len(matches) != 1:
            available = ", ".join(repr(sheet.title) for sheet in worksheets[:20]) or "none"
            workbook.close()
            raise RuntimeError(
                f"Excel worksheet {worksheet_name!r} was not found exactly once. "
                f"Available worksheets: {available}."
            )
        worksheets = matches
    common_header: list[str] | None = None
    common_normalized: list[str] | None = None
    source_sheets = []
    recovered_week_columns: list[str] = []
    removed_metric_label: str | None = None
    preamble_rows_removed = 0
    rows_written = 0
    # Stream into a partial file beside the target and rename it only once the
    # whole workbook has converted, so a failure can never leave a truncated
    # CSV where a complete one belongs. The worker never deletes, so a failed
    # conversion leaves its partial behind as evidence.
    partial = output.with_name(f".{output.name}.partial")
    handle = partial.open("x", encoding="utf-8-sig", newline="")
    try:
        writer = csv.writer(handle, lineterminator="\n")
        skipped_sheets: list[str] = []
        for worksheet in worksheets:
            worksheet = _trim_worksheet(worksheet, excel_trim)
            plan = _xlsx_sheet_plan(
                worksheet, requested_weeks,
                header_mode=header_mode, strict_headers=strict_headers,
            )
            if plan is None:
                skipped_sheets.append(
                    f"{worksheet.title}: {_sheet_header_scan(worksheet)}"
                )
                continue
            header = plan["header"]
            if plan["metric_label"]:
                if removed_metric_label and removed_metric_label != plan["metric_label"]:
                    raise RuntimeError(
                        "Downloaded Excel workbook exposed different metric labels across "
                        f"populated sheets: {removed_metric_label!r} and {plan['metric_label']!r}."
                    )
                removed_metric_label = plan["metric_label"]
            if plan["week_columns"]:
                if recovered_week_columns and recovered_week_columns != plan["week_columns"]:
                    raise RuntimeError(
                        "Downloaded Excel workbook resolved different multi-week columns across "
                        f"its populated sheets: {recovered_week_columns} and {plan['week_columns']}."
                    )
                recovered_week_columns = plan["week_columns"]
            normalized = [
                re.sub(r"\W+", "_", value).strip("_").casefold() or f"col_{index}"
                for index, value in enumerate(header)
            ]
            if common_header is None:
                common_header = header
                common_normalized = normalized
                writer.writerow(common_header)
            elif normalized != common_normalized:
                raise RuntimeError(
                    "Downloaded Excel workbook has populated sheets with different columns: "
                    f"{', '.join(source_sheets)} and {worksheet.title}."
                )
            width = len(common_header)
            drop_index = plan["drop_index"]
            header_index = plan["header_index"]
            for index, row in enumerate(_xlsx_rows(worksheet)):
                if index <= header_index:
                    continue
                if drop_index is not None and len(row) > drop_index:
                    row.pop(drop_index)
                if len(row) > width and any(
                    str(value).strip() for value in row[width:]
                ):
                    raise RuntimeError(
                        "Downloaded Excel row contains populated cells beyond the resolved header "
                        f"on sheet {worksheet.title!r}, row {index + 1}. Header width: {width}; "
                        f"row width: {len(row)}. Refusing to discard data."
                    )
                values = list(row[:width]) + [""] * max(0, width - len(row))
                if any(str(value).strip() for value in values):
                    writer.writerow(values)
                    rows_written += 1
            source_sheets.append(worksheet.title)
            preamble_rows_removed += header_index
    finally:
        workbook.close()
        if not handle.closed:
            handle.close()
    if common_header is None or not rows_written:
        detail = " | ".join(skipped_sheets) or "no populated sheets"
        raise RuntimeError(
            "Downloaded Excel workbook did not contain a usable table with "
            f"data rows (Excel pre-processing: {excel_trim}). Sheet scan, "
            f"after pre-processing: {detail}."
        )
    partial.replace(output)
    return {
        "excel_trim": excel_trim,
        "xlsx_header_mode": header_mode,
        "preamble_rows_removed": preamble_rows_removed,
        "source_encoding": workbook_format,
        "source_delimiter": None,
        "source_sheets": source_sheets,
        "columns": common_header,
        "recovered_week_columns": recovered_week_columns,
        "removed_metric_label": removed_metric_label,
    }


def _script_command(script_path: Path, input_path: Path, output_path: Path) -> list[str]:
    suffix = script_path.suffix.casefold()
    if suffix == ".py":
        return [sys.executable, str(script_path), "--input", str(input_path), "--output", str(output_path)]
    elif suffix == ".ps1":
        return [
            "powershell.exe", "-NoProfile", "-NonInteractive", "-ExecutionPolicy", "Bypass",
            "-File", str(script_path), "-InputPath", str(input_path), "-OutputPath", str(output_path),
        ]
    elif suffix == ".exe":
        return [str(script_path), "--input", str(input_path), "--output", str(output_path)]
    else:
        raise RuntimeError("Transformation script must be a .py, .ps1, or .exe file.")


def _run_transformations(artifacts: list[dict], config: dict) -> list[dict]:
    """Run one configured script once per download and return SQL-ready outputs."""
    if not config.get("enabled"):
        return artifacts
    script_path = Path(str(config.get("script_path") or ""))
    if not script_path.is_file():
        raise RuntimeError(f"Transformation script does not exist: {script_path}")
    results_folder = Path(artifacts[0]["file_path"]).parent / "script_results"
    results_folder.mkdir(parents=True, exist_ok=True)
    transformed = []
    for index, artifact in enumerate(artifacts, start=1):
        input_path = Path(artifact["file_path"])
        output_path = _safe_output_path(results_folder, input_path.name)
        environment = os.environ.copy()
        environment.update({
            "METRONOME_FLOW_INPUT": str(input_path),
            "METRONOME_FLOW_OUTPUT": str(output_path),
            "METRONOME_FLOW_RESULTS_DIR": str(results_folder),
            "METRONOME_FLOW_PERIODS": json.dumps(artifact.get("period_key") or []),
        })
        completed = subprocess.run(
            _script_command(script_path, input_path, output_path),
            cwd=str(script_path.parent), env=environment, capture_output=True,
            text=True, timeout=60 * 60, creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        )
        stdout = (completed.stdout or "").strip()
        stderr = (completed.stderr or "").strip()
        if completed.returncode != 0:
            detail = stderr or stdout or f"exit code {completed.returncode}"
            raise RuntimeError(f"Transformation failed for {input_path.name}: {detail[-4000:]}")
        if not output_path.is_file() or output_path.stat().st_size <= 0:
            raise RuntimeError(
                f"Transformation script completed for {input_path.name} but did not create {output_path}. "
                "The script must write --output (METRONOME_FLOW_OUTPUT)."
            )
        normalization = _normalize_csv(output_path)
        metadata = {**_csv_metadata(output_path), **normalization}
        transformed.append({
            **artifact,
            "file_path": str(output_path), "filename": output_path.name,
            "status": "transformed", "source_file_path": str(input_path),
            "publish_status": "not_applicable",
            "published_file_path": None,
            "published_filename": None,
            "published_file_size": None,
            "published_checksum": None,
            "script_path": str(script_path), "script_index": index,
            "script_stdout": stdout[-4000:], "script_stderr": stderr[-4000:],
            **metadata,
        })
    return transformed


def _csv_metadata(path: Path) -> dict:
    file_size = path.stat().st_size
    if file_size <= 0:
        raise RuntimeError("The downloaded CSV is empty.")
    with path.open("rb") as handle:
        # Sniff the head only. Reading the whole file to inspect 512 bytes
        # costs a second full-size copy in memory on a large export.
        prefix = handle.read(512).lstrip().lower()
    if prefix.startswith(b"<!doctype html") or prefix.startswith(b"<html"):
        raise RuntimeError("The download contains an HTML page instead of CSV data.")
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    row_count = None
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            row_count = max(0, sum(1 for _ in csv.reader(handle)) - 1)
    except (UnicodeDecodeError, csv.Error):
        pass
    return {"file_size": file_size, "checksum": digest.hexdigest(), "row_count": row_count}


def _copy_with_checksum(source_path: Path, output: Path) -> dict:
    """Exclusively copy one file while computing metadata in the same pass."""
    digest = hashlib.sha256()
    copied = 0
    with source_path.open("rb") as source, output.open("xb") as destination:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            destination.write(chunk)
            digest.update(chunk)
            copied += len(chunk)
    return {"file_size": copied, "checksum": digest.hexdigest()}


def _read_size_and_checksum(path: Path) -> dict:
    """One full read pass over a file: its byte count and SHA-256."""
    digest = hashlib.sha256()
    size = 0
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
            size += len(chunk)
    return {"file_size": size, "checksum": digest.hexdigest()}


def _stable_source_snapshot(source_path: Path) -> dict:
    """Read the staged download until two consecutive passes agree.

    The staging profile lives under the signed-in user's profile, which on a
    BI desktop is often redirected to a network share. Right after Edge
    finishes a download there, a second reader can still see a view that is a
    few kilobytes short of the true file while the write-behind tail settles,
    and one ``stat`` or one read pass is not evidence of the final content.
    Two full reads that agree on size and checksum are.
    """
    previous = None
    for attempt in range(6):
        if attempt:
            time.sleep(min(4.0, 0.5 * 2 ** (attempt - 1)))
        current = _read_size_and_checksum(source_path)
        if current == previous:
            return current
        previous = current
    raise RuntimeError(
        f"The staged download kept changing while being read back: {source_path} "
        f"(last pass read {previous['file_size']} bytes)."
    )


def _verify_copied_file(
    output: Path, expected_size: int, expected_checksum: str, *, label: str,
) -> None:
    """Prove the target copy is complete without trusting ``stat`` at all.

    The target folder usually lives on an SMB share, and Windows' SMB client
    can serve stale directory metadata after the written handle closes - in
    either direction: a short size for a complete copy, or the expected size
    for a truncated one. Directory attributes are therefore never evidence.
    Re-read the file and require the bytes on the share to match the copy
    stream's size and checksum.
    """
    observed = 0
    attempts = 5
    for attempt in range(attempts):
        if attempt:
            # Write-behind flushing on a congested share can take several
            # seconds for a large workbook; back off instead of giving up.
            time.sleep(min(4.0, 0.5 * 2 ** (attempt - 1)))
        digest = hashlib.sha256()
        observed = 0
        try:
            with output.open("rb") as handle:
                for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                    digest.update(chunk)
                    observed += len(chunk)
        except OSError:
            if attempt == attempts - 1:
                raise
            continue
        if observed == expected_size and digest.hexdigest() == expected_checksum:
            return
    raise RuntimeError(
        f"{label} was not copied completely to {output}: "
        f"read {observed} of the expected {expected_size} bytes back."
    )


def _raw_xlsx_metadata(
    output: Path,
    original_size: int,
    checksum: str,
    detected: str,
    *,
    normalization_error: str | None = None,
) -> dict:
    """Describe a verified raw Excel workbook when no CSV artifact is required."""
    metadata = {
        "file_size": original_size,
        "checksum": checksum,
        "row_count": None,
        "file_path": str(output),
        "filename": output.name,
        "original_file_path": str(output),
        "original_filename": output.name,
        "original_file_size": original_size,
        "detected_format": detected,
        "source_encoding": detected,
        "source_delimiter": None,
        "source_sheets": [],
        "columns": [],
    }
    if normalization_error is not None:
        metadata["normalization_error"] = normalization_error
    return metadata


def _completed_export_format_label(artifacts: list[dict], configured: str) -> str:
    """Describe what the portal actually delivered for the completion text."""
    detected = {
        str(artifact.get("detected_format") or "").strip().upper()
        for artifact in artifacts
        if str(artifact.get("detected_format") or "").strip()
    }
    return "/".join(sorted(detected)) or str(configured or "csv").upper()


def _validate_excel_container(path: Path, workbook_format: str) -> None:
    """Require a structurally complete workbook before raw-Excel success."""
    if workbook_format == "xls":
        workbook = _open_excel_workbook(path, workbook_format)
        workbook.close()
        return
    if workbook_format not in {"xlsx", "xlsb"}:
        raise RuntimeError(f"Unsupported Excel workbook format: {workbook_format}")
    format_label = workbook_format.upper()
    workbook_member = "xl/workbook.bin" if workbook_format == "xlsb" else "xl/workbook.xml"
    missing: set[str] = set()
    corrupt_member: str | None = None
    try:
        with zipfile.ZipFile(path) as workbook:
            members = set(workbook.namelist())
            missing = {"[Content_Types].xml", workbook_member} - members
            if not missing:
                corrupt_member = workbook.testzip()
    except (OSError, zipfile.BadZipFile, NotImplementedError, RuntimeError) as exc:
        raise RuntimeError(
            f"The downloaded Excel workbook is not a complete {format_label} ZIP container: "
            f"{path.name}"
        ) from exc
    if missing:
        raise RuntimeError(
            f"The downloaded Excel workbook is missing required {format_label} content in "
            f"{path.name}: "
            f"{', '.join(sorted(missing))}"
        )
    if corrupt_member:
        raise RuntimeError(
            f"The downloaded Excel workbook contains a corrupt {format_label} member in "
            f"{path.name}: "
            f"{corrupt_member}"
        )


def _validate_xlsx_container(path: Path) -> None:
    """Backward-compatible wrapper for the existing raw-XLSX validation seam."""
    _validate_excel_container(path, "xlsx")


#: What a downloaded file actually is, read from its first bytes. A portal
#: labels a link, and the flow declares a download type, but neither is
#: evidence: an "(xlsx)" dashboard link can emit HTML, and a CSV-configured
#: flow can receive a workbook.
DOWNLOAD_SIGNATURES = (
    (b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1", "xls"),
    # Standalone BIFF2/3/4/5 workbook streams predate the usual OLE compound
    # wrapper but are still valid .xls/.xlt files that xlrd can read.
    (b"\x09\x00", "xls"),
    (b"\x09\x02", "xls"),
    (b"\x09\x04", "xls"),
    (b"\x09\x08", "xls"),
    (b"%PDF-", "pdf"),
)
HTML_PREFIXES = ("<!doctype html", "<html", "<?xml", "<table", "<meta")


def _detect_download_format(path: Path) -> str:
    """Identify a downloaded file from its content, not from its name.

    Decoding is not a safety net here: ``latin-1`` accepts every byte, so a
    workbook handed to the CSV path decodes into mojibake, gets written back
    out as a "CSV", and only fails much later inside PostgreSQL with a garbled
    column name. Refuse it at the door instead.
    """
    with path.open("rb") as handle:
        head = handle.read(4096)
    if not head:
        raise RuntimeError(f"The downloaded file is empty: {path.name}")
    if head.startswith(b"PK"):
        # OOXML and XLSB are both ZIP packages. workbook.bin is the decisive
        # XLSB marker; corrupt ZIPs retain the suffix-based family only so the
        # later container check can issue the useful integrity error.
        try:
            with zipfile.ZipFile(path) as workbook:
                members = set(workbook.namelist())
            if "xl/workbook.bin" in members:
                return "xlsb"
            return "xlsx"
        except (OSError, zipfile.BadZipFile, NotImplementedError, RuntimeError):
            return "xlsb" if path.suffix.casefold() in XLSB_EXCEL_EXTENSIONS else "xlsx"
    for signature, kind in DOWNLOAD_SIGNATURES:
        if head.startswith(signature):
            return kind
    decoded_head, _encoding = _decode_downloaded_text(
        head, source_label=f"download prefix for {path.name}",
    )
    stripped_text = decoded_head.lstrip("\ufeff \t\r\n").casefold()
    if stripped_text.startswith(HTML_PREFIXES) or re.search(
        r"<(?:[a-z_][\w.-]*:)?(?:html|table|workbook)\b|<\?xml\b",
        stripped_text,
    ):
        return "html"
    if b"\x00" in head and not head.startswith((b"\xff\xfe", b"\xfe\xff")):
        return "binary"
    return "csv"


def _excel_output_suffix(local_path: Path, output: Path, workbook_format: str) -> str:
    """Preserve a compatible original Excel extension, otherwise correct it."""
    compatible = EXCEL_EXTENSIONS_BY_FORMAT[workbook_format]
    for candidate in (output.suffix.casefold(), local_path.suffix.casefold()):
        if candidate in compatible:
            return candidate
    return {"xlsx": ".xlsx", "xls": ".xls", "xlsb": ".xlsb"}[workbook_format]


def _store_completed_download(
    local_path: Path, output: Path, *, file_format: str = "csv",
    asap_download_type: str | None = None,
    requested_period: Any = None,
    allow_raw_xlsx_fallback: bool = False,
    require_normalized_csv: bool = True,
    csv_preamble: str = "asap",
    strict_headers: bool = False,
    xlsx_header_mode: str = "auto",
    excel_trim: str = "none",
    processing_progress: Callable[[str, str], None] | None = None,
) -> dict:
    """Normalize the browser-local file, then copy it to the final target.

    Never pass the final UNC path to Playwright's ``download.save_as``. The
    caller first supplies a completed browser-local path (from Edge's native
    Download object or the staging fallback); this function then copies it to
    the final target in exclusive-create mode so nothing can be overwritten.
    """
    local_path = Path(local_path)
    file_format = str(file_format or "csv").casefold()
    asap_type = (
        resolve_asap_download_type(asap_download_type, legacy_file_format=file_format)
        if asap_download_type is not None else None
    )
    # Settle the staged download before reading anything out of it: format
    # detection, normalization, and the target copy must all see the same
    # final bytes, not a mid-flush view of a share-backed staging folder.
    snapshot = _stable_source_snapshot(local_path)
    detected = _detect_download_format(local_path)
    declared_suffixes = {local_path.suffix.casefold(), output.suffix.casefold()}
    html_excel = (
        detected == "html"
        and (
            bool(declared_suffixes & LEGACY_EXCEL_EXTENSIONS)
            or (asap_type is not None and asap_type.content_family == "excel")
        )
        # Outlook's strict flat-file contract and an explicitly selected ASAP
        # Excel type may opt into validated HTML/XML-as-XLS handling. A portal
        # login/error page remains a failure.
        and require_normalized_csv
        and (
            (strict_headers and csv_preamble == "none")
            or (asap_type is not None and asap_type.content_family == "excel")
        )
    )
    if detected == "xls" and declared_suffixes & (
        OOXML_EXCEL_EXTENSIONS | XLSB_EXCEL_EXTENSIONS
    ):
        raise RuntimeError(
            f"The Excel attachment {local_path.name} uses an OLE encrypted container even "
            "though its filename declares a modern workbook. Password-protected/encrypted "
            "Excel attachments are unsupported; attach an unencrypted workbook instead."
        )
    if detected in {"pdf", "binary"}:
        raise RuntimeError(
            f"The download is not a CSV or an Excel workbook: {local_path.name} "
            f"looks like {detected}. Check what this report's download link "
            "actually produces."
        )
    if asap_type is not None and asap_type.content_family == "html":
        if detected != "html":
            raise RuntimeError(
                f"ASAP was asked for {asap_type.label}, but the download looks like {detected}."
            )
        validation = _validate_asap_html_export(local_path)
        if output.suffix.casefold() not in asap_type.compatible_suffixes:
            output = _safe_output_path(output.parent, f"{output.stem}{asap_type.preferred_suffix}")
        copied = _copy_with_checksum(local_path, output)
        if (
            copied["file_size"] != snapshot["file_size"]
            or copied["checksum"] != snapshot["checksum"]
        ):
            raise RuntimeError("The staged ASAP HTML export changed while it was being copied.")
        _verify_copied_file(
            output, snapshot["file_size"], snapshot["checksum"], label="Downloaded ASAP HTML export",
        )
        return {
            **validation,
            "file_size": copied["file_size"], "checksum": copied["checksum"],
            "file_path": str(output), "filename": output.name,
            "original_file_path": str(output), "original_filename": output.name,
            "original_file_size": copied["file_size"], "detected_format": "html",
        }
    if asap_type is not None and asap_type.content_family == "text":
        if detected != "csv":
            raise RuntimeError(
                f"ASAP was asked for {asap_type.label}, but the download looks like {detected}."
            )
        validation = _validate_asap_plain_text_export(local_path)
        if output.suffix.casefold() not in asap_type.compatible_suffixes:
            output = _safe_output_path(output.parent, f"{output.stem}{asap_type.preferred_suffix}")
        copied = _copy_with_checksum(local_path, output)
        if (
            copied["file_size"] != snapshot["file_size"]
            or copied["checksum"] != snapshot["checksum"]
        ):
            raise RuntimeError("The staged ASAP Plain text export changed while it was being copied.")
        _verify_copied_file(
            output, snapshot["file_size"], snapshot["checksum"], label="Downloaded ASAP Plain text export",
        )
        return {
            **validation,
            "file_size": copied["file_size"], "checksum": copied["checksum"],
            "file_path": str(output), "filename": output.name,
            "original_file_path": str(output), "original_filename": output.name,
            "original_file_size": copied["file_size"], "detected_format": "txt",
        }
    if asap_type is not None and asap_type.content_family == "csv" and detected != "csv":
        raise RuntimeError(
            f"ASAP was asked for {asap_type.label}, but the download looks like {detected}."
        )
    if asap_type is not None and asap_type.content_family == "excel" and (
        detected not in EXCEL_DOWNLOAD_FORMATS and not html_excel
    ):
        raise RuntimeError(
            f"ASAP was asked for {asap_type.label}, but the download looks like {detected}."
        )
    if detected == "html" and not html_excel:
        raise RuntimeError(
            f"The download is an HTML page, not data: {local_path.name}. The "
            "portal most likely returned an error or a login page instead of "
            "the export."
        )
    if html_excel:
        with local_path.open("rb") as handle:
            html_head = handle.read(128 * 1024)
        if looks_like_sign_in(html_head):
            raise RuntimeError(
                "The HTML/XML-based Excel download contains a sign-in, authorization, or portal error page."
            )
        expected_suffix = _excel_output_suffix(local_path, output, "xls")
        if output.suffix.casefold() != expected_suffix:
            output = _safe_output_path(output.parent, f"{output.stem}{expected_suffix}")
        original_size = snapshot["file_size"]
        copied = _copy_with_checksum(local_path, output)
        if copied["file_size"] != original_size or copied["checksum"] != snapshot["checksum"]:
            raise RuntimeError(
                f"The staged legacy Excel attachment changed while copying to {output}: "
                f"streamed {copied['file_size']} of {original_size} settled bytes."
            )
        _verify_copied_file(
            output, original_size, copied["checksum"],
            label="Downloaded legacy Excel attachment",
        )
        normalized_output = _safe_output_path(
            output.parent, f"{output.stem}_normalized.csv",
        )
        if processing_progress is not None:
            processing_progress(
                "file_normalization",
                f"Saved {output.name}; extracting its embedded data table.",
            )
        normalization = _normalize_html_excel(output, normalized_output)
        if processing_progress is not None:
            processing_progress(
                "file_metadata",
                f"Normalized {normalized_output.name}; calculating its checksum and row count.",
            )
        metadata = {**_csv_metadata(normalized_output), **normalization}
        return {
            **metadata,
            "file_path": str(normalized_output),
            "filename": normalized_output.name,
            "original_file_path": str(output),
            "original_filename": output.name,
            "original_file_size": original_size,
            "detected_format": expected_suffix.lstrip("."),
        }
    if detected in EXCEL_DOWNLOAD_FORMATS:
        # The portal decides the format, not the flow's setting. Preserve an
        # original compatible extension (for example .xlsm) and otherwise
        # correct the saved name so a workbook never lands as .csv.
        file_format = detected
        expected_suffix = _excel_output_suffix(local_path, output, detected)
        if output.suffix.casefold() != expected_suffix:
            output = _safe_output_path(output.parent, f"{output.stem}{expected_suffix}")
    elif asap_type is None and detected != file_format:
        file_format = detected
        expected_suffix = f".{detected}"
        if output.suffix.casefold() != expected_suffix:
            output = _safe_output_path(output.parent, f"{output.stem}{expected_suffix}")
    if file_format in EXCEL_DOWNLOAD_FORMATS:
        saved_format = output.suffix.casefold().lstrip(".") or detected
        original_size = snapshot["file_size"]
        if original_size <= 0:
            raise RuntimeError("The downloaded Excel workbook is empty.")
        if not require_normalized_csv:
            # The settled staging file is cheap to inspect compared to the
            # configured (often network-backed) target folder. Check the
            # complete workbook container before creating anything there.
            _validate_excel_container(local_path, detected)
        copied = _copy_with_checksum(local_path, output)
        if copied["file_size"] != original_size or copied["checksum"] != snapshot["checksum"]:
            raise RuntimeError(
                f"The staged Excel workbook changed while copying to {output}: "
                f"streamed {copied['file_size']} of {original_size} settled bytes."
            )
        _verify_copied_file(
            output, original_size, copied["checksum"], label="Downloaded Excel workbook",
        )
        if not require_normalized_csv:
            return _raw_xlsx_metadata(
                output, original_size, copied["checksum"], saved_format,
            )
        normalized_output = _safe_output_path(
            output.parent, f"{output.stem}_normalized.csv",
        )
        requested_weeks = _validated_requested_weeks(requested_period)
        if processing_progress is not None:
            trim_note = (
                " Applying the flow's Excel pre-processing first: dropping "
                "the first row and first column."
                if str(excel_trim or "none").casefold() != "none" else ""
            )
            processing_progress(
                "file_normalization",
                f"Saved {output.name} to the target folder; preparing its "
                f"normalized CSV.{trim_note}",
            )
        try:
            normalization = _normalize_xlsx(
                output, normalized_output, requested_weeks=requested_weeks,
                header_mode=xlsx_header_mode, strict_headers=strict_headers,
                workbook_format=detected, excel_trim=excel_trim,
            )
        except Exception as exc:
            if not allow_raw_xlsx_fallback:
                raise
            # GSCM's native workbook is the requested deliverable. Some GSCM
            # exports omit Excel's default style metadata or lay out their
            # data as a formatted report rather than a rectangular table.
            # openpyxl can warn or reject normalization even though Edge has
            # already produced a complete, usable workbook. When neither a
            # transformation nor SQL handoff needs a CSV, keep that verified
            # XLSX as the successful artifact instead of throwing it away and
            # retrying report navigation.
            _validate_excel_container(local_path, detected)
            return _raw_xlsx_metadata(
                output, original_size, copied["checksum"], saved_format,
                normalization_error=str(exc),
            )
        if processing_progress is not None:
            processing_progress(
                "file_metadata",
                f"Normalized {normalized_output.name}; calculating its checksum and row count.",
            )
        metadata = {**_csv_metadata(normalized_output), **normalization}
        return {
            **metadata,
            "file_path": str(normalized_output),
            "filename": normalized_output.name,
            "original_file_path": str(output),
            "original_filename": output.name,
            "original_file_size": original_size,
            "detected_format": saved_format,
        }
    if file_format != "csv":
        raise RuntimeError(f"Unsupported downloaded file format: {file_format}")
    if asap_type is not None:
        raw_output = _safe_output_path(
            output.parent, f"{output.stem}_raw{output.suffix or '.csv'}",
        )
        raw_copy = _copy_with_checksum(local_path, raw_output)
        if (
            raw_copy["file_size"] != snapshot["file_size"]
            or raw_copy["checksum"] != snapshot["checksum"]
        ):
            raise RuntimeError("The staged raw ASAP CSV changed while it was being copied.")
        _verify_copied_file(
            raw_output, snapshot["file_size"], snapshot["checksum"],
            label="Raw ASAP CSV export",
        )
        normalization = _normalize_csv(
            local_path, output=output, preamble=csv_preamble,
            strict_headers=strict_headers,
        )
        metadata = {**_csv_metadata(output), **normalization}
        return {
            **metadata, "file_path": str(output), "filename": output.name,
            "original_file_path": str(raw_output),
            "original_filename": raw_output.name,
            "original_file_size": raw_copy["file_size"],
            "detected_format": "csv",
        }
    normalization = _normalize_csv(
        local_path, preamble=csv_preamble, strict_headers=strict_headers,
    )
    metadata = {**_csv_metadata(local_path), **normalization}
    copied = _copy_with_checksum(local_path, output)
    if copied["file_size"] != metadata["file_size"]:
        raise RuntimeError(
            f"The staged CSV changed size while copying to {output}: "
            f"streamed {copied['file_size']} of {metadata['file_size']} bytes."
        )
    _verify_copied_file(
        output, metadata["file_size"], copied["checksum"], label="Downloaded CSV",
    )
    return {
        **metadata, "file_path": str(output), "filename": output.name,
        "detected_format": detected,
    }


def _asap_filters_for_export_view(job: dict, export_view: str | None) -> list[dict]:
    """Limit configuration to controls catalogued for the selected export view."""
    definitions = list(job.get("report", {}).get("filters") or [])
    if not export_view:
        return definitions
    catalog_views = (job.get("report", {}).get("automation") or {}).get("export_views") or []
    catalog_view = next(
        (
            item for item in catalog_views
            if isinstance(item, dict)
            and str(item.get("label") or "").casefold() == export_view.casefold()
        ),
        None,
    )
    if catalog_view is None or "filter_keys" not in catalog_view:
        return definitions
    allowed = set(catalog_view.get("filter_keys") or [])
    return [item for item in definitions if item.get("filter_key") in allowed]


def _has_named_control(page: Page | Frame, text: str) -> bool:
    for locator in (
        page.get_by_role("button", name=text, exact=True),
        page.get_by_role("link", name=text, exact=True),
        page.get_by_text(text, exact=True),
    ):
        try:
            if locator.count() and locator.first.is_visible():
                return True
        except Exception:
            continue
    return False


def _export_task_key(export_view, period) -> str:
    """Stable identity of one export file within a bundle."""
    from app.flow_tasks import task_key
    return task_key(export_view, period)


def _job_store_root(job: dict) -> Path | None:
    value = (job.get("paths") or {}).get("artifact_store_root")
    return Path(value) if value else None


def _job_store_id(job: dict, profile_dir: Path) -> str:
    return flow_publish.artifact_store_id(profile_dir, store_root=_job_store_root(job))


def _runtime_store_ids(job: dict, profile_dir: Path) -> set[str]:
    return {flow_publish.artifact_store_id(profile_dir), _job_store_id(job, profile_dir),
            *(flow_publish.artifact_store_id(profile_dir, store_root=Path(root))
              for root in (job.get("resume") or {}).get("artifact_store_roots", []))}


def _decorate_artifact_storage(artifact: dict, job: dict, profile_dir: Path) -> dict:
    """Freeze the private deliverable needed by direct publish and Resume."""
    output_mode = job.get("downloads", {}).get("output_mode", "run_folders")
    direct = output_mode == "direct_replace"
    private = output_mode in {"direct_replace", "private_snapshot"}
    original = artifact.get("original_file_path")
    original_metadata = None
    if original:
        original_path = Path(str(original))
        if not original_path.is_file():
            if direct:
                raise RuntimeError(f"Saved original Flow deliverable is missing: {original_path}")
        else:
            original_metadata = flow_publish.read_size_checksum(original_path)
    original_suffix = Path(str(original or "")).suffix.casefold()
    excel_original = bool(original) and original_suffix in (
        OOXML_EXCEL_EXTENSIONS | LEGACY_EXCEL_EXTENSIONS | XLSB_EXCEL_EXTENSIONS
    )
    deliverable = Path(str(
        original if direct and excel_original else artifact.get("file_path") or ""
    ))
    if not deliverable.is_file():
        if private:
            raise RuntimeError(f"Saved Flow deliverable is missing: {deliverable}")
        artifact.update({
            "storage_scope": "target_run_folder",
            "artifact_store_id": None,
            "publish_status": "not_applicable",
        })
        return artifact
    observed = flow_publish.read_size_checksum(deliverable)
    artifact.update({
        "storage_scope": "worker_private" if private else "target_run_folder",
        "artifact_store_id": _job_store_id(job, profile_dir) if private else None,
        "deliverable_file_path": str(deliverable),
        "deliverable_filename": deliverable.name,
        "deliverable_file_size": observed["file_size"],
        "deliverable_checksum": observed["checksum"],
        "publish_status": "pending" if direct else "not_applicable",
    })
    if original_metadata is not None:
        artifact.update({
            "original_file_size": original_metadata["file_size"],
            "original_checksum": original_metadata["checksum"],
        })
    return artifact


def _validated_resume_artifacts(job: dict) -> dict[str, dict]:
    """Carried artifacts this worker can safely skip and later publish."""
    direct = job.get("downloads", {}).get("output_mode", "run_folders") == "direct_replace"
    valid = {}
    for item in (job.get("resume") or {}).get("completed") or []:
        if not isinstance(item, dict):
            continue
        if direct:
            available = (
                item.get("storage_scope") == "worker_private"
                and bool(job.get("_runtime_artifact_store_id"))
                and item.get("artifact_store_id") in set(job.get("_runtime_artifact_store_ids") or [job.get("_runtime_artifact_store_id")])
                and flow_publish.artifact_file_valid(item, require_deliverable=True)
            )
        else:
            raw_path = item.get("file_path")
            available = not raw_path or Path(str(raw_path)).is_file()
        if available:
            valid[_export_task_key(item.get("export_view"), item.get("period_key"))] = item
    return valid


def _resume_completed_keys(job: dict) -> set[str]:
    """Files a resumed run must skip because a prior run already saved them.

    An entry that names its saved file is only honored while that file still
    exists - a file removed by run-folder retention is downloaded again.
    Entries without a path (queued before paths were carried) keep the old
    always-skip behavior.
    """
    return set(_validated_resume_artifacts(job))


def _export_task_with_retry(
    page: Page, run_task, on_retry, *, max_attempts: int = EXPORT_TASK_ATTEMPTS,
):
    """Restart one export file from scratch after any mid-file failure.

    A bundle of many files must never lose its finished downloads to one
    transient portal hiccup on a later file - a menu item that briefly fails
    to render, or an Browser download that never stabilizes in staging. Each
    attempt restarts the file at report navigation, so the portal is reopened
    in a known state. Files that were already saved are untouched, and a
    partially copied output is never overwritten: the retried download stores
    under the next free suffixed filename.
    """
    failures: list[tuple[int, Exception]] = []
    for attempt in range(1, max_attempts + 1):
        try:
            return run_task(attempt)
        except Exception as exc:
            failures.append((attempt, exc))
            if isinstance(exc, _CompletedDownloadProcessingError):
                raise
            if isinstance(exc, flow_gscm.NotSignedInError):
                # Another attempt cannot sign the profile in; keep the
                # actionable message instead of a retry transcript.
                raise
            if attempt == max_attempts:
                if len(failures) > 1 and any(
                    str(error) != str(failures[0][1]) for _number, error in failures[1:]
                ):
                    # Group identical messages so two attempts that fail the
                    # same way report one message, not two screen dumps.
                    grouped: dict[str, list[int]] = {}
                    for number, error in failures:
                        grouped.setdefault(str(error), []).append(number)
                    detail = " ".join(
                        f"Attempt(s) {','.join(map(str, numbers))}: {message[:2000]}"
                        for message, numbers in grouped.items()
                    )
                    raise RuntimeError(
                        f"Export failed after {max_attempts} attempts. {detail}"
                    ) from exc
                raise
            on_retry(attempt, exc)
            page.wait_for_timeout(5_000)
    raise AssertionError("unreachable")


def _gscm_call(page: Page, headed: bool, notify, operation, profile_dir: Path):
    """Run one GSCM portal operation, recovering a sign-in wall like ASAP does.

    First try the stored SSO credential - the same unattended re-login ASAP
    performs on every run, against the same Samsung SSO. Only when that cannot
    finish does the mode matter: a headed worker has a visible window and a
    human nearby (tell them, wait, retry once); headless raises the actionable
    error.
    """
    try:
        return operation()
    except flow_gscm.NotSignedInError as exc:
        auto_login_failure = None
        try:
            if _gscm_authenticate_if_needed(page, profile_dir):
                notify("GSCM session was recovered automatically with the stored credential.")
                return operation()
        except Exception as auto_exc:
            auto_login_failure = str(auto_exc)
        if headed:
            notify(
                "GSCM sign-in required. Complete Samsung SSO and the Knox MFA "
                "prompt in the visible browser window; the run resumes automatically."
            )
            flow_gscm.wait_for_manual_login(page, report_progress=notify)
            return operation()
        if auto_login_failure:
            raise RuntimeError(
                f"{exc} Automatic sign-in with the stored credential was "
                f"attempted and failed: {auto_login_failure}"
            ) from exc
        raise


def _bounded_detail(detail: dict | None) -> dict | None:
    """Cap the one progress field nothing else truncates before it is sent."""
    if not detail:
        return detail
    message = detail.get("message")
    if isinstance(message, str) and len(message) > PROGRESS_MESSAGE_MAX_CHARS:
        return {
            **detail,
            "message": message[:PROGRESS_MESSAGE_MAX_CHARS] + "… [truncated]",
        }
    return detail


def _failure_screenshot(page, profile_dir: Path, label: str) -> str | None:
    """Best-effort PNG of the live page for a failed run or scan.

    Saved under the profile's ``diagnostics`` folder, never the run folder -
    run folders feed downstream pipelines and retention. The worker never
    deletes files (flow_retention is the one gated deletion site), so growth
    is bounded by overwriting a fixed rotation of slot filenames instead of
    pruning old screenshots. Returns the saved path, or ``None`` when the
    page is already gone.
    """
    try:
        diagnostics = Path(profile_dir) / "diagnostics"
        diagnostics.mkdir(parents=True, exist_ok=True)
        counter_path = diagnostics / ".screenshot_counter"
        try:
            counter = int(counter_path.read_text(encoding="utf-8").strip())
        except (OSError, ValueError):
            counter = 0
        counter_path.write_text(str(counter + 1), encoding="utf-8")
        target = diagnostics / f"failure-{counter % FAILURE_SCREENSHOT_KEEP:02d}.png"
        page.screenshot(path=str(target), full_page=False)
        safe_label = re.sub(r"[^A-Za-z0-9_-]+", "_", str(label))[:60] or "failure"
        return f"{target} ({safe_label}, {datetime.now():%Y-%m-%d %H:%M:%S})"
    except Exception:
        return None


def _prepare_run_folder(
    job: dict, profile_dir: Path, *, run_id: int, register_folder, report_progress,
) -> Path:
    """Create/register a producing run folder and execute assigned retention."""
    assert_job_paths(job)
    if (job.get("paths") or {}).get("flow_folder"):
        flow_layout.ensure_layout(job["paths"]["flow_folder"], job["flow"]["id"])
    output_mode = job.get("downloads", {}).get("output_mode", "run_folders")
    direct = output_mode == "direct_replace"
    private_snapshot = output_mode == "private_snapshot"
    target = Path(job["downloads"]["target_folder"])
    if private_snapshot:
        storage_root = flow_publish.private_local_file_root(
            profile_dir, (job.get("local_file") or {}).get("private_store_key") or "",
            store_root=_job_store_root(job),
        )
    else:
        if not target.is_dir():
            raise RuntimeError(f"Target folder does not exist: {target}")
        storage_root = flow_publish.private_target_root(profile_dir, target, store_root=_job_store_root(job)) if direct else target
    run_folder = flow_retention.create_run_folder(
        storage_root, run_id, job.get("flow", {}).get("id"),
    )
    job["_runtime_run_folder"] = str(run_folder)
    # Registration is idempotent and intentionally precedes all writes into
    # the new folder. Outlook calls this only after it has acquired a new
    # attachment, so a no-match/dedup no-op creates no folder and receives no
    # retention work; pending operations wait for the next producing run.
    assigned = (register_folder(str(run_folder)) or {}).get("ops") or []
    report_progress("running", {
        "stage": "run_folder",
        "message": (
            "Saving this file Flow's immutable source snapshot and normalized CSV in the private worker store."
            if private_snapshot else
            f"Saving immutable artifacts for this run in the private worker store "
            f"before publishing direct files to {target}."
            if direct else
            f"Saving this run's files into {run_folder.name} inside {target}."
        ),
    })
    if assigned:
        results = flow_retention.execute_ops(storage_root, assigned)
        outcomes = "; ".join(
            f"{Path(str(op.get('original_path') or '')).name or 'unknown'}: {result['outcome']}"
            + (f" ({result['detail']})" if result.get("detail") else "")
            for op, result in zip(assigned, results)
        )
        report_progress("running", {
            "stage": "run_folder_retention",
            "message": (
                f"Cleaning up old run folders (keeping the newest "
                f"{flow_retention.RUN_FOLDER_KEEP}): {outcomes}."
            ),
            "retention_results": results,
        })
    return run_folder


def execute_outlook_job(
    job: dict, report_progress, profile_dir: Path, *, run_id: int, register_folder,
) -> tuple[list[dict], list[dict], dict]:
    """Acquire, validate, and store one Outlook Inbox attachment."""
    assert_job_paths(job)
    timings = _Timings()
    source = job.get("outlook_source") or {}
    report_progress("running", {
        "stage": "outlook_search",
        "message": (
            "Searching the default Outlook Inbox for the newest email whose subject "
            f"contains {source.get('subject_contains')!r}."
        ),
    })
    with timings.measure("outlook_acquisition", report_id=job.get("report", {}).get("id")):
        acquisition = flow_outlook.acquire_attachment(
            run_id=run_id,
            profile_dir=profile_dir,
            subject_contains=source.get("subject_contains") or "",
            last_processed_identity=source.get("last_processed_identity"),
            force_reprocess=bool(source.get("force_reprocess")),
        )
    if acquisition["status"] in {"no_match", "already_processed"}:
        message = acquisition.get("message") or (
            "No new qualifying Outlook attachment was found."
        )
        report_progress("running", {
            "stage": "outlook_no_op",
            "message": message,
            "no_op": True,
            "reason": acquisition["status"],
        })
        return [], timings.finish(item_count=0), {
            "no_op": True, "reason": acquisition["status"], "message": message,
        }

    run_folder = _prepare_run_folder(
        job, profile_dir, run_id=run_id, register_folder=register_folder,
        report_progress=report_progress,
    )
    job["_runtime_artifact_store_id"] = _job_store_id(job, profile_dir)
    job["_runtime_artifact_store_ids"] = list(_runtime_store_ids(job, profile_dir))
    job["_runtime_bundle_count"] = 1
    output = _safe_output_path(run_folder, acquisition["filename"])
    report_progress("running", {
        "stage": "outlook_attachment_transfer",
        "message": f"Saving and validating Outlook attachment {acquisition['filename']}.",
    })

    def processing_progress(stage: str, message: str):
        report_progress("running", {"stage": stage, "message": message})

    with timings.measure("file_transfer", report_id=job.get("report", {}).get("id")):
        metadata = _store_completed_download(
            acquisition["path"], output,
            file_format="auto",
            requested_period=None,
            allow_raw_xlsx_fallback=False,
            require_normalized_csv=True,
            csv_preamble="none",
            strict_headers=True,
            xlsx_header_mode="first_row",
            processing_progress=processing_progress,
        )
    receipt = acquisition["receipt"]
    artifact = _decorate_artifact_storage({
        "period_key": None,
        "export_view": None,
        "bundle_index": 1,
        "bundle_count": 1,
        "status": "saved",
        "source_receipt": receipt,
        **metadata,
    }, job, profile_dir)
    return [artifact], timings.finish(item_count=1), {
        "no_op": False, "source_receipt": receipt,
    }


def execute_local_file_job(
    job: dict, report_progress, profile_dir: Path, *, run_id: int, register_folder,
) -> tuple[list[dict], list[dict], dict]:
    """Snapshot one configured file without ever modifying the source path."""
    assert_job_paths(job)
    source = job.get("local_file") or {}
    source_path = Path(str(source.get("path") or ""))
    if not source_path.is_file():
        raise RuntimeError(f"Configured source file does not exist or is not readable: {source_path}")
    suffix = source_path.suffix.casefold()
    expected_format = LOCAL_FILE_FORMAT_BY_EXTENSION.get(suffix)
    if not expected_format:
        raise RuntimeError(f"Configured source file has an unsupported extension: {source_path.name}")
    worksheet = source.get("worksheet")
    if expected_format == "csv" and worksheet is not None:
        raise RuntimeError("CSV file jobs must not specify an Excel worksheet.")
    if expected_format != "csv" and (worksheet is None or not str(worksheet).strip()):
        raise RuntimeError("Excel file jobs must specify the exact worksheet name.")
    normalized_path = flow_publish.normalize_target_path(source_path)
    if normalized_path != source.get("normalized_path"):
        raise RuntimeError("The local-file job path identity is malformed or incomplete.")

    timings = _Timings()
    report_progress("running", {
        "stage": "local_file_snapshot",
        "message": f"Reading and checksumming configured source file {source_path.name}.",
    })
    with timings.measure("local_file_snapshot", report_id=job.get("report", {}).get("id")):
        snapshot = _stable_source_snapshot(source_path)
    identity_payload = json.dumps(
        {
            "checksum": snapshot["checksum"],
            "path": normalized_path,
            "worksheet": worksheet,
        },
        ensure_ascii=False, separators=(",", ":"), sort_keys=True,
    ).encode("utf-8")
    identity = hashlib.sha256(identity_payload).hexdigest()
    previous_identity = source.get("previous_identity")
    if identity == previous_identity and not source.get("force_reprocess"):
        message = f"Source file {source_path.name} is unchanged; transformation and SQL were skipped."
        report_progress("running", {
            "stage": "local_file_no_op", "message": message, "no_op": True,
        })
        return [], timings.finish(item_count=0), {
            "no_op": True, "reason": "already_processed", "message": message,
        }

    run_folder = _prepare_run_folder(
        job, profile_dir, run_id=run_id, register_folder=register_folder,
        report_progress=report_progress,
    )
    job["_runtime_artifact_store_id"] = _job_store_id(job, profile_dir)
    job["_runtime_artifact_store_ids"] = list(_runtime_store_ids(job, profile_dir))
    job["_runtime_bundle_count"] = 1
    source_folder = run_folder / "source"
    source_folder.mkdir()
    if source_folder.is_symlink() or not source_folder.is_dir():
        raise RuntimeError(f"Private source snapshot folder is not a regular directory: {source_folder}")
    raw_output = _safe_output_path(source_folder, source_path.name)
    report_progress("running", {
        "stage": "local_file_copy",
        "message": f"Copying verified source bytes for {source_path.name} into the private run store.",
    })
    with timings.measure("file_transfer", report_id=job.get("report", {}).get("id")):
        copied = _copy_with_checksum(source_path, raw_output)
        if copied != snapshot:
            raise RuntimeError(
                f"Configured source file changed while being copied: {source_path}. Try the run again."
            )
        _verify_copied_file(
            raw_output, snapshot["file_size"], snapshot["checksum"],
            label="Configured source file",
        )
    detected = _detect_download_format(raw_output)
    if detected != expected_format:
        raise RuntimeError(
            f"Configured source {source_path.name} declares {expected_format} but its content looks like {detected}."
        )
    normalized_output = _safe_output_path(
        run_folder, f"{source_path.stem}_normalized.csv",
    )
    report_progress("running", {
        "stage": "file_normalization",
        "message": f"Normalizing {source_path.name} into an immutable CSV artifact.",
    })
    with timings.measure("file_normalization", report_id=job.get("report", {}).get("id")):
        if expected_format == "csv":
            normalization = _normalize_csv(
                raw_output, output=normalized_output,
                preamble="none", strict_headers=True,
            )
        else:
            normalization = _normalize_xlsx(
                raw_output, normalized_output, requested_weeks=[],
                header_mode="first_row", strict_headers=True,
                workbook_format=expected_format, excel_trim="none",
                worksheet_name=str(worksheet),
            )
    metadata = {**_csv_metadata(normalized_output), **normalization}
    try:
        modified_at_ns = source_path.stat().st_mtime_ns
    except OSError:
        modified_at_ns = None
    receipt = {
        "kind": "local_file",
        "identity": identity,
        "previous_identity": previous_identity,
        "raw_checksum": snapshot["checksum"],
        "normalized_path": normalized_path,
        "worksheet": worksheet,
        "config_revision": int(source.get("config_revision") or 0),
        "file_size": snapshot["file_size"],
        "modified_at_ns": modified_at_ns,
    }
    raw_artifact = _decorate_artifact_storage({
        "period_key": None, "export_view": None,
        "bundle_index": 1, "bundle_count": 1,
        "status": "source_snapshot", "source_receipt": receipt,
        "file_path": str(raw_output), "filename": raw_output.name,
        "file_size": snapshot["file_size"], "checksum": snapshot["checksum"],
        "row_count": None, "detected_format": detected,
    }, job, profile_dir)
    normalized_artifact = _decorate_artifact_storage({
        "period_key": None, "export_view": None,
        "bundle_index": 1, "bundle_count": 1,
        "status": "saved", "source_receipt": receipt,
        "original_file_path": str(raw_output),
        "original_filename": raw_output.name,
        "original_file_size": snapshot["file_size"],
        "detected_format": detected,
        "file_path": str(normalized_output), "filename": normalized_output.name,
        **metadata,
    }, job, profile_dir)
    return [raw_artifact, normalized_artifact], timings.finish(item_count=1), {
        "no_op": False,
        "source_receipt": receipt,
        "sql_artifacts": [normalized_artifact],
    }


def execute_job(
    page: Page, job: dict, report_progress, profile_dir: Path,
    download_staging_dir: Path | None = None,
    artifacts: list[dict] | None = None,
    *, run_id: int, register_folder, headed: bool = False, task_assignment: dict | None = None,
) -> tuple[list[dict], list[dict]]:
    # The caller may own the artifact list. Files are appended in place, so
    # everything saved before a mid-bundle failure stays visible to the
    # caller's failure report - the record Resume later relies on.
    assert_job_paths(job)
    timings = _Timings()
    job["_runtime_artifact_store_id"] = _job_store_id(job, profile_dir)
    job["_runtime_artifact_store_ids"] = list(_runtime_store_ids(job, profile_dir))
    report_progress("running", {"stage": "opening_report", "message": "Opening the configured report."})
    is_asap = job["site"].get("adapter") == ASAP_PORTAL_ADAPTER
    is_gscm = job["site"].get("adapter") == GSCM_PORTAL_ADAPTER
    ready_text = job["report"].get("ready_text")
    open_export = job["report"].get("open_export_text")

    if task_assignment is not None:
        from app.flow_parallel_worker import task_output_folder
        target = task_output_folder(job, task_assignment, run_id)
    else:
        target = _prepare_run_folder(
            job, profile_dir, run_id=run_id, register_folder=register_folder,
            report_progress=report_progress,
        )

    periods = job["downloads"].get("periods") or [None]
    # An HTML dashboard has no Export Wizard: its data leaves through the
    # download links the scanner catalogued, and each selected link is one
    # file of the bundle exactly like an export view is.
    report = job.get("report", {})
    dashboard_links = list(report.get("download_links") or [])
    is_dashboard = is_asap and bool(dashboard_links)
    # New scans identify embedded dashboards explicitly. The export-view
    # fallback keeps flows saved before that marker was introduced working.
    is_html_dashboard = is_dashboard and (
        (report.get("automation") or {}).get("kind") == "html_dashboard"
        or not report.get("export_views")
    )
    downstream_requires_csv = bool(
        job.get("transformation", {}).get("enabled")
        or job.get("sql_handoff", {}).get("enabled")
    )
    if is_asap:
        asap_type = resolve_asap_download_type(
            job.get("downloads", {}).get("asap_download_type"),
            legacy_file_format=job.get("downloads", {}).get("file_format"),
        )
        if asap_type.download_only and downstream_requires_csv:
            raise RuntimeError(
                f"{asap_type.label} is download-only; transformation and SQL handoff are unsupported."
            )
    require_normalized_csv = not (
        is_html_dashboard and not downstream_requires_csv
    )
    from app.flow_tasks import task_matrix
    tasks = task_matrix(job)
    if task_assignment is not None and not any(task['ordinal'] == task_assignment['ordinal'] and task['key'] == task_assignment['task_key'] for task in tasks):
        raise RuntimeError('Download task does not belong to the frozen export matrix.')
    job["_runtime_bundle_count"] = len(tasks)
    if artifacts is None:
        artifacts = []

    def _download_task(index: int, task: dict, attempt: int = 1) -> dict:
        period = task["period"]
        export_view = task["export_view"]
        download_link = task.get("download_link")
        with timings.measure("navigation", report_id=job["report"].get("id")):
            if is_gscm:
                # A GSCM bookmark already carries its filters, period, and
                # dimensions. Opening it is the whole configuration step.
                frame = None

                def _notify_auth(message: str):
                    report_progress(
                        "running",
                        {"stage": "authentication", "message": message,
                         "item_index": index, "item_count": len(tasks)},
                        artifacts,
                    )

                def _open_gscm_bookmark():
                    if attempt > 1:
                        # The first failure can leave Setting1 mounted with an
                        # empty/stale virtual grid. A same-host ``open_portal``
                        # deliberately reuses that tree, so make a retry real by
                        # rebuilding Nexacro before opening the bookmark again.
                        flow_gscm.reload_portal(page, job)
                    flow_gscm.open_bookmark(page, job, report_progress=lambda message: report_progress(
                        "running",
                        {"stage": "opening_report", "message": message,
                         "item_index": index, "item_count": len(tasks)},
                        artifacts,
                    ))

                _gscm_call(page, headed, _notify_auth, _open_gscm_bookmark, profile_dir)
                load_buffer_ms = (
                    GSCM_INITIAL_LOAD_BUFFER_MS
                    if attempt == 1 else GSCM_RETRY_LOAD_BUFFER_MS
                )
                report_progress(
                    "running",
                    {
                        "stage": "report_execution",
                        "message": (
                            "Waiting "
                            f"{load_buffer_ms // 1_000} seconds for the GSCM "
                            "report data to finish loading before download."
                        ),
                        "item_index": index,
                        "item_count": len(tasks),
                        "attempt": attempt,
                    },
                    artifacts,
                )
                page.wait_for_timeout(load_buffer_ms)
            elif is_asap:
                frame = _asap_open_report(page, job, profile_dir)
                if download_link:
                    # Dashboard reports have no export-view tabs to activate.
                    _asap_wait_for_loading_clear(page)
                    frame = _asap_frame(page)
                else:
                    frame, selected_view = _asap_activate_export_view(page, frame, export_view)
                    if export_view and selected_view != export_view:
                        raise RuntimeError(
                            f"ASAP activated the wrong export view. Requested: {export_view}. "
                            f"Activated: {selected_view}."
                        )
            else:
                page.goto(job["report"]["url"], wait_until="domcontentloaded", timeout=120_000)
                frame = None
                if ready_text:
                    page.get_by_text(ready_text, exact=False).first.wait_for(
                        state="visible", timeout=120_000,
                    )
                if open_export:
                    _click_named(page, open_export)
        if not is_gscm:
            # A GSCM bookmark has nothing left to configure: its filters were
            # saved inside GSCM and applied when the bookmark opened.
            report_progress(
                "running",
                {
                    "stage": "configuring",
                    "message": f"Configuring export {index} of {len(tasks)}: {export_view or job['report']['name']}.",
                    "period": period, "export_view": export_view,
                    "item_index": index, "item_count": len(tasks),
                },
                artifacts,
            )
        if is_gscm:
            report_progress(
                "running",
                {
                    "stage": "file_export",
                    "message": (
                        "Verifying the rendered GSCM bookmark title and exporting "
                        f"it to Excel: {job['report']['name']}."
                    ),
                    "item_index": index, "item_count": len(tasks),
                },
                artifacts,
            )
            with timings.measure("file_export", report_id=job["report"].get("id")):
                staged_file = _edge_completed_download(
                    page, lambda: flow_gscm.trigger_excel_export(page, job),
                )
                export_pages = []
        elif is_asap:
            view_job = {
                **job,
                "report": {
                    **job["report"],
                    "filters": _asap_filters_for_export_view(job, export_view),
                },
            }
            with timings.measure("configuration", report_id=job["report"].get("id")):
                _asap_apply_configuration(frame, view_job, period)
            if _has_named_control(frame, "RUN"):
                report_progress(
                    "running",
                    {
                        "stage": "report_execution",
                        "message": f"Running export {index} of {len(tasks)}: {export_view or job['report']['name']}.",
                        "period": period, "export_view": export_view,
                    },
                    artifacts,
                )
                def _report_rerun(exc: Exception):
                    report_progress(
                        "running",
                        {
                            "stage": "report_execution",
                            "message": (
                                f"ASAP rendered no rows for export {index} of {len(tasks)}; "
                                f"re-running the report once. {exc}"
                            ),
                            "period": period, "export_view": export_view,
                        },
                        artifacts,
                    )

                with timings.measure("report_execution", report_id=job["report"].get("id")):
                    frame = _asap_run_report_with_retry(page, on_retry=_report_rerun)
            report_progress(
                "running",
                {
                    "stage": "file_export",
                    "message": (
                    f"Downloading {index} of {len(tasks)} from the dashboard: {download_link}."
                    if download_link else
                    f"Exporting {job['downloads'].get('file_format', 'csv').upper()} {index} of {len(tasks)}: {export_view or job['report']['name']}."
                ),
                    "period": period, "export_view": export_view,
                },
                artifacts,
            )
            with timings.measure("file_export", report_id=job["report"].get("id")):
                staging = download_staging_dir or profile_dir / "downloads"
                if download_link:
                    staged_file = _asap_download_dashboard_link(
                        page, download_link, staging, job,
                    )
                    export_pages = []
                else:
                    staged_file, export_pages = _asap_download_with_retry(
                        page, frame, job, staging,
                    )
        else:
            _apply_configuration(page, job, period)
            with page.expect_download(timeout=DOWNLOAD_MAX_TIMEOUT_SECONDS * 1_000) as pending:
                _click_named(page, job["report"]["download_text"])
            download = pending.value
            export_pages = []
        try:
            filename = _render_filename(
                job["downloads"]["filename_template"], job, period, index, export_view,
            )
            output = _safe_output_path(target, filename)
        except Exception as exc:
            if is_asap or is_gscm:
                portal = "GSCM" if is_gscm else "ASAP"
                raise _CompletedDownloadProcessingError(
                    f"Browser completed the {portal} download, but preparing its target path "
                    f"failed. {portal} will not be reopened: {exc}"
                ) from exc
            raise
        if is_asap or is_gscm:
            try:
                report_progress(
                    "running",
                    {
                        "stage": "file_transfer",
                        "message": (
                            f"Browser finished download {index} of {len(tasks)}; saving it "
                            "to the configured target folder."
                        ),
                        "period": period,
                        "export_view": export_view,
                        "item_index": index,
                        "item_count": len(tasks),
                    },
                    artifacts,
                )
            except Exception as exc:
                portal = "GSCM" if is_gscm else "ASAP"
                raise _CompletedDownloadProcessingError(
                    f"Browser completed the {portal} download, but reporting the target-storage "
                    f"stage failed. {portal} will not be reopened: {exc}"
                ) from exc

        def _processing_progress(stage: str, message: str):
            report_progress(
                "running",
                {
                    "stage": stage,
                    "message": message,
                    "period": period,
                    "export_view": export_view,
                    "item_index": index,
                    "item_count": len(tasks),
                },
                artifacts,
            )

        with timings.measure("file_transfer", report_id=job["report"].get("id")):
            # ASAP closes its own export wizard after emitting the download.
            # Do not query or close that vanished page object: Edge can leave
            # it half-detached and any later Playwright call can block forever.
            # The next period reopens the report in the surviving main page.
            if is_asap or is_gscm:
                try:
                    metadata = _store_completed_download(
                        staged_file, output,
                        file_format=job["downloads"].get("file_format") or "csv",
                        asap_download_type=job["downloads"].get("asap_download_type") if is_asap else None,
                        excel_trim=job["downloads"].get("excel_trim") or "none",
                        requested_period=period,
                        allow_raw_xlsx_fallback=(
                            is_gscm
                            and not downstream_requires_csv
                        ),
                        require_normalized_csv=require_normalized_csv,
                        processing_progress=_processing_progress,
                    )
                except Exception as exc:
                    if is_gscm:
                        raise _CompletedDownloadProcessingError(
                            "Browser completed the GSCM workbook download, but local processing "
                            f"failed. GSCM will not be reopened: {exc}"
                        ) from exc
                    if is_asap:
                        raise _CompletedDownloadProcessingError(
                            "Browser completed the ASAP download, but target storage or local "
                            f"processing failed. ASAP will not be reopened: {exc}"
                        ) from exc
                    raise
            else:
                download.save_as(output)
                normalization = _normalize_csv(output)
                metadata = {
                    **_csv_metadata(output), **normalization,
                    "file_path": str(output), "filename": output.name,
                }
        return {
            "period_key": period,
            "export_view": export_view,
            "bundle_index": index,
            "bundle_count": len(tasks),
            "status": "saved",
            "export_transport": "browser",
            **metadata,
        }

    capture_replay = (is_asap or is_gscm) and flow_replay.enabled(job)

    def _record_export_recipe(recorder, task: dict, artifact: dict, index: int) -> None:
        """Best-effort: remember the HTTP request behind a finished export."""
        baseline = artifact.get("original_file_path") or artifact.get("file_path")
        if not baseline:
            return
        task_key = _export_task_key(task["export_view"], task["period"])
        try:
            stored = flow_replay.store_capture(
                profile_dir, job, task_key, recorder, Path(str(baseline)),
            )
        except Exception:
            return
        if stored:
            report_progress(
                "running",
                {
                    "stage": "file_export",
                    "message": (
                        f"Recorded the HTTP request behind export {index} of "
                        f"{len(tasks)}; future runs will try to replay it "
                        "without driving the portal UI."
                    ),
                    "period": task["period"], "export_view": task["export_view"],
                    "item_index": index, "item_count": len(tasks),
                },
                artifacts,
            )

    def _replay_task(index: int, task: dict) -> dict | None:
        """Replay a recorded export request instead of driving the portal UI.

        Attempted once per file and never retried: any rejection - a missing
        or expired recipe, a network error, a sign-in page, a body from the
        wrong file family, or a processing failure downstream - falls back to
        the browser flow, which records a fresh recipe as it succeeds.
        """
        context = getattr(page, "context", None)
        if not capture_replay or context is None:
            return None
        period = task["period"]
        export_view = task["export_view"]
        task_key = _export_task_key(export_view, period)
        recipe = flow_replay.load_recipe(profile_dir, job, task_key)
        if recipe is None:
            return None
        report_progress(
            "running",
            {
                "stage": "file_export",
                "message": (
                    f"Export {index} of {len(tasks)}: replaying the recorded "
                    "HTTP export request instead of driving the portal UI."
                ),
                "period": period, "export_view": export_view,
                "item_index": index, "item_count": len(tasks),
            },
            artifacts,
        )
        staging = download_staging_dir or profile_dir / "downloads"
        with timings.measure("file_export", report_id=job["report"].get("id")):
            staged_file = flow_replay.try_replay(context, recipe, staging)
        if staged_file is None:
            flow_replay.forget_recipe(profile_dir, job, task_key)
            report_progress(
                "running",
                {
                    "stage": "file_export",
                    "message": (
                        f"Export {index} of {len(tasks)}: the recorded request "
                        "did not replay cleanly; falling back to the browser "
                        "export."
                    ),
                    "period": period, "export_view": export_view,
                    "item_index": index, "item_count": len(tasks),
                },
                artifacts,
            )
            return None

        def _processing_progress(stage: str, message: str):
            report_progress(
                "running",
                {
                    "stage": stage, "message": message,
                    "period": period, "export_view": export_view,
                    "item_index": index, "item_count": len(tasks),
                },
                artifacts,
            )

        try:
            filename = _render_filename(
                job["downloads"]["filename_template"], job, period, index, export_view,
            )
            output = _safe_output_path(target, filename)
            with timings.measure("file_transfer", report_id=job["report"].get("id")):
                metadata = _store_completed_download(
                    staged_file, output,
                    file_format=job["downloads"].get("file_format") or "csv",
                    asap_download_type=job["downloads"].get("asap_download_type") if is_asap else None,
                    excel_trim=job["downloads"].get("excel_trim") or "none",
                    requested_period=period,
                    allow_raw_xlsx_fallback=(is_gscm and not downstream_requires_csv),
                    require_normalized_csv=require_normalized_csv,
                    processing_progress=_processing_progress,
                )
        except Exception as exc:
            # Unlike a browser download, a replayed file that fails processing
            # costs nothing to redo: the portal was never opened, so fall back
            # to the full UI export instead of failing the run.
            flow_replay.forget_recipe(profile_dir, job, task_key)
            report_progress(
                "running",
                {
                    "stage": "file_export",
                    "message": (
                        f"Export {index} of {len(tasks)}: the replayed file "
                        "failed processing; falling back to the browser "
                        f"export. {str(exc)[:2000]}"
                    ),
                    "period": period, "export_view": export_view,
                    "item_index": index, "item_count": len(tasks),
                },
                artifacts,
            )
            return None
        return {
            "period_key": period,
            "export_view": export_view,
            "bundle_index": index,
            "bundle_count": len(tasks),
            "status": "saved",
            "export_transport": "http_replay",
            **metadata,
        }

    completed_keys = _resume_completed_keys(job)
    resumed_run_id = (job.get("resume") or {}).get("from_run_id")
    for index, task in enumerate(tasks, start=1):
        if task_assignment is not None and index != task_assignment['ordinal']:
            continue
        if _export_task_key(task["export_view"], task["period"]) in completed_keys:
            report_progress(
                "running",
                {
                    "stage": "resume_skip",
                    "message": (
                        f"Export {index} of {len(tasks)} was already saved by "
                        f"run #{resumed_run_id}; skipping."
                    ),
                    "period": task["period"], "export_view": task["export_view"],
                    "item_index": index, "item_count": len(tasks),
                },
                artifacts,
            )
            continue

        replayed = _replay_task(index, task)
        if replayed is not None:
            artifacts.append(_decorate_artifact_storage(replayed, job, profile_dir))
            continue

        task_attempts = GSCM_EXPORT_TASK_ATTEMPTS if is_gscm else EXPORT_TASK_ATTEMPTS

        def _task_retry(attempt: int, exc: Exception, *, index=index, task=task):
            report_progress(
                "running",
                {
                    "stage": "export_retry",
                    "message": (
                        f"Export {index} of {len(tasks)} failed on attempt "
                        f"{attempt} of {task_attempts}; restarting this file "
                        f"from report navigation. {exc}"
                    ),
                    "period": task["period"], "export_view": task["export_view"],
                    "item_index": index, "item_count": len(tasks), "attempt": attempt,
                },
                artifacts,
            )

        def _run_attempt(attempt: int, *, index=index, task=task) -> dict:
            # The recorder observes the whole attempt so the capture works
            # whichever page or popup the portal routes the download through;
            # matching against the download's own URL keeps it precise.
            # Capture is best-effort: a page without a usable context simply
            # exports without recording a recipe.
            context = getattr(page, "context", None) if capture_replay else None
            recorder = (
                flow_replay.ExportRequestRecorder(context)
                if context is not None else None
            )
            try:
                artifact = _download_task(index, task, attempt)
            finally:
                if recorder is not None:
                    recorder.detach()
            if recorder is not None:
                _record_export_recipe(recorder, task, artifact, index)
            return artifact

        artifacts.append(_export_task_with_retry(
            page,
            _run_attempt,
            _task_retry,
            max_attempts=task_attempts,
        ))
        _decorate_artifact_storage(artifacts[-1], job, profile_dir)
    return artifacts, timings.finish(item_count=len(artifacts))


def _copy_carried_artifact(artifact: dict, run_folder: Path) -> dict:
    """Give a resumed direct run its own immutable copy of a carried task."""
    clone = dict(artifact)
    copied_paths: dict[str, tuple[Path, dict]] = {}

    def copy_field(path_key: str, size_key: str, checksum_key: str) -> None:
        raw = clone.get(path_key)
        if not raw:
            return
        source = Path(str(raw))
        source_key = os.path.normcase(os.path.abspath(str(source)))
        if source_key in copied_paths:
            output, metadata = copied_paths[source_key]
        else:
            output = _safe_output_path(run_folder, source.name)
            metadata = _copy_with_checksum(source, output)
            _verify_copied_file(
                output, metadata["file_size"], metadata["checksum"],
                label="Carried Flow artifact",
            )
            copied_paths[source_key] = (output, metadata)
        clone[path_key] = str(output)
        clone[size_key] = metadata["file_size"]
        clone[checksum_key] = metadata["checksum"]

    copy_field("file_path", "file_size", "checksum")
    copy_field(
        "deliverable_file_path", "deliverable_file_size", "deliverable_checksum",
    )
    copy_field("original_file_path", "original_file_size", "original_checksum")
    clone.update({
        "storage_scope": "worker_private",
        "publish_status": "pending",
        "published_file_path": None,
        "published_filename": None,
        "published_file_size": None,
        "published_checksum": None,
        "status": "saved",
    })
    return clone


def _publish_direct_artifacts(
    job: dict, artifacts: list[dict], *, run_id: int, report_progress,
) -> list[dict]:
    """Publish a complete validated bundle, including Resume-carried tasks."""
    assert_job_paths(job)
    if job.get("downloads", {}).get("output_mode", "run_folders") != "direct_replace":
        return artifacts
    run_folder = Path(str(job.get("_runtime_run_folder") or ""))
    if not run_folder.is_dir():
        raise RuntimeError("Direct publication has no registered private run folder.")
    by_task = {
        _export_task_key(item.get("export_view"), item.get("period_key")): item
        for item in artifacts if item.get("status") == "saved"
    }
    for task_key, carried in _validated_resume_artifacts(job).items():
        if task_key not in by_task:
            by_task[task_key] = _copy_carried_artifact(carried, run_folder)
            by_task[task_key]["artifact_store_id"] = job.get("_runtime_artifact_store_id")
    bundle = sorted(
        by_task.values(),
        key=lambda item: (
            int(item.get("bundle_index") or 0),
            _export_task_key(item.get("export_view"), item.get("period_key")),
        ),
    )
    expected = int(job.get("_runtime_bundle_count") or 0) or max(
        [int(item.get("bundle_count") or 0) for item in bundle] or [len(bundle)]
    )
    if not bundle or len(bundle) != expected:
        raise RuntimeError(
            f"Direct publication requires the complete bundle ({len(bundle)} of {expected} validated)."
        )
    # Keep the caller's list as the authoritative union even if publication
    # raises. The worker's final failed progress post must not overwrite the
    # earlier publish event with only the newly downloaded subset.
    artifacts[:] = bundle
    report_progress("running", {
        "stage": "direct_publish",
        "message": (
            f"Publishing {len(bundle)} validated deliverable(s) directly to "
            f"{job['downloads']['target_folder']}."
        ),
        "item_count": len(bundle),
    }, bundle)
    try:
        published = flow_publish.publish_bundle(
            Path(job["downloads"]["target_folder"]), run_id, bundle,
        )
    except Exception as exc:
        for artifact in bundle:
            artifact["publish_status"] = "failed"
            artifact["publish_error"] = str(exc)
        report_progress("running", {
            "stage": "publish_failed",
            "message": str(exc),
            "item_count": len(bundle),
            "bundle": [
                {
                    "export_view": item.get("export_view"),
                    "period_key": item.get("period_key"),
                    "filename": item.get("deliverable_filename"),
                    "checksum": item.get("deliverable_checksum"),
                    "status": "failed",
                }
                for item in bundle
            ],
        }, bundle)
        raise
    for artifact, result in zip(bundle, published):
        artifact.update(result)
    report_progress("running", {
        "stage": "publish_complete",
        "message": f"Published {len(bundle)} stable direct output file(s).",
        "item_count": len(bundle),
        "published": [
            {
                "export_view": item.get("export_view"),
                "period_key": item.get("period_key"),
                "filename": item["published_filename"],
                "file_path": item["published_file_path"],
                "file_size": item["published_file_size"],
                "checksum": item["published_checksum"],
                "status": item["publish_status"],
            }
            for item in bundle
        ],
    }, bundle)
    return bundle


def _code_version() -> str:
    """The deployment stamp setup.ps1 writes next to this code.

    Every failed run raises the question of whether the desktop worker is
    actually executing the code that was just deployed. Surfacing the stamp at
    startup and in the registration answers it without a remote desktop trip.
    """
    try:
        return (_CODE_DIR / "VERSION").read_text(encoding="utf-8").strip() or "unknown"
    except OSError:
        return "unknown"


def execute_flow(page, job: dict, progress, profile_dir: Path, download_staging_dir=None,
                 *, run_id: int, register_folder, headed: bool = False,
                 artifacts=None, state=None, run_started=None, acquire_bundle=None) -> dict:
    """One acquisition/publication/transform/SQL path for every launch surface.

    Callers own process locks, browser lifetime, progress transport and retention
    registration. Failure state retains partial artifacts and phase timings.
    """
    if job.get('flow', {}).get('execution_method') == 'recorded' and job.get('job_type') != 'sql_retry':
        from app.flow_recording_runtime import execute_recorded_flow
        return execute_recorded_flow(page, job, progress, profile_dir, download_staging_dir,
            run_id=run_id, register_folder=register_folder, headed=headed, artifacts=artifacts,
            state=state, run_started=run_started)
    state = state if state is not None else {}
    artifacts = artifacts if artifacts is not None else []
    timings = []
    transformation_started = sql_started = sql_result = source_receipt = None
    no_op = False
    source_outcome = {}
    run_started = run_started if run_started is not None else time.perf_counter()
    try:
        sql_only = job.get("job_type") == "sql_retry"
        assert_job_paths(job)
        if sql_only:
            if not job.get("sql_handoff", {}).get("enabled"):
                raise RuntimeError("SQL-only retry has no enabled SQL target.")
            sql_artifacts = job.get("sql_retry", {}).get("artifacts") or []
            if not sql_artifacts:
                raise RuntimeError("SQL-only retry has no saved CSV artifacts.")
            required_store = job.get("sql_retry", {}).get(
                "required_artifact_store_id"
            )
            if required_store and required_store not in _runtime_store_ids(job, profile_dir):
                raise RuntimeError(
                    "SQL Retry was claimed by a worker that cannot access its private artifact store."
                )
            invalid = [
                item.get("filename") or str(item.get("file_path") or "artifact")
                for item in sql_artifacts
                if not flow_publish.artifact_file_valid(item)
            ]
            if invalid:
                raise RuntimeError(
                    "Saved SQL artifact is missing or changed in the private worker store: "
                    + ", ".join(invalid[:10])
                )
            artifacts = sql_artifacts
            source_receipt = (
                job.get("source_receipt")
                or job.get("outlook_source_receipt")
            )
            timings = [{"phase": "total", "duration_ms": 0, "status": "running"}]
            progress(
                "running",
                {
                    "stage": "sql_retry",
                    "message": (
                        f"Reusing {len(sql_artifacts)} saved CSV file(s) from run "
                        f"#{job.get('sql_retry', {}).get('source_run_id')}. "
                        "The source will not be opened or downloaded again."
                    ),
                },
                artifacts, timings,
            )
        elif (
            (job.get("flow", {}).get("source_type") or "portal") == "file"
            or bool((job.get("local_file") or {}).get("enabled"))
        ):
            if (
                (job.get("flow", {}).get("source_type") or "portal") != "file"
                or job.get("schema_version") != 3
                or not (job.get("local_file") or {}).get("enabled")
            ):
                raise RuntimeError("Local-file job payload is malformed or uses an unsupported schema version.")
            artifacts, timings, source_outcome = execute_local_file_job(
                job, progress, profile_dir,
                run_id=run_id, register_folder=register_folder,
            )
            no_op = bool(source_outcome.get("no_op"))
            source_receipt = source_outcome.get("source_receipt")
            sql_artifacts = source_outcome.get("sql_artifacts") or []
        elif (job.get("outlook_source") or {}).get("enabled"):
            artifacts, timings, outlook_outcome = execute_outlook_job(
                job, progress, profile_dir,
                run_id=run_id, register_folder=register_folder,
            )
            no_op = bool(outlook_outcome.get("no_op"))
            source_receipt = outlook_outcome.get("source_receipt")
            source_outcome = outlook_outcome
            sql_artifacts = artifacts
        else:
            if (job.get("flow", {}).get("source_type") or "portal") != "portal":
                raise RuntimeError("Flow job names an unsupported source type.")
            # Share the artifact list so a mid-bundle failure still
            # reports every file saved before the error - the final
            # failed progress post below sends this same list.
            artifacts, timings = (acquire_bundle or execute_job)(
                page, job, progress, profile_dir, download_staging_dir,
                artifacts=artifacts,
                run_id=run_id, register_folder=register_folder,
                headed=headed,
            )
            sql_artifacts = artifacts
        if not sql_only and not no_op:
            artifacts = _publish_direct_artifacts(
                job, artifacts, run_id=run_id,
                report_progress=progress,
            )
            if (job.get("flow", {}).get("source_type") or "portal") != "file":
                sql_artifacts = artifacts
        if (
            not sql_only and not no_op
            and job.get("transformation", {}).get("enabled")
        ):
            progress(
                "running",
                {"stage": "transformation", "message": f"Transforming {len(sql_artifacts)} normalized file(s)."},
                artifacts, timings,
            )
            transformation_started = time.perf_counter()
            sql_artifacts = _run_transformations(sql_artifacts, job["transformation"])
            timings.insert(max(0, len(timings) - 1), {
                "phase": "transformation",
                "duration_ms": round((time.perf_counter() - transformation_started) * 1000),
                "status": "succeeded", "item_count": len(sql_artifacts),
            })
            timings[-1]["duration_ms"] = round((time.perf_counter() - run_started) * 1000)
            artifacts = [*artifacts, *sql_artifacts]
            progress(
                "running",
                {
                    "stage": "transformation_complete",
                    "message": f"Created {len(sql_artifacts)} transformed file(s) in script_results.",
                    "results": [
                        {
                            "source": item.get("source_file_path"),
                            "output": item.get("file_path"),
                            "rows": item.get("row_count"),
                            "stdout": item.get("script_stdout"),
                            "stderr": item.get("script_stderr"),
                        }
                        for item in sql_artifacts
                    ],
                },
                artifacts, timings,
            )
        if not no_op and job.get("sql_handoff", {}).get("enabled"):
            from app.flow_sql import load_artifacts
            source_label = "transformed" if job.get("transformation", {}).get("enabled") else "downloaded"
            if sql_only:
                source_label = "saved"
            progress("running", {"stage": "sql_insertion", "message": f"Loading {source_label} files into SQL."}, artifacts, timings)
            sql_started = time.perf_counter()

            def sql_progress(detail: dict):
                progress("running", detail, artifacts, timings)

            sql_result = load_artifacts(
                sql_artifacts, job["sql_handoff"], progress=sql_progress,
            )
            timings.insert(max(0, len(timings) - 1), {
                "phase": "sql_insertion",
                "duration_ms": round((time.perf_counter() - sql_started) * 1000),
                "status": "succeeded",
            })
            timings[-1]["duration_ms"] = round((time.perf_counter() - run_started) * 1000)
            progress(
                "running",
                {
                    "stage": "sql_insertion_complete",
                    "message": f"Inserted {sql_result['rows_written']} row(s) from {sql_result['files_loaded']} file(s).",
                    **sql_result,
                },
                artifacts, timings,
            )
        progress(
            "succeeded", {
                "stage": "complete",
                "no_op": no_op,
                "message": (
                    source_outcome.get("message", "The configured source is unchanged.")
                    if no_op
                    else f"SQL-only retry committed {sql_result['rows_written']} row(s) from {sql_result['files_loaded']} saved file(s)."
                    if sql_only
                    else f"Saved the full {len(sql_artifacts)}-export bundle and committed {sql_result['rows_written']} row(s) to {sql_result['target']}."
                    if sql_result is not None
                    else f"Saved {len(sql_artifacts)} transformed CSV file(s) after {len(artifacts) - len(sql_artifacts)} download(s)."
                    if job.get("transformation", {}).get("enabled")
                    else f"Saved {len(artifacts)} {_completed_export_format_label(artifacts, job.get('downloads', {}).get('file_format') or 'csv')} export(s)."
                ),
            },
            artifacts, timings, source_receipt=source_receipt,
        )
    except Exception:
        if timings:
            if transformation_started is not None and not any(
                item.get("phase") == "transformation" for item in timings
            ):
                timings.insert(-1, {
                    "phase": "transformation",
                    "duration_ms": round((time.perf_counter() - transformation_started) * 1000),
                    "status": "failed",
                })
            if sql_started is not None and not any(
                item.get("phase") == "sql_insertion" for item in timings
            ):
                timings.insert(-1, {
                    "phase": "sql_insertion",
                    "duration_ms": round((time.perf_counter() - sql_started) * 1000),
                    "status": "failed",
                })
            timings[-1].update({
                "duration_ms": round((time.perf_counter() - run_started) * 1000),
                "status": "failed",
            })
        else:
            timings = [{"phase": "total", "duration_ms": round((time.perf_counter() - run_started) * 1000), "status": "failed"}]
        raise
    finally:
        state.update(artifacts=artifacts, timings=timings, transformation_started=transformation_started,
                     sql_started=sql_started, sql_result=sql_result, source_receipt=source_receipt, no_op=no_op)
    return state


def run_worker(server: str, worker_id: str, display_name: str, profile_dir: Path, headed: bool,
               once: bool, idle_exit_seconds: int = 0):
    from app import flow_browser
    code_version = _code_version()
    print(f"Worker {worker_id} starting with code version {code_version}.", flush=True)
    with httpx.Client(base_url=server.rstrip("/"), headers={"User-Agent": "Metronome-Flow-Worker/1"}) as client:
        registration = {
            "worker_id": worker_id,
            "display_name": display_name,
            "capabilities": {"adapters": ["web_export", ASAP_PORTAL_ADAPTER, GSCM_PORTAL_ADAPTER, OUTLOOK_ATTACHMENT_ADAPTER, LOCAL_FILE_ADAPTER], "headed": headed, "process_id": os.getpid(), "delete_existing": False, "overwrite_existing": True, "artifact_store_id": flow_publish.artifact_store_id(profile_dir), "code_version": code_version},
        }
        from app.flow_capacity import slot_number
        from app import flow_tasks
        registration['capabilities'].update(pool='headed' if headed else 'headless', slot=slot_number(worker_id, 'headed' if headed else 'headless'))
        registration['capabilities'][flow_tasks.CAPABILITY] = True
        registration['capabilities'][flow_tasks.HEADED_CAPABILITY] = True
        registration['capabilities'][flow_browser.CAPABILITY] = True
        registration['capabilities']['recorded_flows_v1'] = True
        registration['capabilities']['recorded_date_batches_v1'] = True
        registration['capabilities']['flow_recorder_v1'] = headed
        registration['capabilities']['flow_recorder_controls_v1'] = headed
        # Metronome can take several minutes to boot after an update (service
        # reinstall, migrations, first-request warmup), and this worker now
        # starts first. Outlasting that boot beats dying at two minutes and
        # leaning on the service manager's restart loop.
        for attempt in range(300):
            try:
                registered = _api(client, "POST", "/api/flows/worker/register", registration)
                shared_root = registered.get("flows_root")
                if shared_root:
                    store = Path(shared_root) / ".metronome" / "artifacts"
                    registration["capabilities"]["shared_flow_artifacts"] = True
                    registration["capabilities"]["artifact_store_ids"] = [
                        flow_publish.artifact_store_id(profile_dir),
                        flow_publish.artifact_store_id(profile_dir, store_root=store),
                    ]
                    _api(client, "POST", "/api/flows/worker/register", registration)
                break
            except (httpx.HTTPError, OSError) as exc:
                if attempt == 299:
                    raise RuntimeError(f"Could not register worker after 600 seconds: {exc}") from exc
                time.sleep(2)
        print(f"Worker {worker_id} registered with {server}.", flush=True)
        with sync_playwright() as playwright:
            download_staging_dir = profile_dir / "downloads"
            download_staging_dir.mkdir(parents=True, exist_ok=True)
            context, page, current_channel = None, None, None
            idle_since = time.monotonic()
            while True:
                # Refresh concrete shared roots before claims, including roots
                # frozen in queued recovery jobs after a Paths setting change.
                registered = _api(client, "POST", "/api/flows/worker/register", registration)
                roots = registered.get("artifact_store_roots") or []
                if roots:
                    ids = {flow_publish.artifact_store_id(profile_dir),
                           *(flow_publish.artifact_store_id(profile_dir, store_root=Path(root)) for root in roots)}
                    if ids != set(registration["capabilities"].get("artifact_store_ids") or []):
                        registration["capabilities"].update(shared_flow_artifacts=True, artifact_store_ids=sorted(ids))
                        _api(client, "POST", "/api/flows/worker/register", registration)
                claimed = _api(client, "POST", f"/api/flows/worker/{worker_id}/claim")
                run = claimed.get("run")
                scan = claimed.get("scan")
                work = run or scan or claimed.get('task')
                if work:
                    browser_job = work['job']
                    portal_work = (browser_job.get('flow', {}).get('source_type') not in {'file', 'outlook'}
                                   and browser_job.get('job_type') != 'sql_retry' and not browser_job.get('recording_operation'))
                    requested_channel = flow_browser.channel_for(browser_job)
                    if portal_work and (context is None or requested_channel != current_channel):
                        try:
                            if context is not None:
                                context.close()
                            context = None
                            context = flow_browser.launch(playwright, profile_dir, requested_channel,
                                headed=headed, downloads=download_staging_dir)
                            page = context.pages[0] if context.pages else context.new_page()
                            current_channel = requested_channel
                        except Exception as exc:
                            message = f'Could not open {flow_browser.CHANNELS[requested_channel]}. Check its installation and browser policies. {exc}'
                            detail = {'stage': 'browser_launch', 'message': message}
                            if run:
                                _api(client, 'POST', f'/api/flows/worker/{worker_id}/runs/{run["id"]}/progress',
                                    {'status': 'failed', 'progress': detail, 'error': message})
                            elif scan:
                                _api(client, 'POST', f'/api/flows/worker/{worker_id}/scans/{scan["id"]}/progress',
                                    {'status': 'failed', 'progress': detail, 'error': message, 'complete': False})
                            else:
                                _api(client, 'POST', f'/api/flows/worker/{worker_id}/tasks/{work["id"]}/progress',
                                    {'status': 'failed', 'lease_token': work['lease_token'], 'progress': detail})
                            if once:
                                break
                            continue
                if claimed.get('task'):
                    from app.flow_parallel_worker import execute_task
                    idle_since = time.monotonic()
                    execute_task(client, worker_id, claimed['task'], page, profile_dir, download_staging_dir, headed=headed)
                    idle_since = time.monotonic()
                    if once:
                        break
                    continue
                if not run and not scan:
                    if once:
                        break
                    if headed and registered.get('headed_work_pending'):
                        idle_since = time.monotonic()
                    if idle_exit_seconds and time.monotonic() - idle_since >= idle_exit_seconds:
                        break
                    time.sleep(10)
                    continue
                idle_since = time.monotonic()
                if scan:
                    scan_id = scan["id"]
                    scan_started = time.perf_counter()
                    if scan['job'].get('recording_operation'):
                        from app import flow_recorder_worker
                        def recording_progress(status, detail, **extra):
                            return _api(client, 'POST', f'/api/flows/worker/{worker_id}/scans/{scan_id}/progress',
                                {'status': status, 'progress': detail, 'complete': False, **extra})
                        try:
                            if not headed:
                                raise RuntimeError('Recording requires a visible worker.')
                            if context is not None:
                                context.close()
                            context, page = None, None
                            definition = (scan['job'].get('validation_job') or {}).get('recording', {}).get('definition', {})
                            with flow_recorder_worker.reservation_heartbeat(server, worker_id, scan_id), flow_recorder_worker.browser_session(
                                    playwright, profile_dir, flow_browser.channel_for(scan['job']),
                                    timezone=definition.get('timezone', 'UTC')) as (recording_context, recording_profile):
                                recording_page = recording_context.pages[0] if recording_context.pages else recording_context.new_page()
                                if scan['job']['recording_operation'] == 'record':
                                    result = flow_recorder_worker.record(client, worker_id, scan, recording_page, recording_context, recording_profile, recording_progress)
                                else:
                                    result = flow_recorder_worker.validate(scan, recording_page, recording_profile, recording_progress)
                            recording_progress('succeeded', {'stage': 'complete', 'message': 'Recording operation completed.'}, recording_result=result)
                        except flow_recorder_worker.RecordingCancelled as exc:
                            recording_progress('cancelled', {'stage': 'cancelled', 'message': str(exc)})
                        except Exception as exc:
                            recording_progress('failed', {'stage': 'failed', 'message': str(exc)}, error=str(exc)[:10000])
                        finally:
                            try:
                                context.close()
                            except Exception:
                                pass
                            context, page = None, None
                        idle_since = time.monotonic()
                        if once:
                            break
                        continue

                    def scan_progress(status: str, detail: dict, reports: list | None = None,
                                      timings: list | None = None, error: str | None = None,
                                      complete: bool = True):
                        _api(client, "POST", f"/api/flows/worker/{worker_id}/scans/{scan_id}/progress", {
                            "status": status, "progress": _bounded_detail(detail), "reports": reports or [],
                            "timings": timings or [],
                            "error": error[:ASAP_MAX_ERROR_CHARS] if error else error,
                            "complete": complete,
                        })

                    try:
                        scan_job = scan["job"]
                        if scan_job.get("site", {}).get("adapter") == GSCM_PORTAL_ADAPTER:
                            reports, complete = _gscm_call(
                                page, headed,
                                lambda message: scan_progress(
                                    "running",
                                    {"stage": "authentication", "message": message},
                                ),
                                lambda: flow_gscm.discover_catalog(
                                    page, scan_job, scan_progress,
                                ),
                                profile_dir,
                            )
                            timings = [{
                                "phase": "report_discovery",
                                "duration_ms": round((time.perf_counter() - scan_started) * 1000),
                                "item_count": len(reports), "status": "succeeded",
                            }]
                        else:
                            reports, timings, complete = discover_asap_catalog(
                                page, scan_job, scan_progress, profile_dir,
                            )
                        scan_progress(
                            "succeeded",
                            {"stage": "complete", "message": f"Discovered {len(reports)} report(s)."},
                            reports, timings, complete=complete,
                        )
                    except Exception as exc:
                        failure_message = str(exc)
                        screenshot_path = _failure_screenshot(
                            page, profile_dir, f"scan-{scan_id}",
                        )
                        if screenshot_path:
                            failure_message += f" Screenshot: {screenshot_path}"
                        scan_progress(
                            "failed", {"stage": "failed", "message": failure_message},
                            timings=[{"phase": "total", "duration_ms": round((time.perf_counter() - scan_started) * 1000), "status": "failed"}],
                            error=failure_message, complete=False,
                        )
                    idle_since = time.monotonic()
                    if once:
                        break
                    continue
                run_id = run["id"]
                transport_state = {}
                run_started = time.perf_counter()
                artifacts = []
                timings = []
                transformation_started = None
                sql_started = None
                sql_result = None
                source_receipt = None
                no_op = False
                source_outcome = {}

                def progress(status: str, detail: dict, artifacts: list | None = None,
                             timings: list | None = None, error: str | None = None,
                             traceback_text: str | None = None,
                             source_receipt: dict | None = None):
                    _api(client, "POST", f"/api/flows/worker/{worker_id}/runs/{run_id}/progress", {
                        "status": status, "progress": _bounded_detail(detail), "artifacts": artifacts or [],
                        "timings": timings or [],
                        "retention": (detail or {}).get("retention_results") or [],
                        "error": error[:ASAP_MAX_ERROR_CHARS] if error else error,
                        "traceback": traceback_text[:100_000] if traceback_text else traceback_text,
                        "source_receipt": source_receipt,
                        "finalizer_token": transport_state.get('finalizer_token'),
                    })

                def register_folder(folder: str) -> dict:
                    return _api(
                        client, "POST",
                        f"/api/flows/worker/{worker_id}/runs/{run_id}/register_folder",
                        {"run_folder": folder},
                    )

                heartbeat_stop = threading.Event()

                def heartbeat_loop():
                    with httpx.Client(
                        base_url=server.rstrip("/"),
                        headers={"User-Agent": "Metronome-Flow-Worker/1"},
                    ) as heartbeat_client:
                        while not heartbeat_stop.wait(30):
                            try:
                                _api(
                                    heartbeat_client, "POST",
                                    f"/api/flows/worker/{worker_id}/runs/{run_id}/heartbeat",
                                )
                            except Exception:
                                pass

                heartbeat_thread = threading.Thread(target=heartbeat_loop, daemon=True)
                heartbeat_thread.start()

                from app.flow_execution_lock import job_lock
                execution_locks = job_lock(run["job"])
                execution_state = {}
                try:
                    execution_locks.acquire()
                    acquire_bundle = None
                    if flow_tasks.enabled(run['job']):
                        from functools import partial
                        from app.flow_parallel_worker import acquire_bundle as parallel_bundle
                        acquire_bundle = partial(parallel_bundle, client, worker_id, transport_state)
                    if run['job'].get('flow', {}).get('execution_method') == 'recorded' and run['job'].get('job_type') != 'sql_retry':
                        from app import flow_recorder_worker, flow_portable
                        if run['job'].get('recording', {}).get('engine_hash') != flow_portable.execution_hash():
                            raise RuntimeError('This job is pinned to a different recorded execution core. Validate a new revision.')
                        definition = run['job']['recording']['definition']
                        context.close()
                        try:
                            with flow_recorder_worker.browser_session(playwright, profile_dir, flow_browser.channel_for(run['job']),
                                    headed=headed, timezone=definition['timezone']) as (recording_context, recording_profile):
                                recording_page = recording_context.pages[0] if recording_context.pages else recording_context.new_page()
                                flow_recorder_worker.authenticate(recording_page, run['job'], recording_profile, progress, headed=headed)
                                execute_flow(recording_page, run['job'], progress, recording_profile, recording_profile / 'downloads',
                                    run_id=run_id, register_folder=register_folder, headed=headed,
                                    artifacts=artifacts, state=execution_state, run_started=run_started)
                        finally:
                            context, page = None, None
                    else:
                        execute_flow(page, run["job"], progress, profile_dir, download_staging_dir,
                                     run_id=run_id, register_folder=register_folder, headed=headed,
                                     artifacts=artifacts, state=execution_state, run_started=run_started,
                                     acquire_bundle=acquire_bundle)
                except Exception as exc:
                    artifacts = execution_state.get("artifacts", artifacts)
                    timings = execution_state.get("timings", timings)
                    transformation_started = execution_state.get("transformation_started")
                    sql_started = execution_state.get("sql_started")
                    failure_message = str(exc)
                    failure_traceback = traceback.format_exc()
                    screenshot_path = _failure_screenshot(
                        page, profile_dir, f"run-{run_id}",
                    )
                    if screenshot_path:
                        failure_message += f" Screenshot: {screenshot_path}"
                    failure_detail = {"stage": "failed", "message": failure_message}
                    try:
                        progress(
                            "failed", failure_detail,
                            artifacts=artifacts, timings=timings,
                            error=failure_message, traceback_text=failure_traceback,
                        )
                    except Exception as report_exc:
                        # A rich terminal payload must never leave the run in
                        # ``running`` if one optional diagnostic field is
                        # rejected or a response is lost. A minimal idempotent
                        # terminal update gives the server a second, smaller
                        # chance to close the run. If the first request already
                        # committed, the endpoint returns the terminal state.
                        fallback_message = (
                            f"{failure_message} Terminal diagnostic reporting initially failed: "
                            f"{type(report_exc).__name__}: {report_exc}"
                        )[:10000]
                        print(fallback_message, file=sys.stderr, flush=True)
                        _api(
                            client, "POST",
                            f"/api/flows/worker/{worker_id}/runs/{run_id}/progress",
                            {
                                "status": "failed",
                                "finalizer_token": transport_state.get('finalizer_token'),
                                "progress": {
                                    "stage": "failed",
                                    "message": fallback_message,
                                    "terminal_report_fallback": True,
                                },
                                "artifacts": [],
                                "timings": [],
                                "error": fallback_message,
                                "traceback": None,
                            },
                        )
                finally:
                    execution_locks.release()
                    heartbeat_stop.set()
                    heartbeat_thread.join(timeout=2)
                idle_since = time.monotonic()
                if once:
                    break
            if context is not None:
                context.close()


def _adapter_for_auth_url(auth_url: str) -> str:
    """Pick the portal adapter from the bootstrap URL when none was given."""
    host = re.sub(r"^https?://([^/]+).*$", r"\1", str(auth_url or "")).casefold()
    if "mdscm.sec.samsung.net" in host:
        return GSCM_PORTAL_ADAPTER
    return ASAP_PORTAL_ADAPTER


def authenticate_site(profile_dir: Path, auth_url: str, timeout_minutes: int = 10,
                      adapter: str | None = None, browser_channel: str = "chrome"):
    """Create the automation profile's SSO session in a visible browser window.

    Both portals sit behind the same Samsung SSO, but they prove they are up in
    completely different ways: ASAP renders navigation anchors, GSCM compiles a
    Nexacro component tree that contains no anchors at all.
    """
    adapter = (adapter or "").strip() or _adapter_for_auth_url(auth_url)
    label = "GSCM" if adapter == GSCM_PORTAL_ADAPTER else "ASAP"
    with _exclusive_worker_lock(profile_dir) as acquired:
        if not acquired:
            raise RuntimeError("The Flows worker is still using the automation browser profile.")
        with sync_playwright() as playwright:
            from app import flow_browser
            context = flow_browser.launch(playwright, profile_dir, browser_channel, headed=True)
            page = context.pages[0] if context.pages else context.new_page()
            print(f"Complete {label} sign-in in the browser window if prompted.", flush=True)
            if adapter == GSCM_PORTAL_ADAPTER:
                try:
                    flow_gscm.open_portal(
                        page, auth_url, timeout_ms=timeout_minutes * 60_000,
                    )
                except Exception as exc:
                    current_url = page.url
                    context.close()
                    raise RuntimeError(
                        f"GSCM authentication did not complete within {timeout_minutes} "
                        f"minutes (URL: {current_url}). {exc}"
                    ) from exc
                marker = GSCM_AUTH_MARKER
            else:
                _asap_goto(page, auth_url, profile_dir)
                roots = _wait_for_navigation_roots(page, timeout_minutes * 60_000)
                if not roots:
                    current_url = page.url
                    title = _clean_text(page.title())
                    context.close()
                    raise RuntimeError(
                        f"ASAP authentication did not complete within {timeout_minutes} minutes "
                        f"(URL: {current_url}, title: {title})."
                    )
                marker = AUTH_MARKER
            (profile_dir / marker).write_text(json.dumps({
                "authenticated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
                "host": re.sub(r"^https?://([^/]+).*$", r"\1", page.url),
            }), encoding="utf-8")
            print(f"{label} automation browser authenticated.", flush=True)
            context.close()


def authenticate_asap(profile_dir: Path, auth_url: str, timeout_minutes: int = 10):
    """Backwards-compatible entry point for the ASAP-only bootstrap."""
    authenticate_site(profile_dir, auth_url, timeout_minutes, ASAP_PORTAL_ADAPTER)


def main():
    parser = argparse.ArgumentParser(description="Metronome authenticated download worker")
    parser.add_argument("--server", default=os.environ.get("METRONOME_URL", "http://127.0.0.1:8000"))
    parser.add_argument("--worker-id", default=os.environ.get("METRONOME_FLOW_WORKER_ID", socket.gethostname().lower()))
    parser.add_argument("--name", default=os.environ.get("METRONOME_FLOW_WORKER_NAME", socket.gethostname()))
    parser.add_argument("--profile-dir", default=os.environ.get("METRONOME_FLOW_PROFILE", str(Path.home() / ".metronome-flow-browser")))
    parser.add_argument("--headed", action="store_true", help="Show the browser. Recommended for initial SSO setup.")
    parser.add_argument("--authenticate-url", help="Open a one-time visible portal SSO bootstrap and exit.")
    parser.add_argument("--authenticate-adapter", default=None,
                        help="Portal adapter for the bootstrap. Inferred from the URL when omitted.")
    parser.add_argument("--authentication-timeout-minutes", type=int, default=10)
    parser.add_argument("--once", action="store_true", help="Claim at most one run, then exit.")
    parser.add_argument("--idle-exit-seconds", type=int, default=0, help="Exit after this many idle seconds.")
    args = parser.parse_args()
    profile_dir = Path(args.profile_dir)
    if sys.stdout is None or sys.stderr is None:
        # Interactive scheduled tasks use pythonw: keep diagnostics on disk
        # without opening a console alongside the user's recording browser.
        profile_dir.mkdir(parents=True, exist_ok=True)
        log = (profile_dir / 'worker-console.log').open('a', encoding='utf-8', buffering=1)
        sys.stdout = sys.stdout or log
        sys.stderr = sys.stderr or log
    if args.authenticate_url:
        from app import flow_browser
        with httpx.Client(base_url=args.server.rstrip('/')) as settings_client:
            browser_channel = _api(settings_client, 'GET', '/api/system/flows')['browser_channel']
        authenticate_site(
            profile_dir, args.authenticate_url, args.authentication_timeout_minutes,
            args.authenticate_adapter, browser_channel,
        )
        return
    def _run():
        run_worker(
            args.server, args.worker_id, args.name, profile_dir, args.headed,
            args.once, max(0, args.idle_exit_seconds),
        )

    # During an update the outgoing worker may hold the profile lock for a few
    # more seconds; retry briefly so the handover self-heals. A holder that
    # outlives the retry window is an orphaned process: exit nonzero so the
    # service manager records a failure instead of a healthy-looking exit-0
    # restart loop that never registers a worker.
    deadline = time.monotonic() + LOCK_RETRY_SECONDS
    while True:
        with _exclusive_worker_lock(profile_dir) as acquired:
            if acquired:
                _run()
                return
        if time.monotonic() >= deadline:
            break
        print("Another Metronome flow worker is already running. Retrying...", flush=True)
        time.sleep(5)
    print(
        "Another Metronome flow worker is already running. It holds "
        f"{profile_dir / WORKER_LOCK_FILENAME}. Kill leftover flow_worker.py / "
        "Chrome/Edge processes for this profile, or re-run setup.ps1, which now "
        "clears them.",
        file=sys.stderr, flush=True,
    )
    sys.exit(11)


if __name__ == "__main__":
    main()
